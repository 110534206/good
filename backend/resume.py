from flask import Blueprint, request, jsonify, session, send_file, render_template, redirect, current_app, send_from_directory
from werkzeug.utils import secure_filename
from config import get_db
from semester import get_current_semester_id
from docxtpl import DocxTemplate, InlineImage
from docx.shared import Inches
import os
import traceback
import json
import re
from datetime import datetime, date
from urllib.parse import quote
from notification import create_notification
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import io

# --- 檔案路徑設定：專案根目錄 (good)，使 uploads/resumes 對應 Featured\good\uploads\resumes ---
BASE_UPLOAD_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))

# 定義 uploads/standard_courses
STANDARD_COURSE_UPLOAD_PATH = os.path.join('uploads', 'standard_courses')

# 完整的伺服器儲存目錄 
FULL_STANDARD_COURSE_UPLOAD_DIR = os.path.join(BASE_UPLOAD_DIR, STANDARD_COURSE_UPLOAD_PATH)

# 上傳資料夾設定
UPLOAD_FOLDER = "uploads/resumes"
os.makedirs(os.path.join(BASE_UPLOAD_DIR, UPLOAD_FOLDER), exist_ok=True)

# 缺勤佐證圖片資料夾設定
ABSENCE_PROOF_FOLDER = "uploads/absence_proofs"
os.makedirs(os.path.join(BASE_UPLOAD_DIR, ABSENCE_PROOF_FOLDER), exist_ok=True) 

# 修正：確保 role_map 存在
role_map = {
    "student": "學生",
    "teacher": "指導老師",
    "director": "主任",
    "ta": "科助",
    "admin": "管理員",
    "vendor": "廠商",
    "class_teacher": "班導師",
    "approved": "通過",
    "rejected": "退回"
}

resume_bp = Blueprint("resume_bp", __name__)

def require_login():
    return 'user_id' in session and 'role' in session

# 輔助函數：處理履歷上傳截止時間後的狀態自動更新
def update_resume_status_after_deadline(cursor, conn):
    """
    履歷上傳截止時間後，自動更新狀態：
    1. 將所有 uploaded 狀態的履歷自動改為 approved（班導審核通過）
    2. 將所有班導已通過（status='approved'）的履歷傳給指導老師審核
       使用 status 和 reviewed_by 來判斷履歷狀態
    
    返回: (is_deadline_passed: bool, updated_count: dict)
    """
    try:
        from semester import get_current_semester_deadline
        # 檢查履歷上傳截止時間：優先學期流程表 internship_flows，無則 fallback 公告
        now = datetime.now()
        resume_deadline = get_current_semester_deadline(cursor, 'resume')
        if resume_deadline is None:
            cursor.execute("""
                SELECT end_time 
                FROM announcement 
                WHERE title LIKE '[作業]%上傳履歷截止時間' AND is_published = 1
                ORDER BY created_at DESC 
                LIMIT 1
            """)
            deadline_result = cursor.fetchone()
            if deadline_result and deadline_result.get('end_time'):
                deadline = deadline_result['end_time']
                if isinstance(deadline, datetime):
                    resume_deadline = deadline
                else:
                    try:
                        resume_deadline = datetime.strptime(str(deadline), '%Y-%m-%d %H:%M:%S')
                    except Exception:
                        resume_deadline = datetime.strptime(str(deadline), '%Y-%m-%d %H:%M')
        is_resume_deadline_passed = resume_deadline is not None and now > resume_deadline
        
        # 如果已經過了截止時間，執行狀態更新
        if is_resume_deadline_passed:
            # 檢查 resume_teacher 表是否存在
            resume_teacher_table_exists = False
            try:
                cursor.execute("SHOW TABLES LIKE 'resume_teacher'")
                resume_teacher_table_exists = cursor.fetchone() is not None
            except Exception as e:
                print(f"⚠️ 檢查 resume_teacher 表時發生錯誤: {e}")
                resume_teacher_table_exists = False
            
            # 將所有未退件的履歷（uploaded 狀態）自動改為 approved（班導審核通過）
            # 不處理 rejected 狀態的履歷，保留退件狀態
            cursor.execute("""
                UPDATE resumes 
                SET status = 'approved', updated_at = NOW()
                WHERE status = 'uploaded'
            """)
            uploaded_to_approved_count = cursor.rowcount
            
            # 截止時間後，所有班導已通過（status='approved'）的投遞需要同步到 resume_teacher 表（以 application_id 為單位）
            if resume_teacher_table_exists:
                cursor.execute("""
                    SELECT sja.id AS application_id, ic.advisor_user_id
                    FROM resumes r
                    INNER JOIN student_job_applications sja ON sja.resume_id = r.id AND sja.student_id = r.user_id
                    JOIN internship_companies ic ON sja.company_id = ic.id
                    WHERE r.status = 'approved'
                      AND ic.advisor_user_id IS NOT NULL
                """)
                apps_to_sync = cursor.fetchall()
                print(f"🔍 [DEBUG] 找到 {len(apps_to_sync)} 筆投遞需要同步到 resume_teacher 表")
                synced_count = 0
                for app_info in apps_to_sync:
                    application_id = app_info['application_id']
                    advisor_user_id = app_info['advisor_user_id']
                    if not application_id or not advisor_user_id:
                        continue
                    # 確保清除任何未讀取的結果（防止 "Unread result found" 錯誤）
                    try:
                        cursor.fetchall()
                    except:
                        pass
                    
                    cursor.execute("""
                        SELECT id, review_status FROM resume_teacher 
                        WHERE application_id = %s AND teacher_id = %s
                    """, (application_id, advisor_user_id))
                    existing = cursor.fetchone()
                    if existing:
                        # 只更新狀態為 'uploaded' 的記錄，不要覆蓋已經審核過的記錄（'approved' 或 'rejected'）
                        current_status = existing.get('review_status')
                        if current_status == 'uploaded' or current_status is None:
                            # 確保清除任何未讀取的結果
                            try:
                                cursor.fetchall()
                            except:
                                pass
                            
                            cursor.execute("""
                                UPDATE resume_teacher SET review_status='uploaded', reviewed_at=NULL
                                WHERE application_id = %s AND teacher_id = %s
                            """, (application_id, advisor_user_id))
                            synced_count += 1
                        # 如果已經是 'approved' 或 'rejected'，跳過（不重置）
                    else:
                        # 確保清除任何未讀取的結果
                        try:
                            cursor.fetchall()
                        except:
                            pass
                        
                        cursor.execute("""
                            INSERT INTO resume_teacher (application_id, teacher_id, review_status, comment, reviewed_at, created_at)
                            VALUES (%s, %s, 'uploaded', NULL, NULL, NOW())
                        """, (application_id, advisor_user_id))
                        synced_count += 1
                if synced_count > 0:
                    conn.commit()
                    print(f"✅ 已同步 {synced_count} 筆履歷到 resume_teacher 表，等待指導老師審核")
                else:
                    print(f"⚠️ [DEBUG] 未找到需要同步的履歷")
            
            if uploaded_to_approved_count > 0:
                if synced_count == 0:
                    # 如果沒有同步，也需要提交狀態更新
                    conn.commit()
                print(f"✅ 履歷提交截止時間已過，已將 {uploaded_to_approved_count} 筆未退件的履歷狀態改為 'approved'（班導審核通過），等待指導老師審核")
            
            return is_resume_deadline_passed, {
                'uploaded_to_approved': uploaded_to_approved_count,
                'teacher_review_status_updated': 0
            }
        
        return False, {'uploaded_to_approved': 0, 'teacher_review_status_updated': 0}
    except Exception as e:
        print(f"❌ 更新履歷狀態錯誤: {e}")
        traceback.print_exc()
        return False, {'uploaded_to_approved': 0, 'teacher_review_status_updated': 0}


# 輔助函數：當指導老師通過時，確保 resume_applications 有一筆記錄（廠商才能看到）
def ensure_resume_application_for_teacher_approved(cursor, conn, application_id, job_id):
    """若尚無記錄則新增一筆 resume_applications，使廠商可見該履歷。"""
    max_retries = 3
    retry_delay = 0.1  # 100ms
    
    for attempt in range(max_retries):
        try:
            # 確保清除任何未讀取的結果（防止 "Unread result found" 錯誤）
            try:
                cursor.fetchall()
            except:
                pass
            
            # 使用 INSERT IGNORE 來避免死鎖：如果記錄已存在則忽略，不存在則插入
            # 這比先 SELECT 再 INSERT 更安全，因為是原子操作
            cursor.execute("""
                INSERT IGNORE INTO resume_applications
                (application_id, job_id, apply_status, interview_status, interview_result, company_comment, created_at)
                VALUES (%s, %s, 'uploaded', 'none', 'pending', '', NOW())
            """, (application_id, job_id))
            
            # 檢查是否成功插入（rowcount > 0 表示插入了新記錄）
            if cursor.rowcount > 0:
                print(f"✅ [resume_applications] 指導老師通過後立即建立廠商審核記錄: application_id={application_id}, job_id={job_id}")
            else:
                # 記錄已存在，這是正常的（可能是並發請求或其他原因）
                print(f"ℹ️ [resume_applications] 記錄已存在（可能由並發請求創建）: application_id={application_id}, job_id={job_id}")
            
            return True
            
        except Exception as e:
            error_str = str(e)
            # 檢查是否為死鎖錯誤
            if ("Deadlock" in error_str or "1213" in error_str or "40001" in error_str) and attempt < max_retries - 1:
                # 等待一小段時間後重試
                import time
                time.sleep(retry_delay * (attempt + 1))  # 遞增延遲
                print(f"⚠️ [resume_applications] 檢測到死鎖，重試中 (嘗試 {attempt + 2}/{max_retries}): application_id={application_id}, job_id={job_id}")
                continue
            else:
                # 其他類型的錯誤
                print(f"⚠️ 建立 resume_applications 失敗 (application_id={application_id}, job_id={job_id}): {e}")
                traceback.print_exc()
                return False
    
    return False


# 輔助函數：處理指導老師審核截止時間後的狀態自動更新
def update_resume_applications_after_advisor_deadline(cursor, conn):
    """
    指導老師審核截止時間後，自動將所有指導老師已通過的履歷傳給廠商：
    1. 檢查指導老師審核截止時間是否已過
    2. 將所有 resume_teacher.review_status = 'approved' 的履歷創建 resume_applications 記錄
    3. 通知對應的廠商
    
    返回: (is_deadline_passed: bool, created_count: int)
    """
    try:
        from semester import get_current_semester_deadline
        from notification import create_notification
        
        # 檢查指導老師審核截止時間：優先學期流程表 internship_flows，無則 fallback 公告
        now = datetime.now()
        advisor_deadline = get_current_semester_deadline(cursor, 'advisor')
        if advisor_deadline is None:
            cursor.execute("""
                SELECT end_time 
                FROM announcement 
                WHERE title LIKE '[作業]%指導老師審核履歷截止時間' AND is_published = 1
                ORDER BY created_at DESC 
                LIMIT 1
            """)
            deadline_result = cursor.fetchone()
            if deadline_result and deadline_result.get('end_time'):
                deadline = deadline_result['end_time']
                if isinstance(deadline, datetime):
                    advisor_deadline = deadline
                else:
                    try:
                        advisor_deadline = datetime.strptime(str(deadline), '%Y-%m-%d %H:%M:%S')
                    except Exception:
                        advisor_deadline = datetime.strptime(str(deadline), '%Y-%m-%d %H:%M')
        
        is_advisor_deadline_passed = advisor_deadline is not None and now > advisor_deadline
        
        # 如果已經過了截止時間，執行狀態更新
        if is_advisor_deadline_passed:
            # 檢查 resume_teacher 表是否存在
            resume_teacher_table_exists = False
            try:
                cursor.execute("SHOW TABLES LIKE 'resume_teacher'")
                resume_teacher_table_exists = cursor.fetchone() is not None
            except Exception as e:
                print(f"⚠️ 檢查 resume_teacher 表時發生錯誤: {e}")
                resume_teacher_table_exists = False
            
            if not resume_teacher_table_exists:
                print("⚠️ resume_teacher 表不存在，跳過自動傳給廠商的邏輯")
                return is_advisor_deadline_passed, 0
            
            # 查找所有指導老師已通過但尚未傳給廠商的履歷
            cursor.execute("""
                SELECT 
                    rt.application_id,
                    rt.teacher_id,
                    sja.job_id,
                    sja.company_id,
                    sja.student_id,
                    sja.resume_id,
                    u.name AS student_name,
                    ic.company_name,
                    ic.advisor_user_id
                FROM resume_teacher rt
                INNER JOIN student_job_applications sja ON sja.id = rt.application_id
                INNER JOIN users u ON sja.student_id = u.id
                INNER JOIN internship_companies ic ON sja.company_id = ic.id
                WHERE rt.review_status = 'approved'
                  AND rt.teacher_id IS NOT NULL
                  AND sja.job_id IS NOT NULL
                  AND sja.company_id IS NOT NULL
                  AND NOT EXISTS (
                      SELECT 1 FROM resume_applications ra
                      WHERE ra.application_id = rt.application_id
                        AND ra.job_id = sja.job_id
                  )
            """)
            approved_applications = cursor.fetchall()
            
            if not approved_applications:
                print("✅ 指導老師審核截止時間已過，但沒有需要傳給廠商的履歷")
                return is_advisor_deadline_passed, 0
            
            print(f"🔍 [DEBUG] 找到 {len(approved_applications)} 筆指導老師已通過但尚未傳給廠商的履歷")
            
            created_count = 0
            notified_vendors = set()
            updated_companies = set()
            
            for app in approved_applications:
                application_id = app['application_id']
                job_id = app['job_id']
                company_id = app['company_id']
                student_name = app['student_name']
                company_name = app['company_name']
                advisor_user_id = app['advisor_user_id']
                
                try:
                    # 確保清除任何未讀取的結果（防止 "Unread result found" 錯誤）
                    try:
                        cursor.fetchall()
                    except:
                        pass
                    
                    # 創建 resume_applications 記錄（含 company_comment 以符合 NOT NULL）
                    cursor.execute("""
                        INSERT INTO resume_applications
                        (application_id, job_id, apply_status, interview_status, interview_result, company_comment, created_at)
                        VALUES (%s, %s, 'uploaded', 'none', 'pending', '', NOW())
                    """, (application_id, job_id))
                    created_count += 1
                    updated_companies.add(company_id)
                    print(f"✅ [resume_applications] 指導老師審核截止時間後自動創建廠商審核記錄: application_id={application_id}, job_id={job_id}, apply_status='uploaded'")
                    
                    # 通知對應的廠商
                    if advisor_user_id:
                        # 確保清除任何未讀取的結果
                        try:
                            cursor.fetchall()
                        except:
                            pass
                        
                        cursor.execute("""
                            SELECT id, name FROM users
                            WHERE role = 'vendor' AND teacher_id = %s
                        """, (advisor_user_id,))
                        vendors = cursor.fetchall()
                        
                        for vendor in vendors:
                            vendor_id = vendor['id']
                            if vendor_id not in notified_vendors:
                                create_notification(
                                    user_id=vendor_id,
                                    title="新履歷待審核",
                                    message=(
                                        f"學生 {student_name} 的履歷已由指導老師審核通過，"
                                        f"已投遞至「{company_name}」，請前往審核。"
                                    ),
                                    category="resume",
                                    link_url="/vendor/resumes"
                                )
                                notified_vendors.add(vendor_id)
                except Exception as e:
                    print(f"⚠️ 創建 resume_applications 記錄失敗 (application_id={application_id}, job_id={job_id}): {e}")
                    continue
            
            if created_count > 0:
                conn.commit()
                print(f"✅ 指導老師審核截止時間已過，已將 {created_count} 筆履歷傳給廠商，並通知 {len(notified_vendors)} 位廠商")
            else:
                print(f"⚠️ 指導老師審核截止時間已過，但沒有成功創建任何 resume_applications 記錄")
            
            return is_advisor_deadline_passed, created_count
        
        return False, 0
    except Exception as e:
        print(f"❌ 更新履歷傳給廠商狀態錯誤: {e}")
        traceback.print_exc()
        return False, 0


# -------------------------
# API - 取得指導老師待審核履歷列表
# -------------------------
@resume_bp.route('/api/teacher_review_resumes', methods=['GET'])
def get_teacher_review_resumes():
    # 確保有權限 (teacher, director, class_teacher, admin) 才能進入
    if 'user_id' not in session or session.get('role') not in ['teacher', 'director', 'class_teacher', 'admin']:
        return jsonify({"success": False, "message": "無權限"}), 403

    session_user_id = session['user_id']
    session_role = session['role']
    
    print(f"🔍 [DEBUG] get_teacher_review_resumes: session_role={session_role}, user_id={session_user_id}")
    
    conn = get_db() 
    cursor = conn.cursor(dictionary=True) 
    
    try:
        # 檢查履歷上傳截止時間並自動更新狀態
        is_resume_deadline_passed, update_counts = update_resume_status_after_deadline(cursor, conn)
        
        # 確保清除任何未讀取的結果（防止 "Unread result found" 錯誤）
        try:
            cursor.fetchall()
        except:
            pass
        
        # 檢查指導老師審核截止時間並自動將已通過的履歷傳給廠商
        is_advisor_deadline_passed, advisor_update_count = update_resume_applications_after_advisor_deadline(cursor, conn)
        
        # 確保清除任何未讀取的結果（防止 "Unread result found" 錯誤）
        try:
            cursor.fetchall()
        except:
            pass
        
        # 在截止時間之前，指導老師（包括主任切換身份）不能看到任何履歷，直接返回空結果
        if session_role == 'teacher' and not is_resume_deadline_passed:
            # 獲取履歷上傳截止時間資訊（優先從 internship_flows 表讀取）
            deadline_info = None
            try:
                # 確保清除任何未讀取的結果（防止 "Unread result found" 錯誤）
                try:
                    cursor.fetchall()
                except:
                    pass
                
                from semester import get_current_semester_deadline
                resume_deadline_dt = get_current_semester_deadline(cursor, 'resume')
                
                # 確保清除任何未讀取的結果（get_current_semester_deadline 可能留下未讀取的結果）
                try:
                    cursor.fetchall()
                except:
                    pass
                
                if resume_deadline_dt:
                    deadline_info = resume_deadline_dt.strftime('%Y/%m/%d %H:%M')
                else:
                    # 如果 internship_flows 中沒有，則從 announcement 表查找
                    cursor.execute("""
                        SELECT end_time 
                        FROM announcement 
                        WHERE title LIKE '[作業]%上傳履歷截止時間' AND is_published = 1
                        ORDER BY created_at DESC 
                        LIMIT 1
                    """)
                    deadline_result = cursor.fetchone()
                    if deadline_result and deadline_result.get('end_time'):
                        deadline = deadline_result['end_time']
                        if isinstance(deadline, datetime):
                            deadline_info = deadline.strftime('%Y/%m/%d %H:%M')
                        else:
                            try:
                                deadline_dt = datetime.strptime(str(deadline), '%Y-%m-%d %H:%M:%S')
                                deadline_info = deadline_dt.strftime('%Y/%m/%d %H:%M')
                            except:
                                try:
                                    deadline_dt = datetime.strptime(str(deadline), '%Y-%m-%d %H:%M')
                                    deadline_info = deadline_dt.strftime('%Y/%m/%d %H:%M')
                                except:
                                    deadline_info = str(deadline)
            except Exception as e:
                print(f"⚠️ 無法獲取截止時間資訊: {e}")
            
            print(f"🔒 [DEBUG] 指導老師 (session_role={session_role}, user_id={session_user_id}) 在截止時間前被攔截，返回空數據")
            
            return jsonify({
                "success": True, 
                "data": [],
                "deadline": deadline_info,
                "is_deadline_passed": False,
                "message": "履歷提交截止時間尚未到達，目前無法查看學生履歷。"
            })
        
        # 檢查 resume_teacher 表是否存在
        resume_teacher_table_exists = False
        try:
            cursor.execute("SHOW TABLES LIKE 'resume_teacher'")
            resume_teacher_table_exists = cursor.fetchone() is not None
            if resume_teacher_table_exists:
                # 檢查表是否有必要的欄位
                required_columns = ['application_id', 'teacher_id', 'review_status']
                cursor.execute("SELECT DATABASE() as db_name")
                db_result = cursor.fetchone()
                db_name = db_result['db_name'] if db_result else None
                
                cursor.execute("""
                    SELECT COLUMN_NAME 
                    FROM INFORMATION_SCHEMA.COLUMNS 
                    WHERE TABLE_SCHEMA = %s
                    AND TABLE_NAME = 'resume_teacher' 
                    AND COLUMN_NAME IN ('application_id', 'teacher_id', 'review_status')
                """, (db_name,))
                existing_columns = [row['COLUMN_NAME'] for row in cursor.fetchall()]
                missing_columns = [col for col in required_columns if col not in existing_columns]
                
                if missing_columns:
                    print(f"⚠️ resume_teacher 表缺少必要欄位: {', '.join(missing_columns)}")
                    resume_teacher_table_exists = False
        except Exception as e:
            print(f"⚠️ 檢查 resume_teacher 表時發生錯誤: {e}")
            resume_teacher_table_exists = False
        
        # 指導老師需要 JOIN resume_teacher 表獲取 review_status
        if session_role == 'teacher' and resume_teacher_table_exists:
            sql = """
                SELECT 
                    u.id AS user_id,
                    u.username AS student_id,
                    u.name,
                    c.name AS class_name,
                    c.department,
                    r.id AS resume_id,
                    r.created_at AS upload_time,
                    r.original_filename,
                    r.status AS display_status,
                    rt.review_status AS review_status,
                    rt.comment AS comment,
                    sja.id AS application_id,
                    sja.applied_at,
                    ic.company_name,
                    COALESCE(ij.title, '') AS job_title
                FROM users u
                LEFT JOIN classes c ON u.class_id = c.id
                INNER JOIN student_job_applications sja ON sja.student_id = u.id
                LEFT JOIN resumes r ON r.id = sja.resume_id AND r.user_id = u.id
                LEFT JOIN resume_teacher rt ON rt.application_id = sja.id AND rt.teacher_id = %s
                JOIN internship_companies ic ON sja.company_id = ic.id
                LEFT JOIN internship_jobs ij ON sja.job_id = ij.id
                WHERE u.role = 'student' 
                  AND sja.resume_id IS NOT NULL
                  AND r.id IS NOT NULL
                  AND r.status = 'approved'
                  AND ic.advisor_user_id IS NOT NULL AND ic.advisor_user_id = %s
                ORDER BY c.name, u.username, sja.applied_at DESC, r.created_at DESC
            """
            params = [session_user_id, session_user_id]
        else:
            # 班導和其他角色：不使用 resume_teacher 表；依投遞時間與履歷上傳時間排序，讓「最新投遞／最新上傳」的履歷紀錄排在最前
            sql = """
                SELECT 
                    u.id AS user_id,
                    u.username AS student_id,
                    u.name,
                    c.name AS class_name,
                    c.department,
                    r.id AS resume_id,
                    r.created_at AS upload_time,
                    r.original_filename,
                    r.status AS display_status,
                    NULL AS review_status,
                    sja.id AS application_id,
                    sja.applied_at,
                    ic.company_name,
                    COALESCE(ij.title, '') AS job_title
                FROM users u
                LEFT JOIN classes c ON u.class_id = c.id
                INNER JOIN student_job_applications sja ON sja.student_id = u.id
                LEFT JOIN resumes r ON r.id = sja.resume_id AND r.user_id = u.id
                JOIN internship_companies ic ON sja.company_id = ic.id
                LEFT JOIN internship_jobs ij ON sja.job_id = ij.id
                WHERE u.role = 'student' 
                  AND sja.resume_id IS NOT NULL
                  AND r.id IS NOT NULL
            """
            params = []
            
            # 根據角色過濾資料
            if session_role == 'class_teacher':
                sql += """
                    AND u.class_id IN (
                        SELECT class_id FROM classes_teacher WHERE teacher_id = %s
                    )
                """
                params.append(session_user_id)
            # 依班級、學號、投遞時間與履歷上傳時間排序，讓「最新投遞／最新上傳」的履歷紀錄排在最前（主任、班導查看）
            sql += """
                ORDER BY c.name, u.username, sja.applied_at DESC, r.created_at DESC
            """
        
        try:
            # 確保查詢時能看到最新提交的變更（刷新連接）
            # 這對於確保能看到剛提交的 resume_teacher 記錄很重要
            try:
                conn.commit()  # 提交任何未提交的變更
            except:
                pass
            
            cursor.execute(sql, tuple(params))
            rows = cursor.fetchall()
            # 調試：記錄查詢結果
            if session_role == 'teacher' and resume_teacher_table_exists:
                print(f"🔍 [DEBUG] 查詢結果: 共 {len(rows)} 筆履歷, session_user_id={session_user_id}")
                for row in rows[:5]:  # 只記錄前5筆
                    app_id = row.get('application_id')
                    review_status = row.get('review_status')
                    resume_id = row.get('resume_id')
                    print(f"  - application_id={app_id}, resume_id={resume_id}, review_status={review_status}")
                
                # 調試：驗證特定 application_id 在 resume_teacher 表中的實際狀態
                # 確保清除任何未讀取的結果（防止 "Unread result found" 錯誤）
                try:
                    cursor.fetchall()
                except:
                    pass
                
                test_app_ids = [28, 45, 37]  # 從日誌中看到的 application_id
                for test_app_id in test_app_ids:
                    try:
                        # 確保清除任何未讀取的結果
                        try:
                            cursor.fetchall()
                        except:
                            pass
                        
                        cursor.execute("""
                            SELECT application_id, teacher_id, review_status, reviewed_at 
                            FROM resume_teacher 
                            WHERE application_id = %s AND teacher_id = %s
                        """, (test_app_id, session_user_id))
                        rt_record = cursor.fetchone()
                        if rt_record:
                            print(f"  🔍 [DEBUG] resume_teacher 表實際狀態: application_id={test_app_id}, teacher_id={rt_record.get('teacher_id')}, review_status={rt_record.get('review_status')}, reviewed_at={rt_record.get('reviewed_at')}")
                        else:
                            print(f"  ⚠️ [DEBUG] resume_teacher 表中找不到: application_id={test_app_id}, teacher_id={session_user_id}")
                        # 確保清除任何未讀取的結果
                        try:
                            cursor.fetchall()
                        except:
                            pass
                    except Exception as e:
                        print(f"  ❌ [DEBUG] 查詢 resume_teacher 表失敗: {e}")
                        # 確保清除任何未讀取的結果
                        try:
                            cursor.fetchall()
                        except:
                            pass
                
                # 調試代碼結束後，確保清除所有未讀取的結果
                try:
                    cursor.fetchall()
                except:
                    pass
        except Exception as sql_error:
            if session_role == 'teacher' and 'resume_teacher' in str(sql_error).lower():
                print(f"⚠️ SQL 執行失敗，嘗試使用不包含 resume_teacher 的查詢: {sql_error}")
                sql = """
                    SELECT 
                        u.id AS user_id,
                        u.username AS student_id,
                        u.name,
                        c.name AS class_name,
                        c.department,
                        r.id AS resume_id,
                        r.created_at AS upload_time,
                        r.original_filename,
                        r.status AS display_status,
                        NULL AS review_status,
                        sja.id AS application_id,
                        sja.applied_at,
                        ic.company_name,
                        COALESCE(ij.title, '') AS job_title
                    FROM users u
                    LEFT JOIN classes c ON u.class_id = c.id
                    INNER JOIN student_job_applications sja ON sja.student_id = u.id
                    LEFT JOIN resumes r ON r.id = sja.resume_id AND r.user_id = u.id
                    JOIN internship_companies ic ON sja.company_id = ic.id
                    LEFT JOIN internship_jobs ij ON sja.job_id = ij.id
                    WHERE u.role = 'student' 
                      AND sja.resume_id IS NOT NULL
                      AND r.id IS NOT NULL
                      AND r.status = 'approved'
                      AND ic.advisor_user_id IS NOT NULL AND ic.advisor_user_id = %s
                    ORDER BY c.name, u.username, sja.applied_at DESC, r.created_at DESC
                """
                params = [session_user_id]
                cursor.execute(sql, tuple(params))
                rows = cursor.fetchall()
            else:
                raise
        
        # 整理結果
        result_data = []
        processed_combinations = set()
        
        for row in rows:
            student_id = row['student_id']
            application_id = row.get('application_id')
            
            combo_key = (row['resume_id'], application_id) if row['resume_id'] and application_id else (student_id, application_id)
            
            if not row['resume_id']:
                if combo_key not in processed_combinations:
                    processed_combinations.add(combo_key)
                    result_data.append({
                        'user_id': row['user_id'],
                        'username': student_id,
                        'name': row['name'],
                        'className': row['class_name'] or '—',
                        'upload_time': 'N/A',
                        'original_filename': 'N/A',
                        'company_name': row.get('company_name') or '—',
                        'job_title': row.get('job_title') or '—',
                        'application_id': application_id,
                        'display_company': row.get('company_name') or '—',
                        'display_job': row.get('job_title') or '—',
                        'display_status': 'not_uploaded'
                    })
                continue

            resume_app_key = (row['resume_id'], application_id) if application_id else (row['resume_id'], None)
            
            if resume_app_key not in processed_combinations:
                processed_combinations.add(resume_app_key)
                status = row.get('display_status') or 'uploaded'
                if status not in ['uploaded', 'approved', 'rejected']:
                    status = 'uploaded'
                
                # 根據角色決定顯示的狀態
                if session_role == 'teacher' and resume_teacher_table_exists:
                    teacher_status = row.get('review_status')
                    # 調試：記錄審核狀態
                    if application_id:
                        print(f"🔍 [DEBUG] 查詢結果: application_id={application_id}, resume_id={row['resume_id']}, review_status={teacher_status}, session_user_id={session_user_id}")
                    teacher_status = teacher_status or 'uploaded'
                    if teacher_status in ['uploaded', 'approved', 'rejected']:
                        status = teacher_status
                    else:
                        status = 'uploaded'
                    display_status_for_teacher = status
                else:
                    display_status_for_teacher = row.get('display_status') or 'uploaded'
                
                # 上傳時間：以該筆投遞紀錄的投遞時間 (sja.applied_at) 顯示，與 student_job_applications 一致，主任/班導/指導老師看到的即為「此筆投遞」的時間
                applied_at = row.get('applied_at')
                if applied_at is not None and isinstance(applied_at, datetime):
                    display_upload_time = applied_at.strftime('%Y/%m/%d %H:%M')
                elif applied_at is not None:
                    display_upload_time = str(applied_at)[:16] if len(str(applied_at)) >= 16 else str(applied_at)
                else:
                    display_upload_time = row['upload_time'].strftime('%Y/%m/%d %H:%M') if isinstance(row.get('upload_time'), datetime) else (row['upload_time'] if row.get('upload_time') else 'N/A')
                result_data.append({
                    'id': row['resume_id'],
                    'username': student_id,
                    'name': row['name'],
                    'className': row['class_name'] or '—',
                    'upload_time': display_upload_time,
                    'original_filename': row['original_filename'] or 'N/A',
                    'company_name': row.get('company_name') or '—',
                    'job_title': row.get('job_title') or '—',
                    'application_id': application_id,
                    'display_company': row.get('company_name') or '—',
                    'display_job': row.get('job_title') or '—',
                    'display_status': display_status_for_teacher,
                    'status': row.get('display_status'),
                    'review_status': row.get('review_status') if session_role == 'teacher' else None,
                    'comment': row.get('comment') if session_role == 'teacher' else (row.get('comment') or ''),
                })
        
        # 獲取履歷上傳截止時間資訊（優先從 internship_flows 表讀取）
        deadline_info = None
        teacher_review_deadline_info = None
        try:
            from semester import get_current_semester_deadline
            # 優先從 internship_flows 表讀取履歷提交截止時間
            resume_deadline_dt = get_current_semester_deadline(cursor, 'resume')
            if resume_deadline_dt:
                deadline_info = resume_deadline_dt.strftime('%Y/%m/%d %H:%M')
            else:
                # 如果 internship_flows 中沒有，則從 announcement 表查找
                cursor.execute("""
                    SELECT end_time 
                    FROM announcement 
                    WHERE title LIKE '[作業]%上傳履歷截止時間' AND is_published = 1
                    ORDER BY created_at DESC 
                    LIMIT 1
                """)
                deadline_result = cursor.fetchone()
                if deadline_result and deadline_result.get('end_time'):
                    deadline = deadline_result['end_time']
                    if isinstance(deadline, datetime):
                        deadline_info = deadline.strftime('%Y/%m/%d %H:%M')
                    else:
                        try:
                            deadline_dt = datetime.strptime(str(deadline), '%Y-%m-%d %H:%M:%S')
                            deadline_info = deadline_dt.strftime('%Y/%m/%d %H:%M')
                        except:
                            try:
                                deadline_dt = datetime.strptime(str(deadline), '%Y-%m-%d %H:%M')
                                deadline_info = deadline_dt.strftime('%Y/%m/%d %H:%M')
                            except:
                                deadline_info = str(deadline)
            
            # 獲取指導老師審核截止時間（優先從 internship_flows 表讀取）
            advisor_deadline_dt = get_current_semester_deadline(cursor, 'advisor')
            if advisor_deadline_dt:
                teacher_review_deadline_info = advisor_deadline_dt.strftime('%Y/%m/%d %H:%M')
            else:
                # 如果 internship_flows 中沒有，則從 announcement 表查找
                cursor.execute("""
                    SELECT end_time 
                    FROM announcement 
                    WHERE title LIKE '[作業]%指導老師審核履歷截止時間' AND is_published = 1
                    ORDER BY created_at DESC 
                    LIMIT 1
                """)
                advisor_deadline_result = cursor.fetchone()
                if advisor_deadline_result and advisor_deadline_result.get('end_time'):
                    advisor_deadline = advisor_deadline_result['end_time']
                    if isinstance(advisor_deadline, datetime):
                        teacher_review_deadline_info = advisor_deadline.strftime('%Y/%m/%d %H:%M')
                    else:
                        try:
                            advisor_deadline_dt = datetime.strptime(str(advisor_deadline), '%Y-%m-%d %H:%M:%S')
                            teacher_review_deadline_info = advisor_deadline_dt.strftime('%Y/%m/%d %H:%M')
                        except:
                            try:
                                advisor_deadline_dt = datetime.strptime(str(advisor_deadline), '%Y-%m-%d %H:%M')
                                teacher_review_deadline_info = advisor_deadline_dt.strftime('%Y/%m/%d %H:%M')
                            except:
                                teacher_review_deadline_info = str(advisor_deadline)
        except Exception as e:
            print(f"⚠️ 無法獲取截止時間資訊: {e}")
        
        return jsonify({
            "success": True, 
            "data": result_data,
            "deadline": deadline_info,
            "teacher_review_deadline": teacher_review_deadline_info,
            "is_deadline_passed": is_resume_deadline_passed
        })

    except Exception as e:
        traceback.print_exc()
        print("❌ 取得待審核履歷列表錯誤:", e)
        # 確保清除所有未讀取的結果（防止 "Unread result found" 錯誤）
        try:
            cursor.fetchall()
        except:
            pass
        return jsonify({"success": False, "message": f"伺服器錯誤: {str(e)}"}), 500
    finally:
        # 確保清除所有未讀取的結果（防止 "Unread result found" 錯誤）
        try:
            cursor.fetchall()
        except:
            pass
        try:
            cursor.close()
        except:
            pass
        try:
            conn.close()
        except:
            pass


# -------------------------
# API - 審核履歷 (退件/通過)
# -------------------------
@resume_bp.route('/api/review_resume/<int:resume_id>', methods=['POST'])
def review_resume(resume_id):
    user_id = session.get('user_id')
    user_role = session.get('role')

    # 1. 權限檢查
    ALLOWED_ROLES = ['teacher', 'admin', 'class_teacher']
    if not user_id or user_role not in ALLOWED_ROLES:
        return jsonify({"success": False, "message": "未授權或無權限"}), 403

    data = request.get_json()
    if not data:
        return jsonify({"success": False, "message": "請求資料格式錯誤"}), 400
    
    status = data.get('status')
    comment = data.get('comment', '')
    application_id = data.get('application_id')

    if status not in ['approved', 'rejected']:
        return jsonify({"success": False, "message": "無效的狀態碼"}), 400

    # 班導僅能下載、查看，不能進行審核（通過/退件），與主任設定一致
    if user_role == 'class_teacher':
        return jsonify({"success": False, "message": "班導僅能下載與查看履歷，無法進行審核（通過/退件）。"}), 403

    conn = get_db()
    cursor = conn.cursor(dictionary=True)

    try:
        # 檢查 resume_teacher 表是否存在
        resume_teacher_table_exists = False
        try:
            cursor.execute("SHOW TABLES LIKE 'resume_teacher'")
            resume_teacher_table_exists = cursor.fetchone() is not None
        except Exception as e:
            print(f"⚠️ 檢查 resume_teacher 表時發生錯誤: {e}")
            resume_teacher_table_exists = False
        
        # 2. 查詢履歷並取得學生Email和姓名（指導老師與班導皆需 application_id，以「每筆投遞」為單位審核）
        application_id_int = None
        if resume_teacher_table_exists and (user_role == 'teacher' or user_role == 'class_teacher'):
            if not application_id or application_id == 'null' or str(application_id).strip() == '':
                msg = "指導老師審核請傳 application_id（投遞紀錄 id）" if user_role == 'teacher' else "班導審核請傳 application_id（投遞紀錄 id），以分開紀錄每間公司的狀態"
                return jsonify({"success": False, "message": msg}), 400
            try:
                application_id_int = int(application_id)
            except (TypeError, ValueError):
                print(f"⚠️ [DEBUG] 無效的 application_id: {application_id} (type: {type(application_id)})")
                return jsonify({"success": False, "message": f"無效的 application_id: {application_id}"}), 400
            cursor.execute("""
                SELECT r.id AS resume_id, r.user_id, r.original_filename, r.status AS old_status,
                    rt.review_status AS old_teacher_review_status, r.comment,
                    u.email AS student_email, u.name AS student_name,
                    sja.company_id, sja.job_id
                FROM student_job_applications sja
                JOIN resumes r ON r.id = sja.resume_id AND r.user_id = sja.student_id
                JOIN users u ON u.id = sja.student_id
                LEFT JOIN resume_teacher rt ON rt.application_id = sja.id AND rt.teacher_id = %s
                WHERE sja.id = %s
            """, (user_id, application_id_int))
            resume_data = cursor.fetchone()
            if resume_data:
                resume_id = resume_data['resume_id']
        else:
            cursor.execute("""
                SELECT 
                    r.user_id, r.original_filename, r.status AS old_status,
                    r.reviewed_by AS old_reviewed_by,
                    r.comment,
                    u.email AS student_email, u.name AS student_name
                FROM resumes r
                JOIN users u ON r.user_id = u.id
                WHERE r.id=%s
            """, (resume_id,))
            resume_data = cursor.fetchone()
        
        if not resume_data:
            return jsonify({"success": False, "message": "找不到履歷或投遞紀錄"}), 404

        student_user_id = resume_data['user_id']
        student_email = resume_data['student_email'] 
        student_name = resume_data['student_name']  
        old_status = resume_data['old_status']

        # 3. 更新履歷狀態（同一份履歷投遞不同公司分開紀錄：以 application_id 為單位，不共用 resumes.status）
        if user_role == 'teacher' and resume_teacher_table_exists and application_id_int is not None:
            old_status_for_check = resume_data.get('old_teacher_review_status') or 'uploaded'
            # 確保清除任何未讀取的結果（防止 "Unread result found" 錯誤）
            try:
                cursor.fetchall()
            except:
                pass
            
            # 指導老師：使用 INSERT ... ON DUPLICATE KEY UPDATE 來原子性地更新 resume_teacher 表
            # 這避免了先 UPDATE 再 INSERT 可能導致的死鎖問題
            max_retries = 3
            retry_delay = 0.1  # 100ms
            
            for attempt in range(max_retries):
                try:
                    cursor.execute("""
                        INSERT INTO resume_teacher (application_id, teacher_id, review_status, comment, reviewed_at, created_at)
                        VALUES (%s, %s, %s, %s, NOW(), NOW())
                        ON DUPLICATE KEY UPDATE
                            review_status = VALUES(review_status),
                            comment = VALUES(comment),
                            reviewed_at = NOW()
                    """, (application_id_int, user_id, status, comment))
                    affected_rows = cursor.rowcount if hasattr(cursor, 'rowcount') else 'N/A'
                    print(f"✅ [DEBUG] 指導老師審核完成: application_id={application_id_int}, review_status={status}, teacher_id={user_id}, affected_rows={affected_rows}")
                    # 立即提交事務，確保查詢時能看到最新變更
                    conn.commit()
                    # 驗證更新是否成功
                    try:
                        cursor.execute("""
                            SELECT review_status, reviewed_at FROM resume_teacher 
                            WHERE application_id = %s AND teacher_id = %s
                        """, (application_id_int, user_id))
                        verify_record = cursor.fetchone()
                        if verify_record:
                            print(f"  ✅ [DEBUG] 驗證更新成功: application_id={application_id_int}, review_status={verify_record.get('review_status')}, reviewed_at={verify_record.get('reviewed_at')}")
                        else:
                            print(f"  ⚠️ [DEBUG] 驗證失敗: application_id={application_id_int} 在 resume_teacher 表中找不到記錄")
                    except Exception as verify_error:
                        print(f"  ❌ [DEBUG] 驗證查詢失敗: {verify_error}")
                    break
                except Exception as e:
                    error_str = str(e)
                    # 檢查是否為死鎖錯誤
                    if ("Deadlock" in error_str or "1213" in error_str or "40001" in error_str) and attempt < max_retries - 1:
                        import time
                        time.sleep(retry_delay * (attempt + 1))
                        print(f"⚠️ [DEBUG] 檢測到死鎖，重試中 (嘗試 {attempt + 2}/{max_retries}): application_id={application_id_int}")
                        continue
                    # 如果沒有唯一索引導致 ON DUPLICATE KEY UPDATE 失敗，回退到 UPDATE/INSERT 方法
                    elif "Duplicate entry" not in error_str and "ON DUPLICATE KEY" not in error_str:
                        print(f"⚠️ [DEBUG] INSERT ... ON DUPLICATE KEY UPDATE 失敗，回退到 UPDATE/INSERT 方法: {e}")
                        # 確保清除任何未讀取的結果
                        try:
                            cursor.fetchall()
                        except:
                            pass
                        
                        # 使用帶重試的 UPDATE/INSERT 方法
                        for retry_attempt in range(max_retries):
                            try:
                                cursor.execute("""
                                    UPDATE resume_teacher SET review_status=%s, comment=%s, reviewed_at=NOW()
                                    WHERE application_id=%s AND teacher_id=%s
                                """, (status, comment, application_id_int, user_id))
                                updated_rows = cursor.rowcount
                                if updated_rows == 0:
                                    # 確保清除任何未讀取的結果
                                    try:
                                        cursor.fetchall()
                                    except:
                                        pass
                                    
                                    cursor.execute("""
                                        INSERT INTO resume_teacher (application_id, teacher_id, review_status, comment, reviewed_at, created_at)
                                        VALUES (%s, %s, %s, %s, NOW(), NOW())
                                    """, (application_id_int, user_id, status, comment))
                                print(f"✅ [DEBUG] 指導老師審核完成（回退方法）: application_id={application_id_int}, review_status={status}")
                                break
                            except Exception as retry_error:
                                retry_error_str = str(retry_error)
                                if ("Deadlock" in retry_error_str or "1213" in retry_error_str or "40001" in retry_error_str) and retry_attempt < max_retries - 1:
                                    import time
                                    time.sleep(retry_delay * (retry_attempt + 1))
                                    print(f"⚠️ [DEBUG] 回退方法檢測到死鎖，重試中 (嘗試 {retry_attempt + 2}/{max_retries})")
                                    continue
                                else:
                                    raise
                        break
                    else:
                        # 其他錯誤（如唯一約束衝突），也回退到 UPDATE/INSERT 方法
                        print(f"⚠️ [DEBUG] INSERT ... ON DUPLICATE KEY UPDATE 遇到錯誤，回退到 UPDATE/INSERT 方法: {e}")
                        # 確保清除任何未讀取的結果
                        try:
                            cursor.fetchall()
                        except:
                            pass
                        
                        # 使用帶重試的 UPDATE/INSERT 方法
                        for retry_attempt in range(max_retries):
                            try:
                                cursor.execute("""
                                    UPDATE resume_teacher SET review_status=%s, comment=%s, reviewed_at=NOW()
                                    WHERE application_id=%s AND teacher_id=%s
                                """, (status, comment, application_id_int, user_id))
                                updated_rows = cursor.rowcount
                                if updated_rows == 0:
                                    # 確保清除任何未讀取的結果
                                    try:
                                        cursor.fetchall()
                                    except:
                                        pass
                                    
                                    try:
                                        cursor.execute("""
                                            INSERT INTO resume_teacher (application_id, teacher_id, review_status, comment, reviewed_at, created_at)
                                            VALUES (%s, %s, %s, %s, NOW(), NOW())
                                        """, (application_id_int, user_id, status, comment))
                                    except Exception as insert_error:
                                        # 如果 INSERT 也失敗（可能是唯一約束），嘗試再次 UPDATE
                                        print(f"⚠️ [DEBUG] INSERT 失敗，可能是唯一約束衝突: {insert_error}")
                                        cursor.execute("""
                                            UPDATE resume_teacher SET review_status=%s, comment=%s, reviewed_at=NOW()
                                            WHERE application_id=%s AND teacher_id=%s
                                        """, (status, comment, application_id_int, user_id))
                                print(f"✅ [DEBUG] 指導老師審核完成（回退方法）: application_id={application_id_int}, review_status={status}")
                                break
                            except Exception as retry_error:
                                retry_error_str = str(retry_error)
                                if ("Deadlock" in retry_error_str or "1213" in retry_error_str or "40001" in retry_error_str) and retry_attempt < max_retries - 1:
                                    import time
                                    time.sleep(retry_delay * (retry_attempt + 1))
                                    print(f"⚠️ [DEBUG] 回退方法檢測到死鎖，重試中 (嘗試 {retry_attempt + 2}/{max_retries})")
                                    continue
                                else:
                                    raise
                        break
            # 指導老師通過時立即建立 resume_applications，讓廠商可見該履歷
            if status == 'approved':
                job_id = resume_data.get('job_id')
                if job_id is not None:
                    ensure_resume_application_for_teacher_approved(cursor, conn, application_id_int, job_id)
                    # 確保清除任何未讀取的結果（防止 "Unread result found" 錯誤）
                    try:
                        cursor.fetchall()
                    except:
                        pass
        elif user_role == 'class_teacher' and resume_teacher_table_exists and application_id_int is not None:
            # 班導：只更新「這一筆投遞」的狀態（寫入 resume_teacher，不更新 resumes 表，避免同一份履歷其他公司跟著變）
            old_status_for_check = resume_data.get('old_teacher_review_status') or 'uploaded'
            # 確保清除任何未讀取的結果（防止 "Unread result found" 錯誤）
            try:
                cursor.fetchall()
            except:
                pass
            
            # 使用 INSERT ... ON DUPLICATE KEY UPDATE 來原子性地更新 resume_teacher 表
            # 這避免了先 SELECT 再 UPDATE/INSERT 可能導致的死鎖問題
            max_retries = 3
            retry_delay = 0.1  # 100ms
            
            for attempt in range(max_retries):
                try:
                    cursor.execute("""
                        INSERT INTO resume_teacher (application_id, teacher_id, review_status, comment, reviewed_at, created_at)
                        VALUES (%s, %s, %s, %s, NOW(), NOW())
                        ON DUPLICATE KEY UPDATE
                            review_status = VALUES(review_status),
                            comment = VALUES(comment),
                            reviewed_at = NOW()
                    """, (application_id_int, user_id, status, comment))
                    break
                except Exception as e:
                    error_str = str(e)
                    # 檢查是否為死鎖錯誤
                    if ("Deadlock" in error_str or "1213" in error_str or "40001" in error_str) and attempt < max_retries - 1:
                        import time
                        time.sleep(retry_delay * (attempt + 1))
                        print(f"⚠️ [DEBUG] 班導審核檢測到死鎖，重試中 (嘗試 {attempt + 2}/{max_retries}): application_id={application_id_int}")
                        continue
                    # 如果沒有唯一索引導致 ON DUPLICATE KEY UPDATE 失敗，回退到 SELECT/UPDATE/INSERT 方法
                    elif "Duplicate entry" not in error_str and "ON DUPLICATE KEY" not in error_str:
                        print(f"⚠️ [DEBUG] 班導審核 INSERT ... ON DUPLICATE KEY UPDATE 失敗，回退到 SELECT/UPDATE/INSERT 方法: {e}")
                        # 確保清除任何未讀取的結果
                        try:
                            cursor.fetchall()
                        except:
                            pass
                        
                        # 使用帶重試的 SELECT/UPDATE/INSERT 方法
                        for retry_attempt in range(max_retries):
                            try:
                                cursor.execute("""
                                    SELECT id, review_status FROM resume_teacher
                                    WHERE application_id = %s AND teacher_id = %s
                                """, (application_id_int, user_id))
                                existing_ct = cursor.fetchone()
                                if existing_ct:
                                    # 確保清除任何未讀取的結果
                                    try:
                                        cursor.fetchall()
                                    except:
                                        pass
                                    
                                    cursor.execute("""
                                        UPDATE resume_teacher SET review_status=%s, comment=%s, reviewed_at=NOW()
                                        WHERE application_id=%s AND teacher_id=%s
                                    """, (status, comment, application_id_int, user_id))
                                else:
                                    # 確保清除任何未讀取的結果
                                    try:
                                        cursor.fetchall()
                                    except:
                                        pass
                                    
                                    cursor.execute("""
                                        INSERT INTO resume_teacher (application_id, teacher_id, review_status, comment, reviewed_at, created_at)
                                        VALUES (%s, %s, %s, %s, NOW(), NOW())
                                    """, (application_id_int, user_id, status, comment))
                                break
                            except Exception as retry_error:
                                retry_error_str = str(retry_error)
                                if ("Deadlock" in retry_error_str or "1213" in retry_error_str or "40001" in retry_error_str) and retry_attempt < max_retries - 1:
                                    import time
                                    time.sleep(retry_delay * (retry_attempt + 1))
                                    print(f"⚠️ [DEBUG] 班導審核回退方法檢測到死鎖，重試中 (嘗試 {retry_attempt + 2}/{max_retries})")
                                    continue
                                else:
                                    raise
                        break
                    elif ("Deadlock" in error_str or "1213" in error_str or "40001" in error_str) and attempt < max_retries - 1:
                        # 死鎖錯誤，重試
                        import time
                        time.sleep(retry_delay * (attempt + 1))
                        print(f"⚠️ [DEBUG] 班導審核檢測到死鎖，重試中 (嘗試 {attempt + 2}/{max_retries})")
                        continue
                    else:
                        raise
            if status == 'approved':
                company_id = resume_data.get('company_id')
                # 確保清除任何未讀取的結果
                try:
                    cursor.fetchall()
                except:
                    pass
                
                cursor.execute("""
                    SELECT advisor_user_id FROM internship_companies WHERE id = %s
                """, (company_id,))
                ic = cursor.fetchone()
                advisor_user_id = ic.get('advisor_user_id') if ic else None
                if advisor_user_id:
                    # 確保清除任何未讀取的結果
                    try:
                        cursor.fetchall()
                    except:
                        pass
                    
                    cursor.execute("""
                        SELECT id FROM resume_teacher WHERE application_id = %s AND teacher_id = %s
                    """, (application_id_int, advisor_user_id))
                    if not cursor.fetchone():
                        # 確保清除任何未讀取的結果
                        try:
                            cursor.fetchall()
                        except:
                            pass
                        
                        cursor.execute("""
                            INSERT INTO resume_teacher (application_id, teacher_id, review_status, comment, reviewed_at, created_at)
                            VALUES (%s, %s, 'uploaded', NULL, NULL, NOW())
                        """, (application_id_int, advisor_user_id))
            print(f"✅ [DEBUG] 班導審核完成（僅此筆投遞）: application_id={application_id_int}, review_status={status}")
        else:
            # 非班導/指導老師或無 resume_teacher 表：沿用舊邏輯更新 resumes 表
            old_status_for_check = old_status
            cursor.execute("""
                UPDATE resumes SET 
                    status=%s,
                    comment=%s,
                    reviewed_by=%s,
                    reviewed_at=NOW(),
                    updated_at=NOW()
                WHERE id=%s
            """, (status, comment, user_id, resume_id))
            # 班導審核通過且未傳 application_id 時，為該履歷所有投遞建立指導老師審核記錄（向後相容）
            if status == 'approved' and user_role == 'class_teacher' and resume_teacher_table_exists:
                cursor.execute("""
                    SELECT sja.id AS application_id, ic.advisor_user_id
                    FROM student_job_applications sja
                    JOIN internship_companies ic ON sja.company_id = ic.id
                    WHERE sja.student_id = %s AND sja.resume_id = %s AND ic.advisor_user_id IS NOT NULL
                """, (student_user_id, resume_id))
                for app in cursor.fetchall() or []:
                    # 確保清除任何未讀取的結果（防止 "Unread result found" 錯誤）
                    try:
                        cursor.fetchall()
                    except:
                        pass
                    
                    cursor.execute("""
                        SELECT id FROM resume_teacher WHERE application_id = %s AND teacher_id = %s
                    """, (app['application_id'], app['advisor_user_id']))
                    if not cursor.fetchone():
                        # 確保清除任何未讀取的結果
                        try:
                            cursor.fetchall()
                        except:
                            pass
                        
                        cursor.execute("""
                            INSERT INTO resume_teacher (application_id, teacher_id, review_status, comment, reviewed_at, created_at)
                            VALUES (%s, %s, 'uploaded', NULL, NULL, NOW())
                        """, (app['application_id'], app['advisor_user_id']))
        
        # 4. 取得審核者姓名
        cursor.execute("SELECT name, role FROM users WHERE id = %s", (user_id,))
        reviewer = cursor.fetchone()
        if reviewer:
            reviewer_name = reviewer['name'] if reviewer['name'] else "審核老師"
        else:
            reviewer_name = "審核者"

        # 5. 處理通知 (僅在狀態改變時處理)
        status_changed = (old_status_for_check != status) if old_status_for_check is not None else True
        if status_changed:
            if status == 'rejected':
                # 建立退件通知
                notification_content = (
                    f"您的履歷已被 {reviewer_name} 老師退件。\n\n"
                    f"退件原因：{comment if comment else '請查看老師留言'}\n\n"
                    f"請根據老師的建議修改後重新上傳。"
                )

                create_notification(
                    user_id=student_user_id,
                    title="履歷退件通知",
                    message=notification_content,
                    category="resume"
                )

            elif status == 'approved':
                # 建立通過通知
                notification_content = (
                    f"恭喜您！您的履歷已由 {reviewer_name} 老師審核通過。\n"
                    f"您可以繼續後續的實習申請流程。"
                )

                create_notification(
                    user_id=student_user_id,
                    title="履歷審核通過通知",
                    message=notification_content,
                    category="resume"
                )
                
                # 指導老師通過履歷：已在更新 resume_teacher 時立即建立 resume_applications（見 ensure_resume_application_for_teacher_approved）
                if user_role == 'teacher':
                    print(f"✅ 指導老師通過履歷，已寫入 resume_applications 供廠商審核")

        conn.commit()

        return jsonify({"success": True, "message": "履歷審核狀態更新成功"})

    except Exception as e:
        conn.rollback()
        traceback.print_exc() 
        return jsonify({"success": False, "message": f"伺服器錯誤: {str(e)}，請檢查後台日誌"}), 500

    finally:
        cursor.close()
        conn.close()


# -------------------------
# API - 更新履歷欄位（留言）
# -------------------------
@resume_bp.route('/api/update_resume_field', methods=['POST'])
def update_resume_field():
    if 'user_id' not in session:
        return jsonify({"success": False, "message": "未登入"}), 403
    
    user_id = session['user_id']
    user_role = session.get('role')
    
    data = request.get_json()
    if not data:
        return jsonify({"success": False, "message": "請求資料格式錯誤"}), 400
    
    field = data.get('field')
    value = data.get('value', '')
    resume_id = data.get('resume_id')
    application_id = data.get('application_id')
    
    if field != 'comment':
        return jsonify({"success": False, "message": "不支援的欄位"}), 400

    # 班導僅能下載、查看，不能編輯留言（與主任設定一致）
    if user_role == 'class_teacher':
        return jsonify({"success": False, "message": "班導僅能下載與查看履歷，無法編輯留言。"}), 403
    
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 檢查 resume_teacher 表是否存在
        resume_teacher_table_exists = False
        try:
            cursor.execute("SHOW TABLES LIKE 'resume_teacher'")
            resume_teacher_table_exists = cursor.fetchone() is not None
        except Exception as e:
            resume_teacher_table_exists = False
        
        if user_role == 'teacher' and resume_teacher_table_exists and application_id:
            # 指導老師更新 resume_teacher 表的 comment
            try:
                application_id_int = int(application_id)
            except (TypeError, ValueError):
                return jsonify({"success": False, "message": "無效的 application_id"}), 400
            
            cursor.execute("""
                UPDATE resume_teacher SET comment=%s
                WHERE application_id=%s AND teacher_id=%s
            """, (value, application_id_int, user_id))
            
            if cursor.rowcount == 0:
                cursor.execute("""
                    INSERT INTO resume_teacher (application_id, teacher_id, review_status, comment, created_at)
                    VALUES (%s, %s, 'uploaded', %s, NOW())
                """, (application_id_int, user_id, value))
        else:
            # 其他角色更新 resumes 表的 comment
            if not resume_id:
                return jsonify({"success": False, "message": "請提供 resume_id"}), 400
            
            cursor.execute("""
                UPDATE resumes SET comment=%s, updated_at=NOW()
                WHERE id=%s
            """, (value, resume_id))
        
        conn.commit()
        return jsonify({"success": True, "message": "留言已更新"})
    
    except Exception as e:
        conn.rollback()
        traceback.print_exc()
        return jsonify({"success": False, "message": f"伺服器錯誤: {str(e)}"}), 500
    
    finally:
        cursor.close()
        conn.close()


# -------------------------
# API - 下載履歷
# -------------------------
@resume_bp.route('/api/download_resume/<int:resume_id>', methods=['GET'])
def download_resume(resume_id):
    if 'user_id' not in session:
        return jsonify({"success": False, "message": "未登入"}), 403
    
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        cursor.execute("""
            SELECT filepath, original_filename
            FROM resumes
            WHERE id = %s
        """, (resume_id,))
        resume = cursor.fetchone()
        
        if not resume:
            return jsonify({"success": False, "message": "找不到履歷"}), 404
        
        file_path = resume['filepath']
        original_filename = resume['original_filename']
        
        if not file_path:
            return jsonify({"success": False, "message": "履歷檔案不存在"}), 404
        # 資料庫可能存相對路徑 (uploads/resumes/...)，需對應專案根目錄
        if os.path.isabs(file_path):
            full_path = file_path
        else:
            full_path = os.path.normpath(os.path.join(BASE_UPLOAD_DIR, file_path.replace("\\", "/")))
        if not os.path.exists(full_path):
            return jsonify({"success": False, "message": "履歷檔案不存在"}), 404
        
        return send_file(
            full_path,
            as_attachment=True,
            download_name=original_filename or 'resume.pdf'
        )
    
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": f"下載錯誤: {str(e)}"}), 500
    
    finally:
        cursor.close()
        conn.close()


# -------------------------
# 頁面路由
# -------------------------
@resume_bp.route('/review_resume')
def review_resume_page():
    # 檢查登入狀態
    if not require_login():
        return redirect('/login')
    
    # 統一使用整合後的審核頁面（給指導老師使用）
    return render_template('resume/review_resume.html')



# ============================================
# 以下內容從 resume copy.py 合併而來（獨有部分）
# ============================================

@resume_bp.route('/api/get_cert_authorities', methods=['GET'])
def get_cert_authorities():
    conn = None
    cursor = None
    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("SELECT id, name FROM cert_authorities ORDER BY name")
        authorities = cursor.fetchall()
        
        return jsonify({
            "success": True,
            "authorities": authorities
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": f"伺服器錯誤: {str(e)}"}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@resume_bp.route('/api/get_certificates_by_authority', methods=['GET'])
def get_certificates_by_authority():
    conn = None
    cursor = None
    try:
        authority_id = request.args.get('authority_id')
        if not authority_id:
            return jsonify({"success": False, "message": "缺少 authority_id 參數"}), 400
        
        try:
            authority_id = int(authority_id)
        except ValueError:
            return jsonify({"success": False, "message": "authority_id 必須是數字"}), 400

        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        
        # 檢查是否有 name 欄位（向後兼容）
        cursor.execute("SHOW COLUMNS FROM certificate_codes LIKE 'name'")
        has_name_column = cursor.fetchone() is not None
        
        if has_name_column:
            name_select = "name"
            order_by = "name"
        else:
            # 如果沒有 name 欄位，使用 job_category 和 level 組合
            name_select = "CONCAT(COALESCE(job_category, ''), COALESCE(level, '')) AS name"
            order_by = "COALESCE(job_category, ''), COALESCE(level, '')"
        
        cursor.execute(f"""
            SELECT code, {name_select}, category 
            FROM certificate_codes 
            WHERE authority_id = %s 
            ORDER BY {order_by}
        """, (authority_id,))
        certificates = cursor.fetchall()
        
        return jsonify({
            "success": True,
            "certificates": certificates
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": f"伺服器錯誤: {str(e)}"}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@resume_bp.route('/api/save_resume_data', methods=['POST'])
def save_resume_data():
    if 'user_id' not in session or session.get('role') != 'student':
        return jsonify({"success": False, "message": "未授權"}), 403

    student_id = session['username']
    data = request.get_json()
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 取得目前的學期 ID (如果系統使用學期分流)
        semester_id = get_current_semester_id(cursor)
        resume_id = None
        if data.get("resume_id") is not None and str(data.get("resume_id")).strip():
            try:
                resume_id = int(data.get("resume_id"))
            except (TypeError, ValueError):
                pass

        save_result = save_structured_data(cursor, student_id, data, semester_id, resume_id=resume_id)
        ok = save_result[0] if isinstance(save_result, tuple) else bool(save_result)
        if ok:
            conn.commit()
            return jsonify({"success": True, "message": "履歷資料儲存成功"})
        else:
            conn.rollback()
            return jsonify({"success": False, "message": "履歷資料儲存失敗 (資料庫錯誤)"}), 500

    except Exception as e:
        conn.rollback()
        print("❌ 儲存履歷資料錯誤:", e)
        traceback.print_exc()
        return jsonify({"success": False, "message": f"伺服器錯誤: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

@resume_bp.route('/api/get_resume_data', methods=['GET'])
def get_resume_data():
    if 'user_id' not in session or session.get('role') != 'student':
        return jsonify({"success": False, "message": "未授權"}), 403

    user_id = session['user_id']
    conn = get_db()
    cursor = conn.cursor(dictionary=True)

    try:
        # ===== 1. 決定要載入哪一筆履歷（可傳 resume_id 編輯指定履歷）=====
        resume = None
        resume_id_param = request.args.get("resume_id", "").strip()
        if resume_id_param:
            try:
                rid = int(resume_id_param)
                cursor.execute("SELECT id FROM resumes WHERE id = %s AND user_id = %s", (rid, user_id))
                resume = cursor.fetchone()
                if not resume:
                    return jsonify({"success": False, "message": "找不到該履歷或無權限"}), 404
            except (ValueError, TypeError):
                resume = None
        if not resume:
            cursor.execute("""
                SELECT id FROM resumes 
                WHERE user_id = %s 
                ORDER BY created_at DESC 
                LIMIT 1
            """, (user_id,))
            resume = cursor.fetchone()
        if not resume:
            return jsonify({"success": False, "message": "沒有已提交的履歷"}), 404

        # ===== 2. 抓 StudentID（學號）及 users 的 name, email, avatar_url =====
        cursor.execute("SELECT username, name, email, avatar_url FROM users WHERE id=%s", (user_id,))
        user_result = cursor.fetchone()
        if not user_result:
            return jsonify({"success": False, "message": "找不到使用者"}), 404

        student_id = user_result["username"]

        # 載入該履歷的內容關聯（若有則只顯示勾選的課程/證照/語文/缺勤）
        mapping = _get_resume_content_mapping(cursor, resume["id"])

        # ===== 3. 基本資料 =====
        cursor.execute("SELECT * FROM Student_Info WHERE StuID=%s", (student_id,))
        student_info = cursor.fetchone() or {}

        # ===== 4. 課程資料 =====
        # 檢查是否有 ProofImage 欄位
        cursor.execute("SHOW COLUMNS FROM course_grades LIKE 'ProofImage'")
        has_proof_image = cursor.fetchone() is not None
        
        if has_proof_image:
            cursor.execute("""
                SELECT id, CourseName AS name, Credits AS credits, Grade AS grade, ProofImage AS transcript_path
                FROM course_grades
                WHERE StuID=%s
                ORDER BY CourseName
            """, (student_id,))
        else:
            cursor.execute("""
                SELECT id, CourseName AS name, Credits AS credits, Grade AS grade
                FROM course_grades
                WHERE StuID=%s
                ORDER BY CourseName
            """, (student_id,))
        courses = cursor.fetchall() or []
        if mapping and mapping.get("course_grade_ids"):
            raw_ids = [x.strip() for x in (mapping["course_grade_ids"] or "").split(",") if x.strip()]
            if raw_ids:
                # 向後相容：若為數字則視為 course_grades.id，否則視為課程名稱
                ids_are_int = all((x or "").replace("-", "").isdigit() for x in raw_ids)
                if ids_are_int:
                    ids_set = set(int(x) for x in raw_ids)
                    courses = [c for c in courses if c.get("id") in ids_set]
                else:
                    ids_set = set(raw_ids)
                    courses = [c for c in courses if (c.get("name") or "").strip() in ids_set]
        
        # 提取成績單路徑（從 ProofImage 欄位）
        transcript_path = ''
        for course in courses:
            tp = course.get('transcript_path')
            if tp:
                transcript_path = tp
                break

        # ===== 5. 證照資料 — 單一 SQL，不再三段重複 =====

        sql_cert = """
            SELECT
                sc.id,
                sc.CertPath,
                sc.AcquisitionDate,
                sc.cert_code,
                sc.issuer,
                sc.authority_name,
                sc.job_category AS sc_job_category,
                sc.CreatedAt,
                
                -- 發證中心ID：優先使用 sc.authority_id（如果存在），否則從 certificate_codes 獲取
                COALESCE(
                    sc.authority_id,
                    CASE 
                        WHEN sc.cert_code IS NOT NULL
                        THEN cc.authority_id
                        ELSE NULL
                    END
                ) AS authority_id,

                -- 職類：若 cert_code 有值且不是 OTHER → 取 certificate_codes
                CASE 
                    WHEN sc.cert_code IS NOT NULL
                    THEN COALESCE(cc.job_category, '')
                    ELSE COALESCE(sc.job_category, '')
                END AS job_category,

                -- 等級
                CASE 
                    WHEN sc.cert_code IS NOT NULL
                    THEN COALESCE(cc.level, '')
                    ELSE COALESCE(sc.level, '')
                END AS level,

                -- 組合證照名稱
                CASE 
                    WHEN (
                        CASE 
                            WHEN sc.cert_code IS NOT NULL
                            THEN cc.job_category
                            ELSE sc.job_category
                        END
                    ) IS NOT NULL
                    AND (
                        CASE 
                            WHEN sc.cert_code IS NOT NULL
                            THEN cc.level
                            ELSE sc.level
                        END
                    ) IS NOT NULL
                    AND (
                        CASE 
                            WHEN sc.cert_code IS NOT NULL
                            THEN cc.job_category
                            ELSE sc.job_category
                        END
                    ) != ''
                    AND (
                        CASE 
                            WHEN sc.cert_code IS NOT NULL
                            THEN cc.level
                            ELSE sc.level
                        END
                    ) != ''
                THEN CONCAT(
                    CASE 
                        WHEN sc.cert_code IS NOT NULL
                        THEN cc.job_category
                        ELSE sc.job_category
                    END,
                    CASE 
                        WHEN sc.cert_code IS NOT NULL
                        THEN cc.level
                        ELSE sc.level
                    END
                )
                ELSE ''
                END AS CertName,

                -- 發證中心名稱：優先使用 sc.authority_id 關聯的 cert_authorities，否則使用從 certificate_codes 獲取的 authority_id，最後使用 authority_name
                COALESCE(
                    ca_from_sc.name,
                    ca.name, 
                    sc.authority_name, 
                    'N/A'
                ) AS IssuingBody,
                -- 證照分類：優先使用 authority_id 判斷（勞動部 authority_id=1 為 labor），否則使用 cc.category
                CASE 
                    WHEN COALESCE(sc.authority_id, cc.authority_id) = 1 THEN 'labor'
                    WHEN cc.category IS NOT NULL THEN cc.category
                    ELSE 'other'
                END AS CertType,
                CASE 
                    WHEN COALESCE(sc.authority_id, cc.authority_id) = 1 THEN 'labor'
                    WHEN cc.category IS NOT NULL THEN cc.category
                    ELSE 'other'
                END AS category,
                cc.job_category AS official_name
            FROM student_certifications sc
            LEFT JOIN certificate_codes cc 
                ON sc.cert_code = cc.id
            LEFT JOIN cert_authorities ca 
                ON cc.authority_id = ca.id
            LEFT JOIN cert_authorities ca_from_sc 
                ON sc.authority_id = ca_from_sc.id
            WHERE sc.StuID = %s
            ORDER BY sc.id DESC
        """

        cursor.execute(sql_cert, (student_id,))
        all_certifications = cursor.fetchall() or []
        if mapping and mapping.get("certification_ids"):
            try:
                ids_set = set(int(x.strip()) for x in (mapping["certification_ids"] or "").split(",") if x.strip())
                if ids_set:
                    all_certifications = [c for c in all_certifications if c.get("id") in ids_set]
            except ValueError:
                pass
        
        # 在處理證照分類時
        labor_certs = [c for c in all_certifications if c.get('category') == 'labor' or c.get('authority_id') == 1]
        other_certs = [c for c in all_certifications if c.get('category') not in ['labor', 'intl', 'local']]
        
        # 調試：打印查詢結果，確認 level 字段
        print(f"🔍 查詢證照資料: 共 {len(all_certifications)} 筆")
        for idx, cert in enumerate(all_certifications[:3]):  # 只打印前3筆
            print(f"  證照 {idx+1}: id={cert.get('id')}, cert_code={cert.get('cert_code')}, job_category={cert.get('job_category')}, level={cert.get('level')}, authority_id={cert.get('authority_id')}")

        # ===== 6. 取最新一批證照 =====

        certifications = []
        if all_certifications:
            latest_created_at = all_certifications[0]["CreatedAt"]
            latest_id = all_certifications[0]["id"]

            if latest_created_at:
                certifications = [
                    c for c in all_certifications
                    if c["CreatedAt"] == latest_created_at
                ]
            else:
                max_id = latest_id
                certifications = [
                    c for c in all_certifications
                    if c["id"] >= (max_id - 50)
                ]

            # 過濾空白資料
            certifications = [
                c for c in certifications
                if (
                    (c["job_category"] and c["level"]) or
                    (c["CertName"]) or
                    (c.get("cert_code") is not None)
                )
            ]
            # 依 id 升序排列，使編輯時證照順序與使用者新增順序一致（先新增的在前）
            certifications.sort(key=lambda c: (c.get("id") or 0))

        # ===== 7. 語言能力 =====

        cursor.execute("""
            SELECT id, Language AS language, Level AS level
            FROM student_languageskills
            WHERE StuID=%s
            ORDER BY Language
        """, (student_id,))
        languages = cursor.fetchall() or []
        if mapping and mapping.get("language_skill_ids"):
            try:
                ids_set = set(int(x.strip()) for x in (mapping["language_skill_ids"] or "").split(",") if x.strip())
                if ids_set:
                    languages = [l for l in languages if l.get("id") in ids_set]
            except ValueError:
                pass
        
        # ===== 7.5 缺勤記錄與佐證圖片（依 mapping 篩選，供編輯頁載入）=====
        absence_proof_path = ''
        absence_data = None
        try:
            cursor.execute("SELECT id FROM users WHERE username=%s", (student_id,))
            user_row = cursor.fetchone()
            if user_row:
                user_id = user_row.get('id')
                # 若有 mapping 的 absence_record_ids，只查這些筆；否則查該使用者全部缺勤
                absence_ids_filter = None
                if mapping and mapping.get("absence_record_ids"):
                    try:
                        absence_ids_filter = [int(x.strip()) for x in (mapping["absence_record_ids"] or "").split(",") if x.strip()]
                    except (ValueError, TypeError):
                        pass
                try:
                    cursor.execute("SHOW COLUMNS FROM absence_records LIKE 'semester_id'")
                    has_semester = cursor.fetchone() is not None
                except Exception:
                    has_semester = False
                if absence_ids_filter:
                    placeholders = ",".join(["%s"] * len(absence_ids_filter))
                    cursor.execute(f"""
                        SELECT id, absence_date, absence_type, duration_units, reason, image_path
                        {" , semester_id" if has_semester else ""}
                        FROM absence_records
                        WHERE user_id = %s AND id IN ({placeholders})
                        ORDER BY absence_date DESC, id ASC
                    """, (user_id, *absence_ids_filter))
                else:
                    cursor.execute(f"""
                        SELECT id, absence_date, absence_type, duration_units, reason, image_path
                        {" , semester_id" if has_semester else ""}
                        FROM absence_records
                        WHERE user_id = %s
                        ORDER BY absence_date DESC, id ASC
                    """, (user_id,))
                absence_rows = cursor.fetchall() or []
                # 第一筆有 image_path 的作為佐證圖
                for row in absence_rows:
                    ip = (row.get("image_path") or "").strip()
                    if ip and not absence_proof_path:
                        absence_proof_path = ip
                if not absence_proof_path:
                    cursor.execute("""
                        SELECT image_path FROM absence_records
                        WHERE user_id = %s AND image_path IS NOT NULL AND image_path != ''
                        ORDER BY id DESC LIMIT 1
                    """, (user_id,))
                    fallback = cursor.fetchone()
                    if fallback:
                        absence_proof_path = fallback.get("image_path", "")
                # 組 absence_data 供前端編輯頁顯示列表與佐證圖
                all_types = ["曠課", "遲到", "事假", "病假", "生理假", "公假", "喪假"]
                stats = {t: 0 for t in all_types}
                records = []
                for r in absence_rows:
                    ad = r.get("absence_date")
                    if hasattr(ad, "strftime"):
                        ad = ad.strftime("%Y-%m-%d")
                    records.append({
                        "id": r.get("id"),
                        "absence_date": ad or "",
                        "absence_type": r.get("absence_type") or "",
                        "duration_units": r.get("duration_units") or 0,
                        "reason": r.get("reason") or "",
                        "image_path": (r.get("image_path") or "").strip(),
                    })
                    at = r.get("absence_type")
                    if at in stats:
                        stats[at] = stats.get(at, 0) + int(r.get("duration_units") or 0)
                stat_str = {t: f"{stats.get(t, 0)} 節" for t in all_types}
                absence_data = {
                    "records": records,
                    "stats": stat_str,
                    "start_semester_id": None,
                    "end_semester_id": None,
                }
                if absence_proof_path:
                    print(f"🔍 找到缺勤佐證圖片: {absence_proof_path}")
        except Exception as e:
            print(f"⚠️ 查詢缺勤記錄/佐證圖片失敗: {e}")
            traceback.print_exc()

        # ===== 8. 日期格式轉換 =====
        birth_date = student_info.get("BirthDate")
        if birth_date:
            if isinstance(birth_date, datetime):
                birth_date = birth_date.strftime("%Y-%m-%d")
            else:
                try:
                    birth_date = datetime.strptime(birth_date, "%Y-%m-%d").strftime("%Y-%m-%d")
                except:
                    pass

        # ===== 9. 格式化證照輸出 =====
        formatted_certs = []
        for cert in certifications:
            acquire_date = cert.get("AcquisitionDate")
            formatted_acquire_date = ""
            acquisition_date_str = None  # 用於 JSON 序列化的字符串格式
            
            if acquire_date is not None:
                if isinstance(acquire_date, datetime):
                    formatted_acquire_date = acquire_date.strftime("%Y-%m-%d")
                    acquisition_date_str = formatted_acquire_date
                elif isinstance(acquire_date, date):
                    formatted_acquire_date = acquire_date.strftime("%Y-%m-%d")
                    acquisition_date_str = formatted_acquire_date
                elif acquire_date:
                    try:
                        # 嘗試解析字符串格式的日期
                        if isinstance(acquire_date, str):
                            formatted_acquire_date = datetime.strptime(acquire_date, "%Y-%m-%d").strftime("%Y-%m-%d")
                            acquisition_date_str = formatted_acquire_date
                        else:
                            formatted_acquire_date = str(acquire_date)
                            acquisition_date_str = formatted_acquire_date
                    except Exception as e:
                        print(f"⚠️ 日期格式化失敗: {acquire_date}, 錯誤: {e}")
                        formatted_acquire_date = str(acquire_date) if acquire_date else ""
                        acquisition_date_str = formatted_acquire_date
            
            # 獲取級別字段（SQL 返回的字段名是 level）
            cert_level = cert.get("level", "")
            print(f"🔍 證照資料處理: id={cert.get('id')}, AcquisitionDate={acquire_date}, formatted={formatted_acquire_date}, level={cert_level}, job_category={cert.get('job_category', '')}")
            
            # 獲取證照圖片路徑，並將 Windows 路徑格式（反斜杠）轉換為 Web 路徑格式（正斜杠）
            cert_path_raw = cert.get("CertPath", "")
            cert_path = cert_path_raw.replace("\\", "/") if cert_path_raw else ""
            
            formatted_certs.append({
                "id": cert["id"],
                "cert_code": cert.get("cert_code", ""),
                "cert_path": cert_path,
                "name": cert.get("CertName", ""),
                "job_category": cert.get("job_category", ""),
                "level": cert_level,  # 修正：SQL 返回的字段名是 level，不是 CertLevel
                "authority_name": cert.get("authority_name", ""),
                "issuer": cert.get("issuer", ""),
                "authority_id": cert.get("authority_id") if "authority_id" in cert else None,
                "IssuingBody": cert.get("IssuingBody", ""),
                "CertType": cert.get("CertType", "other"),
                "acquire_date": formatted_acquire_date,
                "AcquisitionDate": acquisition_date_str  # 轉換為字符串格式，確保 JSON 序列化正常
            })

        # ===== 10. 回傳結果（路徑一律正斜線、不回傳 None）=====
        def norm_path(p):
            if p is None or (isinstance(p, str) and p.strip() == ""):
                return ""
            return (p or "").replace("\\", "/").strip()

        # 姓名、電子信箱、個人頭貼優先從 users 表帶入
        user_name = (user_result.get("name") or "").strip() or student_info.get("StuName", "")
        user_email = (user_result.get("email") or "").strip() or student_info.get("Email", "")
        user_avatar = norm_path(user_result.get("avatar_url")) or norm_path(student_info.get("PhotoPath"))

        return jsonify({
            "success": True,
            "data": {
                "resume_id": resume["id"],
                "student_info": {
                    "name": user_name,
                    "birth_date": birth_date or "",
                    "gender": student_info.get("Gender", ""),
                    "phone": student_info.get("Phone", ""),
                    "email": user_email,
                    "address": student_info.get("Address", ""),
                    "conduct_score": student_info.get("ConductScore", ""),
                    "autobiography": student_info.get("Autobiography", ""),
                    "photo_path": user_avatar
                },
                "courses": courses,
                "certifications": formatted_certs,
                "languages": languages,
                "transcript_path": norm_path(transcript_path),
                "absence_proof_path": norm_path(absence_proof_path),
                "absence_data": absence_data
            }
        })

    except Exception as e:
        print("❌ 取得履歷資料錯誤:", e)
        traceback.print_exc()
        return jsonify({"success": False, "message": f"取得履歷資料失敗: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()
        
# -------------------------
# API: 根據發證中心ID取得該中心的職類和級別列表
# -------------------------

@resume_bp.route('/api/get_job_categories_and_levels', methods=['GET'])
def get_job_categories_and_levels():
    conn = None
    cursor = None
    try:
        authority_id = request.args.get('authority_id')
        if not authority_id:
            return jsonify({"success": False, "message": "缺少 authority_id 參數"}), 400
        
        try:
            authority_id = int(authority_id)
        except ValueError:
            return jsonify({"success": False, "message": "authority_id 必須是數字"}), 400

        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        
        # 取得該發證中心的所有證照
        # 使用 certificate_codes 表的 job_category 和 level 字段組合生成 name
        # 檢查是否有 name 欄位（向後兼容）
        cursor.execute("SHOW COLUMNS FROM certificate_codes LIKE 'name'")
        has_name_column = cursor.fetchone() is not None
        
        if has_name_column:
            # 如果還有 name 欄位，使用 COALESCE 向後兼容
            name_select = "COALESCE(CONCAT(job_category, level), name) AS name"
            order_by = "COALESCE(job_category, name), COALESCE(level, '')"
        else:
            # 如果沒有 name 欄位，直接使用 CONCAT
            name_select = "CONCAT(COALESCE(job_category, ''), COALESCE(level, '')) AS name"
            order_by = "COALESCE(job_category, ''), COALESCE(level, '')"
        
        cursor.execute(f"""
            SELECT code, 
                   {name_select},
                   COALESCE(job_category, '') AS job_category,
                   COALESCE(level, '') AS level
            FROM certificate_codes 
            WHERE authority_id = %s 
            ORDER BY {order_by}
        """, (authority_id,))
        certificates = cursor.fetchall()
        
        # 解析職類和級別
        import re
        job_categories = set()  # 使用 set 避免重複
        job_category_levels = {}  # {職類: [級別列表]}
        
        level_pattern = re.compile(r'(甲級|乙級|丙級|丁級|甲|乙|丙|丁)')
        
        print(f"🔍 查詢發證中心 {authority_id} 的證照，共 {len(certificates)} 筆")
        
        for cert in certificates:
            # 優先使用 certificate_codes 表的 job_category 和 level 字段
            job_category = cert.get('job_category', '').strip()
            level = cert.get('level', '').strip()
            cert_name = cert.get('name', '').strip()
            
            # 情況1: job_category 和 level 都有值，直接使用
            if job_category and level:
                job_categories.add(job_category)
                if job_category not in job_category_levels:
                    job_category_levels[job_category] = set()
                job_category_levels[job_category].add(level)
                print(f"  ✅ 使用欄位值: 職類={job_category}, 級別={level}")
            # 情況2: 只有 job_category 有值（即使沒有 level 也顯示職類）
            elif job_category:
                job_categories.add(job_category)
                if job_category not in job_category_levels:
                    job_category_levels[job_category] = set()
                # 嘗試從 name 解析 level（如果有的話）
                if not level and cert_name:
                    match = level_pattern.search(cert_name)
                    if match:
                        parsed_level = match.group(1)
                        level_map = {'甲': '甲級', '乙': '乙級', '丙': '丙級', '丁': '丁級'}
                        full_level = level_map.get(parsed_level, parsed_level)
                        job_category_levels[job_category].add(full_level)
                        print(f"  ✅ 職類有值，從名稱解析級別: 職類={job_category}, 級別={full_level}")
                    else:
                        print(f"  ✅ 職類有值，無級別: 職類={job_category}")
                elif level:
                    job_category_levels[job_category].add(level)
                    print(f"  ✅ 職類和級別都有值: 職類={job_category}, 級別={level}")
                else:
                    print(f"  ✅ 職類有值，無級別: 職類={job_category}")
            # 情況3: 只有 level 有值，嘗試從 name 解析 job_category
            elif level and not job_category and cert_name:
                # 從名稱中移除級別，剩下的作為職類
                parsed_job_category = level_pattern.sub('', cert_name).strip()
                if parsed_job_category:
                    job_categories.add(parsed_job_category)
                    if parsed_job_category not in job_category_levels:
                        job_category_levels[parsed_job_category] = set()
                    job_category_levels[parsed_job_category].add(level)
                    print(f"  ✅ 級別有值，從名稱解析職類: 職類={parsed_job_category}, 級別={level}")
            # 情況4: 都沒有值，從 name 字段解析職類和級別（向後兼容）
            elif cert_name:
                match = level_pattern.search(cert_name)
                if match:
                    parsed_level = match.group(1)
                    level_map = {'甲': '甲級', '乙': '乙級', '丙': '丙級', '丁': '丁級'}
                    full_level = level_map.get(parsed_level, parsed_level)
                    
                    # 提取職類（移除級別後的部分）
                    parsed_job_category = level_pattern.sub('', cert_name).strip()
                    
                    if parsed_job_category:
                        job_categories.add(parsed_job_category)
                        if parsed_job_category not in job_category_levels:
                            job_category_levels[parsed_job_category] = set()
                        job_category_levels[parsed_job_category].add(full_level)
                        print(f"  ✅ 從名稱解析: 職類={parsed_job_category}, 級別={full_level}")
                else:
                    # 如果無法解析級別，但名稱不為空，將整個名稱作為職類（無級別）
                    job_categories.add(cert_name)
                    if cert_name not in job_category_levels:
                        job_category_levels[cert_name] = set()
                    print(f"  ✅ 從名稱解析（無級別）: 職類={cert_name}")
            else:
                print(f"  ⚠️ 跳過無效證照記錄: code={cert.get('code')}, name={cert_name}")
        
        # 轉換為列表並排序
        job_categories_list = sorted(list(job_categories))
        # 將級別集合轉換為排序列表
        for job_category in job_category_levels:
            job_category_levels[job_category] = sorted(list(job_category_levels[job_category]))
        
        return jsonify({
            "success": True,
            "job_categories": job_categories_list,
            "job_category_levels": job_category_levels  # {職類: [級別列表]}
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": f"伺服器錯誤: {str(e)}"}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@resume_bp.route('/api/submit_and_generate', methods=['POST'])
def submit_and_generate_api():
    context = {}
    conn = None
    cursor = None

    try:
        if session.get('role') != 'student' or not session.get('user_id'):
            return jsonify({"success": False, "message": "只有學生可以提交"}), 403

        user_id = session['user_id']
        conn = get_db()
        cursor = conn.cursor(dictionary=True)

        data = request.form.to_dict()
        courses = json.loads(data.get('courses', '[]'))
        photo = request.files.get('photo')
        transcript_file = request.files.get('transcript_file')
        cert_files = request.files.getlist('cert_photos[]')
        cert_names = request.form.getlist('cert_names[]')

        ALLOWED_IMAGE_MIMES = ['image/jpeg', 'image/png', 'image/gif', 'image/webp']

        # 儲存照片
        photo_path = None
        if photo and photo.filename:
            if photo.mimetype not in ALLOWED_IMAGE_MIMES:
                return jsonify({"success": False, "message": f"照片檔案格式錯誤 ({photo.mimetype})"}), 400
            filename = secure_filename(photo.filename)
            photo_dir = os.path.join(BASE_UPLOAD_DIR, UPLOAD_FOLDER, "photos")
            os.makedirs(photo_dir, exist_ok=True)
            ext = os.path.splitext(filename)[1]
            new_filename = f"{user_id}_photo_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
            photo_path = os.path.join(photo_dir, new_filename)
            photo.save(photo_path)
            photo_path = photo_path.replace("\\", "/")
        else:
            # 未上傳新照片時保留資料庫既有路徑
            cursor.execute("SELECT PhotoPath FROM Student_Info WHERE StuID = (SELECT username FROM users WHERE id=%s)", (user_id,))
            row = cursor.fetchone()
            if row and row.get('PhotoPath'):
                photo_path = (row['PhotoPath'] or '').replace('\\', '/')

        # 儲存成績單檔案（先儲存檔案，再 update 到 course_grades 的 transcript_path）
        transcript_path = None
        if transcript_file and transcript_file.filename:
            if transcript_file.mimetype not in ALLOWED_IMAGE_MIMES:
                return jsonify({"success": False, "message": f"成績單檔案格式錯誤 ({transcript_file.mimetype})"}), 400
            filename = secure_filename(transcript_file.filename)
            transcript_dir = os.path.join(BASE_UPLOAD_DIR, UPLOAD_FOLDER, "transcripts")
            os.makedirs(transcript_dir, exist_ok=True)
            ext = os.path.splitext(filename)[1]
            new_filename = f"{user_id}_transcript_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
            transcript_path = os.path.join(transcript_dir, new_filename)
            transcript_file.save(transcript_path)
            transcript_path = transcript_path.replace("\\", "/")

        # 儲存多張證照
        cert_photo_paths = []
        if cert_files:
            cert_dir = os.path.join(BASE_UPLOAD_DIR, UPLOAD_FOLDER, "cert_photos")
            os.makedirs(cert_dir, exist_ok=True)

        for idx, file in enumerate(cert_files, start=1):
            if file and file.filename:
                if file.mimetype not in ALLOWED_IMAGE_MIMES:
                    print(f"⚠️ 證照檔案格式錯誤已跳過: {file.filename} ({file.mimetype})")
                    continue
                ext = os.path.splitext(secure_filename(file.filename))[1]
                new_filename = f"{user_id}_cert_{idx}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
                file_path = os.path.join(cert_dir, new_filename)
                file.save(file_path)
                cert_photo_paths.append(file_path)

        # 處理單張證照圖片（certificate_image + certificate_description）
        certificate_image_file = request.files.get('certificate_image')
        certificate_description = request.form.get('certificate_description', '')
        image_path_for_template = None
        if certificate_image_file and certificate_image_file.filename != '' and 'user_id' in session:
            try:
                cert_folder = os.path.join(BASE_UPLOAD_DIR, UPLOAD_FOLDER, 'certificates')
                os.makedirs(cert_folder, exist_ok=True)
                filename = secure_filename(certificate_image_file.filename)
                file_extension = os.path.splitext(filename)[1] or '.png'
                unique_filename = f"{session['user_id']}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{os.urandom(4).hex()}{file_extension}"
                image_save_path = os.path.join(cert_folder, unique_filename)
                certificate_image_file.save(image_save_path)
                image_path_for_template = image_save_path
            except Exception as e:
                print(f"❌ 儲存單一證照圖片失敗: {e}")
                traceback.print_exc()
                image_path_for_template = None

        # 先解析 structured_certifications，才能決定是否插入單張證照（避免多張證照時索引錯位、圖片重複）
        _structured_from_form = []
        _structured_json = request.form.get('structured_certifications', '')
        if _structured_json:
            try:
                _structured_from_form = json.loads(_structured_json)
            except Exception:
                pass
        # 僅在「沒有」使用新格式多張證照時，才在 index 0 插入單張證照
        if (image_path_for_template or certificate_description) and not _structured_from_form:
            if cert_photo_paths is None:
                cert_photo_paths = []
            if cert_names is None:
                cert_names = []
            cert_photo_paths.insert(0, image_path_for_template or "")
            cert_names.insert(0, certificate_description or "")

        # 組合缺勤統計（支援學期範圍篩選）
        absence_stats = {}
        
        # 獲取學期範圍參數
        start_semester_id = request.form.get("start_semester_id", None)
        end_semester_id = request.form.get("end_semester_id", None)
        
        # 構建查詢條件
        where_conditions = ["user_id = %s"]
        query_params = [user_id]
        
        # 如果有學期範圍，添加學期篩選
        if start_semester_id and end_semester_id:
            # 獲取所有在範圍內的學期ID
            cursor.execute("""
                SELECT id FROM semesters 
                WHERE code >= (SELECT code FROM semesters WHERE id = %s)
                AND code <= (SELECT code FROM semesters WHERE id = %s)
                ORDER BY code
            """, (start_semester_id, end_semester_id))
            semester_ids_in_range = [row['id'] for row in cursor.fetchall()]
            if semester_ids_in_range:
                placeholders = ','.join(['%s'] * len(semester_ids_in_range))
                where_conditions.append(f"semester_id IN ({placeholders})")
                query_params.extend(semester_ids_in_range)
        
        where_clause = " AND ".join(where_conditions)
        
        cursor.execute(f"""
            SELECT 
                absence_type, 
                SUM(duration_units) AS total_units 
            FROM absence_records
            WHERE {where_clause}
            GROUP BY absence_type
        """, tuple(query_params))
        results = cursor.fetchall()
        all_types = ["曠課", "遲到", "事假", "病假", "生理假", "公假", "喪假"]
        db_stats = {t: 0 for t in all_types}
        for row in results:
            typ = row.get('absence_type')
            if typ in db_stats:
                try:
                    db_stats[typ] = int(row.get('total_units') or 0)
                except Exception:
                    db_stats[typ] = 0
        for t in all_types:
            absence_stats[f"absence_{t}_units"] = f"{db_stats.get(t,0)} 節"

        incoming_stats_json = request.form.get("absence_stats_json", None)
        if incoming_stats_json:
            try:
                incoming = json.loads(incoming_stats_json)
                for t in all_types:
                    val = incoming.get(t)
                    if val is not None:
                        try:
                            val_int = int(val)
                        except Exception:
                            try:
                                val_int = int(str(val).replace("節","").strip())
                            except Exception:
                                val_int = db_stats.get(t, 0)
                        absence_stats[f"absence_{t}_units"] = f"{val_int} 節"
            except Exception as e:
                print("⚠️ 無法解析 absence_stats_json，忽略前端傳入值:", e)

        total = 0
        for t in all_types:
            v = absence_stats.get(f"absence_{t}_units", "0 節")
            try:
                total += int(str(v).replace("節","").strip())
            except Exception:
                pass
        absence_stats["absence_總計_units"] = f"{total} 節"
        
        # 調試輸出：確認缺勤統計數據
        print("📊 缺勤統計數據:", absence_stats)
        
        context.update(absence_stats)
        
        # 調試輸出：確認 context 中的缺勤統計數據
        print("📊 context 中的缺勤統計數據:", {k: v for k, v in context.items() if k.startswith("absence_")})

        # 處理並儲存缺勤佐證圖片（與你原邏輯一致）
        absence_image_path = None
        try:
            uploaded_proof = request.files.get('proof_image') or request.files.get('absence_proof')
            if uploaded_proof and uploaded_proof.filename:
                if uploaded_proof.mimetype in ALLOWED_IMAGE_MIMES:
                    abs_dir = os.path.join(BASE_UPLOAD_DIR, ABSENCE_PROOF_FOLDER)
                    os.makedirs(abs_dir, exist_ok=True)
                    ext = os.path.splitext(secure_filename(uploaded_proof.filename))[1] or ".png"
                    fname = f"{user_id}_absence_proof_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
                    savep = os.path.join(abs_dir, fname)
                    uploaded_proof.save(savep)
                    # 存進 context 用相對路徑，產生 Word 時 resolve_upload_path 可正確找到
                    absence_image_path = (ABSENCE_PROOF_FOLDER + "/" + fname).replace("\\", "/")
                else:
                    print(f"⚠️ 上傳的缺勤佐證圖片格式不支援: {uploaded_proof.mimetype}")
        except Exception as e:
            print("⚠️ 儲存上傳的缺勤佐證圖片失敗:", e)
            traceback.print_exc()

        if not absence_image_path:
            try:
                ar_json = request.form.get("absence_records_json", None)
                if ar_json:
                    try:
                        ar_list = json.loads(ar_json)
                        for rec in reversed(ar_list):
                            img = rec.get("image_filename") or rec.get("image_path")
                            if img:
                                absence_image_path = img
                                break
                    except Exception as e:
                        print("⚠️ 解析 absence_records_json 失敗:", e)
            except Exception as e:
                print("⚠️ 嘗試讀取 absence_records_json 失敗:", e)

        if not absence_image_path:
            try:
                cursor.execute("""
                    SELECT image_path
                    FROM absence_records
                    WHERE user_id = %s AND image_path IS NOT NULL AND image_path != ''
                    ORDER BY created_at DESC
                    LIMIT 1
                """, (user_id,))
                row = cursor.fetchone()
                if row:
                    absence_image_path = row.get('image_path')
            except Exception as e:
                print(f"Error fetching latest absence proof path from DB: {e}")

        context['Absence_Proof_Path'] = absence_image_path

        # 更新缺勤記錄的佐證圖片（從資料庫讀取的記錄，只需更新圖片）
        try:
            # 1. 處理個別記錄的佐證圖片上傳
            absence_records_with_images_json = request.form.get("absence_records_with_images", None)
            if absence_records_with_images_json:
                try:
                    records_with_images = json.loads(absence_records_with_images_json)
                    print(f"📝 準備更新 {len(records_with_images)} 筆缺勤記錄的佐證圖片")
                    
                    for record_info in records_with_images:
                        record_id = record_info.get("record_id")
                        if not record_id:
                            continue
                        
                        # 獲取對應的圖片文件
                        image_key = f"proof_image_{record_id}"
                        uploaded_image = request.files.get(image_key)
                        
                        if uploaded_image and uploaded_image.filename:
                            try:
                                abs_dir = os.path.join(BASE_UPLOAD_DIR, ABSENCE_PROOF_FOLDER)
                                os.makedirs(abs_dir, exist_ok=True)
                                ext = os.path.splitext(secure_filename(uploaded_image.filename))[1] or ".png"
                                fname = f"{user_id}_record_{record_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
                                save_path = os.path.join(abs_dir, fname)
                                uploaded_image.save(save_path)
                                save_path_db = (ABSENCE_PROOF_FOLDER + "/" + fname).replace("\\", "/")
                                # 更新資料庫中對應記錄的 image_path
                                cursor.execute("""
                                    UPDATE absence_records 
                                    SET image_path = %s, updated_at = NOW()
                                    WHERE id = %s AND user_id = %s
                                """, (save_path_db, record_id, user_id))
                                
                                print(f"✅ 已更新缺勤記錄 {record_id} 的佐證圖片: {save_path}")
                            except Exception as e:
                                print(f"⚠️ 更新缺勤記錄 {record_id} 的佐證圖片失敗: {e}")
                                traceback.print_exc()
                    
                    conn.commit()
                    print(f"✅ 所有缺勤記錄的佐證圖片已成功更新")
                except Exception as e:
                    print(f"⚠️ 解析 absence_records_with_images 失敗: {e}")
                    traceback.print_exc()
            
            # 2. 如果有整體佐證圖片，更新到該學期或全部沒有圖片的記錄
            if absence_image_path:
                semester_id = request.form.get("semester_id", None)
                try:
                    if semester_id:
                        # 檢查是否有 semester_id 欄位
                        cursor.execute("SHOW COLUMNS FROM absence_records LIKE 'semester_id'")
                        has_semester_id = cursor.fetchone() is not None
                        
                        if has_semester_id:
                            cursor.execute("""
                                UPDATE absence_records 
                                SET image_path = %s, updated_at = NOW()
                                WHERE user_id = %s AND semester_id = %s 
                                AND (image_path IS NULL OR image_path = '')
                            """, (absence_image_path, user_id, semester_id))
                        else:
                            cursor.execute("""
                                UPDATE absence_records 
                                SET image_path = %s, updated_at = NOW()
                                WHERE user_id = %s 
                                AND (image_path IS NULL OR image_path = '')
                            """, (absence_image_path, user_id))
                    else:
                        # 未傳學期時：先更新所有「尚無圖片」的記錄
                        cursor.execute("""
                            UPDATE absence_records 
                            SET image_path = %s, updated_at = NOW()
                            WHERE user_id = %s 
                            AND (image_path IS NULL OR image_path = '')
                        """, (absence_image_path, user_id))
                    # 若沒有任何一筆被更新，改為更新該使用者最新一筆缺勤記錄
                    if cursor.rowcount == 0:
                        cursor.execute("""
                            UPDATE absence_records ar
                            JOIN (SELECT id FROM absence_records WHERE user_id = %s ORDER BY id DESC LIMIT 1) t ON ar.id = t.id
                            SET ar.image_path = %s, ar.updated_at = NOW()
                        """, (user_id, absence_image_path))
                    conn.commit()
                    print(f"✅ 已將整體佐證圖片更新到缺勤記錄")
                except Exception as e:
                    print(f"⚠️ 更新整體佐證圖片失敗: {e}")
                    traceback.print_exc()
        except Exception as e:
            print(f"⚠️ 處理缺勤記錄圖片失敗: {e}")
            traceback.print_exc()

        # 查學生學號 (username)
        cursor.execute("SELECT username FROM users WHERE id=%s", (user_id,))
        result = cursor.fetchone()
        if not result:
            return jsonify({"success": False, "message": "找不到使用者"}), 404
        student_id = result['username']

        # 確保 courses 中的 grade 欄位存在
        for c in courses:
            c['grade'] = c.get('grade', '')

        # 解析文本證照資料（非圖片）
        structured_certifications = []
        
        # 優先從 JSON 字串解析（新格式）
        structured_certifications_json = request.form.get('structured_certifications', '')
        if structured_certifications_json:
            try:
                structured_certifications = json.loads(structured_certifications_json)
                print(f"✅ 從 JSON 解析到 {len(structured_certifications)} 項證照")
            except Exception as e:
                print(f"⚠️ 解析 structured_certifications JSON 失敗: {e}")
                structured_certifications = []
        
        # 如果 JSON 解析失敗或為空，則使用舊格式（向後兼容）
        if not structured_certifications:
            cert_names_text = request.form.getlist('cert_name[]')
            cert_types = request.form.getlist('cert_type[]')
            cert_codes_text = request.form.getlist('cert_code[]')  # 新增：證照代碼
            cert_issuers_text = request.form.getlist('cert_issuer[]')  # 新增：發證人

            for n, t, code, issuer in zip(cert_names_text, cert_types, cert_codes_text, cert_issuers_text):
               if n.strip():
                    structured_certifications.append({
                    "name": n.strip(),
                    "type": t.strip() if t else "other",
                    "code": code.strip().upper() if code else "",  # 新增：證照代碼
                    "issuer": issuer.strip() if issuer else ""  # 新增：發證人
            })

        # 將表單上傳的證照圖片路徑依序寫入 structured_certifications（與前端證照項目順序 1:1 對應，避免重複或遺漏）
        if cert_photo_paths and structured_certifications:
            for i in range(len(structured_certifications)):
                if i < len(cert_photo_paths) and cert_photo_paths[i]:
                    path = cert_photo_paths[i]
                    normalized = (path or "").replace("\\", "/")
                    if "uploads" in normalized:
                        parts = normalized.split("/")
                        idx_u = parts.index("uploads")
                        rel = "/".join(parts[idx_u:])
                        structured_certifications[i]["cert_path"] = rel
                    else:
                        structured_certifications[i]["cert_path"] = normalized

        # 解析語言能力資料
        structured_languages = []
        # 前端使用 lang_en_level, lang_tw_level, lang_jp_level, lang_hk_level
        lang_mapping = {
            'lang_en_level': '英語',
            'lang_tw_level': '台語',
            'lang_jp_level': '日語',
            'lang_hk_level': '客語'
        }
        
        for form_field, lang_name in lang_mapping.items():
            level = request.form.get(form_field, '').strip() or '略懂'
            structured_languages.append({
                "language": lang_name,
                "level": level
            })

        # 收集證照代碼和發證人（從前端表單）
        cert_codes = request.form.getlist('cert_code[]')
        cert_issuers = request.form.getlist('cert_issuer[]')  # 新增：發證人列表
        
        # 若未上傳新成績單，從資料庫取既有 ProofImage 寫回各課程（供 save_structured_data 使用）
        if not transcript_path and result and student_id:
            cursor.execute("SHOW COLUMNS FROM course_grades LIKE 'ProofImage'")
            if cursor.fetchone():
                cursor.execute("""
                    SELECT ProofImage FROM course_grades
                    WHERE StuID = %s AND ProofImage IS NOT NULL AND ProofImage != ''
                    LIMIT 1
                """, (student_id,))
                row = cursor.fetchone()
                if row and row.get('ProofImage'):
                    existing_proof = (row['ProofImage'] or '').replace('\\', '/')
                    for c in courses:
                        if not c.get('proof_image'):
                            c['proof_image'] = existing_proof

        # 建立結構化資料（傳入 save_structured_data）
        semester_id = get_current_semester_id(cursor)
        structured_data = {
            "name": data.get("name"),
            "birth_date": data.get("birth_date"),
            "gender": data.get("gender"),
            "phone": data.get("phone"),
            "email": data.get("email"),
            "address": data.get("address"),
            "conduct_score": score_to_grade(data.get("conduct_score")),
            "autobiography": data.get("autobiography"),
            "courses": courses,
            "photo_path": photo_path,
            "structured_certifications": structured_certifications,
            "structured_languages": structured_languages,
            "cert_photo_paths": cert_photo_paths,
            "cert_names": cert_names,
            "cert_codes": cert_codes,  # 新增：證照代碼列表
            "cert_issuers": cert_issuers  # 新增：發證人列表
        }

        # 將表單數據和結構化數據也加入 context (以便套版)
        context.update(data)
        context.update(structured_data)

        # 不論新增或編輯，都準備 selected_* 供 resume_content_mapping 使用
        resume_id_for_save = None
        if request.form.get("resume_id", "").strip():
            try:
                resume_id_for_save = int(request.form.get("resume_id", "").strip())
            except (ValueError, TypeError):
                pass
        structured_data["selected_course_grade_ids"] = [c.get("name") for c in courses if c.get("name")]
        try:
            structured_data["selected_certification_ids"] = json.loads(request.form.get("selected_certification_ids", "[]") or "[]")
        except Exception:
            structured_data["selected_certification_ids"] = [c.get("id") for c in structured_certifications if c.get("id")]
        try:
            structured_data["selected_language_skill_ids"] = json.loads(request.form.get("selected_language_skill_ids", "[]") or "[]")
        except Exception:
            structured_data["selected_language_skill_ids"] = None
        try:
            structured_data["selected_absence_record_ids"] = json.loads(request.form.get("selected_absence_record_ids", "[]") or "[]")
        except Exception:
            structured_data["selected_absence_record_ids"] = None

        # 儲存結構化資料（包含 language / Certs / course_grades）；回傳 mapping IDs 供 resume_content_mapping 寫入
        save_result = save_structured_data(cursor, student_id, structured_data, semester_id=semester_id, resume_id=resume_id_for_save)
        save_ok = save_result[0] if isinstance(save_result, tuple) else bool(save_result)
        mapping_ids = save_result[1] if isinstance(save_result, tuple) and len(save_result) > 1 else {}
        if not save_ok:
            conn.rollback()
            return jsonify({"success": False, "message": "資料儲存失敗"}), 500

        # 將成績單圖片路徑更新到 course_grades 表的 ProofImage 欄位
        if transcript_path:
            try:
                # 檢查表是否有 SemesterID 和 ProofImage 列
                cursor.execute("SHOW COLUMNS FROM course_grades LIKE 'SemesterID'")
                has_semester_id = cursor.fetchone() is not None
                cursor.execute("SHOW COLUMNS FROM course_grades LIKE 'ProofImage'")
                has_proof_image = cursor.fetchone() is not None
                
                if has_proof_image:
                    if has_semester_id and semester_id:
                        # 嘗試 update 同學該學期的 course_grades（若沒有，插入一筆佔位紀錄）
                        cursor.execute("""
                            UPDATE course_grades
                            SET ProofImage = %s
                            WHERE StuID = %s AND SemesterID = %s
                        """, (transcript_path, student_id, semester_id))
                        if cursor.rowcount == 0:
                            # 沒有更新到任何列，插入一筆僅含 ProofImage 的占位
                            cursor.execute("""
                                INSERT INTO course_grades (StuID, CourseName, Credits, Grade, SemesterID, ProofImage)
                                VALUES (%s, %s, %s, %s, %s, %s)
                            """, (student_id, '', 0, '', semester_id, transcript_path))
                    else:
                        # 沒有 SemesterID 列，只根據 StuID 更新
                        cursor.execute("""
                            UPDATE course_grades
                            SET ProofImage = %s
                            WHERE StuID = %s
                            LIMIT 1
                        """, (transcript_path, student_id))
                else:
                    # 如果沒有 ProofImage 列，嘗試使用 transcript_path（兼容舊結構）
                    cursor.execute("SHOW COLUMNS FROM course_grades LIKE 'transcript_path'")
                    has_transcript_path = cursor.fetchone() is not None
                    if has_transcript_path:
                        if has_semester_id and semester_id:
                            cursor.execute("""
                                UPDATE course_grades
                                SET transcript_path = %s
                                WHERE StuID = %s AND SemesterID = %s
                            """, (transcript_path, student_id, semester_id))
                        else:
                            cursor.execute("""
                                UPDATE course_grades
                                SET transcript_path = %s
                                WHERE StuID = %s
                                LIMIT 1
                            """, (transcript_path, student_id))
            except Exception as e:
                print("⚠️ 更新 course_grades.ProofImage 失敗:", e)
                traceback.print_exc()

        # 生成 Word 文件
        student_data_for_doc = get_student_info_for_doc(cursor, student_id, semester_id=semester_id, resume_id=resume_id_for_save)
        # PhotoPath & ConductScoreNumeric（未上傳新照片時保留既有路徑）
        if photo_path is not None:
            student_data_for_doc["info"]["PhotoPath"] = photo_path
        student_data_for_doc["info"]["ConductScoreNumeric"] = data.get("conduct_score_numeric")
        # 傳遞證照圖片與名稱清單（generate 會自行從 certs 讀）
        student_data_for_doc["cert_photo_paths"] = cert_photo_paths
        student_data_for_doc["cert_names"] = cert_names
        # 合併 context（包含缺勤統計數據）
        student_data_for_doc.update(context)
        
        # 調試輸出：確認 student_data_for_doc 中的缺勤統計數據
        absence_keys_in_doc = {k: v for k, v in student_data_for_doc.items() if k.startswith("absence_")}
        print("📊 student_data_for_doc 中的缺勤統計數據:", absence_keys_in_doc)

        # 編輯模式：若有 resume_id 則覆蓋既有檔案並只更新該筆履歷，畫面只保留一個檔案；新增則產生新檔並 INSERT
        resume_id_param = request.form.get("resume_id", "").strip()
        had_resume_id = bool(resume_id_param)  # 僅在「從未帶 resume_id」時才 INSERT，避免編輯時誤新增一筆
        existing_filepath = None
        if resume_id_param:
            try:
                rid = int(resume_id_param)
                cursor.execute(
                    "SELECT filepath, original_filename FROM resumes WHERE id = %s AND user_id = %s",
                    (rid, user_id)
                )
                row = cursor.fetchone()
                if row and row.get("filepath"):
                    existing_filepath = row["filepath"]
            except (ValueError, TypeError):
                pass

        if existing_filepath:
            # 覆蓋既有檔案：DB 可能存相對路徑，轉成絕對路徑再寫入
            if os.path.isabs(existing_filepath):
                save_path = existing_filepath
            else:
                save_path = os.path.normpath(os.path.join(BASE_UPLOAD_DIR, existing_filepath.replace("\\", "/")))
            filename = os.path.basename(save_path) or f"{student_id}_履歷_{datetime.now().strftime('%Y%m%d_%H%M%S')}.docx"
        else:
            filename = f"{student_id}_履歷_{datetime.now().strftime('%Y%m%d_%H%M%S')}.docx"
            save_path = os.path.join(BASE_UPLOAD_DIR, UPLOAD_FOLDER, filename) if not os.path.isabs(UPLOAD_FOLDER) else os.path.join(UPLOAD_FOLDER, filename)

        if not generate_application_form_docx(student_data_for_doc, save_path):
            conn.rollback()
            return jsonify({"success": False, "message": "文件生成失敗"}), 500

        if had_resume_id and resume_id_param:
            # 編輯：一律只更新該筆履歷，不新增列，並更新 resume_content_mapping（含 absence_record_ids）
            try:
                rid = int(resume_id_param)
                filepath_for_db = existing_filepath if existing_filepath else (os.path.join(UPLOAD_FOLDER, filename)).replace("\\", "/")
                cursor.execute("""
                    UPDATE resumes
                    SET filepath = %s, original_filename = %s, updated_at = NOW()
                    WHERE id = %s AND user_id = %s
                """, (filepath_for_db, filename, rid, user_id))
                _upsert_resume_content_mapping(
                    cursor, rid, student_id,
                    course_grade_ids=mapping_ids.get("course_grade_ids"),
                    certification_ids=mapping_ids.get("certification_ids"),
                    language_skill_ids=mapping_ids.get("language_skill_ids"),
                    absence_record_ids=mapping_ids.get("absence_record_ids"),
                )
            except (ValueError, TypeError):
                pass  # 不將 resume_id_param 清空，避免誤執行下面的 INSERT
        if not had_resume_id:
            filepath_for_db = (os.path.join(UPLOAD_FOLDER, filename)).replace("\\", "/")
            cursor.execute("""
                INSERT INTO resumes
                (user_id, filepath, original_filename, status, category, semester_id, created_at)
                VALUES (%s, %s, %s, %s, %s, %s, NOW())
            """, (
                user_id,
                filepath_for_db,
                filename,
                'uploaded',
                'draft',
                semester_id
            ))
            # 新增履歷也要寫入 resume_content_mapping（使用 save_structured_data 回傳的 ID，確保存的是 id 而非課程名稱）
            new_resume_id = cursor.lastrowid
            if new_resume_id:
                _upsert_resume_content_mapping(
                    cursor, new_resume_id, student_id,
                    course_grade_ids=mapping_ids.get("course_grade_ids"),
                    certification_ids=mapping_ids.get("certification_ids"),
                    language_skill_ids=mapping_ids.get("language_skill_ids"),
                    absence_record_ids=mapping_ids.get("absence_record_ids")
                )

        conn.commit()
        return jsonify({
            "success": True,
            "message": "履歷已成功提交並生成文件",
            "file_path": save_path,
            "filename": filename
        })

    except Exception as e:
        print("❌ submit_and_generate_api 發生錯誤:", e)
        traceback.print_exc()
        if conn:
            conn.rollback()
        return jsonify({"success": False, "message": f"系統錯誤: {str(e)}"}), 500

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@resume_bp.route('/api/upload_transcript', methods=['POST'])
def upload_transcript():
    if 'user_id' not in session or session.get('role') != 'student':
        return jsonify({"success": False, "message": "未授權"}), 403

    student_id = session['username']
    
    if 'file' not in request.files:
        return jsonify({"success": False, "message": "缺少文件"}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({"success": False, "message": "未選擇文件"}), 400

    # 檢查文件類型 (圖片)
    allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'pdf'}
    if '.' not in file.filename or file.filename.rsplit('.', 1)[1].lower() not in allowed_extensions:
        return jsonify({"success": False, "message": "不支援的文件類型"}), 400

    filename = secure_filename(file.filename)
    # 儲存路徑：uploads/resumes/StuID/transcript_timestamp.ext
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    student_dir = os.path.join(BASE_UPLOAD_DIR, UPLOAD_FOLDER, student_id)
    os.makedirs(student_dir, exist_ok=True)
    
    # 儲存名稱
    ext = filename.rsplit('.', 1)[1].lower()
    save_filename = f"transcript_{timestamp}.{ext}"
    save_path_abs = os.path.join(student_dir, save_filename)
    
    file.save(save_path_abs)

    # 相對路徑（用於資料庫儲存）
    relative_path = os.path.join(UPLOAD_FOLDER, student_id, save_filename).replace('\\', '/')

    conn = get_db()
    cursor = conn.cursor()
    try:
        # 更新成績單路徑到 student_info.transcript_path (舊欄位，兼容)
        # 這裡改為更新到 course_grades 的 ProofImage 欄位（以最新的成績單圖片為主）
        
        # 1. 確保 course_grades 表有 ProofImage 欄位
        cursor.execute("SHOW COLUMNS FROM course_grades LIKE 'ProofImage'")
        if not cursor.fetchone():
            conn.rollback()
            return jsonify({"success": False, "message": "資料庫缺少 course_grades.ProofImage 欄位"}), 500

        # 2. 取得目前的學期 ID (如果系統使用學期分流)
        semester_id = get_current_semester_id(cursor)
        
        # 3. 儲存路徑到 course_grades 的所有課程記錄 (該學期或所有)
        if semester_id:
            # 只更新該學期的記錄
            cursor.execute("""
                UPDATE course_grades 
                SET ProofImage = %s 
                WHERE StuID = %s AND IFNULL(SemesterID,'') = %s
            """, (relative_path, student_id, semester_id))
        else:
            # 更新所有記錄 (如果沒有學期分流)
            cursor.execute("""
                UPDATE course_grades 
                SET ProofImage = %s 
                WHERE StuID = %s
            """, (relative_path, student_id))

        conn.commit()
        return jsonify({"success": True, "message": "成績單圖片上傳成功", "path": relative_path})

    except Exception as e:
        conn.rollback()
        print("❌ 上傳成績單圖片錯誤:", e)
        traceback.print_exc()
        return jsonify({"success": False, "message": f"伺服器錯誤: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

@resume_bp.route('/api/upload_photo', methods=['POST'])
def upload_photo():
    if 'user_id' not in session or session.get('role') != 'student':
        return jsonify({"success": False, "message": "未授權"}), 403

    student_id = session['username']
    
    if 'file' not in request.files:
        return jsonify({"success": False, "message": "缺少文件"}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({"success": False, "message": "未選擇文件"}), 400

    # 檢查文件類型 (圖片)
    allowed_extensions = {'png', 'jpg', 'jpeg'}
    if '.' not in file.filename or file.filename.rsplit('.', 1)[1].lower() not in allowed_extensions:
        return jsonify({"success": False, "message": "不支援的文件類型"}), 400

    filename = secure_filename(file.filename)
    # 儲存路徑：uploads/resumes/StuID/photo.ext
    student_dir = os.path.join(BASE_UPLOAD_DIR, UPLOAD_FOLDER, student_id)
    os.makedirs(student_dir, exist_ok=True)
    
    # 儲存名稱 (固定名稱，會覆蓋舊的)
    ext = filename.rsplit('.', 1)[1].lower()
    save_filename = f"photo.{ext}"
    save_path_abs = os.path.join(student_dir, save_filename)
    
    file.save(save_path_abs)

    # 相對路徑（用於資料庫儲存）
    relative_path = os.path.join(UPLOAD_FOLDER, student_id, save_filename).replace('\\', '/')

    conn = get_db()
    cursor = conn.cursor()
    try:
        # 更新照片路徑到 Student_Info.PhotoPath
        cursor.execute("""
            INSERT INTO Student_Info (StuID, PhotoPath)
            VALUES (%s, %s)
            ON DUPLICATE KEY UPDATE PhotoPath = VALUES(PhotoPath), UpdatedAt = NOW()
        """, (student_id, relative_path))

        conn.commit()
        return jsonify({"success": True, "message": "照片上傳成功", "path": relative_path})

    except Exception as e:
        conn.rollback()
        print("❌ 上傳學生照片錯誤:", e)
        traceback.print_exc()
        return jsonify({"success": False, "message": f"伺服器錯誤: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

@resume_bp.route('/api/upload_certificate_photo', methods=['POST'])
def upload_certificate_photo():
    if 'user_id' not in session or session.get('role') != 'student':
        return jsonify({"success": False, "message": "未授權"}), 403

    student_id = session['username']
    
    if 'file' not in request.files:
        return jsonify({"success": False, "message": "缺少文件"}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({"success": False, "message": "未選擇文件"}), 400

    # 檢查文件類型 (圖片)
    allowed_extensions = {'png', 'jpg', 'jpeg', 'pdf'}
    if '.' not in file.filename or file.filename.rsplit('.', 1)[1].lower() not in allowed_extensions:
        return jsonify({"success": False, "message": "不支援的文件類型"}), 400

    filename = secure_filename(file.filename)
    # 儲存路徑：uploads/resumes/StuID/certs/cert_timestamp.ext
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    student_certs_dir = os.path.join(BASE_UPLOAD_DIR, UPLOAD_FOLDER, student_id, "certs")
    os.makedirs(student_certs_dir, exist_ok=True)
    
    # 儲存名稱
    ext = filename.rsplit('.', 1)[1].lower()
    save_filename = f"cert_{timestamp}.{ext}"
    save_path_abs = os.path.join(student_certs_dir, save_filename)
    
    file.save(save_path_abs)

    # 相對路徑（用於資料庫儲存）
    relative_path = os.path.join(UPLOAD_FOLDER, student_id, "certs", save_filename).replace('\\', '/')

    # 不直接在這邊寫入 student_certifications 表，而是返回路徑供前端更新 structured_certifications
    return jsonify({"success": True, "message": "證照圖片上傳成功", "path": relative_path})

@resume_bp.route('/api/get_standard_courses', methods=['GET'])
def get_standard_courses():
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("""
            SELECT 
                course_name AS name, 
                credits 
            FROM standard_courses 
            WHERE is_active = 1 
            ORDER BY order_index
        """)
        courses = cursor.fetchall()
        return jsonify({"success": True, "courses": courses})
    except Exception as e:
        print("❌ 取得標準核心科目錯誤:", e)
        traceback.print_exc()
        return jsonify({"success": False, "message": "取得標準核心科目失敗"}), 500
    finally:
        cursor.close()
        conn.close()

@resume_bp.route('/api/save_personal_template', methods=['POST'])
def save_personal_template():
    if 'user_id' not in session or session.get('role') != 'student':
        return jsonify({"success": False, "message": "未授權"}), 403

    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    try:
        data = request.get_json()
        student_id = session['username']
        template_name = data.get('display_name', '我的課程模板')
        
        # 確保 courses 列表不為 None
        courses_data = data.get('courses', [])
        
        # 檢查 courses 數據結構並將其轉換為 JSON
        valid_courses = []
        for course in courses_data:
            if course.get('name') and course.get('credits') is not None:
                valid_courses.append({
                    'name': course['name'],
                    'credits': format_credits(course['credits']), # 使用格式化函數
                    'grade': course.get('grade', ''),
                    'isNotTaken': course.get('isNotTaken', False)  # 保存未修課狀態
                })
        
        courses_json = json.dumps(valid_courses, ensure_ascii=False)
        
        # 儲存或更新模板
        cursor.execute("""
            INSERT INTO templates (template_type, content, display_name, is_active, uploaded_by, uploaded_at)
            VALUES (%s, %s, %s, %s, %s, NOW())
            ON DUPLICATE KEY UPDATE 
                content=VALUES(content), 
                display_name=VALUES(display_name), 
                updated_at=NOW()
        """, ('student_custom', courses_json, template_name, 1, student_id))
        
        conn.commit()
        return jsonify({"success": True, "message": "個人課程模板儲存成功"})
        
    except Exception as e:
        conn.rollback()
        print("❌ 儲存個人課程模板錯誤:", e)
        traceback.print_exc()
        return jsonify({"success": False, "message": f"伺服器錯誤: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

@resume_bp.route('/api/get_personal_template', methods=['GET'])
def get_personal_template():
    if 'user_id' not in session or session.get('role') != 'student':
        return jsonify({"success": False, "message": "未授權"}), 403

    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    try:
        student_id = session['username']
        
        cursor.execute("""
            SELECT 
                content, 
                display_name
            FROM templates
            WHERE uploaded_by = %s AND template_type = 'student_custom' AND is_active = 1
            ORDER BY uploaded_at DESC
            LIMIT 1
        """, (student_id,))
        
        template = cursor.fetchone()
        
        if template:
            courses = json.loads(template['content'])
            return jsonify({
                "success": True, 
                "display_name": template['display_name'],
                "courses": courses
            })
        else:
            return jsonify({"success": False, "message": "未找到個人課程模板"})
            
    except Exception as e:
        print("❌ 取得個人課程模板錯誤:", e)
        traceback.print_exc()
        return jsonify({"success": False, "message": f"伺服器錯誤: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

@resume_bp.route('/api/upload_course_grade_excel', methods=['POST'])
def upload_course_grade_excel():
    if 'user_id' not in session or session.get('role') != 'ta':
        return jsonify({"success": False, "message": "未授權"}), 403

    if 'file' not in request.files:
        return jsonify({"success": False, "message": "缺少文件"}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({"success": False, "message": "未選擇文件"}), 400

    allowed_extensions = {'xlsx', 'xls'}
    if '.' not in file.filename or file.filename.rsplit('.', 1)[1].lower() not in allowed_extensions:
        return jsonify({"success": False, "message": "不支援的文件類型"}), 400
    
    # 使用 BytesIO 讀取文件，不直接儲存到磁碟
    file_stream = io.BytesIO(file.read())
    
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 1. 載入工作簿
        workbook = load_workbook(file_stream)
        sheet = workbook.active
        
        # 2. 獲取標頭（假設第一行是標頭）
        headers = [cell.value for cell in sheet[1]]
        
        # 預期的欄位名稱
        student_id_col = None
        course_name_col = None
        credits_col = None
        grade_col = None

        # 找出欄位索引
        for i, header in enumerate(headers):
            if header and '學號' in str(header):
                student_id_col = i + 1
            elif header and ('課程名稱' in str(header) or '科目名稱' in str(header)):
                course_name_col = i + 1
            elif header and '學分' in str(header):
                credits_col = i + 1
            elif header and ('成績' in str(header) or '等第' in str(header)):
                grade_col = i + 1

        if not student_id_col or not course_name_col or not credits_col or not grade_col:
            return jsonify({"success": False, "message": "Excel 檔案缺少必要的欄位（學號、課程名稱/科目名稱、學分、成績/等第）"}), 400

        # 3. 處理數據
        data_to_import = {} # { student_id: [ {course_name, credits, grade}, ... ] }
        for row_index in range(2, sheet.max_row + 1):
            try:
                student_id = str(sheet.cell(row=row_index, column=student_id_col).value or '').strip()
                course_name = str(sheet.cell(row=row_index, column=course_name_col).value or '').strip()
                credits = str(sheet.cell(row=row_index, column=credits_col).value or '').strip()
                grade = str(sheet.cell(row=row_index, column=grade_col).value or '').strip()

                if not student_id or not course_name:
                    continue

                if student_id not in data_to_import:
                    data_to_import[student_id] = []
                
                # 簡單格式化學分
                try:
                    credits = float(credits)
                    if credits.is_integer():
                        credits = int(credits)
                except ValueError:
                    # 保持原始字串格式，例如 "2/2"
                    pass

                data_to_import[student_id].append({
                    'name': course_name,
                    'credits': credits,
                    'grade': grade
                })

            except Exception as row_e:
                print(f"⚠️ 處理 Excel 第 {row_index} 行錯誤: {row_e}")
                continue

        if not data_to_import:
            return jsonify({"success": False, "message": "Excel 檔案中未找到有效成績資料"}), 400
        
        # 4. 寫入資料庫
        semester_id = get_current_semester_id(cursor)
        imported_count = 0
        
        # 檢查 course_grades 表中是否有 SemesterID 欄位
        cursor.execute("SHOW COLUMNS FROM course_grades LIKE 'SemesterID'")
        has_semester_id = cursor.fetchone() is not None
        
        for student_id, courses in data_to_import.items():
            try:
                # 刪除該學期或全部舊資料
                if has_semester_id and semester_id:
                    cursor.execute(
                        "DELETE FROM course_grades WHERE StuID=%s AND IFNULL(SemesterID,'')=%s",
                        (student_id, semester_id)
                    )
                else:
                    cursor.execute("DELETE FROM course_grades WHERE StuID=%s", (student_id,))

                # 批量插入新資料
                for c in courses:
                    if has_semester_id and semester_id:
                        cursor.execute("""
                            INSERT INTO course_grades
                                (StuID, CourseName, Credits, Grade, SemesterID)
                            VALUES (%s,%s,%s,%s,%s)
                        """, (student_id, c['name'], c['credits'], c['grade'], semester_id))
                    else:
                        cursor.execute("""
                            INSERT INTO course_grades
                                (StuID, CourseName, Credits, Grade)
                            VALUES (%s,%s,%s,%s)
                        """, (student_id, c['name'], c['credits'], c['grade']))
                
                imported_count += 1
                
            except Exception as db_e:
                print(f"❌ 匯入學生 {student_id} 成績資料失敗: {db_e}")
                conn.rollback() # 確保操作可以被撤銷，但這裡應該使用更細粒度的錯誤處理
                # 這裡為了簡化，如果一個學生失敗就繼續下一個學生，並在外面做一次大提交
                continue

        conn.commit()
        return jsonify({"success": True, "message": f"成功匯入 {imported_count} 位學生的成績資料"})
        
    except Exception as e:
        conn.rollback()
        print("❌ 匯入成績 Excel 錯誤:", e)
        traceback.print_exc()
        return jsonify({"success": False, "message": f"伺服器錯誤: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

@resume_bp.route('/api/import_standard_courses', methods=['POST'])
def import_standard_courses():
    if 'user_id' not in session or session.get('role') != 'ta':
        return jsonify({"success": False, "message": "未授權"}), 403

    if 'file' not in request.files:
        return jsonify({"success": False, "message": "缺少文件"}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({"success": False, "message": "未選擇文件"}), 400

    allowed_extensions = {'xlsx', 'xls'}
    if '.' not in file.filename or file.filename.rsplit('.', 1)[1].lower() not in allowed_extensions:
        return jsonify({"success": False, "message": "不支援的文件類型"}), 400
    
    file_stream = io.BytesIO(file.read())
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        workbook = load_workbook(file_stream)
        sheet = workbook.active
        
        headers = [cell.value for cell in sheet[1]]
        
        course_name_col = None
        credits_col = None
        
        for i, header in enumerate(headers):
            if header and ('課程名稱' in str(header) or '科目名稱' in str(header)):
                course_name_col = i + 1
            elif header and '學分' in str(header):
                credits_col = i + 1

        if not course_name_col or not credits_col:
            return jsonify({"success": False, "message": "Excel 檔案缺少必要的欄位（課程名稱/科目名稱、學分）"}), 400

        # 清空現有核心科目（避免重複或過時資料）
        cursor.execute("UPDATE standard_courses SET is_active = 0")

        imported_count = 0
        for row_index in range(2, sheet.max_row + 1):
            try:
                course_name = str(sheet.cell(row=row_index, column=course_name_col).value or '').strip()
                credits_value = str(sheet.cell(row=row_index, column=credits_col).value or '').strip()

                if not course_name or not credits_value:
                    continue

                # 嘗試將學分轉換為數字
                try:
                    credits = float(credits_value)
                except ValueError:
                    credits = 0.0 # 無效學分設為 0

                # 檢查是否已存在，如果存在則更新 is_active 和 credits
                cursor.execute("""
                    SELECT id FROM standard_courses WHERE course_name = %s LIMIT 1
                """, (course_name,))
                existing_course = cursor.fetchone()
                
                if existing_course:
                    cursor.execute("""
                        UPDATE standard_courses 
                        SET credits = %s, is_active = 1, updated_at = NOW() 
                        WHERE id = %s
                    """, (credits, existing_course['id']))
                else:
                    cursor.execute("""
                        INSERT INTO standard_courses 
                            (course_name, credits, is_active, uploaded_by, uploaded_at)
                        VALUES (%s, %s, 1, %s, NOW())
                    """, (course_name, credits, session['username']))
                
                imported_count += 1
                
            except Exception as row_e:
                print(f"⚠️ 處理 Excel 第 {row_index} 行錯誤: {row_e}")
                continue

        conn.commit()
        return jsonify({"success": True, "message": f"成功匯入 {imported_count} 筆核心科目資料"})
        
    except Exception as e:
        conn.rollback()
        print("❌ 匯入核心科目 Excel 錯誤:", e)
        traceback.print_exc()
        return jsonify({"success": False, "message": f"伺服器錯誤: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

@resume_bp.route('/ta/upload_standard_courses')
def upload_standard_courses_page():
    if 'user_id' not in session or session.get('role') != 'ta':
        return redirect('/login')
    return render_template('ta/upload_standard_courses.html')

@resume_bp.route('/api/company_positions', methods=['GET'])
def get_company_positions():
    try:
        company_name = request.args.get('company_name', '')
        if not company_name:
            return jsonify({"success": False, "message": "請提供公司名稱"}), 400
            
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        
        # 查詢該公司的所有職缺
        cursor.execute("""
            SELECT 
                id,
                title,
                description,
                salary,
                period,
                work_time,
                slots
            FROM internship_jobs
            WHERE company_id IN (
                SELECT id FROM companies WHERE name = %s AND status = 'approved'
            )
            AND is_active = 1
            ORDER BY title
        """, (company_name,))
        
        positions = cursor.fetchall()
        cursor.close()
        conn.close()
        
        return jsonify({
            "success": True,
            "positions": positions
        })
        
    except Exception as e:
        print(f"Error fetching company positions: {str(e)}")
        traceback.print_exc()
        return jsonify({"success": False, "message": "無法取得公司職缺列表"}), 500

@resume_bp.route('/class_review_resume')
def class_review_resume_page():
    # 檢查登入狀態
    if not require_login():
        return redirect('/login')
    
    # 班導審核履歷頁面
    return render_template('resume/class_review_resume.html')


@resume_bp.route("/api/get_class_resumes", methods=["GET"])
def get_class_resumes():
    # 驗證登入
    if not require_login():
        return jsonify({"success": False, "message": "未授權"}), 403

    user_id = session['user_id']
    role = session['role']
    mode = request.args.get('mode', '').strip().lower()
    target_company_id = request.args.get('company_id', type=int)

    conn = get_db()
    cursor = conn.cursor(dictionary=True)

    try:
        # 檢查履歷上傳截止時間並自動更新狀態
        is_resume_deadline_passed, update_counts = update_resume_status_after_deadline(cursor, conn)
        
        resumes = []  # 初始化結果列表
        sql_query = ""
        sql_params = tuple()

        print(f"🔍 [DEBUG] get_class_resumes called - user_id: {user_id}, role: {role}, company_id: {target_company_id}")

        # ------------------------------------------------------------------
        # 1. 班導 / 教師 (role == "teacher" or "class_teacher")
        # ------------------------------------------------------------------
        if role in ["teacher", "class_teacher"]:
            # 對於指導老師（teacher），只顯示選擇了該老師管理的公司的學生履歷
            # 對於班導（class_teacher），顯示班導的學生履歷
            if role == "teacher":
                # 修改：返回所有投遞履歷記錄
                # 這樣每個投遞記錄都會對應一行履歷資料
                sql_query = """
                    SELECT DISTINCT
                        r.id,
                        u.id AS user_id,
                        u.name AS student_name,
                        u.username AS student_number,
                        c.name AS class_name,
                        c.department,
                        r.original_filename,
                        r.filepath,
                        r.status,
                        r.reviewed_by,
                        r.comment,
                        r.note,
                        r.created_at,
                        pref.company_name AS company_name,
                        pref.job_title AS job_title,
                        pref.application_id AS preference_id,
                        sja.applied_at,
                        pref.application_status AS preference_status,
                        NULL AS vendor_comment
                    FROM resumes r
                    JOIN users u ON r.user_id = u.id
                    LEFT JOIN classes c ON u.class_id = c.id
                    INNER JOIN student_job_applications sja ON sja.resume_id = r.id AND sja.student_id = u.id
                    INNER JOIN (
                        SELECT 
                            sja.student_id,
                            sja.id AS application_id,
                            sja.applied_at,
                            ic.company_name,
                            COALESCE(ij.title, '') AS job_title,
                            ij.id AS job_id,
                            sja.status AS application_status
                        FROM student_job_applications sja
                        JOIN internship_companies ic ON sja.company_id = ic.id
                        LEFT JOIN internship_jobs ij ON sja.job_id = ij.id
                        WHERE ic.advisor_user_id = %s
                        AND sja.status = 'submitted'
                    ) pref ON pref.student_id = u.id AND pref.application_id = sja.id
                    WHERE r.status = 'approved'
                    -- 只顯示班導已通過（status='approved'）的履歷，供指導老師審核
                    -- 只顯示選擇了該指導老師管理的公司的學生履歷
                    AND EXISTS (
                        SELECT 1
                        FROM student_job_applications sja2
                        JOIN internship_companies ic2 ON sja2.company_id = ic2.id
                        WHERE sja2.student_id = u.id 
                            AND ic2.advisor_user_id = %s
                            AND sja2.status = 'submitted'
                    )
                    ORDER BY sja.applied_at DESC
                """
                sql_params = (user_id, user_id)
            else:
                # class_teacher 角色：顯示班導的學生履歷（返回所有志願序）
                sql_query = """
                    SELECT DISTINCT
                        r.id,
                        u.id AS user_id,
                        u.name AS student_name,
                        u.username AS student_number,
                        c.name AS class_name,
                        c.department,
                        r.original_filename,
                        r.filepath,
                        r.status,
                        r.comment,
                        r.note,
                        r.created_at,
                        pref.company_name AS company_name,
                        pref.job_title AS job_title,
                        pref.application_id AS preference_id,
                        sja.applied_at,
                        pref.application_status AS preference_status,
                        NULL AS vendor_comment
                    FROM resumes r
                    JOIN users u ON r.user_id = u.id
                    LEFT JOIN classes c ON u.class_id = c.id
                    INNER JOIN student_job_applications sja ON sja.resume_id = r.id AND sja.student_id = u.id
                    LEFT JOIN (
                        SELECT 
                            sja.student_id,
                            sja.id AS application_id,
                            sja.applied_at,
                            ic.company_name,
                            COALESCE(ij.title, '') AS job_title,
                            ij.id AS job_id,
                            sja.status AS application_status
                        FROM student_job_applications sja
                        JOIN internship_companies ic ON sja.company_id = ic.id
                        LEFT JOIN internship_jobs ij ON sja.job_id = ij.id
                        WHERE sja.status = 'submitted'
                    ) pref ON pref.student_id = u.id AND pref.application_id = sja.id
                    WHERE r.status IN ('uploaded', 'approved')
                    AND r.status != 'rejected'
                    AND sja.resume_id IS NOT NULL
                    -- 確保只選擇每個 application_id 對應的最新履歷
                    AND r.created_at = (
                        SELECT MAX(r2.created_at)
                        FROM resumes r2
                        INNER JOIN student_job_applications sja2 ON sja2.resume_id = r2.id
                        WHERE sja2.id = sja.id
                        AND sja2.resume_id IS NOT NULL
                        AND r2.status IN ('uploaded', 'approved')
                        AND r2.status != 'rejected'
                    )
                    AND EXISTS (
                        SELECT 1
                        FROM classes c2
                        JOIN classes_teacher ct ON ct.class_id = c2.id
                        WHERE c2.id = u.class_id AND ct.teacher_id = %s
                    )
                    ORDER BY sja.applied_at DESC
                """
                sql_params = (user_id,)
            
            # 如果指定了 company_id，添加額外的篩選條件
            if target_company_id:
                # 在 WHERE 子句結束前添加 company_id 篩選
                sql_query = sql_query.replace(
                    "ORDER BY pref.preference_order ASC",
                    "AND pref.application_id IN (SELECT id FROM student_job_applications WHERE company_id = %s) ORDER BY sja.applied_at DESC"
                )
                sql_params = sql_params + (target_company_id,)

            cursor.execute(sql_query, sql_params)
            resumes = cursor.fetchall()

            if resumes:
                print(f"✅ [DEBUG] Teacher/class_teacher user {user_id} found {len(resumes)} resumes")
            else:
                print(f"⚠️ [DEBUG] Teacher/class_teacher user {user_id} has no assigned classes or advisor students.")

        # ------------------------------------------------------------------
        # 2. 主任 (role == "director")
        # ------------------------------------------------------------------
        elif role == "director":
            if mode == "director":
                department = get_director_department(cursor, user_id)
                if not department:
                    resumes = []
                else:
                    sql_query = """
                        SELECT 
                            r.id,
                            u.name AS student_name,
                            u.username AS student_number,
                            c.name AS class_name,
                            c.department,
                            r.original_filename,
                            r.filepath,
                            r.status,
                            r.comment,
                            r.note,
                            r.created_at
                        FROM resumes r
                        JOIN users u ON r.user_id = u.id
                        JOIN classes c ON u.class_id = c.id
                        WHERE c.department = %s
                        AND r.status != 'rejected'
                        ORDER BY c.name, u.name
                    """
                    sql_params = (department,)
                    cursor.execute(sql_query, sql_params)
                    resumes = cursor.fetchall()
            else:
                sql_query = """
                    SELECT 
                        r.id,
                        u.name AS student_name,
                        u.username AS student_number,
                        c.name AS class_name,
                        c.department,
                        r.original_filename,
                        r.filepath,
                        r.status,
                        r.comment,
                        r.note,
                        r.created_at
                    FROM resumes r
                    JOIN users u ON r.user_id = u.id
                    LEFT JOIN classes c ON u.class_id = c.id
                    JOIN classes_teacher ct ON ct.class_id = c.id
                    WHERE ct.teacher_id = %s
                    ORDER BY c.name, u.name
                """
                sql_params = (user_id,)
                cursor.execute(sql_query, sql_params)
                resumes = cursor.fetchall()

        # ------------------------------------------------------------------
        # 3. TA 或 Admin (role == "ta" or "admin")
        # ------------------------------------------------------------------
        elif role in ["ta", "admin"]:
            sql_query = """
                SELECT 
                    r.id,
                    u.name AS student_name,
                    u.username AS student_number,
                    c.name AS class_name,
                    c.department,
                    r.original_filename,
                    r.filepath,
                    r.status,
                    r.comment,
                    r.note,
                    r.created_at
                FROM resumes r
                JOIN users u ON r.user_id = u.id
                LEFT JOIN classes c ON u.class_id = c.id
                ORDER BY c.name, u.name
            """
            cursor.execute(sql_query, tuple())
            resumes = cursor.fetchall()

        # ------------------------------------------------------------------
        # 4. Vendor (role == "vendor")
        # ------------------------------------------------------------------
        elif role == "vendor":
            sql_query = """
                SELECT DISTINCT
                    r.id,
                    u.name AS student_name,
                    u.username AS student_number,
                    c.name AS class_name,
                    c.department,
                    r.original_filename,
                    r.filepath,
                    r.status,
                    r.comment,
                    r.note,
                    r.created_at
                FROM resumes r
                JOIN users u ON r.user_id = u.id
                LEFT JOIN classes c ON u.class_id = c.id
                WHERE EXISTS (
                    SELECT 1 FROM student_preferences sp
                    JOIN internship_companies ic ON sp.company_id = ic.id
                    WHERE sp.student_id = u.id
                    AND ic.uploaded_by_user_id = %s
                ) OR EXISTS (
                    SELECT 1 FROM internship_experiences ie
                    JOIN internship_companies ic ON ie.company_id = ic.id
                    WHERE ie.user_id = u.id
                    AND ic.uploaded_by_user_id = %s
                )
                ORDER BY c.name, u.name
            """
            cursor.execute(sql_query, (user_id, user_id))
            resumes = cursor.fetchall()

        else:
            return jsonify({"success": False, "message": "無效的角色或權限"}), 403

        # 格式化日期時間並統一字段名稱
        for r in resumes:
            if isinstance(r.get('created_at'), datetime):
                r['created_at'] = r['created_at'].strftime("%Y/%m/%d %H:%M")
            # 統一字段名稱，確保前端能正確訪問
            if 'student_name' in r:
                r['name'] = r['student_name']
            if 'student_number' in r:
                r['username'] = r['student_number']
            if 'class_name' in r:
                r['className'] = r['class_name']
            if 'created_at' in r:
                r['upload_time'] = r['created_at']
            # 處理志願序狀態：統一使用 status 欄位
            # 統一使用 resumes.status 欄位
            status = r.get('status') or 'uploaded'
            if status not in ['uploaded', 'approved', 'rejected']:
                status = 'uploaded'
            r['application_statuses'] = status
            r['display_status'] = status
            if False:  # 移除角色判斷，統一處理
                # 其他角色（class_teacher, director, ta, admin）：使用 status 欄位（班導/主任的審核狀態）
                # status 欄位也使用 enum('uploaded', 'approved', 'rejected')
                if 'preference_status' in r and r.get('preference_status'):
                    r['application_statuses'] = r['preference_status']
                    r['display_status'] = r['preference_status']
                else:
                    # 確保 status 符合 enum('uploaded', 'approved', 'rejected')
                    status_value = r.get('status') or 'uploaded'
                    if status_value not in ['uploaded', 'approved', 'rejected']:
                        status_value = 'uploaded'
                    r['application_statuses'] = status_value
                    r['display_status'] = status_value
            # 使用履歷的 comment

        print(f"✅ [DEBUG] Returning {len(resumes)} resumes for role {role}")
        return jsonify({"success": True, "resumes": resumes})

    except Exception as e:
        print("❌ 取得班級履歷資料錯誤：", traceback.print_exc())
        return jsonify({"success": False, "message": "伺服器錯誤"}), 500

    finally:
        cursor.close()
        conn.close()


@resume_bp.route('/api/get_absence_stats', methods=['GET'])
def get_absence_stats():
    if 'user_id' not in session:
        return jsonify({"success": False, "message": "請先登入"}), 401

    user_id = session['user_id']
    semester_id = request.args.get('semester_id', None)
    start_semester_id = request.args.get('start_semester_id', None)
    end_semester_id = request.args.get('end_semester_id', None)
    
    conn = get_db()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("SHOW COLUMNS FROM absence_records LIKE 'semester_id'")
        has_semester_id = cursor.fetchone() is not None
        
        if has_semester_id:
            if start_semester_id and end_semester_id:
                cursor.execute("""
                    SELECT id FROM semesters 
                    WHERE code >= (SELECT code FROM semesters WHERE id = %s)
                    AND code <= (SELECT code FROM semesters WHERE id = %s)
                    ORDER BY code
                """, (start_semester_id, end_semester_id))
                semester_ids_in_range = [row['id'] for row in cursor.fetchall()]
                if semester_ids_in_range:
                    placeholders = ','.join(['%s'] * len(semester_ids_in_range))
                    cursor.execute(f"""
                        SELECT 
                            ar.absence_type, 
                            SUM(ar.duration_units) AS total_units
                        FROM absence_records ar
                        LEFT JOIN semesters s ON ar.semester_id = s.id
                        WHERE ar.user_id = %s AND ar.semester_id IN ({placeholders})
                        GROUP BY ar.absence_type
                    """, (user_id, *semester_ids_in_range))
                else:
                    cursor.execute("""
                        SELECT 
                            ar.absence_type, 
                            SUM(ar.duration_units) AS total_units
                        FROM absence_records ar
                        WHERE ar.user_id = %s AND 1=0
                        GROUP BY ar.absence_type
                    """, (user_id,))
            elif semester_id:
                cursor.execute("""
                    SELECT 
                        ar.absence_type, 
                        SUM(ar.duration_units) AS total_units
                    FROM absence_records ar
                    LEFT JOIN semesters s ON ar.semester_id = s.id
                    WHERE ar.user_id = %s AND ar.semester_id = %s
                    GROUP BY ar.absence_type
                """, (user_id, semester_id))
            else:
                current_semester_id = get_current_semester_id(cursor)
                if current_semester_id:
                    cursor.execute("""
                        SELECT 
                            ar.absence_type, 
                            SUM(ar.duration_units) AS total_units
                        FROM absence_records ar
                        LEFT JOIN semesters s ON ar.semester_id = s.id
                        WHERE ar.user_id = %s AND ar.semester_id = %s
                        GROUP BY ar.absence_type
                    """, (user_id, current_semester_id))
                else:
                    cursor.execute("""
                        SELECT 
                            ar.absence_type, 
                            SUM(ar.duration_units) AS total_units
                        FROM absence_records ar
                        LEFT JOIN semesters s ON ar.semester_id = s.id
                        WHERE ar.user_id = %s
                        GROUP BY ar.absence_type
                    """, (user_id,))
        else:
            cursor.execute("""
                SELECT 
                    absence_type, 
                    SUM(duration_units) AS total_units 
                FROM absence_records
                WHERE user_id = %s
                GROUP BY absence_type
            """, (user_id,))
        
        results = cursor.fetchall()
        stats = {}
        for row in results:
            stats[row['absence_type']] = int(row['total_units'])

        return jsonify({"success": True, "stats": stats})

    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": f"伺服器錯誤: {str(e)}"}), 500

    finally:
        cursor.close()
        conn.close()

@resume_bp.route('/api/absence/available_semesters', methods=['GET'])
def get_absence_available_semesters():
    """取得缺勤可用的學期列表（根據預設範圍過濾）"""
    if 'user_id' not in session:
        return jsonify({"success": False, "message": "請先登入"}), 401
    
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        user_id = session['user_id']
        cursor.execute("SELECT username FROM users WHERE id = %s", (user_id,))
        user_result = cursor.fetchone()
        
        admission_year = None
        if user_result and user_result.get('username'):
            username = user_result['username']
            if len(username) >= 3:
                try:
                    admission_year = int(username[:3])
                except ValueError:
                    pass
        
        cursor.execute("SHOW TABLES LIKE 'absence_default_semester_range'")
        table_exists = cursor.fetchone() is not None
        
        start_semester_code = None
        end_semester_code = None
        
        if table_exists:
            cursor.execute("SHOW COLUMNS FROM absence_default_semester_range LIKE 'admission_year'")
            has_admission_year = cursor.fetchone() is not None
            
            if has_admission_year and admission_year:
                cursor.execute("""
                    SELECT start_semester_code, end_semester_code
                    FROM absence_default_semester_range
                    WHERE admission_year = %s
                    ORDER BY id DESC
                    LIMIT 1
                """, (admission_year,))
            else:
                cursor.execute("""
                    SELECT start_semester_code, end_semester_code
                    FROM absence_default_semester_range
                    ORDER BY id DESC
                    LIMIT 1
                """)
            
            range_result = cursor.fetchone()
            if range_result:
                start_semester_code = range_result.get('start_semester_code')
                end_semester_code = range_result.get('end_semester_code')
        
        if start_semester_code and end_semester_code:
            cursor.execute("""
                SELECT id, code, start_date, end_date, is_active
                FROM semesters
                WHERE code >= %s AND code <= %s
                ORDER BY code ASC
            """, (start_semester_code, end_semester_code))
        else:
            cursor.execute("""
                SELECT id, code, start_date, end_date, is_active
                FROM semesters
                ORDER BY code DESC
            """)
        
        semesters = cursor.fetchall()
        
        for s in semesters:
            if isinstance(s.get('start_date'), datetime):
                s['start_date'] = s['start_date'].strftime("%Y-%m-%d")
            if isinstance(s.get('end_date'), datetime):
                s['end_date'] = s['end_date'].strftime("%Y-%m-%d")
        
        return jsonify({
            "success": True,
            "semesters": semesters
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": f"取得學期列表失敗: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

@resume_bp.route('/api/get_absence_default_range', methods=['GET'])
def get_absence_default_range():
    """取得缺勤預設學期範圍"""
    if 'user_id' not in session:
        return jsonify({"success": False, "message": "請先登入"}), 401
    
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        cursor.execute("SHOW TABLES LIKE 'absence_default_semester_range'")
        table_exists = cursor.fetchone() is not None
        
        if not table_exists:
            return jsonify({
                "success": True,
                "defaultStart": "",
                "defaultEnd": ""
            })
        
        user_id = session['user_id']
        cursor.execute("SELECT username FROM users WHERE id = %s", (user_id,))
        user_result = cursor.fetchone()
        
        admission_year = None
        if user_result and user_result.get('username'):
            username = user_result['username']
            if len(username) >= 3:
                try:
                    admission_year = int(username[:3])
                except ValueError:
                    pass
        
        cursor.execute("SHOW COLUMNS FROM absence_default_semester_range LIKE 'admission_year'")
        has_admission_year = cursor.fetchone() is not None
        
        if has_admission_year and admission_year:
            cursor.execute("""
                SELECT start_semester_code, end_semester_code
                FROM absence_default_semester_range
                WHERE admission_year = %s
                ORDER BY id DESC
                LIMIT 1
            """, (admission_year,))
        else:
            cursor.execute("""
                SELECT start_semester_code, end_semester_code
                FROM absence_default_semester_range
                ORDER BY id DESC
                LIMIT 1
            """)
        
        result = cursor.fetchone()
        
        if result:
            return jsonify({
                "success": True,
                "defaultStart": result.get('start_semester_code', ''),
                "defaultEnd": result.get('end_semester_code', '')
            })
        else:
            return jsonify({
                "success": True,
                "defaultStart": "",
                "defaultEnd": ""
            })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": f"取得預設學期範圍失敗: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

@resume_bp.route('/api/get_semester_absence_records', methods=['GET'])
def get_semester_absence_records():
    """獲取學生的學期出勤記錄，用於自動填充表單"""
    if 'user_id' not in session:
        return jsonify({"success": False, "message": "請先登入"}), 401

    user_id = session['user_id']
    semester_id = request.args.get('semester_id', None)
    start_semester_id = request.args.get('start_semester_id', None)
    end_semester_id = request.args.get('end_semester_id', None)
    start_date = request.args.get('start_date', None)
    end_date = request.args.get('end_date', None)
    
    conn = get_db()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("SHOW COLUMNS FROM absence_records LIKE 'semester_id'")
        has_semester_id = cursor.fetchone() is not None
        
        where_conditions = ["ar.user_id = %s"]
        query_params = [user_id]
        
        if has_semester_id:
            if start_semester_id and end_semester_id:
                cursor.execute("""
                    SELECT id FROM semesters 
                    WHERE code >= (SELECT code FROM semesters WHERE id = %s)
                    AND code <= (SELECT code FROM semesters WHERE id = %s)
                    ORDER BY code
                """, (start_semester_id, end_semester_id))
                semester_ids_in_range = [row['id'] for row in cursor.fetchall()]
                if semester_ids_in_range:
                    placeholders = ','.join(['%s'] * len(semester_ids_in_range))
                    where_conditions.append(f"ar.semester_id IN ({placeholders})")
                    query_params.extend(semester_ids_in_range)
            elif semester_id:
                where_conditions.append("ar.semester_id = %s")
                query_params.append(semester_id)
        
        if not (start_semester_id and end_semester_id):
            if start_date:
                where_conditions.append("ar.absence_date >= %s")
                query_params.append(start_date)
            if end_date:
                where_conditions.append("ar.absence_date <= %s")
                query_params.append(end_date)
        
        where_clause = " AND ".join(where_conditions)
        
        if has_semester_id:
            query = f"""
                SELECT 
                    ar.id,
                    ar.absence_date,
                    ar.absence_type,
                    ar.duration_units,
                    ar.reason,
                    ar.image_path,
                    ar.created_at,
                    s.code AS semester_code,
                    s.id AS semester_id,
                    u.username AS student_id,
                    u.name AS student_name
                FROM absence_records ar
                LEFT JOIN semesters s ON ar.semester_id = s.id
                LEFT JOIN users u ON ar.user_id = u.id
                WHERE {where_clause}
                ORDER BY ar.absence_date DESC, ar.created_at DESC
            """
            cursor.execute(query, tuple(query_params))
        else:
            query = f"""
                SELECT 
                    ar.id,
                    ar.absence_date,
                    ar.absence_type,
                    ar.duration_units,
                    ar.reason,
                    ar.image_path,
                    ar.created_at,
                    NULL AS semester_code,
                    NULL AS semester_id,
                    u.username AS student_id,
                    u.name AS student_name
                FROM absence_records ar
                LEFT JOIN users u ON ar.user_id = u.id
                WHERE {where_clause}
                ORDER BY ar.absence_date DESC, ar.created_at DESC
            """
            cursor.execute(query, tuple(query_params))
        
        records = cursor.fetchall()
        
        for record in records:
            if record.get('absence_date'):
                absence_date = record['absence_date']
                if isinstance(absence_date, datetime):
                    record['absence_date'] = absence_date.strftime("%Y-%m-%d")
                elif isinstance(absence_date, str):
                    date_match = re.search(r'(\d{4})-(\d{2})-(\d{2})', absence_date)
                    if date_match:
                        record['absence_date'] = date_match.group(0)
                    elif 'T' in absence_date:
                        record['absence_date'] = absence_date.split('T')[0]
        
        return jsonify({"success": True, "records": records})

    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": f"取得缺勤記錄失敗: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

@resume_bp.route('/api/get_absence_records', methods=['GET'])
def get_absence_records():
    if 'user_id' not in session:
        return jsonify({"success": False, "message": "請先登入"}), 401

    user_id = session['user_id']
    conn = get_db()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("SHOW COLUMNS FROM absence_records LIKE 'semester_id'")
        has_semester_id = cursor.fetchone() is not None

        if has_semester_id:
            cursor.execute("""
                SELECT 
                    ar.id,
                    ar.absence_date,
                    ar.absence_type,
                    ar.duration_units,
                    ar.reason,
                    ar.image_path,
                    ar.created_at,
                    s.code AS semester_code,
                    s.id AS semester_id
                FROM absence_records ar
                LEFT JOIN semesters s ON ar.semester_id = s.id
                WHERE ar.user_id = %s
                ORDER BY ar.absence_date DESC, ar.created_at DESC
            """, (user_id,))
        else:
            cursor.execute("""
                SELECT 
                    ar.id,
                    ar.absence_date,
                    ar.absence_type,
                    ar.duration_units,
                    ar.reason,
                    ar.image_path,
                    ar.created_at,
                    NULL AS semester_code,
                    NULL AS semester_id
                FROM absence_records ar
                WHERE ar.user_id = %s
                ORDER BY ar.absence_date DESC, ar.created_at DESC
            """, (user_id,))

        records = cursor.fetchall()

        # 格式化日期
        for record in records:
            if record.get('absence_date'):
                absence_date = record['absence_date']
                if isinstance(absence_date, datetime):
                    record['absence_date'] = absence_date.strftime("%Y-%m-%d")
                elif isinstance(absence_date, str):
                    date_match = re.search(r'(\d{4})-(\d{2})-(\d{2})', absence_date)
                    if date_match:
                        record['absence_date'] = date_match.group(0)
                    elif 'T' in absence_date:
                        record['absence_date'] = absence_date.split('T')[0]

        return jsonify({"success": True, "records": records})

    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": f"取得缺勤記錄失敗: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

@resume_bp.route('/api/submit_absence_record', methods=['POST'])
def submit_absence_record():
    if 'user_id' not in session:
        return jsonify({"success": False, "message": "請先登入"}), 401

    user_id = session['user_id']
    conn = get_db()
    cursor = conn.cursor(dictionary=True)

    try:
        absence_date = request.form.get('absence_date')
        absence_type = request.form.get('absence_type')
        duration_units = request.form.get('duration_units')
        reason = request.form.get('reason')
        proof_image = request.files.get('proof_image')

        if not all([absence_date, absence_type, duration_units, reason]):
            return jsonify({"success": False, "message": "請填寫所有必填欄位"}), 400

        duration_units_int = int(duration_units)
        if duration_units_int <= 0:
            return jsonify({"success": False, "message": "節數必須為正整數"}), 400

        # 獲取當前學期ID
        semester_id = get_current_semester_id(cursor)

        # 處理佐證圖片
        image_path = None
        if proof_image and proof_image.filename:
            filename = secure_filename(proof_image.filename)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{user_id}_{timestamp}_{filename}"
            filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
            proof_image.save(filepath)
            image_path = f"/uploads/{filename}"

        # 檢查是否有 semester_id 欄位
        cursor.execute("SHOW COLUMNS FROM absence_records LIKE 'semester_id'")
        has_semester_id = cursor.fetchone() is not None

        if has_semester_id and semester_id:
            cursor.execute("""
                INSERT INTO absence_records 
                (user_id, absence_date, absence_type, duration_units, reason, image_path, semester_id)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (user_id, absence_date, absence_type, duration_units_int, reason, image_path, semester_id))
        else:
            cursor.execute("""
                INSERT INTO absence_records 
                (user_id, absence_date, absence_type, duration_units, reason, image_path)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (user_id, absence_date, absence_type, duration_units_int, reason, image_path))

        conn.commit()
        return jsonify({"success": True, "message": "缺勤記錄已保存"})

    except Exception as e:
        conn.rollback()
        traceback.print_exc()
        return jsonify({"success": False, "message": f"保存失敗: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

@resume_bp.route('/upload_resume')
def upload_resume_page():
    return render_template('resume/upload_resume.html')  


@resume_bp.route('/ai_edit_resume')
def ai_edit_resume_page():
    return render_template('resume/ai_edit_resume.html')

from flask import Blueprint, request, jsonify, session, send_file, render_template, redirect, current_app,send_from_directory
from werkzeug.utils import secure_filename
from config import get_db
from semester import get_current_semester_id
from docxtpl import DocxTemplate, InlineImage
from docx.shared import Inches
import os
import traceback
import json
import re
from datetime import datetime, date
from urllib.parse import quote
from notification import create_notification
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import io

# resume_bp 已在文件開頭定義，不需要重複定義

@resume_bp.route('/api/resume_status', methods=['GET'])
def resume_status():
    resume_id = request.args.get('resume_id')
    if not resume_id:
        return jsonify({"success": False, "message": "缺少 resume_id"}), 400

    try:
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT status FROM resumes WHERE id = %s", (resume_id,))
        resume = cursor.fetchone()
        cursor.close()
        conn.close()

        if not resume:
            return jsonify({"success": False, "message": "找不到該履歷"}), 404

        return jsonify({"success": True, "status": resume['status']})

    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": f"伺服器錯誤: {str(e)}"}), 500

@resume_bp.route("/api/my_resumes", methods=["GET"])
def get_my_resumes():
    if not require_login(): 
        return jsonify({"success": False, "message": "未登入"}), 401
    user_id = session['user_id']
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 獲取所有履歷，包含分類資訊
        cursor.execute("""
            SELECT id, original_filename, status, category, created_at, updated_at, comment, note
            FROM resumes
            WHERE user_id = %s
            ORDER BY created_at DESC
        """, (user_id,))
        resumes = cursor.fetchall()
        
        # 格式化日期
        from datetime import datetime
        for r in resumes:
            if isinstance(r.get('created_at'), datetime):
                r['created_at'] = r['created_at'].strftime("%Y-%m-%d %H:%M:%S")
            if isinstance(r.get('updated_at'), datetime):
                r['updated_at'] = r['updated_at'].strftime("%Y-%m-%d %H:%M:%S")
            # 確保有 category 欄位，預設為 draft
            if not r.get('category'):
                r['category'] = 'draft'
        
        return jsonify({"success": True, "resumes": resumes})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500
    finally:
        cursor.close()
        conn.close()

@resume_bp.route("/api/resumes/<int:resume_id>/filename", methods=["PUT"])
def update_resume_filename(resume_id):
    if not require_login(): 
        return jsonify({"success": False, "message": "未登入"}), 401
    
    user_id = session['user_id']
    data = request.get_json()
    new_filename = data.get('filename', '').strip()
    
    if not new_filename:
        return jsonify({"success": False, "message": "文件名不能為空"}), 400
    
    # 內建預設為 .docx：若使用者只改檔名沒加副檔名，自動補上，避免刪掉 .docx 導致無法下載/開啟
    if not (new_filename.lower().endswith('.docx') or new_filename.lower().endswith('.doc')):
        new_filename = new_filename.rstrip('.') + '.docx'
    
    # 驗證文件名長度
    if len(new_filename) > 255:
        return jsonify({"success": False, "message": "文件名過長（最多255個字符）"}), 400
    
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 驗證履歷屬於當前用戶
        cursor.execute("""
            SELECT id, original_filename FROM resumes 
            WHERE id = %s AND user_id = %s
        """, (resume_id, user_id))
        resume = cursor.fetchone()
        
        if not resume:
            return jsonify({"success": False, "message": "履歷不存在或無權限"}), 403
        
        # 更新文件名
        cursor.execute("""
            UPDATE resumes 
            SET original_filename = %s,
                updated_at = NOW()
            WHERE id = %s AND user_id = %s
        """, (new_filename, resume_id, user_id))
        
        conn.commit()
        return jsonify({"success": True, "message": "文件名已更新"})
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "message": str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@resume_bp.route("/api/resumes/<int:resume_id>/category", methods=["PUT"])
def update_resume_category(resume_id):
    if not require_login(): 
        return jsonify({"success": False, "message": "未登入"}), 401
    
    user_id = session['user_id']
    data = request.json
    new_category = data.get('category')
    
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 驗證履歷屬於當前用戶
        cursor.execute("""
            SELECT id, category, status, note FROM resumes 
            WHERE id = %s AND user_id = %s
        """, (resume_id, user_id))
        resume = cursor.fetchone()
        
        if not resume:
            return jsonify({"success": False, "message": "履歷不存在或無權限"}), 403
        
        resume_category = resume.get('category', 'draft')
        resume_status = resume.get('status', 'uploaded')
        resume_note = resume.get('note') or ''
        
        # 驗證分類值
        valid_categories = ['draft', 'ready']
        if new_category not in valid_categories:
            return jsonify({"success": False, "message": "無效的分類值"}), 400
        
        # 如果選擇「正式版本」（ready），執行提交邏輯
        if new_category == 'ready':
            # 只允許提交草稿（category='draft'）
            if resume_category != 'draft':
                return jsonify({
                    "success": False, 
                    "message": "只能將草稿改為正式版本。此履歷可能已經是正式版本。"
                }), 403
            
            # 將履歷從草稿改為正式版本（可以投遞）
            cursor.execute("""
                UPDATE resumes 
                SET category = 'ready',
                    status = 'uploaded',
                    updated_at = NOW()
                WHERE id = %s AND user_id = %s 
                AND category = 'draft'
            """, (resume_id, user_id))
            
            if cursor.rowcount == 0:
                return jsonify({"success": False, "message": "提交失敗，履歷狀態可能已改變"}), 400
            
            conn.commit()
            return jsonify({"success": True, "message": "履歷已提交為正式版本，可以投遞"})
        
        # 如果選擇「草稿」（draft），只允許從正式版本改回草稿（且必須是未投遞的）
        if new_category == 'draft':
            # 只允許將正式版本改回草稿，且必須是未投遞的（status='uploaded' 且沒有審核記錄）
            if resume_category != 'ready':
                return jsonify({
                    "success": False, 
                    "message": "只能將正式版本改回草稿。"
                }), 403
            
            # 只有未投遞的正式版本才能改回草稿（status='uploaded' 且沒有 reviewed_by）
            cursor.execute("""
                UPDATE resumes 
                SET category = 'draft',
                    updated_at = NOW()
                WHERE id = %s AND user_id = %s 
                AND category = 'ready' 
                AND status = 'uploaded'
                AND reviewed_by IS NULL
            """, (resume_id, user_id))
            
            if cursor.rowcount == 0:
                return jsonify({
                    "success": False, 
                    "message": "無法改回草稿。此履歷可能已經投遞或正在審核中。"
                }), 400
            
            conn.commit()
            return jsonify({"success": True, "message": "履歷已改回草稿"})
        conn.commit()
        return jsonify({"success": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "message": str(e)}), 500
    finally:
        cursor.close()
        conn.close()

@resume_bp.route("/api/resumes/<int:resume_id>/submit", methods=["POST"])
def submit_resume(resume_id):
    if not require_login(): 
        return jsonify({"success": False, "message": "未登入"}), 401
    
    user_id = session['user_id']
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 驗證履歷屬於當前用戶，且為真正的草稿（未提交）
        cursor.execute("""
            SELECT id, category, status, note FROM resumes 
            WHERE id = %s AND user_id = %s
        """, (resume_id, user_id))
        resume = cursor.fetchone()
        
        if not resume:
            return jsonify({"success": False, "message": "履歷不存在或無權限"}), 403
        
        # 只允許提交草稿（category='draft'）
        resume_category = resume.get('category', 'draft')
        
        if resume_category != 'draft':
            return jsonify({
                "success": False, 
                "message": "只能提交草稿履歷。此履歷可能已經是正式版本。"
            }), 403
        
        # 將履歷從草稿改為正式版本（可以投遞）
        cursor.execute("""
            UPDATE resumes 
            SET category = 'ready',
                status = 'uploaded',
                updated_at = NOW()
            WHERE id = %s AND user_id = %s 
            AND category = 'draft'
        """, (resume_id, user_id))
        
        if cursor.rowcount == 0:
            return jsonify({"success": False, "message": "提交失敗，履歷狀態可能已改變"}), 400
        
        conn.commit()
        return jsonify({"success": True, "message": "履歷已提交為正式版本，可以投遞"})
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "message": str(e)}), 500
    finally:
        cursor.close()
        conn.close()

@resume_bp.route("/api/resumes/<int:resume_id>", methods=["DELETE"])
def delete_resume(resume_id):
    if not require_login(): 
        return jsonify({"success": False, "message": "未登入"}), 401
    
    user_id = session['user_id']
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 驗證履歷屬於當前用戶，且為真正的草稿（未提交）
        cursor.execute("""
            SELECT id, category, status, filepath FROM resumes 
            WHERE id = %s AND user_id = %s
        """, (resume_id, user_id))
        resume = cursor.fetchone()
        
        if not resume:
            return jsonify({"success": False, "message": "履歷不存在或無權限"}), 403
        
        # 只允許刪除草稿（category='draft'）
        resume_category = resume.get('category', 'draft')
        
        if resume_category != 'draft':
            return jsonify({
                "success": False, 
                "message": "只能刪除草稿履歷。已提交為正式版本的履歷無法刪除。"
            }), 403
        
        # 刪除檔案（如果存在）
        import os
        filepath = resume.get('filepath')
        if filepath:
            try:
                full_path = os.path.join(current_app.config['UPLOAD_FOLDER'], filepath)
                if os.path.exists(full_path):
                    os.remove(full_path)
            except Exception as e:
                print(f"刪除檔案失敗: {e}")
        
        # 刪除資料庫記錄（只刪除草稿）
        cursor.execute("""
            DELETE FROM resumes 
            WHERE id = %s AND user_id = %s 
            AND category = 'draft'
        """, (resume_id, user_id))
        
        if cursor.rowcount == 0:
            return jsonify({"success": False, "message": "刪除失敗，履歷狀態可能已改變"}), 400
        
        conn.commit()
        return jsonify({"success": True, "message": "履歷已刪除"})
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "message": str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@resume_bp.route("/api/resumes/<int:resume_id>/duplicate", methods=["POST"])
def duplicate_resume(resume_id):
    if not require_login():
        return jsonify({"success": False, "message": "未登入"}), 401

    user_id = session['user_id']
    conn = get_db()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("""
            SELECT id, user_id, filepath, original_filename, status, category, semester_id
            FROM resumes
            WHERE id = %s AND user_id = %s
        """, (resume_id, user_id))
        src = cursor.fetchone()
        if not src:
            return jsonify({"success": False, "message": "履歷不存在或無權限"}), 403

        src_path = src.get('filepath')
        if not src_path:
            return jsonify({"success": False, "message": "原履歷無檔案路徑，無法複製"}), 400

        base_dir = BASE_UPLOAD_DIR
        full_src = os.path.normpath(os.path.join(base_dir, src_path.replace("\\", "/")))
        if not os.path.exists(full_src):
            return jsonify({"success": False, "message": "伺服器上找不到原履歷檔案"}), 400

        # 新檔名：原檔名（去掉副檔名）+ _副本_時間戳.docx
        base_name = (src.get('original_filename') or '履歷').rsplit('.', 1)[0]
        ext = (src.get('original_filename') or '').rsplit('.', 1)[-1].lower() or 'docx'
        if ext not in ('docx', 'doc'):
            ext = 'docx'
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        new_filename = f"{base_name}_副本_{ts}.{ext}"
        # 新檔案放在同一目錄，檔名加時間戳避免覆蓋
        src_dir = os.path.dirname(full_src)
        new_full_path = os.path.join(src_dir, new_filename)
        new_relative = os.path.relpath(new_full_path, base_dir).replace("\\", "/")

        import shutil
        shutil.copy2(full_src, new_full_path)

        semester_id = src.get('semester_id')
        cursor.execute("""
            INSERT INTO resumes
            (user_id, filepath, original_filename, status, category, semester_id, created_at)
            VALUES (%s, %s, %s, %s, %s, %s, NOW())
        """, (
            user_id,
            new_relative,
            new_filename,
            'uploaded',
            'draft',
            semester_id
        ))
        new_resume_id = cursor.lastrowid
        if not new_resume_id:
            conn.rollback()
            return jsonify({"success": False, "message": "建立副本記錄失敗"}), 500

        # 複製 resume_content_mapping 與關聯表（若有）
        cursor.execute(
            "SELECT id, stu_info_id FROM resume_content_mapping WHERE resume_id = %s LIMIT 1",
            (resume_id,),
        )
        old_mapping = cursor.fetchone()
        if old_mapping:
            old_mapping_id = old_mapping["id"]
            cursor.execute(
                "INSERT INTO resume_content_mapping (resume_id, stu_info_id) VALUES (%s, %s)",
                (new_resume_id, old_mapping["stu_info_id"]),
            )
            new_mapping_id = cursor.lastrowid
            cursor.execute(
                "INSERT INTO resume_grade_rel (mapping_id, grade_id) SELECT %s, grade_id FROM resume_grade_rel WHERE mapping_id = %s",
                (new_mapping_id, old_mapping_id),
            )
            cursor.execute(
                "INSERT INTO resume_cert_rel (mapping_id, cert_id) SELECT %s, cert_id FROM resume_cert_rel WHERE mapping_id = %s",
                (new_mapping_id, old_mapping_id),
            )
            cursor.execute(
                "INSERT INTO resume_lang_rel (mapping_id, lang_skill_id) SELECT %s, lang_skill_id FROM resume_lang_rel WHERE mapping_id = %s",
                (new_mapping_id, old_mapping_id),
            )
            cursor.execute(
                "INSERT INTO resume_absence_rel (mapping_id, absence_id) SELECT %s, absence_id FROM resume_absence_rel WHERE mapping_id = %s",
                (new_mapping_id, old_mapping_id),
            )

        conn.commit()
        return jsonify({
            "success": True,
            "message": "已建立履歷副本",
            "resume_id": new_resume_id,
            "original_filename": new_filename,
        })
    except Exception as e:
        if conn:
            conn.rollback()
        print(f"❌ 複製履歷失敗: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@resume_bp.route("/resume_folders")
def resume_folders_page():
    return render_template("resume/resume_folders.html")

@resume_bp.route('/api/download_course_template', methods=['GET'])
def download_course_template():
    """學生下載已修習專業核心科目Excel範本"""
    if 'user_id' not in session or session.get('role') != 'student':
        return jsonify({"success": False, "message": "未授權"}), 403
    
    try:
        template_file_name = "已修習專業核心科目範本.xlsx"
        backend_dir = current_app.root_path
        project_root = os.path.dirname(backend_dir)
        file_path = os.path.join(project_root, 'frontend', 'static', 'examples', template_file_name)
        
        if not os.path.exists(file_path):
            return jsonify({"success": False, "message": "找不到範本檔案"}), 404
        
        return send_file(
            file_path,
            as_attachment=True,
            download_name=template_file_name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": "下載範本失敗"}), 500

@resume_bp.route('/api/student/upload_course_excel', methods=['POST'])
def student_upload_course_excel():
    """學生上傳已修習專業核心科目Excel，根據是否有成績自動設置狀態"""
    if 'user_id' not in session or session.get('role') != 'student':
        return jsonify({"success": False, "message": "未授權"}), 403
    
    if 'file' not in request.files:
        return jsonify({"success": False, "message": "缺少文件"}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({"success": False, "message": "未選擇文件"}), 400
    
    allowed_extensions = {'xlsx', 'xls'}
    if '.' not in file.filename or file.filename.rsplit('.', 1)[1].lower() not in allowed_extensions:
        return jsonify({"success": False, "message": "不支援的文件類型"}), 400
    
    file_stream = io.BytesIO(file.read())
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 1. 載入工作簿（先讀取原始值，再讀取顯示值）
        file_stream.seek(0)
        workbook = load_workbook(file_stream, data_only=False)
        sheet = workbook.active
        
        # 同時讀取顯示值版本（用於獲取被誤解析為日期的分數）
        file_stream.seek(0)
        workbook_display = load_workbook(file_stream, data_only=True)
        sheet_display = workbook_display.active
        
        # 2. 獲取標頭（假設第一行是標頭）
        headers = [cell.value for cell in sheet[1]]
        
        # 找出欄位索引
        course_name_col = None
        credits_col = None
        grade_col = None
        
        for i, header in enumerate(headers):
            if header and ('課程名稱' in str(header) or '科目名稱' in str(header)):
                course_name_col = i + 1
            elif header and '學分' in str(header):
                credits_col = i + 1
            elif header and ('成績' in str(header) or '等第' in str(header)):
                grade_col = i + 1
        
        if not course_name_col or not credits_col:
            return jsonify({"success": False, "message": "Excel 檔案缺少必要的欄位（課程名稱/科目名稱、學分）"}), 400
        
        # 3. 處理數據
        courses = []
        for row_index in range(2, sheet.max_row + 1):
            try:
                course_name = str(sheet.cell(row=row_index, column=course_name_col).value or '').strip()
                credits_cell = sheet.cell(row=row_index, column=credits_col)
                credits_value = credits_cell.value
                credits_display_value = sheet_display.cell(row=row_index, column=credits_col).value
                
                grade_cell = sheet.cell(row=row_index, column=grade_col) if grade_col else None
                grade_value = grade_cell.value if grade_cell else None
                
                if not course_name:
                    continue
                
                # 處理學分數（特殊處理：如果被解析為日期，嘗試恢復為分數格式）
                credits_str = ''
                if credits_value is not None:
                    # 檢查是否為日期類型（可能是 Excel 將 "3/3" 誤解析為日期）
                    if isinstance(credits_value, datetime) or isinstance(credits_display_value, datetime):
                        # 如果是日期類型，檢查是否可能是分數格式（如 3/3 被解析為 3月3日）
                        date_obj = credits_value if isinstance(credits_value, datetime) else credits_display_value
                        month = date_obj.month
                        day = date_obj.day
                        
                        # 檢查是否可能是分數格式：
                        # 1. 月份和日期相同（如 3/3）
                        # 2. 或者月份和日期都在 1-12 範圍內（常見的分數格式）
                        if month == day and 1 <= month <= 12:
                            # 很可能是分數格式，轉換為 "3/3" 格式
                            credits_str = f"{month}/{day}"
                        elif 1 <= month <= 12 and 1 <= day <= 12:
                            # 可能是分數格式，轉換為 "月/日" 格式
                            credits_str = f"{month}/{day}"
                        else:
                            # 不太可能是分數，保持日期格式或轉換為字符串
                            credits_str = f"{month}/{day}"
                    elif isinstance(credits_value, (int, float)):
                        # 數字類型
                        if isinstance(credits_value, float) and credits_value.is_integer():
                            credits_str = str(int(credits_value))
                        else:
                            credits_str = str(credits_value)
                    else:
                        # 字符串類型
                        credits_str = str(credits_value).strip()
                
                # 處理成績
                grade_str = ''
                if grade_value is not None:
                    grade_str = str(grade_value).strip()
                
                # 預設狀態為"已修課"（isNotTaken = false）
                # 只有當明確需要標記為未修課時才設為 true
                # 注意：這裡預設所有課程都是已修課，用戶可以在前端點擊按鈕切換為未修課
                isNotTaken = False
                
                courses.append({
                    'name': course_name,
                    'credits': credits_str,
                    'grade': grade_str,
                    'isNotTaken': isNotTaken
                })
                
            except Exception as row_e:
                print(f"⚠️ 處理 Excel 第 {row_index} 行錯誤: {row_e}")
                continue
        
        if not courses:
            return jsonify({"success": False, "message": "Excel 檔案中未找到有效課程資料"}), 400
        
        return jsonify({
            "success": True,
            "courses": courses,
            "message": f"成功解析 {len(courses)} 門課程"
        })
        
    except Exception as e:
        print("❌ 匯入課程 Excel 錯誤:", e)
        traceback.print_exc()
        return jsonify({"success": False, "message": f"伺服器錯誤: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()


def _normalize_cert_category(category):
    """
    將資料庫 category 正規化為 labor / intl / local / other，
    對應履歷：1.勞動部 / 2.國際證照 / 3.國內證照 / 4.其他證照
    """
    if not category:
        return "other"
    c = str(category).strip().lower()
    if c == "labor":
        return "labor"
    if c == "intl":
        return "intl"
    if c == "local":
        return "local"
    return "other"


def _upsert_resume_content_mapping(cursor, resume_id, stu_id, course_grade_ids=None, certification_ids=None, language_skill_ids=None, absence_record_ids=None):
    """依 resume_id 更新或新增一筆 resume_content_mapping，並以關聯表 resume_*_rel 儲存勾選的 ID。"""
    cursor.execute("SELECT id FROM resume_content_mapping WHERE resume_id = %s LIMIT 1", (resume_id,))
    row = cursor.fetchone()
    if row:
        mapping_id = row["id"]
        cursor.execute(
            "UPDATE resume_content_mapping SET stu_info_id = %s, updated_at = CURRENT_TIMESTAMP WHERE resume_id = %s",
            (stu_id, resume_id),
        )
    else:
        cursor.execute(
            "INSERT INTO resume_content_mapping (resume_id, stu_info_id) VALUES (%s, %s)",
            (resume_id, stu_id),
        )
        mapping_id = cursor.lastrowid

    # 先刪除該 mapping 在四張關聯表中的舊資料，再寫入新勾選的 ID
    cursor.execute("DELETE FROM resume_grade_rel WHERE mapping_id = %s", (mapping_id,))
    cursor.execute("DELETE FROM resume_cert_rel WHERE mapping_id = %s", (mapping_id,))
    cursor.execute("DELETE FROM resume_lang_rel WHERE mapping_id = %s", (mapping_id,))
    cursor.execute("DELETE FROM resume_absence_rel WHERE mapping_id = %s", (mapping_id,))

    def _to_id_list(v):
        if not v:
            return []
        if isinstance(v, (list, tuple)):
            return [int(x) for x in v if x is not None and str(x).strip() != ""]
        return [int(x.strip()) for x in str(v).split(",") if x and x.strip()]

    grade_ids = _to_id_list(course_grade_ids)
    cert_ids = _to_id_list(certification_ids)
    lang_ids = _to_id_list(language_skill_ids)
    abs_ids = _to_id_list(absence_record_ids)

    for gid in grade_ids:
        cursor.execute("INSERT INTO resume_grade_rel (mapping_id, grade_id) VALUES (%s, %s)", (mapping_id, gid))
    for cid in cert_ids:
        cursor.execute("INSERT INTO resume_cert_rel (mapping_id, cert_id) VALUES (%s, %s)", (mapping_id, cid))
    for lid in lang_ids:
        cursor.execute("INSERT INTO resume_lang_rel (mapping_id, lang_skill_id) VALUES (%s, %s)", (mapping_id, lid))
    for aid in abs_ids:
        cursor.execute("INSERT INTO resume_absence_rel (mapping_id, absence_id) VALUES (%s, %s)", (mapping_id, aid))



def _get_resume_content_mapping(cursor, resume_id):
    """依 resume_id 從關聯表讀取勾選的 course_grade_ids, certification_ids, language_skill_ids, absence_record_ids。
    回傳與舊版相容的 dict，鍵值為逗號分隔字串。若無 mapping 則回傳 None。"""
    cursor.execute("SELECT id FROM resume_content_mapping WHERE resume_id = %s LIMIT 1", (resume_id,))
    row = cursor.fetchone()
    if not row:
        return None
    mapping_id = row["id"]

    cursor.execute("SELECT grade_id FROM resume_grade_rel WHERE mapping_id = %s", (mapping_id,))
    grade_ids = ",".join(str(r["grade_id"]) for r in (cursor.fetchall() or []))
    cursor.execute("SELECT cert_id FROM resume_cert_rel WHERE mapping_id = %s", (mapping_id,))
    cert_ids = ",".join(str(r["cert_id"]) for r in (cursor.fetchall() or []))
    cursor.execute("SELECT lang_skill_id FROM resume_lang_rel WHERE mapping_id = %s", (mapping_id,))
    lang_ids = ",".join(str(r["lang_skill_id"]) for r in (cursor.fetchall() or []))
    cursor.execute("SELECT absence_id FROM resume_absence_rel WHERE mapping_id = %s", (mapping_id,))
    abs_ids = ",".join(str(r["absence_id"]) for r in (cursor.fetchall() or []))

    return {
        "course_grade_ids": grade_ids or None,
        "certification_ids": cert_ids or None,
        "language_skill_ids": lang_ids or None,
        "absence_record_ids": abs_ids or None,
    }


def is_valid_image_file(file_path):
    """
    驗證圖片文件是否有效
    返回 True 如果文件是有效的圖片，否則返回 False
    """
    if not file_path or not os.path.exists(file_path):
        return False
    
    try:
        from PIL import Image
        # 嘗試打開並驗證圖片
        with Image.open(file_path) as img:
            img.verify()  # 驗證圖片是否損壞
        
        # verify() 後需要重新打開圖片（因為 verify 會關閉文件）
        with Image.open(file_path) as img:
            # 檢查圖片格式是否被支持
            if img.format not in ['JPEG', 'PNG', 'GIF', 'BMP', 'TIFF']:
                print(f"⚠️ 不支持的圖片格式: {img.format} (路徑: {file_path})")
                return False
        return True
    except ImportError:
        # 如果 PIL 未安裝，跳過驗證（向後兼容）
        print("⚠️ PIL/Pillow 未安裝，跳過圖片驗證")
        return True  # 返回 True 讓程序繼續運行
    except Exception as e:
        print(f"⚠️ 圖片驗證失敗 {file_path}: {e}")
        return False

def safe_create_inline_image(doc, file_path, width, description=""):
    """
    安全地創建 InlineImage 對象，如果失敗則返回 None。
    先用 PIL 與 python-docx Image.from_file 驗證，避免 render 時 UnrecognizedImageError。
    """
    if not file_path or not os.path.exists(file_path):
        return None

    if not is_valid_image_file(file_path):
        print(f"⚠️ {description}圖片無效或損壞，跳過: {file_path}")
        return None

    abs_path = os.path.abspath(file_path)
    try:
        from docx.image.image import Image as DocxImage
        DocxImage.from_file(abs_path)
    except Exception as e:
        print(f"⚠️ {description}圖片格式不被 Word 支援，跳過: {file_path} ({e})")
        return None

    try:
        image_obj = InlineImage(doc, abs_path, width=width)
        return image_obj
    except Exception as e:
        print(f"⚠️ {description}圖片載入錯誤 (路徑: {file_path}): {e}")
        traceback.print_exc()
        return None



def resolve_upload_path(path):
    """
    將相對路徑（如 uploads/absence_proofs/xxx）轉為絕對路徑，
    便於 os.path.exists 與 InlineImage 在產生 Word 時正確找到檔案。
    """
    if not path or not str(path).strip():
        return ""
    path = str(path).replace("\\", "/").strip()
    if os.path.isabs(path) and os.path.exists(path):
        return path
    # 相對於 BASE_UPLOAD_DIR（backend 目錄）
    abs_path = os.path.join(BASE_UPLOAD_DIR, path)
    if os.path.exists(abs_path):
        return os.path.normpath(abs_path)
    if os.path.exists(path):
        return os.path.abspath(path)
    return path

# resume_bp 已在文件開頭定義，不需要重複定義

def format_credits(credits_value):
    """格式化學分數，整數顯示為整數格式"""
    if credits_value is None:
        return ''
    
    # 如果是字符串，嘗試解析
    if isinstance(credits_value, str):
        credits_value = credits_value.strip()
        # 如果包含分數符號（如"2/2"），直接返回
        if '/' in credits_value:
            return credits_value
        # 嘗試轉換為數字
        try:
            num_value = float(credits_value)
            # 如果是整數，返回整數格式
            if num_value.is_integer():
                return str(int(num_value))
            return str(num_value)
        except (ValueError, TypeError):
            # 無法轉換為數字，返回原字符串
            return credits_value
    
    # 如果是數字類型
    if isinstance(credits_value, (int, float)):
        # 如果是整數，返回整數格式
        if isinstance(credits_value, float) and credits_value.is_integer():
            return str(int(credits_value))
        elif isinstance(credits_value, int):
            return str(credits_value)
        else:
            return str(credits_value)
    
    # 其他類型，轉換為字符串
    return str(credits_value)

def score_to_grade(score):
    # 若已經是等第，直接回傳
    if str(score).strip() in ['優', '甲', '乙', '丙', '丁']:
        return str(score).strip()

    # 若是分數才做數字轉換
    try:
        score = int(str(score).strip())
    except (ValueError, TypeError):
        return '丁'

    if score >= 90:
        return '優'
    elif score >= 80:
        return '甲'
    elif score >= 70:
        return '乙'
    elif score >= 60:
        return '丙'
    else:
        return '丁'

def generate_language_marks(level):
    marks = {'Jing': '□', 'Zhong': '□', 'Lue': '□'}
    level_map = {'精通': 'Jing', '中等': 'Zhong', '略懂': 'Lue'}
    level_key = level_map.get(level)
    if level_key in marks:
        marks[level_key] = '■'
    return marks

def get_user_by_username(cursor, username):
    cursor.execute("SELECT * FROM users WHERE username = %s", (username,))
    return cursor.fetchone()


def get_user_by_id(cursor, user_id):
    cursor.execute("SELECT * FROM users WHERE id = %s", (user_id,))
    return cursor.fetchone()


def get_director_department(cursor, user_id):
    """
    取得主任所屬 department（透過 classes_teacher -> classes.department）
    若管理多個班級，只回傳第一個有 department 的值（可擴充回傳 list）
    """
    cursor.execute("""
        SELECT DISTINCT c.department
        FROM classes c
        JOIN classes_teacher ct ON ct.class_id = c.id
        WHERE ct.teacher_id = %s
        LIMIT 1
    """, (user_id,))
    r = cursor.fetchone()
    return r['department'] if r and r.get('department') else None


def teacher_manages_class(cursor, teacher_id, class_id):
    cursor.execute("""
        SELECT 1 FROM classes_teacher
        WHERE teacher_id = %s AND class_id = %s
        LIMIT 1
    """, (teacher_id, class_id))
    return cursor.fetchone() is not None


def can_access_target_resume(cursor, session_user_id, session_role, target_user_id):
    # admin 可以
    if session_role == "admin":
        return True

    # student 只能自己
    if session_role == "student":
        return session_user_id == target_user_id

    # ta 可以讀所有
    if session_role == "ta":
        return True

    # vendor 角色不允許訪問履歷
    if session_role == "vendor":
        return False

    # 取得 target student's class_id
    cursor.execute("SELECT class_id FROM users WHERE id = %s", (target_user_id,))
    u = cursor.fetchone()
    if not u:
        return False
    target_class_id = u.get('class_id')

    if session_role == "teacher":
        return teacher_manages_class(cursor, session_user_id, target_class_id)

    if session_role == "class_teacher":
        return teacher_manages_class(cursor, session_user_id, target_class_id)

    if session_role == "director":
        director_dept = get_director_department(cursor, session_user_id)
        if not director_dept:
            return False
        cursor.execute("SELECT department FROM classes WHERE id = %s", (target_class_id,))
        cd = cursor.fetchone()
        if not cd:
            return False
        return cd.get('department') == director_dept

    return False


def load_student_certifications(cursor, student_id):
    """
    回傳該學生所有證照完整資訊
    """
    sql = """
        SELECT
            CONCAT(COALESCE(cc.job_category, ''), COALESCE(cc.level, '')) AS cert_name,
            cc.category AS cert_category,
            CONCAT(CONCAT(COALESCE(cc.job_category, ''), COALESCE(cc.level, '')), ' (', ca.name, ')') AS full_name,
            sc.CertPath AS cert_path,
            sc.AcquisitionDate AS acquire_date,
            sc.cert_code AS cert_code
        FROM student_certifications sc
        LEFT JOIN certificate_codes cc 
            ON sc.cert_code = cc.id
        LEFT JOIN cert_authorities ca 
            ON cc.authority_id = ca.id
        WHERE sc.StuID = %s
        ORDER BY sc.AcquisitionDate DESC, sc.id ASC
    """
    cursor.execute(sql, (student_id,))
    rows = cursor.fetchall()
    # 轉為 Python dict（cursor.fetchall() 已返回字典，因為使用了 dictionary=True）
    results = []
    for r in rows:
        if r:  # 確保 r 不是 None
            cert_name_from_join = r.get('cert_name', '')
            cert_category_from_join = r.get('cert_category', '')
            # cert_code 現為 certificate_codes.id（整數），JOIN 成功即有 category；NULL 為其他證照
            category = cert_category_from_join if cert_category_from_join else 'other'

            results.append({
                "cert_name": cert_name_from_join or '',
                "category": category,        # labor / intl / local / other
                "full_name": r.get('full_name', '') or '',       # 表格區使用 → 例: 電腦軟體乙級 (勞動部)
                "cert_path": r.get('cert_path', '') or '',       # 圖片路徑
                "acquire_date": r.get('acquire_date', '') or '',    # 日期
            })
    return results


def categorize_certifications(cert_list):
    """
    分類證照 → 放到四種類別（對應履歷 1.勞動部 2.國際證照 3.國內證照 4.其他證照）
    """
    labor = []
    international = []
    local = []
    other = []
    for c in cert_list:
        cert_name = c.get("cert_name") or c.get("CertName") or f"{c.get('job_category', '')}{c.get('level', '')}"
        cert_path = c.get("cert_path") or c.get("CertPath", "")
        acquire_date = c.get("acquire_date") or c.get("AcquisitionDate", "")
        item = {
            "table_name": cert_name,
            "photo_name": cert_name,
            "photo_path": cert_path,
            "date": acquire_date,
        }
        raw = c.get("CertCategory") or c.get("category", "other")
        category = _normalize_cert_category(raw)
        if category == "labor":
            labor.append(item)
        elif category == "intl":
            international.append(item)
        elif category == "local":
            local.append(item)
        else:
            other.append(item)
    return labor, international, local, other


def fill_certificates_to_doc(context, prefix, items, max_count):
    """
    填入 Word 模板（表格區）
    prefix 例如: LaborCerts_  → LaborCerts_1, LaborCerts_2 …
    空欄位與 None 一律顯示空白。
    """
    for i in range(1, max_count + 1):
        if i <= len(items):
            name = (items[i-1].get("table_name") or "").strip()
            if name == "None":
                name = ""
            context[f"{prefix}{i}"] = name
        else:
            context[f"{prefix}{i}"] = ""


def fill_certificate_photos(context, doc, items, start_index, max_count=8):
    """
    圖片區（依順序放，不分類）
    start_index → 從第幾張開始，例如 1、9、17、25
    max_count → 最多填充幾張（實際填充的數量可能少於此值）
    空欄位一律填空白，不顯示 None。
    """
    image_size = Inches(3.0)
    actual_count = min(len(items), max_count)
    
    # 填充實際有的證照
    for idx, item in enumerate(items[:max_count], start=start_index):
        photo_path = item.get("photo_path", "") or ""
        photo_name = (item.get("photo_name", "") or "").strip()
        if photo_name == "None":
            photo_name = ""
        
        if photo_path:
            image_obj = safe_create_inline_image(doc, photo_path, image_size, "證照")
            if image_obj:
                context[f"CertPhotoImages_{idx}"] = image_obj
            else:
                context[f"CertPhotoImages_{idx}"] = ""
        else:
            context[f"CertPhotoImages_{idx}"] = ""
        context[f"CertPhotoName_{idx}"] = photo_name or ""

    # 未使用的欄位：圖片與名稱都設為空白，避免模板顯示 None
    if actual_count < max_count:
        for idx in range(start_index + actual_count, start_index + max_count):
            context[f"CertPhotoImages_{idx}"] = ""
            context[f"CertPhotoName_{idx}"] = ""

def save_structured_data(cursor, student_id, data, semester_id=None, resume_id=None):
    try:
        # -------------------------------------------------------------
        # 1) 儲存 Student_Info（基本資料）
        # -------------------------------------------------------------
        cursor.execute("""
            INSERT INTO Student_Info 
                (StuID, StuName, BirthDate, Gender, Phone, Email, Address, 
                 ConductScore, Autobiography, PhotoPath, UpdatedAt)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())
            ON DUPLICATE KEY UPDATE 
                StuName=VALUES(StuName),
                BirthDate=VALUES(BirthDate),
                Gender=VALUES(Gender),
                Phone=VALUES(Phone),
                Email=VALUES(Email),
                Address=VALUES(Address),
                ConductScore=VALUES(ConductScore),
                Autobiography=VALUES(Autobiography),
                PhotoPath=VALUES(PhotoPath),
                UpdatedAt=NOW()
        """, (
            student_id,
            data.get("name"),
            data.get("birth_date"),
            data.get("gender"),
            data.get("phone"),
            data.get("email"),
            data.get("address"),
            data.get("conduct_score"),
            data.get("autobiography"),
            data.get("photo_path")
        ))

        # -------------------------------------------------------------
        # 2) 儲存 course_grades
        # -------------------------------------------------------------
        cursor.execute("SHOW COLUMNS FROM course_grades LIKE 'SemesterID'")
        has_semester_id = cursor.fetchone() is not None

        if has_semester_id and semester_id:
            cursor.execute(
                "DELETE FROM course_grades WHERE StuID=%s AND IFNULL(SemesterID,'')=%s",
                (student_id, semester_id)
            )
        else:
            cursor.execute("DELETE FROM course_grades WHERE StuID=%s", (student_id,))

        seen_courses = set()
        course_grade_ids_saved = []  # 儲存剛插入的 course_grades.id，供 resume_content_mapping 使用
        for c in data.get("courses", []):
            cname = (c.get("name") or "").strip()
            if not cname:
                continue
            if cname in seen_courses:
                continue
            seen_courses.add(cname)

            if has_semester_id and semester_id:
                cursor.execute("""
                    INSERT INTO course_grades
                        (StuID, CourseName, Credits, Grade, SemesterID, ProofImage)
                    VALUES (%s,%s,%s,%s,%s,%s)
                """, (student_id, cname, c.get("credits"), c.get("grade"), semester_id, c.get("proof_image")))
            else:
                cursor.execute("""
                    INSERT INTO course_grades
                        (StuID, CourseName, Credits, Grade, ProofImage)
                    VALUES (%s,%s,%s,%s,%s)
                """, (student_id, cname, c.get("credits"), c.get("grade"), c.get("proof_image")))
            course_grade_ids_saved.append(cursor.lastrowid)
        
        # -------------------------------------------------------------
        # 3) 儲存 student_certifications
        # -------------------------------------------------------------
        
        # 檢查 student_certifications 表的欄位，以確定要插入哪些數據
        cursor.execute("SHOW COLUMNS FROM student_certifications")
        cert_columns = cursor.fetchall()
        known_columns = {c['Field']: c for c in cert_columns}

        cert_rows = []
        processed_certs = set() # 用於去重 (job_category, level)

        # (3) 處理結構化的證照資料 (structured_certifications)
        struct_certs = data.get("structured_certifications", [])
        print(f"📋 收到 structured_certifications: {len(struct_certs)} 筆")
        for cert in struct_certs:
            row = {"StuID": student_id}
            db_job_category = None
            db_level = None
            db_authority_id = None

            # student_certifications.cert_code 存 certificate_codes.id（整數）
            raw = (cert.get("cert_code") or "").strip() if cert.get("cert_code") is not None else ""
            raw_upper = raw.upper() if isinstance(raw, str) else str(raw)
            cert_code_id = None  # 要寫入 DB 的 certificate_codes.id

            if not raw or raw_upper == 'OTHER':
                # 其他證照：cert_code 存 NULL
                if not cert.get("name"):
                    continue
                cert_code_id = None
                db_job_category = (cert.get("job_category") or "").strip() or None
                db_level = (cert.get("level") or "").strip() or None
            elif raw.isdigit():
                # 數字可能是 id（如 2）或 code（如 14901、11800）。先查 id 是否存在，否則當 code 查
                try:
                    as_id = int(raw)
                    cursor.execute("SELECT id FROM certificate_codes WHERE id = %s LIMIT 1", (as_id,))
                    if cursor.fetchone():
                        cert_code_id = as_id
                    else:
                        level_from_cert = (cert.get("level") or "").strip()
                        if level_from_cert:
                            cursor.execute("""
                                SELECT id, job_category, level, authority_id 
                                FROM certificate_codes 
                                WHERE code = %s AND (level = %s OR level IS NULL)
                                LIMIT 1
                            """, (raw_upper, level_from_cert))
                        else:
                            cursor.execute("""
                                SELECT id, job_category, level, authority_id 
                                FROM certificate_codes 
                                WHERE code = %s
                                LIMIT 1
                            """, (raw_upper,))
                        cert_info = cursor.fetchone()
                        if cert_info:
                            cert_code_id = cert_info.get('id')
                except (ValueError, TypeError) as e:
                    cert_code_id = None
                except Exception as e:
                    print(f"⚠️ 查詢 certificate_codes 失敗: {e}")
                    cert_code_id = None
            else:
                # 前端傳的是證照代碼 code（如 TQC-WORD），查詢對應的 id
                level_from_cert = (cert.get("level") or "").strip()
                try:
                    if level_from_cert:
                        cursor.execute("""
                            SELECT id, job_category, level, authority_id 
                            FROM certificate_codes 
                            WHERE code = %s AND (level = %s OR level IS NULL)
                            LIMIT 1
                        """, (raw_upper, level_from_cert))
                    else:
                        cursor.execute("""
                            SELECT id, job_category, level, authority_id 
                            FROM certificate_codes 
                            WHERE code = %s
                            LIMIT 1
                        """, (raw_upper,))
                    cert_info = cursor.fetchone()
                    if cert_info:
                        cert_code_id = cert_info.get('id')
                except Exception as e:
                    print(f"⚠️ 查詢 certificate_codes 失敗: {e}")

            # 編輯時若表單未帶 cert_code（或為 OTHER/空），保留既有 student_certifications 的 cert_code，避免日文等證照被誤存為「其他」
            existing_cert_id = cert.get("id")
            if (cert_code_id is None or cert_code_id == 0) and existing_cert_id and str(existing_cert_id).strip():
                try:
                    eid = int(existing_cert_id)
                    cursor.execute("SELECT cert_code FROM student_certifications WHERE id = %s AND StuID = %s LIMIT 1", (eid, student_id))
                    existing_row = cursor.fetchone()
                    if existing_row and existing_row.get("cert_code") is not None:
                        cert_code_id = existing_row["cert_code"]
                        print(f"✅ 保留既有證照 cert_code（id={eid}）: {cert_code_id}")
                except (ValueError, TypeError, Exception):
                    pass
            row["cert_code"] = cert_code_id

            # 優先使用前端傳來的 authority_id
            if cert.get("authority_id") and str(cert.get("authority_id")).strip() not in ('', 'OTHER'):
                try:
                    db_authority_id = int(cert.get("authority_id"))
                    print(f"✅ 使用前端傳來的 authority_id: {db_authority_id}")
                except (ValueError, TypeError):
                    pass

            # 若有 cert_code_id，從 certificate_codes 取 job_category / level / authority_id
            if cert_code_id and not db_authority_id:
                try:
                    cursor.execute("""
                        SELECT job_category, level, authority_id 
                        FROM certificate_codes 
                        WHERE id = %s
                        LIMIT 1
                    """, (cert_code_id,))
                    cert_info = cursor.fetchone()
                    if cert_info:
                        db_job_category = (cert_info.get('job_category') or '').strip()
                        db_level = (cert_info.get('level') or '').strip()
                        db_authority_id = cert_info.get('authority_id')
                except Exception as e:
                    print(f"⚠️ 查詢 certificate_codes 失敗: {e}")
            
            # 如果前端傳來了 authority_name（其他證照），且 authority_id 為 OTHER，則設置為 NULL
            if cert.get("authority_id") == 'OTHER' or (cert.get("authority_name") and not db_authority_id):
                db_authority_id = None

            # 證照名稱：優先使用前端傳來的 name（確保不因 DB 查不到而跳過），否則用 DB 職類+級別
            cert_name = (cert.get("name") or "").strip()
            if not cert_name and db_job_category and db_level:
                cert_name = f"{db_job_category}{db_level}"

            if not cert_name:
                print(f"⚠️ 忽略無名稱證照記錄: {cert}")
                continue

            # 檢查是否重複（使用 job_category, level 作為唯一標識）
            if db_job_category and db_level:
                cert_identifier = (db_job_category, db_level)
                if cert_identifier in processed_certs:
                    print(f"⚠️ 跳過重複的結構化證照記錄: cert_code_id={cert_code_id}")
                    continue
                processed_certs.add(cert_identifier)

            # 填入欄位
            if "CertName" in known_columns:
                row["CertName"] = cert_name
            if "job_category" in known_columns:
                # 優先使用前端傳來的 job_category（如果是其他證照），否則使用資料庫查到的
                row["job_category"] = cert.get("job_category") if cert.get("job_category") else (db_job_category if db_job_category else None)
            if "level" in known_columns:
                # 優先使用前端傳來的 level（如果是其他證照），否則使用資料庫查到的
                row["level"] = cert.get("level") if cert.get("level") else (db_level if db_level else None)
            if "authority_id" in known_columns:
                row["authority_id"] = int(db_authority_id) if db_authority_id else None
            if "authority_name" in known_columns:
                # 如果是其他證照，保存 authority_name
                row["authority_name"] = cert.get("authority_name") if cert.get("authority_name") else None
            if "issuer" in known_columns:
                row["issuer"] = cert.get("issuer") if cert.get("issuer") else None
            if "AcquisitionDate" in known_columns and cert.get("acquire_date"):
                # 嘗試將日期轉為 YYYY-MM-DD 格式
                try:
                    date_obj = datetime.strptime(cert["acquire_date"].split('T')[0], "%Y-%m-%d")
                    row["AcquisitionDate"] = date_obj.strftime("%Y-%m-%d")
                except:
                    row["AcquisitionDate"] = cert["acquire_date"] # 保持原樣
            
            # 處理路徑
            path = cert.get("cert_path")
            if "CertPath" in known_columns and path:
                # 將 Windows 路徑格式（反斜杠）轉換為 Web 路徑格式（正斜杠）
                normalized_path = path.replace("\\", "/") 
                # 確保路徑是相對路徑格式
                if normalized_path.startswith("uploads/"):
                    row["CertPath"] = normalized_path
                else:
                    # 如果路徑包含絕對路徑，提取相對路徑部分
                    parts = normalized_path.split("/")
                    if "uploads" in parts:
                        idx_uploads = parts.index("uploads")
                        row["CertPath"] = "/".join(parts[idx_uploads:])
                    else:
                        row["CertPath"] = normalized_path
            else:
                row["CertPath"] = None
            
            cert_rows.append(row)

        # (4) 處理上傳證照圖片（舊的圖片上傳方式，向後兼容） - 這裡為了程式碼完整性省略，因為前端應主要傳遞 structured_certifications

        # (5) 實際寫入資料庫；有 resume_id 時不刪除既有證照，僅新增或更新（ON DUPLICATE KEY UPDATE 避免重複鍵）
        if cert_rows:
            print(f"📋 準備寫入 student_certifications: {len(cert_rows)} 筆")
            if not resume_id:
                cursor.execute("DELETE FROM student_certifications WHERE StuID=%s", (student_id,))
            insert_failed = False
            for row in cert_rows:
                cols = list(row.keys())
                values = list(row.values())
                cols.append("CreatedAt")
                placeholders = ", ".join(["%s"] * (len(values) + 1))
                sql = f"INSERT INTO student_certifications ({','.join(cols)}) VALUES ({placeholders})"
                if resume_id:
                    # 唯一鍵 (StuID, cert_code, level)：重複時只更新可變欄位
                    update_cols = [c for c in cols if c not in ('StuID', 'cert_code', 'level', 'CreatedAt')]
                    if update_cols:
                        sql += " ON DUPLICATE KEY UPDATE " + ", ".join(f"{c}=VALUES({c})" for c in update_cols)
                try:
                    cursor.execute(sql, (*values, datetime.now()))
                except Exception as e:
                    print(f"⚠️ 插入證照記錄失敗: {e}")
                    print(f" 記錄內容: {row}")
                    insert_failed = True
                    raise
            if not insert_failed:
                print(f"✅ student_certifications 已寫入 {len(cert_rows)} 筆")
        else:
            if struct_certs:
                print(f"⚠️ 有 {len(struct_certs)} 筆證照但 cert_rows 為空（可能全部被跳過），不刪除既存證照")
        
        # -------------------------------------------------------------
        # 4) 儲存語言能力 student_languageskills
        # -------------------------------------------------------------
        cursor.execute("DELETE FROM student_languageskills WHERE StuID=%s", (student_id,))
        lang_ids_saved = []
        for row in data.get("structured_languages", []):
            if row.get("language") and row.get("level"):
                cursor.execute("""
                    INSERT INTO student_languageskills (StuID, Language, Level, CreatedAt)
                    VALUES (%s,%s,%s,NOW())
                """, (student_id, row["language"], row["level"]))
                lang_ids_saved.append(cursor.lastrowid)

        # 一律計算 mapping 用的 ID（course_grades.id、證照 id、語文 id），供呼叫端寫入 resume_content_mapping
        course_ids = course_grade_ids_saved if course_grade_ids_saved else None
        data["selected_course_grade_ids"] = course_ids
        cert_ids_for_mapping = data.get("selected_certification_ids")
        if cert_rows:
            cursor.execute("SELECT id, cert_code, level FROM student_certifications WHERE StuID=%s", (student_id,))
            db_certs = {(r.get("cert_code"), r.get("level")): r.get("id") for r in cursor.fetchall()}
            ids_in_order = []
            for r in cert_rows:
                key = (r.get("cert_code"), r.get("level"))
                if key in db_certs and db_certs[key]:
                    ids_in_order.append(db_certs[key])
            if ids_in_order:
                cert_ids_for_mapping = ids_in_order
        lang_ids_for_mapping = data.get("selected_language_skill_ids")
        if lang_ids_saved:
            lang_ids_for_mapping = lang_ids_saved
        mapping_ids = {
            "course_grade_ids": course_ids,
            "certification_ids": cert_ids_for_mapping,
            "language_skill_ids": lang_ids_for_mapping,
            "absence_record_ids": data.get("selected_absence_record_ids"),
        }
        if resume_id:
            _upsert_resume_content_mapping(cursor, resume_id, student_id, course_grade_ids=course_ids,
                certification_ids=cert_ids_for_mapping,
                language_skill_ids=lang_ids_for_mapping,
                absence_record_ids=data.get("selected_absence_record_ids"))

        return (True, mapping_ids)

    except Exception as e:
        print("❌ 儲存結構化資料錯誤:", e)
        traceback.print_exc()
        return False


def get_student_info_for_doc(cursor, student_id, semester_id=None, resume_id=None):
    data = {}
    cursor.execute("SELECT * FROM Student_Info WHERE StuID=%s", (student_id,))
    data['info'] = cursor.fetchone() or {}

    mapping = None
    if resume_id:
        mapping = _get_resume_content_mapping(cursor, resume_id)

    # 檢查表是否有 SemesterID、ProofImage 和 transcript_path 列
    try:
        cursor.execute("SHOW COLUMNS FROM course_grades LIKE 'SemesterID'")
        has_semester_id = cursor.fetchone() is not None
        cursor.execute("SHOW COLUMNS FROM course_grades LIKE 'ProofImage'")
        has_proof_image = cursor.fetchone() is not None
        cursor.execute("SHOW COLUMNS FROM course_grades LIKE 'transcript_path'")
        has_transcript_path = cursor.fetchone() is not None
    except:
        has_semester_id = False
        has_proof_image = False
        has_transcript_path = False

    # 優先使用 ProofImage 欄位，如果沒有則使用 transcript_path（兼容舊結構）
    transcript_field = 'ProofImage' if has_proof_image else ('transcript_path' if has_transcript_path else None)

    if semester_id is not None and has_semester_id:
        if transcript_field:
            cursor.execute(f"""
                SELECT id, CourseName, Credits, Grade, IFNULL({transcript_field}, '') AS transcript_path, SemesterID 
                FROM course_grades 
                WHERE StuID=%s AND SemesterID=%s
                ORDER BY CourseName
            """, (student_id, semester_id))
        else:
            cursor.execute("""
                SELECT id, CourseName, Credits, Grade, SemesterID 
                FROM course_grades 
                WHERE StuID=%s AND SemesterID=%s
                ORDER BY CourseName
            """, (student_id, semester_id))
    else:
        if transcript_field:
            cursor.execute(f"""
                SELECT id, CourseName, Credits, Grade, IFNULL({transcript_field}, '') AS transcript_path 
                FROM course_grades 
                WHERE StuID=%s 
                ORDER BY CourseName
            """, (student_id,))
        else:
            cursor.execute("""
                SELECT id, CourseName, Credits, Grade 
                FROM course_grades 
                WHERE StuID=%s 
                ORDER BY CourseName
            """, (student_id,))
    
    grades_rows = cursor.fetchall() or []
    if mapping and mapping.get("course_grade_ids"):
        raw_ids = [x.strip() for x in (mapping["course_grade_ids"] or "").split(",") if x.strip()]
        if raw_ids:
            ids_are_int = all((x or "").replace("-", "").isdigit() for x in raw_ids)
            if ids_are_int:
                ids_set = set(int(x) for x in raw_ids)
                grades_rows = [r for r in grades_rows if r.get("id") in ids_set]
            else:
                ids_set = set(raw_ids)
                grades_rows = [r for r in grades_rows if (r.get("CourseName") or "").strip() in ids_set]
    
    data['grades'] = grades_rows
    data['transcript_path'] = ''
    # 嘗試從成績記錄中找到路徑
    for row in grades_rows:
        tp = row.get('transcript_path')
        if tp:
            data['transcript_path'] = tp
            break

    # 證照 - authority 以 student_certifications.authority_id 優先對應 cert_authorities.id，無則用 certificate_codes.authority_id；分類用 cc.category (labor/local/intl/other)
    cursor.execute("""
        SELECT 
            sc.id, sc.StuID, sc.cert_code,
            COALESCE(sc.authority_id, cc.authority_id) AS authority_id,
            cc.job_category AS CertName, sc.AcquisitionDate, sc.CertPath,
            sc.issuer, 
            cc.job_category, cc.level, cc.category AS CertCategory,
            ca.name AS authority_name
        FROM student_certifications sc
        LEFT JOIN certificate_codes cc 
            ON sc.cert_code = cc.id
        LEFT JOIN cert_authorities ca 
            ON ca.id = COALESCE(sc.authority_id, cc.authority_id)
        WHERE sc.StuID = %s
        ORDER BY sc.AcquisitionDate DESC, sc.id ASC
    """, (student_id,))
    certs_rows = cursor.fetchall() or []
    if mapping and mapping.get("certification_ids"):
        try:
            ids_set = set(int(x.strip()) for x in (mapping["certification_ids"] or "").split(",") if x.strip())
            if ids_set:
                certs_rows = [c for c in certs_rows if c.get("id") in ids_set]
        except ValueError:
            pass
    data['certifications'] = certs_rows

    # 語言能力
    cursor.execute(""" 
        SELECT id, Language AS language, Level AS level 
        FROM student_languageskills 
        WHERE StuID=%s 
        ORDER BY Language
    """, (student_id,))
    lang_rows = cursor.fetchall() or []
    if mapping and mapping.get("language_skill_ids"):
        try:
            ids_set = set(int(x.strip()) for x in (mapping["language_skill_ids"] or "").split(",") if x.strip())
            if ids_set:
                lang_rows = [l for l in lang_rows if l.get("id") in ids_set]
        except ValueError:
            pass
    data['languages'] = lang_rows

    # 缺勤記錄佐證圖片（僅返回最新的）
    absence_proof_path = ''
    try:
        cursor.execute("SELECT id FROM users WHERE username=%s", (student_id,))
        user_row = cursor.fetchone()
        if user_row:
            user_id = user_row.get('id')
            # 嘗試使用 created_at 排序
            try:
                cursor.execute("""
                    SELECT image_path 
                    FROM absence_records 
                    WHERE user_id = %s AND image_path IS NOT NULL AND image_path != '' 
                    ORDER BY created_at DESC 
                    LIMIT 1
                """, (user_id,))
            except:
                # 如果 created_at 欄位不存在，使用 id 排序
                cursor.execute("""
                    SELECT image_path 
                    FROM absence_records 
                    WHERE user_id = %s AND image_path IS NOT NULL AND image_path != '' 
                    ORDER BY id DESC 
                    LIMIT 1
                """, (user_id,))
            absence_row = cursor.fetchone()
            if absence_row:
                absence_proof_path = absence_row.get('image_path', '')
    except Exception as e:
        print(f"⚠️ 查詢缺勤佐證圖片失敗: {e}")
        traceback.print_exc()
    data['absence_proof_path'] = absence_proof_path

    return data

def format_data_for_doc(student_data, doc_path=None):
    context = {}
    doc = DocxTemplate(doc_path) if doc_path else None

    # 1. 基本資料 (Info)
    info = student_data['info']
    context['StuID'] = info.get('StuID', '')
    context['StuName'] = info.get('StuName', '')
    context['Gender'] = info.get('Gender', '')
    context['Phone'] = info.get('Phone', '')
    context['Email'] = info.get('Email', '')
    context['Address'] = info.get('Address', '')
    context['ConductScore'] = info.get('ConductScore', '')
    context['Autobiography'] = info.get('Autobiography', '')
    
    # 生日格式化
    birth_date = info.get('BirthDate')
    if isinstance(birth_date, (datetime, date)):
        context['BirthDate'] = birth_date.strftime("%Y-%m-%d")
    elif birth_date:
        try:
            context['BirthDate'] = datetime.strptime(str(birth_date).split(' ')[0], "%Y-%m-%d").strftime("%Y-%m-%d")
        except:
            context['BirthDate'] = str(birth_date)
    else:
        context['BirthDate'] = ''
    
    # 學生照片
    photo_path = info.get('PhotoPath')
    if photo_path and doc:
        image_size = Inches(1.5)
        image_obj = safe_create_inline_image(doc, photo_path, image_size, "學生照片")
        context['StudentPhoto'] = image_obj if image_obj else ""
    else:
        context['StudentPhoto'] = ""

    # 2. 核心科目 (Core Courses) - 假設所有課程都是核心科目
    core_courses = []
    for c in student_data['grades']:
        core_courses.append({
            'CourseName': c.get('CourseName', ''),
            'Credits': format_credits(c.get('Credits')),
            'Grade': score_to_grade(c.get('Grade')),
        })
    context['core_courses'] = core_courses

    # 3. 證照 (Certifications)
    all_certs = student_data['certifications']
    labor, international, local, other = categorize_certifications(all_certs)
    
    # 填入表格區（每個類別最多 4 個）
    fill_certificates_to_doc(context, "LaborCerts_", labor, 4)
    fill_certificates_to_doc(context, "IntlCerts_", international, 4)
    fill_certificates_to_doc(context, "LocalCerts_", local, 4)
    fill_certificates_to_doc(context, "OtherCerts_", other, 4)
    
    # 先將所有證照圖片/名稱佔位符設為空白，避免模板顯示 None
    for i in range(1, 33):
        context[f"CertPhotoImages_{i}"] = ""
        context[f"CertPhotoName_{i}"] = ""
    
    # 圖片區（不分類，按順序最多 32 張）
    certs_for_photos = [
        {'photo_path': c.get('CertPath'), 'photo_name': f"{c.get('job_category', '')}{c.get('level', '')}" if c.get('job_category') else c.get('CertName')}
        for c in all_certs if c.get('CertPath')
    ]

    if doc:
        # 第一頁圖片 (1-8)
        fill_certificate_photos(context, doc, certs_for_photos, 1, 8)
        # 第二頁圖片 (9-16)
        fill_certificate_photos(context, doc, certs_for_photos[8:], 9, 8)
        # 第三頁圖片 (17-24)
        fill_certificate_photos(context, doc, certs_for_photos[16:], 17, 8)
        # 第四頁圖片 (25-32)
        fill_certificate_photos(context, doc, certs_for_photos[24:], 25, 8)

    # 4. 語言能力 (Languages)
    # 根據模板格式，使用語言名稱作為前綴
    language_map = {
        '英語': 'En',
        '日語': 'Jp',
        '台語': 'Tw',
        '客語': 'Hk'
    }
    
    # 初始化所有語言為未選（□）
    for lang_name, prefix in language_map.items():
        context[f'{prefix}_Jing'] = '□'
        context[f'{prefix}_Zhong'] = '□'
        context[f'{prefix}_Lue'] = '□'
    
    # 根據實際數據填充
    for lang in student_data['languages']:
        lang_name = lang.get('language', '')
        level = lang.get('level', '')
        prefix = language_map.get(lang_name)
        
        if prefix and level:
            marks = generate_language_marks(level)
            context[f'{prefix}_Jing'] = marks['Jing']
            context[f'{prefix}_Zhong'] = marks['Zhong']
            context[f'{prefix}_Lue'] = marks['Lue']
    
    return context, doc

def generate_application_form_docx(student_data, output_path):
    try:
        base_dir = os.path.dirname(__file__)
        template_path = os.path.abspath(os.path.join(base_dir, "..", "frontend", "static", "examples", "實習履歷(空白).docx"))
        if not os.path.exists(template_path):
            print("❌ 找不到模板：", template_path)
            return False

        doc = DocxTemplate(template_path)
        info = student_data.get("info", {})
        grades = student_data.get("grades", [])
        certs = student_data.get("certifications", [])

        # 格式化出生日期
        def fmt_date(val):
            if hasattr(val, 'strftime'):
                return val.strftime("%Y-%m-%d")
            if isinstance(val, str) and len(val) >= 10:
                return val.split("T")[0]
            return ""

        bdate = fmt_date(info.get("BirthDate"))
        year, month, day = ("", "", "")
        if bdate:
            try:
                year, month, day = bdate.split("-")
            except:
                pass

        # 照片（使用 safe_create 驗證格式，相對路徑需轉絕對路徑）
        image_obj = None
        photo_path = resolve_upload_path(info.get("PhotoPath") or "")
        if photo_path and os.path.exists(photo_path):
            image_obj = safe_create_inline_image(doc, photo_path, Inches(1.2), "學生照片")

        # 處理課程資料（保留原邏輯）
        MAX_COURSES = 30
        padded_grades = grades[:MAX_COURSES]
        padded_grades += [{'CourseName': '', 'Credits': ''}] * (MAX_COURSES - len(padded_grades))

        context_courses = {}
        NUM_ROWS = 10
        NUM_COLS = 3
        for i in range(NUM_ROWS):
            for j in range(NUM_COLS):
                index = i * NUM_COLS + j
                if index < MAX_COURSES:
                    course = padded_grades[index]
                    row_num = i + 1
                    col_num = j + 1
                    context_courses[f'CourseName_{row_num}_{col_num}'] = (course.get('CourseName') or '')
                    context_courses[f'Credits_{row_num}_{col_num}'] = (course.get('Credits') or '')

        # 插入成績單圖片：相對路徑需轉絕對路徑
        transcript_obj = None
        transcript_path = resolve_upload_path(student_data.get("transcript_path") or info.get("TranscriptPath") or '')
        if transcript_path and os.path.exists(transcript_path):
            transcript_obj = safe_create_inline_image(doc, transcript_path, Inches(6.0), "成績單")

        # 缺勤佐證圖片（相對路徑需轉絕對路徑；相容 Absence_Proof_Path / absence_proof_path）
        absence_proof_obj = None
        absence_raw = (student_data.get("Absence_Proof_Path") or student_data.get("absence_proof_path") or "").strip()
        absence_proof_path = resolve_upload_path(absence_raw)
        image_size = Inches(6.0)
        if absence_proof_path and os.path.exists(absence_proof_path):
            absence_proof_obj = safe_create_inline_image(doc, absence_proof_path, image_size, "缺勤佐證")

        # 操行等級
        conduct_score = info.get('ConductScore', '')
        conduct_marks = {k: '□' for k in ['C_You', 'C_Jia', 'C_Yi', 'C_Bing', 'C_Ding']}
        mapping = {'優': 'C_You', '甲': 'C_Jia', '乙': 'C_Yi', '丙': 'C_Bing', '丁': 'C_Ding'}
        if conduct_score in mapping:
            conduct_marks[mapping[conduct_score]] = '■'

        # 證照分類 - 使用新的分類邏輯
        # certs 已經從 get_student_info_for_doc 返回，格式統一
        # 優先使用前端提交的證照名稱（如果有的話）
        # 這樣可以確保只顯示用戶實際選擇的證照，而不是數據庫中所有相關記錄
        cert_photo_paths_from_form = student_data.get("cert_photo_paths", []) or []
        structured_certs = student_data.get("structured_certifications", [])

        # 以 structured_certifications 為準疊代，確保每筆都有 authority_id / category，避免索引錯位
        # 同一 photo_path 只使用一次，避免編輯後重新上傳時出現兩張一樣的證照圖
        if structured_certs:
            certs_from_form = []
            used_photo_paths = set()
            for idx, struct in enumerate(structured_certs):
                name = (struct.get("name") or "").strip()
                if not name:
                    continue
                path = (struct.get("cert_path") or "")
                if not path and idx < len(cert_photo_paths_from_form):
                    path = cert_photo_paths_from_form[idx] or ""
                path = (path or "").replace("\\", "/").strip()
                if path and path in used_photo_paths:
                    path = ""
                elif path:
                    used_photo_paths.add(path)
                matching_cert = certs[idx] if idx < len(certs) else None
                if not matching_cert and path:
                    for c in certs:
                        if (c.get("cert_path") or "").replace("\\", "/") == path:
                            matching_cert = c
                            break
                # 類別：優先使用 DB（certificate_codes）的 category，確保如 JLPT 等 intl 正確歸類
                aid = struct.get("authority_id")
                db_category = None
                if matching_cert:
                    db_category = (matching_cert.get("CertCategory") or matching_cert.get("category") or "").strip().lower()
                if aid is not None and str(aid).strip() in ("1",):
                    category = "labor"
                elif db_category and db_category in ("labor", "intl", "local", "other"):
                    category = db_category
                else:
                    try:
                        ai = int(aid)
                        if ai == 1:
                            category = "labor"
                        else:
                            category = (struct.get("category") or "").strip().lower() or "other"
                            if category not in ("labor", "intl", "local", "other"):
                                category = "other"
                    except (TypeError, ValueError):
                        cat = (struct.get("category") or "").strip().lower()
                        category = cat if cat in ("labor", "intl", "local", "other") else (db_category if db_category in ("labor", "intl", "local", "other") else "other")
                certs_from_form.append({
                    "cert_name": name,
                    "category": category,
                    "cert_path": path or (matching_cert.get("cert_path", "") if matching_cert else ""),
                    "acquire_date": (matching_cert.get("acquire_date") or struct.get("acquire_date") or "") if matching_cert else (struct.get("acquire_date") or ""),
                })
            if certs_from_form:
                certs = certs_from_form
        
        # 分類證照
        labor_list, intl_list, local_list, other_list = categorize_certifications(certs)

        def pad_list(lst, length=5):
            lst = lst[:length]
            lst += [''] * (length - len(lst))
            return lst

        # 建 context
        # 處理自傳：移除多餘的換行符，避免產生空白行
        autobiography = info.get('Autobiography', '').strip()
        if autobiography:
            # 將多個連續換行符替換為單個換行符，移除開頭和結尾的換行符
            autobiography = re.sub(r'\n{3,}', '\n\n', autobiography)
            autobiography = autobiography.strip('\n')
        
        context = {
            'StuID': (info.get('StuID') or ''),
            'StuName': (info.get('StuName') or ''),
            'BirthYear': year, 'BirthMonth': month, 'BirthDay': day,
            'Gender': (info.get('Gender') or ''),
            'Phone': (info.get('Phone') or ''),
            'Email': (info.get('Email') or ''),
            'Address': (info.get('Address') or ''),
            'ConductScoreNumeric': (info.get('ConductScoreNumeric') or ''),
            'ConductScore': conduct_score,
            'Autobiography': (autobiography or ''),
        }
        if image_obj:
            context['Image_1'] = image_obj
        if transcript_obj:
            context['transcript_path'] = transcript_obj
        if absence_proof_obj:
            context['Absence_Proof_Image'] = absence_proof_obj

        empty_vars_to_clear = [
            'empty_line_1', 'empty_line_2', 'empty_line_3',
            'blank_line_1', 'blank_line_2', 'blank_line_3',
            'spacer_1', 'spacer_2', 'spacer_3',
            'extra_line_1', 'extra_line_2', 'extra_line_3',
            'blank_1', 'blank_2', 'blank_3',
        ]
        for var in empty_vars_to_clear:
            context[var] = ""

        # 加入缺勤統計
        # 只填充這8個標準字段，確保沒有多餘的空白行
        absence_fields = ['曠課', '遲到', '事假', '病假', '生理假', '公假', '喪假', '總計']
        for t in absence_fields:
            key = f"absence_{t}_units"
            value = (student_data.get(key) or "0 節")
            context[key] = value
            # 調試輸出
            if value == "0 節" and t != "總計":
                print(f"⚠️ 缺勤統計 {key} 未找到，使用預設值: {value}")
            else:
                print(f"✅ 缺勤統計 {key} = {value}")
        
        # 如果模板中有額外的行（例如第9、10、11行），將它們設為空字符串
        # 常見的額外變數名可能是：absence_row_9, absence_row_10, absence_row_11 等
        # 或者：absence_9_units, absence_10_units, absence_11_units 等
        # 清空可能的額外行變數
        for i in range(9, 12):  # 第9、10、11行
            # 嘗試多種可能的變數名格式
            possible_keys = [
                f"absence_row_{i}",
                f"absence_{i}_units",
                f"absence_row_{i}_units",
                f"absence_item_{i}",
                f"absence_type_{i}",
            ]
            for key in possible_keys:
                context[key] = ""
        
        # 清空可能存在的其他缺勤類型變數（防止模板中有額外的空白行）
        # 例如：absence_其他_units, absence_其他1_units 等
        # 只保留標準的8個字段，其他都設為空字符串
        standard_keys = [f"absence_{t}_units" for t in absence_fields]
        for key in list(context.keys()):
            if key.startswith("absence_") and key.endswith("_units"):
                if key not in standard_keys:
                    context[key] = ""  # 清空非標準字段

        # 加入操行等級勾選
        context.update(conduct_marks)

        # 加入課程資料
        context.update(context_courses)

        # 加入證照文字清單 - 使用新的填充函數
        fill_certificates_to_doc(context, "LaborCerts_", labor_list, 5)
        fill_certificates_to_doc(context, "IntlCerts_", intl_list, 5)
        fill_certificates_to_doc(context, "LocalCerts_", local_list, 5)
        fill_certificates_to_doc(context, "OtherCerts_", other_list, 5)

        # 證照圖片（不分類，依順序塞）- 使用新的填充函數
        # 將四類組裝成一個大 list（圖片不分類）
        flat_list = labor_list + intl_list + local_list + other_list
        
        # 分頁顯示證照圖片：每頁8張，最多32張（4頁）
        # 使用區塊變數控制頁面顯示/隱藏
        certs_per_page = 8
        max_total = 32  # 最多32張（4頁）
        
        # 將相對路徑轉為絕對路徑，否則 os.path.exists 會失敗、導致只顯示部分證照圖（如 DB 有 3 張路徑卻只生成 2 張）
        resolved_flat = []
        for c in flat_list:
            p = (c.get("photo_path") or "").strip()
            resolved_p = resolve_upload_path(p) if p else ""
            resolved_flat.append({**c, "photo_path": resolved_p or p})
        # 只處理實際有圖片的證照（最多32張）
        certs_with_photos = [c for c in resolved_flat if c.get("photo_path") and os.path.exists(c.get("photo_path", ""))]
        certs_to_display = certs_with_photos[:max_total]
        total_certs = len(certs_to_display)
        
        # 初始化證照名稱為空；圖片僅在有圖時才設 key，未輸入不設以免顯示 None
        for idx in range(1, 33):
            context[f"CertPhotoName_{idx}"] = ""
        
        # 初始化所有頁面區塊為 False（不顯示）
        # 使用布林值控制頁面顯示，模板中使用 {% if cert_page_2_block %} ... {% endif %}
        context["cert_page_2_block"] = False
        context["cert_page_3_block"] = False
        context["cert_page_4_block"] = False
        
        if total_certs > 0:
            # 第一頁（1-8）：總是填充（如果有證照）
            first_page_certs = certs_to_display[:min(8, total_certs)]
            if first_page_certs:
                fill_certificate_photos(context, doc, first_page_certs, start_index=1, max_count=8)
            
            # 第二頁（9-16）：如果 total_certs > 8 則顯示
            if total_certs > 8:
                context["cert_page_2_block"] = True  # 設置為 True 以顯示區塊
                second_page_certs = certs_to_display[8:min(16, total_certs)]
                if second_page_certs:
                    fill_certificate_photos(context, doc, second_page_certs, start_index=9, max_count=8)
            
            # 第三頁（17-24）：如果 total_certs > 16 則顯示
            if total_certs > 16:
                context["cert_page_3_block"] = True  # 設置為 True 以顯示區塊
                third_page_certs = certs_to_display[16:min(24, total_certs)]
                if third_page_certs:
                    fill_certificate_photos(context, doc, third_page_certs, start_index=17, max_count=8)
            
            # 第四頁（25-32）：如果 total_certs > 24 則顯示
            if total_certs > 24:
                context["cert_page_4_block"] = True  # 設置為 True 以顯示區塊
                fourth_page_certs = certs_to_display[24:min(32, total_certs)]
                if fourth_page_certs:
                    fill_certificate_photos(context, doc, fourth_page_certs, start_index=25, max_count=8)

        # 語文能力
        lang_context = {}
        lang_codes = ['En', 'Jp', 'Tw', 'Hk']
        level_codes = ['Jing', 'Zhong', 'Lue']
        for code in lang_codes:
            for level_code in level_codes:
                lang_context[f'{code}_{level_code}'] = '□'

        lang_code_map = {'英語': 'En', '日語': 'Jp', '台語': 'Tw', '客語': 'Hk'}
        level_code_map = {'精通': 'Jing', '中等': 'Zhong', '略懂': 'Lue'}

        for lang_skill in student_data.get('languages', []):
            lang = lang_skill.get('Language') or lang_skill.get('language')
            level = lang_skill.get('Level') or lang_skill.get('level')
            lang_code = lang_code_map.get(lang)
            level_code = level_code_map.get(level)
            if lang_code and level_code:
                key = f'{lang_code}_{level_code}'
                if key in lang_context:
                    lang_context[key] = '■'

        # 未填寫的語文能力自動代入「略懂」
        for code in lang_codes:
            if all(lang_context.get(f'{code}_{lc}', '□') == '□' for lc in level_codes):
                lang_context[f'{code}_Lue'] = '■'

        context.update(lang_context)
        
        # 渲染前：所有 None 改為 ""，避免 Word 顯示 "None"；未輸入欄位顯示空白
        for key in list(context.keys()):
            if context[key] is None:
                context[key] = ""

        doc.render(context)
        # 確保輸出路徑所在目錄存在（不同電腦路徑不同時可正常寫入）
        output_dir = os.path.dirname(output_path)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)
        doc.save(output_path)
        print(f"✅ 履歷文件已生成: {output_path}")
        return True

    except Exception as e:
        print("❌ 生成 Word 檔錯誤:", e)
        traceback.print_exc()
        return False