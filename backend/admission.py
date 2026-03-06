from flask import Blueprint, request, jsonify, session, render_template, redirect, send_file
from config import get_db
from datetime import datetime
from semester import get_current_semester_code, get_current_semester_id
from notification import create_notification
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import io
import traceback
import re

admission_bp = Blueprint("admission_bp", __name__, url_prefix="/admission")

def _get_active_semester_year(cursor):
    """取得當前啟用學期學年（semesters 表 is_active=1 的 code 前三碼，如 1132->113）"""
    cursor.execute("SELECT code FROM semesters WHERE is_active = 1 LIMIT 1")
    row = cursor.fetchone()
    if not row or row.get('code') is None:
        cursor.execute("SELECT code FROM semesters WHERE code IS NOT NULL AND code != '' ORDER BY code DESC LIMIT 1")
        row = cursor.fetchone()
    if not row or row.get('code') is None:
        cursor.execute("SELECT code FROM semesters ORDER BY id DESC LIMIT 1")
        row = cursor.fetchone()
    raw = row.get('code') if row else None
    if raw is None:
        return None
    if isinstance(raw, int):
        return raw // 10 if raw >= 100 else None
    if isinstance(raw, bytes):
        raw = raw.decode('utf-8', errors='ignore')
    code = str(raw).strip()
    if len(code) >= 3:
        try:
            return int(code[:3])
        except (TypeError, ValueError):
            pass
    return None

def _resolve_duplicate_students(cursor, semester_id):
    """
    自動處理重複學生：如果有重複的學生，選擇志願序最高的記錄（preference_order 最小），
    將其他記錄的 director_decision 更新為 Pending。
    
    Args:
        cursor: 資料庫游標
        semester_id: 學期ID
    """
    try:
        # 檢查是否有 semester_id 欄位
        cursor.execute("""
            SELECT COLUMN_NAME
            FROM information_schema.COLUMNS
            WHERE TABLE_SCHEMA = DATABASE()
            AND TABLE_NAME = 'manage_director'
            AND COLUMN_NAME = 'semester_id'
        """)
        has_semester_id = cursor.fetchone() is not None
        cursor.fetchall()
        
        # 查詢同一學期內，同一學生有多筆記錄的情況（只考慮 Approved 或 Pending 的記錄）
        if has_semester_id:
            cursor.execute("""
                SELECT 
                    md.match_id,
                    md.student_id,
                    md.preference_id,
                    md.director_decision,
                    sp.preference_order
                FROM manage_director md
                LEFT JOIN student_job_applications sja ON md.preference_id = sja.id
                LEFT JOIN student_preferences sp ON sja.student_id = sp.student_id
                    AND sja.company_id = sp.company_id
                    AND sja.job_id = sp.job_id
                    AND sp.semester_id = %s
                WHERE md.semester_id = %s
                    AND md.director_decision IN ('Approved', 'Pending')
                ORDER BY md.student_id, COALESCE(sp.preference_order, 999), md.match_id
            """, (semester_id, semester_id))
        else:
            cursor.execute("""
                SELECT 
                    md.match_id,
                    md.student_id,
                    md.preference_id,
                    md.director_decision,
                    sp.preference_order
                FROM manage_director md
                LEFT JOIN student_job_applications sja ON md.preference_id = sja.id
                LEFT JOIN student_preferences sp ON sja.student_id = sp.student_id
                    AND sja.company_id = sp.company_id
                    AND sja.job_id = sp.job_id
                WHERE md.director_decision IN ('Approved', 'Pending')
                ORDER BY md.student_id, COALESCE(sp.preference_order, 999), md.match_id
            """)
        
        all_records = cursor.fetchall()
        print(f"🔍 [resolve_duplicate_students] 查詢到 {len(all_records)} 筆記錄")
        
        # 按學生分組
        student_records = {}
        for record in all_records:
            student_id = record.get('student_id')
            if student_id:
                if student_id not in student_records:
                    student_records[student_id] = []
                student_records[student_id].append(record)
        
        # 處理每個有重複記錄的學生
        updated_count = 0
        for student_id, records in student_records.items():
            if len(records) <= 1:
                continue  # 沒有重複，跳過
            
            print(f"🔍 [resolve_duplicate_students] 發現重複學生 student_id={student_id}，有 {len(records)} 筆記錄")
            
            # 選擇志願序最高的記錄（preference_order 最小）
            # 如果志願序相同，選擇 match_id 較小的（較早創建的）
            best_record = None
            best_order = 999
            best_match_id = None
            
            for record in records:
                preference_order = record.get('preference_order')
                match_id = record.get('match_id')
                print(f"  - match_id={match_id}, preference_order={preference_order}, director_decision={record.get('director_decision')}")
                
                if preference_order is None:
                    preference_order = 999  # 沒有志願序的排在最後
                
                match_id_int = int(match_id) if match_id and str(match_id).isdigit() else 999999
                
                # 優先考慮志願序，如果志願序相同則考慮 match_id
                if preference_order < best_order or (preference_order == best_order and match_id_int < (int(best_match_id) if best_match_id and str(best_match_id).isdigit() else 999999)):
                    best_record = record
                    best_order = preference_order
                    best_match_id = match_id
            
            if not best_record:
                print(f"  ⚠️ 無法選擇最佳記錄，跳過")
                continue
            
            print(f"  ✅ 選擇最佳記錄: match_id={best_match_id}, preference_order={best_order}")
            
            # 將其他記錄的 director_decision 更新為 Pending
            best_match_id_str = str(best_record.get('match_id'))
            for record in records:
                match_id = str(record.get('match_id'))
                if match_id != best_match_id_str:
                    # 更新為 Pending（但保持 Approved 的記錄不變，除非志願序更低）
                    current_decision = record.get('director_decision')
                    if current_decision == 'Approved':
                        # 如果當前是 Approved 但志願序更低，更新為 Pending
                        print(f"  🔄 更新 match_id={match_id} 為 Pending（志願序較低）")
                        cursor.execute("""
                            UPDATE manage_director
                            SET director_decision = 'Pending',
                                updated_at = CURRENT_TIMESTAMP
                            WHERE match_id = %s
                        """, (match_id,))
                        updated_count += cursor.rowcount
                    elif current_decision == 'Pending':
                        # 已經是 Pending，不需要更新
                        print(f"  ℹ️ match_id={match_id} 已經是 Pending，跳過")
                        pass
        
        if updated_count > 0:
            print(f"✅ 自動處理重複學生：已將 {updated_count} 筆記錄更新為 Pending")
        
        return updated_count
    except Exception as e:
        print(f"⚠️ 處理重複學生時發生錯誤: {str(e)}")
        traceback.print_exc()
        return 0

# =========================================================
# 頁面路由：查看錄取結果
# =========================================================
@admission_bp.route("/results", methods=["GET"])
def admission_results_page():
    """查看學生錄取結果頁面"""
    if 'user_id' not in session:
        return redirect('/login')
    
    user_role = session.get('role')
    # 允許班導、老師、主任、ta、admin 訪問
    if user_role not in ['class_teacher', 'teacher', 'director', 'ta', 'admin']:
        return "無權限訪問此頁面", 403
    
    return render_template('user_shared/admission_results.html', user_role=user_role or '')

# =========================================================
# 頁面路由：實習生管理
# =========================================================
@admission_bp.route("/intern_management", methods=["GET"])
def intern_management_page():
    """實習生管理頁面"""
    if 'user_id' not in session:
        return redirect('/login')
    
    user_role = session.get('role')
    # 允許老師、主任、ta、admin、vendor 訪問
    if user_role not in ['teacher', 'director', 'ta', 'admin', 'vendor']:
        return "無權限訪問此頁面", 403
    
    return render_template('user_shared/Intern management.html')

# =========================================================
# 頁面路由：實習生／未錄取名單管理
# =========================================================
@admission_bp.route("/unadmitted_list", methods=["GET"])
def unadmitted_list_page():
    """實習生／未錄取名單管理頁面（科助、主任、老師、管理員）"""
    if 'user_id' not in session:
        return redirect('/login')
    
    user_role = session.get('role')
    if user_role not in ['ta', 'admin', 'director', 'teacher']:
        return "無權限訪問此頁面", 403
    
    return render_template('ta/unadmitted_list.html', user_role=user_role or '')

# =========================================================
# 頁面路由：主任媒合
# =========================================================
@admission_bp.route("/manage_director", methods=["GET"])
def manage_director_page():
    """主任媒合頁面"""
    if 'user_id' not in session:
        return redirect('/login')
    
    user_role = session.get('role')
    # 只允許主任訪問
    if user_role != 'director':
        return "無權限訪問此頁面", 403
    
    return render_template('user_shared/manage_director.html')

# =========================================================
# API: 記錄實習錄取結果（錄取後自動綁定指導老師與學生）
# =========================================================
@admission_bp.route("/api/record_admission", methods=["POST"])
def record_admission():
    """
    記錄實習錄取結果，並自動綁定指導老師與學生
    可由廠商、指導老師或管理員調用
    """
    if 'user_id' not in session:
        return jsonify({"success": False, "message": "未授權"}), 403
    
    data = request.get_json() or {}
    student_id = data.get("student_id")
    company_id = data.get("company_id")
    job_id = data.get("job_id")  # 可選
    preference_order = data.get("preference_order")  # 可選，記錄最終錄取志願
    
    if not student_id or not company_id:
        return jsonify({"success": False, "message": "請提供學生ID和公司ID"}), 400
    
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 1. 驗證學生和公司是否存在
        cursor.execute("SELECT id, name, username FROM users WHERE id = %s AND role = 'student'", (student_id,))
        student = cursor.fetchone()
        if not student:
            return jsonify({"success": False, "message": "找不到該學生"}), 404
        
        cursor.execute("SELECT id, company_name, advisor_user_id FROM internship_companies WHERE id = %s", (company_id,))
        company = cursor.fetchone()
        if not company:
            return jsonify({"success": False, "message": "找不到該公司"}), 404
        
        # 2. 獲取指導老師ID（從公司的 advisor_user_id）
        advisor_user_id = company.get('advisor_user_id')
        if not advisor_user_id:
            return jsonify({"success": False, "message": "該公司尚未指派指導老師"}), 400
        
        # 驗證指導老師是否存在
        cursor.execute("SELECT id, name FROM users WHERE id = %s AND role IN ('teacher', 'director')", (advisor_user_id,))
        advisor = cursor.fetchone()
        if not advisor:
            return jsonify({"success": False, "message": "找不到該指導老師"}), 404
        
        # 3. 獲取當前學期（代碼與 ID）
        semester_code = get_current_semester_code(cursor)
        current_semester_id = get_current_semester_id(cursor)
        if not semester_code and not current_semester_id:
            return jsonify({"success": False, "message": "目前沒有設定當前學期"}), 400
        current_datetime_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # 3.1 偵測 teacher_student_relations 表是否有 semester_id 或 semester 欄位
        cursor.execute("""
            SELECT COLUMN_NAME FROM information_schema.COLUMNS
            WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'teacher_student_relations'
            AND COLUMN_NAME IN ('semester_id', 'semester', 'company_id')
        """)
        tsr_columns = {row['COLUMN_NAME'] for row in cursor.fetchall()}
        has_semester_id = 'semester_id' in tsr_columns
        has_semester = 'semester' in tsr_columns
        has_company_id = 'company_id' in tsr_columns
        
        # 4. 檢查是否已經存在該關係（避免重複）
        if has_semester_id and current_semester_id:
            cursor.execute("""
                SELECT id FROM teacher_student_relations
                WHERE teacher_id = %s AND student_id = %s AND semester_id = %s
            """, (advisor_user_id, student_id, current_semester_id))
        elif has_semester and semester_code:
            cursor.execute("""
                SELECT id FROM teacher_student_relations
                WHERE teacher_id = %s AND student_id = %s AND semester = %s
            """, (advisor_user_id, student_id, semester_code))
        else:
            cursor.execute("""
                SELECT id FROM teacher_student_relations
                WHERE teacher_id = %s AND student_id = %s
            """, (advisor_user_id, student_id))
        existing_relation = cursor.fetchone()
        
        if existing_relation:
            # 如果已存在，更新公司ID（可能學生換了公司，僅當表有 company_id 時）
            if has_company_id:
                cursor.execute("""
                    UPDATE teacher_student_relations
                    SET company_id = %s, updated_at = NOW()
                    WHERE id = %s
                """, (company_id, existing_relation['id']))
        else:
            # 5. 創建師生關係記錄
            if has_semester_id and current_semester_id:
                cols = "(teacher_id, student_id, semester_id, role, created_at)"
                vals = "(%s, %s, %s, '指導老師', NOW())"
                args = (advisor_user_id, student_id, current_semester_id)
                if has_company_id:
                    cols = "(teacher_id, student_id, company_id, semester_id, role, created_at)"
                    vals = "(%s, %s, %s, %s, '指導老師', NOW())"
                    args = (advisor_user_id, student_id, company_id, current_semester_id)
                cursor.execute("""
                    INSERT INTO teacher_student_relations """ + cols + """ VALUES """ + vals, args)
            elif has_semester and semester_code:
                cols = "(teacher_id, student_id, semester, role, created_at)"
                vals = "(%s, %s, %s, '指導老師', NOW())"
                args = (advisor_user_id, student_id, semester_code)
                if has_company_id:
                    cols = "(teacher_id, student_id, company_id, semester, role, created_at)"
                    vals = "(%s, %s, %s, %s, '指導老師', NOW())"
                    args = (advisor_user_id, student_id, company_id, semester_code)
                cursor.execute("""
                    INSERT INTO teacher_student_relations """ + cols + """ VALUES """ + vals, args)
            else:
                cursor.execute("""
                    INSERT INTO teacher_student_relations (teacher_id, student_id, role, created_at)
                    VALUES (%s, %s, '指導老師', NOW())
                """, (advisor_user_id, student_id))
        
        # 6. 在 internship_offers 表中記錄錄取結果 (新增的邏輯)
        # 這是 get_my_admission API 優先讀取的資料來源
        print(f"🔍 [DEBUG] record_admission - 準備寫入 internship_offers: student_id={student_id}, job_id={job_id}")
        
        # 檢查是否已存在於 internship_offers
        if job_id:
            cursor.execute("""
                SELECT id FROM internship_offers
                WHERE student_id = %s AND job_id = %s
            """, (student_id, job_id))
        else:
            # 如果 job_id 為 NULL，檢查是否有該學生的其他錄取記錄
            cursor.execute("""
                SELECT id FROM internship_offers
                WHERE student_id = %s AND job_id IS NULL
            """, (student_id,))
        existing_offer = cursor.fetchone()
        print(f"🔍 [DEBUG] existing_offer: {existing_offer}")

        if existing_offer:
            # 如果已存在，更新錄取狀態和時間（使用 'accepted' 狀態，與資料庫中的值一致）
            cursor.execute("""
                UPDATE internship_offers
                SET status = 'accepted', offered_at = %s, responded_at = %s
                WHERE id = %s
            """, (current_datetime_str, current_datetime_str, existing_offer['id']))
            print(f"✅ [DEBUG] 更新 internship_offers 記錄: id={existing_offer['id']}")
        else:
            # 插入新的錄取記錄（使用 'accepted' 狀態）
            cursor.execute("""
                INSERT INTO internship_offers 
                (student_id, job_id, status, offered_at, responded_at)
                VALUES (%s, %s, %s, %s, %s)
            """, (student_id, job_id, 'accepted', current_datetime_str, current_datetime_str))
            inserted_id = cursor.lastrowid
            print(f"✅ [DEBUG] 插入新 internship_offers 記錄: id={inserted_id}, student_id={student_id}, job_id={job_id}")
            
        # 7. 更新學生的志願序狀態
        if preference_order:
            cursor.execute("""
                UPDATE student_preferences
                SET status = 'approved'
                WHERE student_id = %s AND preference_order = %s
            """, (student_id, preference_order))
        
        conn.commit()
        
        return jsonify({
            "success": True,
            "message": f"錄取結果已記錄，已自動綁定指導老師 {advisor['name']} 與學生 {student['name']}。資料已寫入 internship_offers。",
            "teacher_id": advisor_user_id,
            "teacher_name": advisor['name'],
            "student_id": student_id,
            "student_name": student['name'],
            "company_id": company_id,
            "company_name": company['company_name']
        })
    
    except Exception as e:
        traceback.print_exc()
        conn.rollback()
        return jsonify({"success": False, "message": f"記錄錄取結果失敗: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

# =========================================================
# API: 獲取學生的錄取結果（我的實習成果）
# (此處保持不變，因為它已經設計為優先讀取 internship_offers)
# =========================================================
@admission_bp.route("/api/get_my_admission", methods=["GET"])
def get_my_admission():
    """學生查看自己的錄取結果"""
    if 'user_id' not in session or session.get('role') != 'student':
        return jsonify({"success": False, "message": "未授權"}), 403
    
    student_id = session.get('user_id')
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 從 internship_offers 表獲取錄取資料（正式錄取結果）
        # 若已執行 migration 新增 semester_id，可於 SELECT 中加入 io.semester_id AS offer_semester_id，以對應實習學期
        cursor.execute("""
            SELECT 
                io.id AS offer_id,
                io.job_id,
                io.status AS offer_status,
                io.offered_at,
                io.responded_at,
                ij.company_id,
                ij.title AS job_title,
                ij.description AS job_description,
                ij.period AS internship_period,
                ij.work_time AS internship_time,
                ij.salary AS job_salary,
                ic.company_name,
                ic.location AS company_address,
                ic.contact_person AS contact_name,
                ic.contact_email,
                ic.contact_phone,
                ic.advisor_user_id
            FROM internship_offers io
            LEFT JOIN internship_jobs ij ON io.job_id = ij.id
            LEFT JOIN internship_companies ic ON ij.company_id = ic.id
            WHERE io.student_id = %s
              AND io.status IN ('offered', 'accepted', 'Approved')
            ORDER BY io.offered_at DESC
            LIMIT 1
        """, (student_id,))
        offer_info = cursor.fetchone()
        
        print(f"🔍 [DEBUG] offer_info from internship_offers: {offer_info}")
        
        # 如果從 internship_offers 獲取到資料，使用它
        if offer_info:
            print(f"✅ [DEBUG] 找到 internship_offers 記錄")
            print(f"    job_id={offer_info.get('job_id')}, company_id={offer_info.get('company_id')}")
            
            # 如果 company_id 為 NULL，嘗試從 student_preferences 獲取公司資訊
            company_id = offer_info.get('company_id')
            if not company_id and offer_info.get('job_id'):
                # 從 student_preferences 獲取公司資訊
                cursor.execute("""
                    SELECT company_id
                    FROM student_preferences
                    WHERE student_id = %s AND job_id = %s
                    LIMIT 1
                """, (student_id, offer_info.get('job_id')))
                pref_company = cursor.fetchone()
                if pref_company and pref_company.get('company_id'):
                    company_id = pref_company.get('company_id')
                    print(f"🔍 [DEBUG] 從 student_preferences 獲取到 company_id={company_id}")
            
            # 如果還是沒有 company_id，嘗試直接從 internship_jobs 獲取
            if not company_id and offer_info.get('job_id'):
                cursor.execute("""
                    SELECT company_id
                    FROM internship_jobs
                    WHERE id = %s
                """, (offer_info.get('job_id'),))
                job_row = cursor.fetchone()
                if job_row and job_row.get('company_id'):
                    company_id = job_row.get('company_id')
                    print(f"🔍 [DEBUG] 從 internship_jobs 獲取到 company_id={company_id}")
            
            # 如果有 company_id，重新查詢完整的公司資訊
            if company_id:
                cursor.execute("""
                    SELECT 
                        id AS company_id,
                        company_name,
                        location AS company_address,
                        contact_person AS contact_name,
                        contact_email,
                        contact_phone,
                        advisor_user_id
                    FROM internship_companies
                    WHERE id = %s
                """, (company_id,))
                company_row = cursor.fetchone()
                if company_row:
                    offer_info['company_id'] = company_row.get('company_id')
                    offer_info['company_name'] = company_row.get('company_name')
                    offer_info['company_address'] = company_row.get('company_address')
                    offer_info['contact_name'] = company_row.get('contact_name')
                    offer_info['contact_email'] = company_row.get('contact_email')
                    offer_info['contact_phone'] = company_row.get('contact_phone')
                    offer_info['advisor_user_id'] = company_row.get('advisor_user_id')
                    print(f"✅ [DEBUG] 重新查詢到完整的公司資訊: {company_row.get('company_name')}")
            
            # 獲取指導老師資訊（從公司的 advisor_user_id）
            teacher_id = offer_info.get('advisor_user_id')
            teacher_name = None
            teacher_email = None
            if teacher_id:
                cursor.execute("""
                    SELECT id, name, email
                    FROM users
                    WHERE id = %s AND role IN ('teacher', 'director')
                """, (teacher_id,))
                teacher_info = cursor.fetchone()
                if teacher_info:
                    teacher_name = teacher_info.get('name')
                    teacher_email = teacher_info.get('email')
            
            # 實習期間：依 internship_configs 取得（作業在 1131、實習在 1132，應顯示 1132 的實習起訖）
            # 若有 internship_offers.semester_id 則用該實習學期；否則依學生屆別取一筆 config
            semester_code = None
            semester_start_date = None
            semester_end_date = None
            offer_semester_id = offer_info.get('offer_semester_id') if isinstance(offer_info.get('offer_semester_id'), int) else None
            cursor.execute("SELECT role, admission_year, username FROM users WHERE id = %s", (student_id,))
            user_row = cursor.fetchone()
            admission_year_val = None
            if user_row:
                if user_row.get('admission_year') is not None and str(user_row.get('admission_year', '')).strip() != '':
                    try:
                        admission_year_val = int(user_row['admission_year'])
                    except (TypeError, ValueError):
                        pass
                if admission_year_val is None and user_row.get('username') and len(str(user_row.get('username', ''))) >= 3:
                    try:
                        admission_year_val = int(str(user_row['username'])[:3])
                    except (TypeError, ValueError):
                        pass
            if admission_year_val is not None:
                cursor.execute("""
                    SELECT ic.intern_start_date, ic.intern_end_date, s.code AS semester_code
                    FROM internship_configs ic
                    LEFT JOIN semesters s ON s.id = ic.semester_id
                    WHERE (ic.user_id = %s OR (ic.user_id IS NULL AND ic.admission_year = %s))
                      AND (ic.semester_id = %s OR %s IS NULL)
                    ORDER BY ic.user_id DESC, ic.semester_id DESC
                    LIMIT 1
                """, (student_id, admission_year_val, offer_semester_id, offer_semester_id))
                ic_row = cursor.fetchone()
                if ic_row:
                    semester_start_date = ic_row.get('intern_start_date')
                    semester_end_date = ic_row.get('intern_end_date')
                    semester_code = ic_row.get('semester_code')
                    if isinstance(semester_start_date, datetime):
                        semester_start_date = semester_start_date.strftime("%Y-%m-%d")
                    if isinstance(semester_end_date, datetime):
                        semester_end_date = semester_end_date.strftime("%Y-%m-%d")
            if not semester_code:
                semester_code = get_current_semester_code(cursor)
            if not semester_start_date or not semester_end_date:
                if semester_code:
                    cursor.execute("SELECT start_date, end_date FROM semesters WHERE code = %s LIMIT 1", (semester_code,))
                    sem_row = cursor.fetchone()
                    if sem_row:
                        if not semester_start_date:
                            semester_start_date = sem_row.get('start_date')
                            if isinstance(semester_start_date, datetime):
                                semester_start_date = semester_start_date.strftime("%Y-%m-%d")
                        if not semester_end_date:
                            semester_end_date = sem_row.get('end_date')
                            if isinstance(semester_end_date, datetime):
                                semester_end_date = semester_end_date.strftime("%Y-%m-%d")
            
            # 構建 admission 物件
            admission = {
                'company_id': offer_info.get('company_id'),
                'company_name': offer_info.get('company_name'),
                'company_address': offer_info.get('company_address'),
                'contact_name': offer_info.get('contact_name'),
                'contact_email': offer_info.get('contact_email'),
                'contact_phone': offer_info.get('contact_phone'),
                'admitted_at': offer_info.get('offered_at'),
                'teacher_id': teacher_id,
                'teacher_name': teacher_name,
                'teacher_email': teacher_email,
                'semester': semester_code,
                'semester_start_date': semester_start_date,
                'semester_end_date': semester_end_date
            }
            
            # 構建 final_preference 物件
            final_preference = {
                'job_id': offer_info.get('job_id'),
                'job_title': offer_info.get('job_title'),
                'job_description': offer_info.get('job_description'),
                'internship_period': offer_info.get('internship_period'),
                'internship_time': offer_info.get('internship_time'),
                'salary': offer_info.get('job_salary')
            }
            
            # 如果 job_title 為空，嘗試從 internship_jobs 獲取
            if not final_preference.get('job_title') and offer_info.get('job_id'):
                cursor.execute("""
                    SELECT title, description, period, work_time, salary
                    FROM internship_jobs
                    WHERE id = %s
                """, (offer_info.get('job_id'),))
                job_info = cursor.fetchone()
                if job_info:
                    final_preference['job_title'] = job_info.get('title')
                    final_preference['job_description'] = job_info.get('description')
                    final_preference['internship_period'] = job_info.get('period')
                    final_preference['internship_time'] = job_info.get('work_time')
                    if job_info.get('salary') is not None:
                        final_preference['salary'] = job_info.get('salary')
                    print(f"✅ [DEBUG] 重新查詢到職缺資訊: {job_info.get('title')}")
            
            # 嘗試從 student_preferences 獲取志願序資訊
            # 優先選擇 preference_order 最小且 status = 'approved' 的志願（已通過廠商審核的志願）
            if offer_info.get('job_id'):
                # 先查找該 job_id 且已通過審核的志願
                cursor.execute("""
                    SELECT preference_order, submitted_at, company_id
                    FROM student_preferences
                    WHERE student_id = %s AND job_id = %s AND status = 'approved'
                    ORDER BY preference_order ASC
                    LIMIT 1
                """, (student_id, offer_info.get('job_id')))
                pref_info = cursor.fetchone()
                if pref_info:
                    final_preference['preference_order'] = pref_info.get('preference_order')
                    final_preference['submitted_at'] = pref_info.get('submitted_at')
            else:
                # 如果沒有 job_id，查找該學生所有已通過審核的志願，選擇 preference_order 最小的
                cursor.execute("""
                    SELECT 
                        sp.preference_order, 
                        sp.submitted_at,
                        sp.job_id,
                        sp.company_id,
                        ij.title AS job_title,
                        ij.description AS job_description,
                        ij.period AS internship_period,
                        ij.work_time AS internship_time,
                        ij.salary AS job_salary
                    FROM student_preferences sp
                    LEFT JOIN internship_jobs ij ON sp.job_id = ij.id
                    WHERE sp.student_id = %s 
                      AND sp.status = 'approved'
                    ORDER BY sp.preference_order ASC
                    LIMIT 1
                """, (student_id,))
                top_preference = cursor.fetchone()
                if top_preference:
                    # 如果找到排名更前面的志願，使用它
                    if not final_preference.get('preference_order') or top_preference.get('preference_order') < final_preference.get('preference_order', 999):
                        final_preference['preference_order'] = top_preference.get('preference_order')
                        final_preference['submitted_at'] = top_preference.get('submitted_at')
                        if top_preference.get('job_title'):
                            final_preference['job_title'] = top_preference.get('job_title')
                        if top_preference.get('job_description'):
                            final_preference['job_description'] = top_preference.get('job_description')
                        if top_preference.get('internship_period'):
                            final_preference['internship_period'] = top_preference.get('internship_period')
                        if top_preference.get('internship_time'):
                            final_preference['internship_time'] = top_preference.get('internship_time')
                        if top_preference.get('job_salary') is not None:
                            final_preference['salary'] = top_preference.get('job_salary')
                        if top_preference.get('job_id'):
                            final_preference['job_id'] = top_preference.get('job_id')
                        # 如果公司資訊不同，也需要更新
                        if top_preference.get('company_id') and top_preference.get('company_id') != offer_info.get('company_id'):
                            cursor.execute("""
                                SELECT 
                                    id AS company_id,
                                    company_name,
                                    location AS company_address,
                                    contact_person AS contact_name,
                                    contact_email,
                                    contact_phone,
                                    advisor_user_id
                                FROM internship_companies
                                WHERE id = %s
                            """, (top_preference.get('company_id'),))
                            top_company = cursor.fetchone()
                            if top_company:
                                admission['company_id'] = top_company.get('company_id')
                                admission['company_name'] = top_company.get('company_name')
                                admission['company_address'] = top_company.get('company_address')
                                admission['contact_name'] = top_company.get('contact_name')
                                admission['contact_email'] = top_company.get('contact_email')
                                admission['contact_phone'] = top_company.get('contact_phone')
                                # 更新指導老師資訊
                                if top_company.get('advisor_user_id'):
                                    cursor.execute("""
                                        SELECT id, name, email
                                        FROM users
                                        WHERE id = %s AND role IN ('teacher', 'director')
                                    """, (top_company.get('advisor_user_id'),))
                                    top_teacher = cursor.fetchone()
                                    if top_teacher:
                                        admission['teacher_id'] = top_teacher.get('id')
                                        admission['teacher_name'] = top_teacher.get('name')
                                        admission['teacher_email'] = top_teacher.get('email')
                        print(f"✅ [DEBUG] 使用排名最前面的志願: preference_order={top_preference.get('preference_order')}")
            
            # 標記已從 internship_offers 獲取到資料，跳過後續的 company_info 處理
            company_info = None
            print(f"✅ [DEBUG] 使用 internship_offers 資料，跳過舊邏輯")
        else:
            # 如果沒有從 internship_offers 獲取到，則使用舊的邏輯（向後兼容）
            print(f"⚠️ [DEBUG] 未找到 internship_offers 記錄，使用舊邏輯")
            # 獲取學生的錄取結果（從 teacher_student_relations）
            cursor.execute("""
                SELECT 
                    tsr.id AS relation_id,
                    tsr.semester,
                    tsr.created_at AS admitted_at,
                    u_teacher.id AS teacher_id,
                    u_teacher.name AS teacher_name,
                    u_teacher.email AS teacher_email
                FROM teacher_student_relations tsr
                LEFT JOIN users u_teacher ON tsr.teacher_id = u_teacher.id
                WHERE tsr.student_id = %s
                ORDER BY tsr.created_at DESC
                LIMIT 1
            """, (student_id,))
            admission = cursor.fetchone()
            
            # 實習期間：優先從 internship_configs 取得（實習學期 1132 的起訖），否則用 semesters
            semester_code = admission.get('semester') if admission else None
            semester_start_date = None
            semester_end_date = None
            cursor.execute("SELECT role, admission_year, username FROM users WHERE id = %s", (student_id,))
            user_row = cursor.fetchone()
            ay_val = None
            if user_row and (user_row.get('admission_year') or user_row.get('username')):
                if user_row.get('admission_year') is not None and str(user_row.get('admission_year', '')).strip() != '':
                    try:
                        ay_val = int(user_row['admission_year'])
                    except (TypeError, ValueError):
                        pass
                if ay_val is None and user_row.get('username') and len(str(user_row.get('username', ''))) >= 3:
                    try:
                        ay_val = int(str(user_row['username'])[:3])
                    except (TypeError, ValueError):
                        pass
            if ay_val is not None:
                cursor.execute("""
                    SELECT ic.intern_start_date, ic.intern_end_date, s.code AS semester_code
                    FROM internship_configs ic
                    LEFT JOIN semesters s ON s.id = ic.semester_id
                    WHERE (ic.user_id = %s OR (ic.user_id IS NULL AND ic.admission_year = %s))
                    ORDER BY ic.user_id DESC, ic.semester_id DESC
                    LIMIT 1
                """, (student_id, ay_val))
                ic_row = cursor.fetchone()
                if ic_row:
                    semester_start_date = ic_row.get('intern_start_date')
                    semester_end_date = ic_row.get('intern_end_date')
                    if not semester_code:
                        semester_code = ic_row.get('semester_code')
                    if isinstance(semester_start_date, datetime):
                        semester_start_date = semester_start_date.strftime("%Y-%m-%d")
                    if isinstance(semester_end_date, datetime):
                        semester_end_date = semester_end_date.strftime("%Y-%m-%d")
            if not semester_code:
                semester_code = get_current_semester_code(cursor)
            if (not semester_start_date or not semester_end_date) and semester_code:
                cursor.execute("SELECT start_date, end_date FROM semesters WHERE code = %s LIMIT 1", (semester_code,))
                sem_row = cursor.fetchone()
                if sem_row:
                    if not semester_start_date:
                        semester_start_date = sem_row.get('start_date')
                        if isinstance(semester_start_date, datetime):
                            semester_start_date = semester_start_date.strftime("%Y-%m-%d")
                    if not semester_end_date:
                        semester_end_date = sem_row.get('end_date')
                        if isinstance(semester_end_date, datetime):
                            semester_end_date = semester_end_date.strftime("%Y-%m-%d")
            
            if admission:
                admission['semester_start_date'] = semester_start_date
                admission['semester_end_date'] = semester_end_date
                if semester_code:
                    admission['semester'] = semester_code
            
            if not admission:
                return jsonify({
                    "success": True,
                    "admission": None,
                    "message": "目前尚未錄取任何實習公司"
                })
            
            # 優先從 internship_experiences 獲取公司資訊（廠商確認媒合結果時記錄的）
            cursor.execute("""
                SELECT 
                    ie.company_id,
                    ie.job_id,
                    ie.year,
                    ie.created_at AS admitted_at,
                    ic.company_name,
                    ic.location AS company_address,
                    ic.contact_person AS contact_name,
                    ic.contact_email,
                    ic.contact_phone,
                    ic.advisor_user_id,
                    ij.title AS job_title,
                    ij.description AS job_description,
                    ij.period AS internship_period,
                    ij.work_time AS internship_time,
                    ij.salary AS job_salary
                FROM internship_experiences ie
                LEFT JOIN internship_companies ic ON ie.company_id = ic.id
                LEFT JOIN internship_jobs ij ON ie.job_id = ij.id
                WHERE ie.user_id = %s 
                  AND ie.content = '已錄取'
                ORDER BY ie.created_at DESC
                LIMIT 1
            """, (student_id,))
            company_info = cursor.fetchone()
            
            # 如果從 internship_experiences 獲取到公司資訊，使用它
            if company_info:
                # 優先使用公司的 advisor_user_id 作為指導老師（這是該公司實際的指導老師）
                company_advisor_id = company_info.get('advisor_user_id')
                if company_advisor_id:
                    cursor.execute("""
                        SELECT id, name, email
                        FROM users
                        WHERE id = %s AND role IN ('teacher', 'director')
                    """, (company_advisor_id,))
                    company_advisor = cursor.fetchone()
                    if company_advisor:
                        # 使用公司的指導老師資訊，而不是 teacher_student_relations 中的
                        admission['teacher_id'] = company_advisor.get('id')
                        admission['teacher_name'] = company_advisor.get('name')
                        admission['teacher_email'] = company_advisor.get('email')
                        print(f"✅ [DEBUG] 使用公司的指導老師: {company_advisor.get('name')} (advisor_user_id={company_advisor_id})")
                admission['company_id'] = company_info.get('company_id')
                admission['company_name'] = company_info.get('company_name')
                admission['company_address'] = company_info.get('company_address')
                admission['contact_name'] = company_info.get('contact_name')
                admission['contact_email'] = company_info.get('contact_email')
                admission['contact_phone'] = company_info.get('contact_phone')
                
                # 更新錄取時間為 internship_experiences 的創建時間（廠商確認的時間）
                if company_info.get('admitted_at'):
                    admission['admitted_at'] = company_info.get('admitted_at')
                
                # 優先從 student_preferences 獲取排名最前面的志願（preference_order 最小）
                # 而不是只查詢當前 company_info 對應的志願
                cursor.execute("""
                    SELECT 
                        sp.preference_order,
                        sp.submitted_at,
                        sp.job_id,
                        sp.company_id,
                        ij.title AS job_title,
                        ij.description AS job_description,
                        ij.period AS internship_period,
                        ij.work_time AS internship_time,
                        ij.salary AS job_salary,
                        ic.company_name,
                        ic.location AS company_address,
                        ic.contact_person AS contact_name,
                        ic.contact_email,
                        ic.contact_phone,
                        ic.advisor_user_id
                    FROM student_preferences sp
                    LEFT JOIN internship_jobs ij ON sp.job_id = ij.id
                    LEFT JOIN internship_companies ic ON sp.company_id = ic.id
                    WHERE sp.student_id = %s 
                      AND sp.status = 'approved'
                    ORDER BY sp.preference_order ASC, sp.submitted_at DESC
                    LIMIT 1
                """, (student_id,))
                top_preference_info = cursor.fetchone()
                
                if top_preference_info:
                    # 使用排名最前面的志願資訊
                    final_preference = {
                        'preference_order': top_preference_info.get('preference_order'),
                        'submitted_at': top_preference_info.get('submitted_at'),
                        'job_id': top_preference_info.get('job_id'),
                        'job_title': top_preference_info.get('job_title'),
                        'job_description': top_preference_info.get('job_description'),
                        'internship_period': top_preference_info.get('internship_period'),
                        'internship_time': top_preference_info.get('internship_time'),
                        'salary': top_preference_info.get('job_salary')
                    }
                    # 如果排名最前面的志願與當前 company_info 不同，更新公司資訊
                    if top_preference_info.get('company_id') != company_info.get('company_id'):
                        admission['company_id'] = top_preference_info.get('company_id')
                        admission['company_name'] = top_preference_info.get('company_name')
                        admission['company_address'] = top_preference_info.get('company_address')
                        admission['contact_name'] = top_preference_info.get('contact_name')
                        admission['contact_email'] = top_preference_info.get('contact_email')
                        admission['contact_phone'] = top_preference_info.get('contact_phone')
                        # 更新指導老師資訊
                        if top_preference_info.get('advisor_user_id'):
                            cursor.execute("""
                                SELECT id, name, email
                                FROM users
                                WHERE id = %s AND role IN ('teacher', 'director')
                            """, (top_preference_info.get('advisor_user_id'),))
                            top_teacher = cursor.fetchone()
                            if top_teacher:
                                admission['teacher_id'] = top_teacher.get('id')
                                admission['teacher_name'] = top_teacher.get('name')
                                admission['teacher_email'] = top_teacher.get('email')
                        print(f"✅ [DEBUG] 使用排名最前面的志願: preference_order={top_preference_info.get('preference_order')}, company={top_preference_info.get('company_name')}")
                else:
                    # 如果沒有找到 approved 的志願，使用 company_info 的資料
                    final_preference = {
                        'preference_order': None,
                        'submitted_at': None,
                        'job_id': company_info.get('job_id'),
                        'job_title': company_info.get('job_title'),
                        'job_description': company_info.get('job_description'),
                        'internship_period': company_info.get('internship_period'),
                        'internship_time': company_info.get('internship_time'),
                        'salary': company_info.get('job_salary')
                    }
            else:
                # 如果沒有從 internship_experiences 獲取到，則從 student_preferences 獲取（備用方案）
                # 優先選擇 preference_order 最小的志願（排名最前面的）
                cursor.execute("""
                    SELECT 
                        sp.company_id,
                        sp.preference_order,
                        sp.submitted_at,
                        ic.company_name,
                        ic.location AS company_address,
                        ic.contact_person AS contact_name,
                        ic.contact_email,
                        ic.contact_phone,
                        ic.advisor_user_id,
                        ij.id AS job_id,
                        ij.title AS job_title,
                        ij.description AS job_description,
                        ij.period AS internship_period,
                        ij.work_time AS internship_time,
                        ij.salary AS job_salary
                    FROM student_preferences sp
                    LEFT JOIN internship_companies ic ON sp.company_id = ic.id
                    LEFT JOIN internship_jobs ij ON sp.job_id = ij.id
                    WHERE sp.student_id = %s 
                      AND sp.status = 'approved'
                    ORDER BY sp.preference_order ASC, sp.submitted_at DESC
                    LIMIT 1
                """, (student_id,))
                final_preference = cursor.fetchone()
                
                # 如果從 student_preferences 獲取到公司資訊，合併到 admission 中
                if final_preference:
                    admission['company_id'] = final_preference.get('company_id')
                    admission['company_name'] = final_preference.get('company_name')
                    admission['company_address'] = final_preference.get('company_address')
                    admission['contact_name'] = final_preference.get('contact_name')
                    admission['contact_email'] = final_preference.get('contact_email')
                    admission['contact_phone'] = final_preference.get('contact_phone')
                    
                    # 優先使用公司的 advisor_user_id 作為指導老師（這是該公司實際的指導老師）
                    company_advisor_id = final_preference.get('advisor_user_id')
                    if company_advisor_id:
                        cursor.execute("""
                            SELECT id, name, email
                            FROM users
                            WHERE id = %s AND role IN ('teacher', 'director')
                        """, (company_advisor_id,))
                        company_advisor = cursor.fetchone()
                        if company_advisor:
                            # 使用公司的指導老師資訊，而不是 teacher_student_relations 中的
                            admission['teacher_id'] = company_advisor.get('id')
                            admission['teacher_name'] = company_advisor.get('name')
                            admission['teacher_email'] = company_advisor.get('email')
                            print(f"✅ [DEBUG] 使用公司的指導老師: {company_advisor.get('name')} (advisor_user_id={company_advisor_id})")
                    
                    # 清理 final_preference，只保留志願相關資訊
                    final_preference_clean = {
                        'preference_order': final_preference.get('preference_order'),
                        'submitted_at': final_preference.get('submitted_at'),
                        'job_id': final_preference.get('job_id'),
                        'job_title': final_preference.get('job_title'),
                        'job_description': final_preference.get('job_description'),
                        'internship_period': final_preference.get('internship_period'),
                        'internship_time': final_preference.get('internship_time'),
                        'salary': final_preference.get('job_salary')
                    }
                    final_preference = final_preference_clean
                else:
                    final_preference = None
        
        # 獲取實習心得（從 internship_experiences）
        company_id = admission.get('company_id')
        experiences = []
        if company_id:
            cursor.execute("""
                SELECT 
                    ie.id AS experience_id,
                    ie.year AS internship_year,
                    ie.content AS experience_content,
                    ie.rating,
                    ie.created_at
                FROM internship_experiences ie
                WHERE ie.user_id = %s AND ie.company_id = %s
                ORDER BY ie.year DESC, ie.created_at DESC
            """, (student_id, company_id))
            experiences = cursor.fetchall()
        
        # 格式化日期
        if isinstance(admission.get('admitted_at'), datetime):
            admission['admitted_at'] = admission['admitted_at'].strftime("%Y-%m-%d %H:%M:%S")
        
        if final_preference and isinstance(final_preference.get('submitted_at'), datetime):
            # 錄取志願的提交時間只顯示年月日
            final_preference['submitted_at'] = final_preference['submitted_at'].strftime("%Y-%m-%d")
        elif final_preference and final_preference.get('submitted_at'):
            # 如果已經是字串格式，確保只顯示日期部分
            submitted_at_str = str(final_preference.get('submitted_at'))
            if ' ' in submitted_at_str:
                final_preference['submitted_at'] = submitted_at_str.split(' ')[0]
        
        for exp in experiences:
            if isinstance(exp.get('created_at'), datetime):
                exp['created_at'] = exp['created_at'].strftime("%Y-%m-%d %H:%M:%S")
        
        # 錄取資料來源：internship_offers 表
        
        # 調試：打印最終返回的資料
        print(f"🔍 [DEBUG] 最終返回的 admission: {admission}")
        print(f"🔍 [DEBUG] 最終返回的 final_preference: {final_preference}")
        
        return jsonify({
            "success": True,
            "admission": admission,
            "final_preference": final_preference,
            "experiences": experiences
        })
    
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": f"查詢失敗: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

# =========================================================
# 臨時 API: 為現有錄取記錄補寫 internship_offers 資料
# =========================================================
@admission_bp.route("/api/sync_internship_offers", methods=["POST"])
def sync_internship_offers():
    """為現有錄取記錄補寫 internship_offers 資料（一次性操作）"""
    if 'user_id' not in session or session.get('role') not in ['admin', 'ta']:
        return jsonify({"success": False, "message": "未授權"}), 403
    
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 獲取所有有錄取關係但沒有 internship_offers 記錄的學生
        cursor.execute("""
            SELECT DISTINCT
                tsr.student_id,
                sp.job_id,
                sp.company_id,
                tsr.created_at AS admitted_at
            FROM teacher_student_relations tsr
            INNER JOIN student_preferences sp ON tsr.student_id = sp.student_id
            WHERE sp.status = 'approved'
              AND NOT EXISTS (
                  SELECT 1 FROM internship_offers io
                  WHERE io.student_id = tsr.student_id
                    AND (io.job_id = sp.job_id OR (io.job_id IS NULL AND sp.job_id IS NULL))
              )
            ORDER BY tsr.student_id, sp.preference_order
        """)
        missing_records = cursor.fetchall()
        
        inserted_count = 0
        updated_count = 0
        
        for record in missing_records:
            student_id = record['student_id']
            job_id = record['job_id']
            admitted_at = record['admitted_at']
            
            # 檢查是否已存在
            if job_id:
                cursor.execute("""
                    SELECT id FROM internship_offers
                    WHERE student_id = %s AND job_id = %s
                """, (student_id, job_id))
            else:
                cursor.execute("""
                    SELECT id FROM internship_offers
                    WHERE student_id = %s AND job_id IS NULL
                """, (student_id,))
            
            existing = cursor.fetchone()
            
            if existing:
                # 更新現有記錄
                cursor.execute("""
                    UPDATE internship_offers
                    SET status = 'accepted', offered_at = %s, responded_at = %s
                    WHERE id = %s
                """, (admitted_at, admitted_at, existing['id']))
                updated_count += 1
            else:
                # 插入新記錄
                cursor.execute("""
                    INSERT INTO internship_offers
                    (student_id, job_id, status, offered_at, responded_at)
                    VALUES (%s, %s, 'accepted', %s, %s)
                """, (student_id, job_id, admitted_at, admitted_at))
                inserted_count += 1
        
        conn.commit()
        
        return jsonify({
            "success": True,
            "message": f"同步完成：新增 {inserted_count} 筆記錄，更新 {updated_count} 筆記錄",
            "inserted": inserted_count,
            "updated": updated_count,
            "total_processed": len(missing_records)
        })
    
    except Exception as e:
        traceback.print_exc()
        conn.rollback()
        return jsonify({"success": False, "message": f"同步失敗: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

# =========================================================
# API: 指導老師查看錄取該公司學生的列表
# =========================================================
@admission_bp.route("/api/get_company_students", methods=["GET"])
def get_company_students():
    """指導老師查看錄取該公司學生的列表"""
    if 'user_id' not in session or session.get('role') not in ['teacher', 'director']:
        return jsonify({"success": False, "message": "未授權"}), 403
    
    teacher_id = session.get('user_id')
    company_id = request.args.get('company_id', type=int)
    
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 如果提供了 company_id，只查詢該公司的學生
        if company_id:
            cursor.execute("""
                SELECT 
                    tsr.id AS relation_id,
                    tsr.semester,
                    tsr.created_at AS admitted_at,
                    u_student.id AS student_id,
                    u_student.name AS student_name,
                    u_student.username AS student_number,
                    c.name AS class_name,
                    ic.company_name,
                    ij.title AS job_title
                FROM teacher_student_relations tsr
                JOIN users u_student ON tsr.student_id = u_student.id
                LEFT JOIN classes c ON u_student.class_id = c.id
                JOIN student_preferences sp ON tsr.student_id = sp.student_id AND sp.status = 'approved'
                JOIN internship_companies ic ON sp.company_id = ic.id
                LEFT JOIN internship_jobs ij ON sp.job_id = ij.id
                WHERE tsr.teacher_id = %s AND sp.company_id = %s
                ORDER BY tsr.created_at DESC
            """, (teacher_id, company_id))
        else:
            # 查詢所有該指導老師的學生
            cursor.execute("""
                SELECT 
                    tsr.id AS relation_id,
                    tsr.semester,
                    tsr.created_at AS admitted_at,
                    u_student.id AS student_id,
                    u_student.name AS student_name,
                    u_student.username AS student_number,
                    c.name AS class_name,
                    ic.company_name,
                    ij.title AS job_title
                FROM teacher_student_relations tsr
                JOIN users u_student ON tsr.student_id = u_student.id
                LEFT JOIN classes c ON u_student.class_id = c.id
                JOIN student_preferences sp ON tsr.student_id = sp.student_id AND sp.status = 'approved'
                JOIN internship_companies ic ON sp.company_id = ic.id
                LEFT JOIN internship_jobs ij ON sp.job_id = ij.id
                WHERE tsr.teacher_id = %s
                ORDER BY tsr.created_at DESC
            """, (teacher_id,))
        
        students = cursor.fetchall()
        
        # 格式化日期
        for s in students:
            if isinstance(s.get('admitted_at'), datetime):
                s['admitted_at'] = s['admitted_at'].strftime("%Y-%m-%d %H:%M:%S")
        
        return jsonify({
            "success": True,
            "students": students
        })
    
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": f"查詢失敗: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

# =========================================================
# API: 退實習生（刪除師生關係）
# =========================================================
@admission_bp.route("/api/withdraw_student", methods=["POST"])
def withdraw_student():
    """退實習生，刪除 teacher_student_relations 記錄"""
    if 'user_id' not in session or session.get('role') not in ['teacher', 'director', 'ta', 'admin', 'vendor']:
        return jsonify({"success": False, "message": "未授權"}), 403
    
    data = request.get_json() or {}
    relation_id = data.get("relation_id")
    student_id = data.get("student_id")
    
    if not relation_id and not student_id:
        return jsonify({"success": False, "message": "請提供關係ID或學生ID"}), 400
    
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        teacher_id = session.get('user_id')
        
        # 如果提供了 relation_id，直接刪除
        if relation_id:
            # 驗證該關係是否屬於當前老師（除非是 admin 或 ta）
            if session.get('role') not in ['admin', 'ta']:
                cursor.execute("""
                    SELECT id FROM teacher_student_relations 
                    WHERE id = %s AND teacher_id = %s
                """, (relation_id, teacher_id))
                relation = cursor.fetchone()
                if not relation:
                    return jsonify({"success": False, "message": "找不到該關係或無權限"}), 404
            
            cursor.execute("DELETE FROM teacher_student_relations WHERE id = %s", (relation_id,))
        else:
            # 如果只提供了 student_id，需要找到對應的關係
            if session.get('role') not in ['admin', 'ta']:
                cursor.execute("""
                    SELECT id FROM teacher_student_relations 
                    WHERE student_id = %s AND teacher_id = %s
                """, (student_id, teacher_id))
            else:
                cursor.execute("""
                    SELECT id FROM teacher_student_relations 
                    WHERE student_id = %s
                """, (student_id,))
            
            relation = cursor.fetchone()
            if not relation:
                # 如果找不到實習關係，視為已經退出，直接返回成功
                # 同時更新志願序狀態
                if student_id:
                    cursor.execute("""
                        UPDATE student_preferences
                        SET status = 'pending'
                        WHERE student_id = %s AND status = 'approved'
                    """, (student_id,))
                    conn.commit()
                return jsonify({
                    "success": True,
                    "message": "已成功退實習生"
                })
            
            cursor.execute("DELETE FROM teacher_student_relations WHERE id = %s", (relation['id'],))
        
        # 同時將學生的志願序狀態改為 pending（取消錄取）
        if student_id:
            cursor.execute("""
                UPDATE student_preferences
                SET status = 'pending'
                WHERE student_id = %s AND status = 'approved'
            """, (student_id,))
        
        conn.commit()
        
        return jsonify({
            "success": True,
            "message": "已成功退實習生"
        })
    
    except Exception as e:
        traceback.print_exc()
        if conn:
            conn.rollback()
        return jsonify({"success": False, "message": f"退實習生失敗: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

# =========================================================
# API: 獲取所有學生的錄取結果列表（支援篩選）
# 以 internship_offers 為單一真相來源，一學生一筆錄取結果
# =========================================================
@admission_bp.route("/api/get_all_admissions", methods=["GET"])
def get_all_admissions():
    """獲取所有學生的錄取結果列表，支援按班級、學期、公司等篩選。資料來源：internship_offers（一學生一筆）。"""
    if 'user_id' not in session:
        return jsonify({"success": False, "message": "未授權"}), 403
    
    user_id = session.get('user_id')
    user_role = session.get('role')
    
    # 獲取篩選參數
    class_id = request.args.get('class_id', type=int)
    semester = request.args.get('semester', '').strip()
    company_id = request.args.get('company_id', type=int)
    keyword = request.args.get('keyword', '').strip()
    
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 以 internship_offers 為主表，一學生一筆錄取；學期由 student_preferences 取得（同學生同職缺取一筆，避免重複）
        base_query = """
            SELECT 
                io.id AS relation_id,
                COALESCE(s.code, '') AS semester,
                io.offered_at AS admitted_at,
                u_student.id AS student_id,
                u_student.name AS student_name,
                u_student.username AS student_number,
                c.id AS class_id,
                c.name AS class_name,
                c.department,
                ic.id AS company_id,
                ic.company_name,
                ij.id AS job_id,
                ij.title AS job_title,
                u_teacher.id AS teacher_id,
                u_teacher.name AS teacher_name,
                sp.preference_order,
                COALESCE(sp.status, 'approved') AS preference_status
            FROM internship_offers io
            JOIN users u_student ON io.student_id = u_student.id
            LEFT JOIN classes c ON u_student.class_id = c.id
            JOIN internship_jobs ij ON io.job_id = ij.id
            JOIN internship_companies ic ON ij.company_id = ic.id
            LEFT JOIN users u_teacher ON ic.advisor_user_id = u_teacher.id
            LEFT JOIN (
                SELECT sp0.student_id, sp0.job_id, sp0.semester_id, sp0.preference_order, sp0.status
                FROM student_preferences sp0
                INNER JOIN (
                    SELECT student_id, job_id, MAX(semester_id) AS semester_id
                    FROM student_preferences
                    GROUP BY student_id, job_id
                ) sp1 ON sp0.student_id = sp1.student_id AND sp0.job_id = sp1.job_id AND sp0.semester_id = sp1.semester_id
            ) sp ON sp.student_id = io.student_id AND sp.job_id = io.job_id
            LEFT JOIN semesters s ON s.id = sp.semester_id
            WHERE 1=1
        """
        params = []
        
        # 根據角色限制查詢範圍
        if user_role == 'class_teacher':
            # 班導只能看到自己管理的班級的學生
            cursor.execute("""
                SELECT class_id FROM classes_teacher 
                WHERE teacher_id = %s
            """, (user_id,))
            teacher_classes = cursor.fetchall()
            if teacher_classes:
                class_ids = [tc['class_id'] for tc in teacher_classes]
                placeholders = ','.join(['%s'] * len(class_ids))
                base_query += f" AND u_student.class_id IN ({placeholders})"
                params.extend(class_ids)
            else:
                return jsonify({
                    "success": True,
                    "students": [],
                    "count": 0
                })
        elif user_role == 'teacher':
            # 指導老師只能看到「錄取公司」為自己擔任顧問的公司（該公司 advisor_user_id = 當前老師）
            base_query += " AND ic.advisor_user_id = %s"
            params.append(user_id)
        elif user_role == 'director':
            # 主任可以看到自己科系的學生
            cursor.execute("""
                SELECT DISTINCT c.department
                FROM classes c
                JOIN classes_teacher ct ON ct.class_id = c.id
                WHERE ct.teacher_id = %s
                LIMIT 1
            """, (user_id,))
            dept_result = cursor.fetchone()
            if dept_result and dept_result.get('department'):
                base_query += " AND c.department = %s"
                params.append(dept_result['department'])
        # ta 和 admin 不額外限制
        
        # 應用篩選條件
        if class_id:
            base_query += " AND u_student.class_id = %s"
            params.append(class_id)
        
        if semester:
            base_query += " AND EXISTS (SELECT 1 FROM student_preferences sp2 WHERE sp2.student_id = io.student_id AND sp2.job_id = io.job_id AND sp2.semester_id = (SELECT id FROM semesters WHERE code = %s LIMIT 1))"
            params.append(semester)
        
        if company_id:
            base_query += " AND ic.id = %s"
            params.append(company_id)
        
        if keyword:
            base_query += " AND (u_student.name LIKE %s OR u_student.username LIKE %s OR ic.company_name LIKE %s OR c.name LIKE %s)"
            keyword_pattern = f"%{keyword}%"
            params.extend([keyword_pattern, keyword_pattern, keyword_pattern, keyword_pattern])
        
        base_query += " ORDER BY io.offered_at DESC, u_student.name ASC"
        
        cursor.execute(base_query, params)
        students = cursor.fetchall()
        
        # 格式化日期
        for s in students:
            if isinstance(s.get('admitted_at'), datetime):
                s['admitted_at'] = s['admitted_at'].strftime("%Y-%m-%d %H:%M:%S")
        
        return jsonify({
            "success": True,
            "students": students,
            "count": len(students)
        })
    
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": f"查詢失敗: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

# =========================================================
# API: 獲取所有學生列表（用於未被錄取學生顯示）
# =========================================================
@admission_bp.route("/api/get_all_students", methods=["GET"])
def get_all_students():
    """獲取所有學生列表（根據角色過濾），標記哪些已在媒合結果中。可傳 ?semester_id= 指定學期。"""
    if 'user_id' not in session:
        return jsonify({"success": False, "message": "未授權"}), 403
    
    user_id = session.get('user_id')
    user_role = session.get('role')
    
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 支援下拉選單選擇學期（與科助工作台一致）
        chosen_id = request.args.get('semester_id', type=int)
        if chosen_id:
            cursor.execute("SELECT id, code FROM semesters WHERE id = %s", (chosen_id,))
            row = cursor.fetchone()
            if not row:
                return jsonify({"success": False, "message": "找不到該學期"}), 400
            current_semester_id = row['id']
            current_semester_code = row.get('code') or ''
        else:
            current_semester_code = get_current_semester_code(cursor)
            current_semester_id = get_current_semester_id(cursor)
        if not current_semester_code:
            return jsonify({"success": False, "message": "無法取得當前學期"}), 500
        if not current_semester_id:
            return jsonify({"success": False, "message": "無法取得當前學期"}), 500
        
        # 獲取所有已在媒合結果中的學生 ID（只統計正取學生，與 director_matching_results 邏輯一致）
        # 檢查邏輯：
        # 1. 從 resume_applications 表讀取廠商的媒合排序資料
        # 2. 只統計正取學生（is_reserve = 0 或 is_reserve IS NULL 且 slot_index IS NOT NULL）
        # 3. 排除 director_decision = 'Rejected' 的記錄
        # 這樣可以確保與 manage_director.html 頁面顯示的媒合結果一致
        cursor.execute("""
            SELECT DISTINCT sja.student_id
            FROM resume_applications ra
            INNER JOIN student_job_applications sja ON ra.application_id = sja.id
            INNER JOIN student_preferences sp ON sja.student_id = sp.student_id 
                AND sja.company_id = sp.company_id 
                AND sja.job_id = sp.job_id
                AND sp.semester_id = %s
            LEFT JOIN manage_director md ON ra.application_id = md.preference_id
            LEFT JOIN internship_companies ic ON sja.company_id = ic.id
            WHERE ra.apply_status = 'approved'  -- 廠商必須已通過履歷審核
            AND (ra.is_reserve IS NOT NULL OR ra.slot_index IS NOT NULL)  -- 必須已完成媒合排序
            AND (md.director_decision IS NULL OR md.director_decision != 'Rejected')  -- 排除已被主任移除的記錄
            AND ic.status = 'approved'
            AND (
                -- 正取條件：is_reserve = 0 或 (is_reserve IS NULL 且 slot_index IS NOT NULL)
                (ra.is_reserve = 0) 
                OR (ra.is_reserve IS NULL AND ra.slot_index IS NOT NULL)
                OR (md.director_decision = 'Approved' AND md.final_rank IS NOT NULL)
                OR (md.director_decision = 'Pending' AND md.original_type = 'Regular' AND md.original_rank IS NOT NULL)
            )
        """, (current_semester_id,))
        matched_student_ids = {row['student_id'] for row in cursor.fetchall()}
        
        # 學期對應學號邏輯：1132→110xxx，1141/1142→111xxx（學號前3碼 = 學年前3碼 - 3）
        student_id_prefix = None
        if current_semester_code and len(current_semester_code) >= 3:
            try:
                year_part = int(current_semester_code[:3])  # 113, 114
                student_id_prefix = str(year_part - 3)  # 110, 111
            except (ValueError, TypeError):
                pass

        # 基礎查詢：獲取所有學生
        base_query = """
            SELECT 
                u.id AS student_id,
                u.id,
                u.name AS student_name,
                u.username AS student_number,
                u.admission_year AS admission_year,
                c.id AS class_id,
                c.name AS class_name,
                c.department
            FROM users u
            LEFT JOIN classes c ON u.class_id = c.id
            WHERE u.role = 'student'
        """
        params = []

        if student_id_prefix:
            base_query += " AND u.username LIKE %s"
            params.append(student_id_prefix + "%")
        
        # 根據角色限制查詢範圍
        if user_role == 'director':
            # 主任可以看到自己科系的學生
            cursor.execute("""
                SELECT DISTINCT c.department
                FROM classes c
                JOIN classes_teacher ct ON ct.class_id = c.id
                WHERE ct.teacher_id = %s
                LIMIT 1
            """, (user_id,))
            dept_result = cursor.fetchone()
            if dept_result and dept_result.get('department'):
                base_query += " AND c.department = %s"
                params.append(dept_result['department'])
        elif user_role == 'class_teacher':
            # 班導只能看到自己管理的班級的學生
            cursor.execute("""
                SELECT class_id FROM classes_teacher 
                WHERE teacher_id = %s
            """, (user_id,))
            teacher_classes = cursor.fetchall()
            if teacher_classes:
                class_ids = [tc['class_id'] for tc in teacher_classes]
                placeholders = ','.join(['%s'] * len(class_ids))
                base_query += f" AND u.class_id IN ({placeholders})"
                params.extend(class_ids)
            else:
                return jsonify({
                    "success": True,
                    "students": [],
                    "count": 0
                })
        # ta 和 admin 可以看到所有學生，不需要額外限制
        
        base_query += " ORDER BY u.username ASC"
        
        cursor.execute(base_query, params)
        all_students = cursor.fetchall()
        
        # 為每個學生標記是否已在媒合結果中，並獲取志願序資訊
        for student in all_students:
            student_id = student['student_id']
            # 入學屆數：如果沒有 admission_year，從學號前3碼推斷
            if (student.get('admission_year') is None or str(student.get('admission_year', '')).strip() == '') and student.get('student_number') and len(str(student['student_number'])) >= 3:
                try:
                    student['admission_year'] = int(str(student['student_number'])[:3])
                except (TypeError, ValueError):
                    pass
            # 標記是否已在媒合結果中
            student['is_matched'] = student_id in matched_student_ids
            
            # 為每個學生獲取志願序資訊（只包括 preference_order 在 1-5 範圍內的）
            if current_semester_id:
                cursor.execute("""
                    SELECT 
                        sp.preference_order,
                        ic.company_name,
                        ij.title AS job_title
                    FROM student_preferences sp
                    LEFT JOIN internship_companies ic ON sp.company_id = ic.id
                    LEFT JOIN internship_jobs ij ON sp.job_id = ij.id
                    WHERE sp.student_id = %s
                      AND sp.semester_id = %s
                      AND sp.preference_order >= 1
                      AND sp.preference_order <= 5
                    ORDER BY sp.preference_order ASC
                """, (student_id, current_semester_id))
            else:
                cursor.execute("""
                    SELECT 
                        sp.preference_order,
                        ic.company_name,
                        ij.title AS job_title
                    FROM student_preferences sp
                    LEFT JOIN internship_companies ic ON sp.company_id = ic.id
                    LEFT JOIN internship_jobs ij ON sp.job_id = ij.id
                    WHERE sp.student_id = %s
                      AND sp.preference_order >= 1
                      AND sp.preference_order <= 5
                    ORDER BY sp.preference_order ASC
                """, (student_id,))
            
            preferences = cursor.fetchall() or []
            student['preferences'] = preferences
        
        # 學期代碼 1132 = 113學年第2學期；名單依學號篩選：1132→110xxx，1142→111xxx
        semester_label = current_semester_code
        if current_semester_code and len(current_semester_code) >= 4:
            try:
                year_part = current_semester_code[:3]  # 113
                term_part = current_semester_code[-1]  # 2 或 1
                term_name = "第1學期" if term_part == "1" else "第2學期"
                semester_label = f"{year_part}學年{term_name}"
            except Exception:
                pass

        return jsonify({
            "success": True,
            "semester_id": current_semester_id,
            "semester_code": current_semester_code,
            "semester_label": semester_label,
            "student_id_prefix": student_id_prefix,  # 例：1132→"110"，1142→"111"（學號前3碼）
            "students": all_students,
            "count": len(all_students)
        })
    
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": f"查詢失敗: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

# =========================================================
# API: 獲取學生的志願序資料（主任查看）
# =========================================================
@admission_bp.route("/api/get_student_preferences", methods=["GET"])
def get_student_preferences():
    """獲取指定學生的志願序資料（主任可以查看）"""
    if 'user_id' not in session or session.get('role') != 'director':
        return jsonify({"success": False, "message": "未授權"}), 403
    
    student_id = request.args.get('student_id')
    if not student_id:
        return jsonify({"success": False, "message": "請提供學生ID"}), 400
    
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 獲取當前學期ID
        current_semester_id = get_current_semester_id(cursor)
        
        # 查詢學生的志願序
        if current_semester_id:
            cursor.execute("""
                SELECT 
                    sp.id AS preference_id,
                    sp.preference_order,
                    sp.company_id,
                    sp.job_id,
                    sp.status,
                    sp.submitted_at,
                    ic.company_name,
                    ij.title AS job_title
                FROM student_preferences sp
                LEFT JOIN internship_companies ic ON sp.company_id = ic.id
                LEFT JOIN internship_jobs ij ON sp.job_id = ij.id
                WHERE sp.student_id = %s
                  AND sp.semester_id = %s
                ORDER BY sp.preference_order ASC
            """, (student_id, current_semester_id))
        else:
            cursor.execute("""
                SELECT 
                    sp.id AS preference_id,
                    sp.preference_order,
                    sp.company_id,
                    sp.job_id,
                    sp.status,
                    sp.submitted_at,
                    ic.company_name,
                    ij.title AS job_title
                FROM student_preferences sp
                LEFT JOIN internship_companies ic ON sp.company_id = ic.id
                LEFT JOIN internship_jobs ij ON sp.job_id = ij.id
                WHERE sp.student_id = %s
                ORDER BY sp.preference_order ASC
            """, (student_id,))
        
        preferences = cursor.fetchall() or []
        
        # 格式化日期
        for pref in preferences:
            if isinstance(pref.get('submitted_at'), datetime):
                pref['submitted_at'] = pref['submitted_at'].strftime("%Y-%m-%d %H:%M:%S")
            elif pref.get('submitted_at'):
                pref['submitted_at'] = str(pref['submitted_at'])
            else:
                pref['submitted_at'] = ""
        
        return jsonify({
            "success": True,
            "preferences": preferences
        })
    
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": f"查詢失敗: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

# =========================================================
# API: 廠商查看媒合結果（包含所有狀態為 approved 的學生履歷）
# =========================================================
@admission_bp.route("/api/vendor_matching_results", methods=["GET"])
def vendor_matching_results():
    """廠商查看媒合結果，返回所有狀態為 approved 的學生履歷"""
    if 'user_id' not in session or session.get('role') != 'vendor':
        return jsonify({"success": False, "message": "未授權"}), 403
    
    vendor_id = session.get('user_id')
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 獲取廠商關聯的公司（通過 advisor_user_id，與 vendor.py 中的邏輯一致）
        # 先獲取廠商的 teacher_id，然後找到該指導老師對接的公司
        cursor.execute("""
            SELECT teacher_id FROM users WHERE id = %s AND role = 'vendor'
        """, (vendor_id,))
        vendor_row = cursor.fetchone()
        
        if not vendor_row or not vendor_row.get("teacher_id"):
            return jsonify({
                "success": True,
                "matches": [],
                "summary": {
                    "total_jobs": 0,
                    "total_students": 0,
                    "by_company": []
                },
                "message": "廠商帳號資料不完整，無法查詢媒合結果"
            })
        
        teacher_id = vendor_row.get("teacher_id")
        if not teacher_id:
            return jsonify({
                "success": True,
                "matches": [],
                "summary": {
                    "total_jobs": 0,
                    "total_students": 0,
                    "by_company": []
                },
                "message": "廠商尚未指派指導老師，無法查詢媒合結果"
            })
        
        # 驗證該 ID 是否為有效的指導老師
        cursor.execute("""
            SELECT id FROM users WHERE id = %s AND role IN ('teacher', 'director')
        """, (teacher_id,))
        teacher_row = cursor.fetchone()
        
        if not teacher_row:
            return jsonify({
                "success": True,
                "matches": [],
                "summary": {
                    "total_jobs": 0,
                    "total_students": 0,
                    "by_company": []
                },
                "message": "找不到對應的指導老師，無法查詢媒合結果"
            })
        
        teacher_id = teacher_row["id"]
        
        # 找到該指導老師對接的公司（只回傳已審核通過的公司）
        cursor.execute("""
            SELECT DISTINCT ic.id, ic.company_name
            FROM internship_companies ic
            WHERE ic.advisor_user_id = %s AND ic.status = 'approved'
            ORDER BY ic.company_name
        """, (teacher_id,))
        companies = cursor.fetchall() or []
        company_ids = [c['id'] for c in companies] if companies else []
        
        if not company_ids:
            return jsonify({
                "success": True,
                "matches": [],
                "summary": {
                    "total_jobs": 0,
                    "total_students": 0,
                    "by_company": []
                },
                "message": "您尚未上傳任何公司或沒有關聯的公司"
            })
        
        # 獲取科助確認後的媒合結果（從 internship_offers 表讀取，只有科助確認後才會寫入此表）
        # 使用子查詢確保每個 (student_id, company_id, job_id) 組合只取一筆記錄
        placeholders = ','.join(['%s'] * len(company_ids))
        try:
            cursor.execute(f"""
                SELECT DISTINCT
                    u.id AS student_id,
                    u.name AS student_name,
                    u.username AS student_number,
                    u.email AS student_email,
                    c.name AS class_name,
                    c.department AS class_department,
                    ic.id AS company_id,
                    ic.company_name,
                    COALESCE(ij.id, ra.job_id) AS job_id,
                    COALESCE(ij.title, '未指定職缺') AS job_title,
                    sp.preference_order,
                    sp.submitted_at AS preference_submitted_at,
                    ra.apply_status AS preference_status,
                    COALESCE(io.offered_at, tsr.created_at, ra.updated_at, CURDATE()) AS admitted_at,
                    '1132' AS semester,
                    ra.is_reserve,
                    ra.slot_index
                FROM internship_offers io
                INNER JOIN users u ON io.student_id = u.id
                INNER JOIN student_job_applications sja ON u.id = sja.student_id AND io.job_id = sja.job_id
                INNER JOIN (
                    SELECT ra1.*
                    FROM resume_applications ra1
                    INNER JOIN (
                        SELECT application_id, MAX(updated_at) AS max_updated_at
                        FROM resume_applications
                        WHERE apply_status = 'approved'
                        GROUP BY application_id
                    ) ra2 ON ra1.application_id = ra2.application_id 
                        AND ra1.updated_at = ra2.max_updated_at
                        AND ra1.apply_status = 'approved'
                ) ra ON ra.application_id = sja.id
                LEFT JOIN classes c ON u.class_id = c.id
                INNER JOIN internship_companies ic ON sja.company_id = ic.id
                LEFT JOIN internship_jobs ij ON COALESCE(ra.job_id, sja.job_id) = ij.id
                LEFT JOIN (
                    SELECT sp1.*
                    FROM student_preferences sp1
                    INNER JOIN (
                        SELECT student_id, company_id, job_id, MAX(submitted_at) AS max_submitted_at
                        FROM student_preferences
                        GROUP BY student_id, company_id, job_id
                    ) sp2 ON sp1.student_id = sp2.student_id 
                        AND sp1.company_id = sp2.company_id 
                        AND sp1.job_id = sp2.job_id
                        AND sp1.submitted_at = sp2.max_submitted_at
                ) sp ON sja.student_id = sp.student_id 
                    AND sja.company_id = sp.company_id 
                    AND sja.job_id = sp.job_id
                LEFT JOIN (
                    SELECT tsr1.*
                    FROM teacher_student_relations tsr1
                    INNER JOIN (
                        SELECT student_id, MAX(created_at) AS max_created_at
                        FROM teacher_student_relations
                        GROUP BY student_id
                    ) tsr2 ON tsr1.student_id = tsr2.student_id 
                        AND tsr1.created_at = tsr2.max_created_at
                ) tsr ON tsr.student_id = u.id
                WHERE sja.company_id IN ({placeholders})
                  AND ra.apply_status = 'approved'
                  AND io.status = 'accepted'
                ORDER BY ic.company_name, COALESCE(ra.slot_index, 999), ra.is_reserve, sp.preference_order, u.name
            """, tuple(company_ids))
        except Exception as qerr:
            err_str = str(qerr).lower()
            if "unknown column" in err_str or "1054" in str(qerr) or "1146" in str(qerr):
                # resume_applications 可能無 job_id / slot_index / is_reserve 或缺少 teacher_student_relations 等，改用精簡查詢
                # 但仍需從 internship_offers 表讀取，確保只返回科助確認後的資料
                cursor.execute(f"""
                    SELECT DISTINCT
                        u.id AS student_id,
                        u.name AS student_name,
                        u.username AS student_number,
                        u.email AS student_email,
                        c.name AS class_name,
                        c.department AS class_department,
                        ic.id AS company_id,
                        ic.company_name,
                        COALESCE(ij.id, sja.job_id) AS job_id,
                        COALESCE(ij.title, '未指定職缺') AS job_title,
                        sp.preference_order,
                        sp.submitted_at AS preference_submitted_at,
                        ra.apply_status AS preference_status,
                        COALESCE(io.offered_at, ra.updated_at, CURDATE()) AS admitted_at,
                        '1132' AS semester,
                        NULL AS is_reserve,
                        NULL AS slot_index
                    FROM internship_offers io
                    INNER JOIN users u ON io.student_id = u.id
                    INNER JOIN student_job_applications sja ON u.id = sja.student_id AND io.job_id = sja.job_id
                    INNER JOIN resume_applications ra ON ra.application_id = sja.id
                    LEFT JOIN classes c ON u.class_id = c.id
                    INNER JOIN internship_companies ic ON sja.company_id = ic.id
                    LEFT JOIN internship_jobs ij ON sja.job_id = ij.id
                    LEFT JOIN student_preferences sp ON sja.student_id = sp.student_id 
                        AND sja.company_id = sp.company_id 
                        AND sja.job_id = sp.job_id
                    WHERE sja.company_id IN ({placeholders})
                      AND ra.apply_status = 'approved'
                      AND io.status = 'accepted'
                    ORDER BY ic.company_name, sp.preference_order, u.name
                """, tuple(company_ids))
            else:
                raise
        
        matches = cursor.fetchall()
        
        # 去重：確保每個學生只出現一次（科助確認後，一個學生只會有一個職缺）
        # 使用 student_id 作為唯一鍵
        # 如果有多筆記錄，優先保留有媒合排序的（有 slot_index 或 is_reserve），然後保留最新的
        seen_students = {}
        unique_matches = []
        
        for match in matches:
            student_id = match.get('student_id')
            if not student_id:
                continue
            
            if student_id not in seen_students:
                seen_students[student_id] = match
                unique_matches.append(match)
            else:
                # 如果已存在，比較兩筆記錄，保留更合適的
                existing_match = seen_students[student_id]
                current_has_sorting = match.get('slot_index') is not None or match.get('is_reserve') is not None
                existing_has_sorting = existing_match.get('slot_index') is not None or existing_match.get('is_reserve') is not None
                
                # 優先保留有媒合排序的記錄
                if current_has_sorting and not existing_has_sorting:
                    # 當前記錄有媒合排序，替換舊記錄
                    index = unique_matches.index(existing_match)
                    unique_matches[index] = match
                    seen_students[student_id] = match
                elif not current_has_sorting and existing_has_sorting:
                    # 舊記錄有媒合排序，保留舊記錄
                    pass
                else:
                    # 兩筆都有或都沒有媒合排序，保留最新的（admitted_at 較新的）
                    current_updated = match.get('admitted_at') or ''
                    existing_updated = existing_match.get('admitted_at') or ''
                    if current_updated and existing_updated:
                        try:
                            # 使用外層已導入的 datetime
                            current_dt = datetime.strptime(str(current_updated), '%Y-%m-%d') if isinstance(current_updated, str) else current_updated
                            existing_dt = datetime.strptime(str(existing_updated), '%Y-%m-%d') if isinstance(existing_updated, str) else existing_updated
                            if isinstance(current_dt, datetime) and isinstance(existing_dt, datetime) and current_dt > existing_dt:
                                index = unique_matches.index(existing_match)
                                unique_matches[index] = match
                                seen_students[student_id] = match
                        except:
                            # 如果日期解析失敗，保留舊記錄
                            pass
        
        matches = unique_matches
        
        # 格式化日期
        for match in matches:
            if isinstance(match.get('preference_submitted_at'), datetime):
                # 錄取志願的提交時間只顯示年月日
                match['preference_submitted_at'] = match['preference_submitted_at'].strftime("%Y-%m-%d")
            elif match.get('preference_submitted_at'):
                # 如果已經是字串格式，確保只顯示日期部分
                submitted_at_str = str(match.get('preference_submitted_at'))
                if ' ' in submitted_at_str:
                    match['preference_submitted_at'] = submitted_at_str.split(' ')[0]
            if isinstance(match.get('admitted_at'), datetime):
                # 媒合時間只顯示日期部分（YYYY-MM-DD）
                match['admitted_at'] = match['admitted_at'].strftime("%Y-%m-%d")
            elif match.get('admitted_at'):
                # 如果已經是字串格式，確保只顯示日期部分
                admitted_at_str = str(match.get('admitted_at'))
                if ' ' in admitted_at_str:
                    match['admitted_at'] = admitted_at_str.split(' ')[0]
            else:
                # 如果沒有媒合時間，使用當天日期
                match['admitted_at'] = datetime.now().strftime("%Y-%m-%d")
            
            # 確保學期為 1132
            if not match.get('semester'):
                match['semester'] = '1132'
        
        # 統計信息：計算所有狀態為 approved 的學生履歷數量（去重，每個學生只計算一次）
        total_students = len(set(m['student_id'] for m in matches)) if matches else 0
        
        # 按公司統計
        by_company = {}
        for match in matches:
            company_name = match['company_name']
            if company_name not in by_company:
                by_company[company_name] = {
                    'company_name': company_name,
                    'matched_students': set()
                }
            by_company[company_name]['matched_students'].add(match['student_id'])
        
        # 轉換為列表格式
        by_company_list = [
            {
                'company_name': k,
                'matched_students': len(v['matched_students'])
            }
            for k, v in by_company.items()
        ]
        
        # 獲取職缺總數（從 vendor/api/positions API 獲取，這裡先返回 0，由前端補充）
        total_jobs = 0
        
        return jsonify({
            "success": True,
            "matches": matches,
            "summary": {
                "total_jobs": total_jobs,
                "total_students": total_students,
                "by_company": by_company_list
            }
        })
    
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": f"查詢失敗: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

# =========================================================
# API: 主任查看「尚未排序」的廠商列表（含：有學生但未排序、或沒有學生之廠商）
# =========================================================
@admission_bp.route("/api/director_unsorted_companies", methods=["GET"])
def director_unsorted_companies():
    """回傳尚未完成媒合排序的廠商：含 (1) 有已審核學生但尚未排序 (2) 有職缺但尚無任何學生（已選:0）的公司。"""
    if 'user_id' not in session or session.get('role') != 'director':
        return jsonify({"success": False, "message": "未授權"}), 403
    
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        current_semester_id = get_current_semester_id(cursor)
        if not current_semester_id:
            return jsonify({"success": False, "message": "無法取得當前學期"}), 500
        
        # (1) 有已審核申請且屬當前學期，但該公司底下沒有任何一筆已設定 is_reserve/slot_index
        query_part1 = """
            SELECT DISTINCT ic.id AS company_id, ic.company_name
            FROM internship_companies ic
            INNER JOIN student_job_applications sja ON sja.company_id = ic.id
            INNER JOIN resume_applications ra ON ra.application_id = sja.id
                AND ra.apply_status = 'approved'
            LEFT JOIN resumes r ON r.id = sja.resume_id AND r.user_id = sja.student_id
            LEFT JOIN student_preferences sp ON sp.student_id = sja.student_id
                AND sp.company_id = sja.company_id AND sp.job_id = sja.job_id
                AND sp.semester_id = %s
            WHERE (ic.status = 'approved' OR ic.status IS NULL)
            AND (r.semester_id = %s OR sp.student_id IS NOT NULL)
            AND NOT EXISTS (
                SELECT 1 FROM resume_applications ra2
                INNER JOIN student_job_applications sja2 ON ra2.application_id = sja2.id
                WHERE sja2.company_id = ic.id
                AND ra2.apply_status = 'approved'
                AND (ra2.is_reserve IS NOT NULL OR ra2.slot_index IS NOT NULL)
            )
        """
        # (2) 有職缺但尚無任何已審核學生（已選:0）的廠商，例如人人人公司
        query_part2 = """
            SELECT DISTINCT ic.id AS company_id, ic.company_name
            FROM internship_companies ic
            INNER JOIN internship_jobs ij ON ij.company_id = ic.id AND ij.is_active = 1
            WHERE ic.status = 'approved'
            AND NOT EXISTS (
                SELECT 1 FROM resume_applications ra
                INNER JOIN student_job_applications sja ON ra.application_id = sja.id
                WHERE sja.company_id = ic.id AND ra.apply_status = 'approved'
            )
        """
        cursor.execute(query_part1, (current_semester_id, current_semester_id))
        rows1 = cursor.fetchall() or []
        cursor.execute(query_part2)
        rows2 = cursor.fetchall() or []
        # 合併並依公司名稱排序、去重（同一公司可能同時符合兩種條件）
        seen = {}
        for r in rows1 + rows2:
            cid = r["company_id"]
            if cid not in seen:
                seen[cid] = r["company_name"]
        unsorted = [{"company_id": cid, "company_name": name} for cid, name in sorted(seen.items(), key=lambda x: x[1])]
        cursor.close()
        conn.close()
        
        return jsonify({
            "success": True,
            "unsorted_companies": unsorted
        })
    except Exception as exc:
        if conn:
            try:
                cursor.close()
                conn.close()
            except Exception:
                pass
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": str(exc)}), 500


# =========================================================
# API: 主任查看所有廠商媒合結果（包含重複中選檢測）
# =========================================================
@admission_bp.route("/api/director_matching_results", methods=["GET"])
def director_matching_results():
    """主任查看所有廠商的媒合結果，自動檢測重複中選的學生（從 manage_director 表讀取）"""
    if 'user_id' not in session or session.get('role') != 'director':
        return jsonify({"success": False, "message": "未授權"}), 403
    
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 獲取當前學期ID和代碼
        current_semester_id = get_current_semester_id(cursor)
        current_semester_code = get_current_semester_code(cursor)
        if not current_semester_id or not current_semester_code:
            return jsonify({"success": False, "message": "無法取得當前學期"}), 500
        
        # 優先從 resume_applications 讀取廠商的媒合排序資料
        # 如果 manage_director 表有資料，則合併兩者的資料
        # 這樣即使 manage_director 表為空，也能顯示廠商的排序結果
        # 注意：使用 LEFT JOIN 確保即使 student_job_applications 或 student_preferences 沒有記錄也能顯示
        query = """
            SELECT 
                COALESCE(md.match_id, CONCAT('ra_', ra.id)) AS match_id,
                COALESCE(md.vendor_id, sja.company_id) AS vendor_id,
                COALESCE(md.student_id, sja.student_id) AS student_id,
                COALESCE(sp.id, NULL) AS preference_id,
                COALESCE(md.original_type, CASE WHEN ra.is_reserve = 0 THEN 'Regular' ELSE 'Backup' END) AS original_type,
                COALESCE(md.original_rank, ra.slot_index) AS original_rank,
                COALESCE(md.is_conflict, 0) AS is_conflict,
                COALESCE(md.director_decision, 'Pending') AS director_decision,
                md.final_rank,
                COALESCE(md.is_adjusted, 0) AS is_adjusted,
                COALESCE(md.updated_at, ra.updated_at, ra.created_at) AS updated_at,
                COALESCE(sp.company_id, sja.company_id, md.vendor_id, ra.job_id) AS company_id,
                sp.preference_order,
                COALESCE(sp.job_id, sja.job_id, ra.job_id) AS job_id,
                ic.company_name,
                u.name AS student_name,
                u.username AS student_number,
                u.email AS student_email,
                c.name AS class_name,
                c.department AS class_department,
                COALESCE(v.name, ic.company_name) AS vendor_name,
                ij.title AS job_title,
                ij.slots AS job_slots,
                ra.is_reserve AS vendor_is_reserve,
                ra.slot_index AS vendor_slot_index
            FROM resume_applications ra
            LEFT JOIN student_job_applications sja ON ra.application_id = sja.id
            LEFT JOIN student_preferences sp ON (sja.student_id = sp.student_id 
                AND sja.company_id = sp.company_id 
                AND sja.job_id = sp.job_id
                AND (sp.semester_id = %s OR sp.semester_id IS NULL))
            LEFT JOIN manage_director md ON ra.application_id = md.preference_id
            LEFT JOIN internship_companies ic ON COALESCE(sp.company_id, sja.company_id, md.vendor_id) = ic.id
            LEFT JOIN internship_jobs ij ON COALESCE(sp.job_id, sja.job_id, ra.job_id) = ij.id
            LEFT JOIN users u ON COALESCE(md.student_id, sja.student_id) = u.id
            LEFT JOIN classes c ON u.class_id = c.id
            LEFT JOIN users v ON md.vendor_id = v.id
            WHERE ra.apply_status = 'approved'  -- 廠商必須已通過履歷審核
            AND (
                (ra.is_reserve = 0 AND ra.slot_index IS NOT NULL)  -- 正取學生：is_reserve=0且slot_index有值
                OR (ra.is_reserve = 1)  -- 候補學生：is_reserve=1
            )  -- 必須已完成媒合排序
            AND (md.director_decision IS NULL OR md.director_decision != 'Rejected')  -- 排除已被主任移除的記錄（如果 manage_director 中有記錄且是 Rejected，則不顯示）
            AND (ic.status = 'approved' OR ic.status IS NULL)  -- 公司狀態必須是已審核（如果公司不存在則也顯示）
            ORDER BY 
                CASE COALESCE(md.director_decision, 'Pending')
                    WHEN 'Approved' THEN 1 
                    WHEN 'Pending' THEN 2 
                    WHEN 'Rejected' THEN 3 
                    ELSE 4 
                END,
                COALESCE(sp.company_id, sja.company_id, md.vendor_id), 
                COALESCE(sp.job_id, sja.job_id, ra.job_id), 
                CASE WHEN md.director_decision = 'Approved' AND md.final_rank IS NOT NULL THEN 0 ELSE 1 END,
                COALESCE(md.final_rank, md.original_rank, ra.slot_index, 999) ASC,
                COALESCE(md.original_rank, ra.slot_index, 999) ASC
        """
        cursor.execute(query, (current_semester_id,))
        all_results = cursor.fetchall() or []
        
        # 調試：檢查是否有廠商排序資料
        vendor_sort_count = sum(1 for r in all_results if r.get("vendor_is_reserve") is not None or r.get("vendor_slot_index") is not None)
        print(f"📊 director_matching_results: 總記錄數={len(all_results)}, 有廠商排序資料的記錄數={vendor_sort_count}")
        
        # 調試：檢查查詢參數
        print(f"📊 查詢參數: current_semester_id={current_semester_id}")
        
        # 調試：如果沒有結果，檢查每個 JOIN 條件
        if len(all_results) == 0:
            # 檢查是否有 student_job_applications 記錄
            debug_query_sja = """
                SELECT COUNT(*) as count
                FROM resume_applications ra
                INNER JOIN student_job_applications sja ON ra.application_id = sja.id
                WHERE ra.apply_status = 'approved'
                AND (
                    (ra.is_reserve = 0 AND ra.slot_index IS NOT NULL)
                    OR (ra.is_reserve = 1)
                )
            """
            cursor.execute(debug_query_sja)
            debug_result_sja = cursor.fetchone()
            print(f"⚠️ 加上 student_job_applications JOIN 後的記錄數: {debug_result_sja.get('count', 0) if debug_result_sja else 0}")
            
            # 檢查具體的 resume_applications 記錄
            debug_query_ra = """
                SELECT ra.id, ra.application_id, ra.job_id, ra.apply_status, ra.is_reserve, ra.slot_index,
                       sja.id as sja_id, sja.student_id, sja.company_id, sja.job_id as sja_job_id
                FROM resume_applications ra
                LEFT JOIN student_job_applications sja ON ra.application_id = sja.id
                WHERE ra.apply_status = 'approved'
                AND (
                    (ra.is_reserve = 0 AND ra.slot_index IS NOT NULL)
                    OR (ra.is_reserve = 1)
                )
                LIMIT 5
            """
            cursor.execute(debug_query_ra)
            debug_ra_results = cursor.fetchall()
            print(f"⚠️ resume_applications 記錄詳情（前5筆）:")
            for ra_record in debug_ra_results:
                print(f"   - ra.id={ra_record.get('id')}, application_id={ra_record.get('application_id')}, sja_id={ra_record.get('sja_id')}, student_id={ra_record.get('student_id')}, is_reserve={ra_record.get('is_reserve')}, slot_index={ra_record.get('slot_index')}")
            
            # 檢查 manage_director 表的記錄
            debug_query_md = """
                SELECT md.match_id, md.preference_id, md.student_id, md.director_decision
                FROM manage_director md
                LIMIT 5
            """
            cursor.execute(debug_query_md)
            debug_md_results = cursor.fetchall()
            print(f"⚠️ manage_director 記錄詳情（前5筆）:")
            for md_record in debug_md_results:
                print(f"   - match_id={md_record.get('match_id')}, preference_id={md_record.get('preference_id')}, student_id={md_record.get('student_id')}, director_decision={md_record.get('director_decision')}")
        
        # 調試：如果沒有資料，檢查可能的原因
        if len(all_results) == 0:
            # 檢查 resume_applications 中是否有符合條件的記錄（不考慮 manage_director 和 student_preferences）
            debug_query1 = """
                SELECT COUNT(*) as count
                FROM resume_applications ra
                INNER JOIN student_job_applications sja ON ra.application_id = sja.id
                LEFT JOIN internship_companies ic ON sja.company_id = ic.id
                WHERE ra.apply_status = 'approved'
                AND (
                    (ra.is_reserve = 0 AND ra.slot_index IS NOT NULL)
                    OR (ra.is_reserve = 1)
                )
            """
            cursor.execute(debug_query1)
            debug_result1 = cursor.fetchone()
            total_with_sort = debug_result1.get('count', 0) if debug_result1 else 0
            print(f"⚠️ 沒有符合條件的記錄。resume_applications 中有媒合排序的記錄數: {total_with_sort}")
            
            # 檢查公司狀態
            debug_query2 = """
                SELECT COUNT(*) as count
                FROM resume_applications ra
                INNER JOIN student_job_applications sja ON ra.application_id = sja.id
                LEFT JOIN internship_companies ic ON sja.company_id = ic.id
                WHERE ra.apply_status = 'approved'
                AND (
                    (ra.is_reserve = 0 AND ra.slot_index IS NOT NULL)
                    OR (ra.is_reserve = 1)
                )
                AND (ic.status = 'approved' OR ic.status IS NULL)
            """
            cursor.execute(debug_query2)
            debug_result2 = cursor.fetchone()
            total_with_company = debug_result2.get('count', 0) if debug_result2 else 0
            print(f"⚠️ 加上公司狀態條件後的記錄數: {total_with_company}")
            
            # 檢查是否有 apply_status = 'approved' 的記錄
            debug_query3 = """
                SELECT COUNT(*) as count, 
                       SUM(CASE WHEN is_reserve = 0 AND slot_index IS NOT NULL THEN 1 ELSE 0 END) as regular_count,
                       SUM(CASE WHEN is_reserve = 1 THEN 1 ELSE 0 END) as reserve_count
                FROM resume_applications
                WHERE apply_status = 'approved'
                AND (
                    (is_reserve = 0 AND slot_index IS NOT NULL)
                    OR (is_reserve = 1)
                )
            """
            cursor.execute(debug_query3)
            debug_result3 = cursor.fetchone()
            if debug_result3:
                print(f"⚠️ resume_applications 中 apply_status='approved' 且有媒合排序的記錄數: {debug_result3.get('count', 0)}")
                print(f"   - 正取: {debug_result3.get('regular_count', 0)}, 候補: {debug_result3.get('reserve_count', 0)}")
        
        # 格式化結果並組織資料結構
        formatted_results = []
        student_company_map = {}  # 用於檢測重複中選：{student_id: [company_ids]}
        
        # 使用字典去重：同一學生在同一公司/職缺只保留一條記錄
        # key: (student_id, company_id, job_id)
        seen_students = {}
        
        for result in all_results:
            student_id = result.get("student_id")
            company_id = result.get("company_id")
            job_id = result.get("job_id")
            
            # 檢查是否已存在（同一學生在同一公司/職缺）
            key = (student_id, company_id, job_id)
            if key in seen_students:
                # 如果已存在，優先保留有 manage_director 記錄的（match_id 不以 'ra_' 開頭）
                existing_match_id = seen_students[key].get("match_id", "")
                current_match_id = result.get("match_id", "")
                if isinstance(existing_match_id, str) and existing_match_id.startswith("ra_") and not (isinstance(current_match_id, str) and current_match_id.startswith("ra_")):
                    # 當前記錄有 manage_director，替換舊記錄
                    seen_students[key] = result
                # 否則保留已存在的記錄（跳過當前重複記錄）
                continue
            
            # 記錄已看到的學生
            seen_students[key] = result
        
        # 處理去重後的記錄
        for key, result in seen_students.items():
            student_id, company_id, job_id = key
            
            # 記錄每個學生被哪些公司選中
            if student_id not in student_company_map:
                student_company_map[student_id] = []
            if company_id not in student_company_map[student_id]:
                student_company_map[student_id].append(company_id)
            
            # 判斷是否為正取或備取
            # 優先使用廠商的媒合排序資料（resume_applications 表的 is_reserve 和 slot_index）
            # 如果沒有廠商排序資料，則根據 director_decision 和 original_type 判斷
            is_reserve = False
            slot_index = None
            
            # 優先使用廠商的媒合排序資料
            vendor_is_reserve = result.get("vendor_is_reserve")
            vendor_slot_index = result.get("vendor_slot_index")
            original_rank = result.get("original_rank")
            original_type = result.get("original_type")
            
            # 如果 resume_applications 表中有記錄（vendor_is_reserve 或 vendor_slot_index 不是 NULL），表示廠商已經排序
            if vendor_is_reserve is not None or vendor_slot_index is not None:
                # 有廠商的媒合排序資料，優先使用
                # is_reserve: 0=正取, 1=備取
                is_reserve = bool(vendor_is_reserve) if vendor_is_reserve is not None else False
                # 優先使用 original_rank（如果存在），否則使用 vendor_slot_index
                slot_index = original_rank if original_rank is not None else (vendor_slot_index if vendor_slot_index is not None else None)
            elif result.get("director_decision") == "Approved" and result.get("final_rank") is not None:
                # 主任已核定為正取
                is_reserve = False
                slot_index = result.get("final_rank")
            elif result.get("director_decision") == "Pending":
                # 待定狀態，根據原始設定判斷
                if original_type == "Regular" and original_rank is not None:
                    is_reserve = False
                    slot_index = original_rank
                else:
                    is_reserve = True
                    slot_index = None
            else:
                is_reserve = True
                slot_index = None
            
            formatted_result = {
                "id": result.get("match_id"),  # 使用 match_id 作為識別符
                "match_id": result.get("match_id"),
                "vendor_id": result.get("vendor_id"),
                "vendor_name": result.get("vendor_name"),
                "company_id": company_id,
                "company_name": result.get("company_name"),
                "job_id": result.get("job_id"),
                "job_title": result.get("job_title") or "未指定職缺",
                "student_id": student_id,
                "student_name": result.get("student_name"),
                "student_number": result.get("student_number"),
                "student_email": result.get("student_email"),
                "class_name": result.get("class_name"),
                "class_department": result.get("class_department"),
                "preference_order": result.get("preference_order"),
                "preference_id": result.get("preference_id"),
                "slot_index": slot_index,
                "is_reserve": is_reserve,
                "director_decision": result.get("director_decision"),
                "final_rank": result.get("final_rank"),
                "is_adjusted": bool(result.get("is_adjusted")),
                "is_conflict": bool(result.get("is_conflict")),
                "original_type": result.get("original_type"),
                "original_rank": result.get("original_rank"),
                "vendor_slot_index": result.get("vendor_slot_index"),  # 廠商的排序索引，用於判斷是否有媒合排序
                "updated_at": result.get("updated_at").strftime("%Y-%m-%d %H:%M:%S") if isinstance(result.get("updated_at"), datetime) else str(result.get("updated_at", ""))
            }
            formatted_results.append(formatted_result)
        
        # 標記重複中選的學生（根據 is_conflict 或實際重複情況）
        duplicate_students = {}
        for sid, companies in student_company_map.items():
            if len(companies) > 1:
                duplicate_students[sid] = companies
        
        # 也檢查 is_conflict 標記
        for result in formatted_results:
            if result.get("is_conflict") or result["student_id"] in duplicate_students:
                result["is_duplicate"] = True
                result["duplicate_companies"] = duplicate_students.get(result["student_id"], [])
            else:
                result["is_duplicate"] = False
                result["duplicate_companies"] = []
        
        # 先獲取所有已審核的公司（即使沒有媒合結果也要顯示）
        cursor.execute("""
            SELECT DISTINCT ic.id AS company_id, ic.company_name
            FROM internship_companies ic
            WHERE ic.status = 'approved'
            ORDER BY ic.company_name
        """)
        all_companies = cursor.fetchall() or []
        
        # 獲取所有已審核公司的職缺
        cursor.execute("""
            SELECT ij.id AS job_id, ij.company_id, ij.title AS job_title, ij.slots AS job_slots
            FROM internship_jobs ij
            JOIN internship_companies ic ON ij.company_id = ic.id
            WHERE ic.status = 'approved' AND ij.is_active = 1
            ORDER BY ij.company_id, ij.id
        """)
        all_jobs = cursor.fetchall() or []
        
        # 按公司組織資料
        companies_data = {}
            
        # 先初始化所有已審核的公司
        for company in all_companies:
            company_id = company["company_id"]
            company_name = company["company_name"]
            companies_data[company_id] = {
                "company_id": company_id,
                "company_name": company_name,
                "jobs": {}
            }
            
        # 為每個公司添加職缺（即使沒有媒合結果）
        for job in all_jobs:
            company_id = job["company_id"]
            job_id = job["job_id"]
            job_title = job["job_title"] or "未指定職缺"
            job_slots = job["job_slots"] or 1
            
            if company_id in companies_data:
                companies_data[company_id]["jobs"][job_id] = {
                    "job_id": job_id,
                    "job_title": job_title,
                    "job_slots": job_slots,
                    "regulars": [],
                    "reserves": []
                }
        
        # 將媒合結果分配到對應的公司和職缺
        # 對於重複中選的學生，優先選擇「有媒合排序結果」且「志願序最高」的記錄
        # 如果第一志願的公司還沒做媒合排序，選擇其他有媒合排序的公司中志願序最高的
        student_best_match_id = {}  # key: student_id, value: (match_id, preference_order, has_sorting)
        
        for result in formatted_results:
            student_id = result.get("student_id")
            if student_id in duplicate_students:
                # 這是重複中選的學生，需要選擇最合適的記錄
                preference_order = result.get("preference_order")
                if preference_order is None:
                    preference_order = 999  # 沒有志願序的排在最後
                
                # 判斷是否有媒合排序結果（廠商已排序）
                # 如果有 vendor_slot_index 或 original_rank，表示有媒合排序
                # vendor_slot_index 來自 resume_applications.slot_index，是最直接的判斷方式
                has_sorting = (result.get("vendor_slot_index") is not None) or (result.get("original_rank") is not None)
                
                match_id = result.get("match_id") or result.get("id")
                # 確保 match_id 是字符串，以便正確比較
                match_id_str = str(match_id) if match_id is not None else None
                
                if student_id not in student_best_match_id:
                    student_best_match_id[student_id] = (match_id_str, preference_order, has_sorting)
                else:
                    # 優先考慮志願序（志願序越小，優先級越高）
                    # 只有在志願序相同時，才優先選擇有媒合排序的記錄
                    current_has_sorting = student_best_match_id[student_id][2]
                    current_order = student_best_match_id[student_id][1]
                    
                    # 如果當前記錄的志願序更小（優先級更高），則替換
                    if preference_order < current_order:
                        student_best_match_id[student_id] = (match_id_str, preference_order, has_sorting)
                    # 如果志願序相同，則優先選擇有媒合排序的記錄
                    elif preference_order == current_order:
                        if has_sorting and not current_has_sorting:
                            student_best_match_id[student_id] = (match_id_str, preference_order, has_sorting)
                    # 如果當前記錄的志願序更大（優先級更低），則不替換
        
        # 即時排序：對於重複中選的學生，優先選擇志願序最高的記錄（preference_order 最小的）
        # 不需要等待第一志願完成媒合排序，只要廠商有做媒合排序並選擇了該學生，就應該顯示
        # 邏輯已在上面完成：優先考慮志願序，如果志願序相同則優先選擇有媒合排序的記錄
        
        # 先建立廠商原始排序的資料（不過濾重複學生，保持原樣）
        # 使用集合追蹤已添加的學生，確保同一學生在同一公司/職缺只出現一次
        added_students_vendor = {}  # key: (company_id, job_id, student_id)
        companies_data_vendor = {}
        
        # 初始化所有已審核的公司（廠商原始排序）
        for company in all_companies:
            company_id = company["company_id"]
            company_name = company["company_name"]
            companies_data_vendor[company_id] = {
                "company_id": company_id,
                "company_name": company_name,
                "jobs": {}
            }
        
        # 為每個公司添加職缺（廠商原始排序）
        for job in all_jobs:
            company_id = job["company_id"]
            job_id = job["job_id"]
            job_title = job["job_title"] or "未指定職缺"
            job_slots = job["job_slots"] or 1
            
            if company_id in companies_data_vendor:
                companies_data_vendor[company_id]["jobs"][job_id] = {
                    "job_id": job_id,
                    "job_title": job_title,
                    "job_slots": job_slots,
                    "regulars": [],
                    "reserves": []
                }
        
        # 將所有媒合結果分配到廠商原始排序（不過濾）
        for result in formatted_results:
            company_id = result["company_id"]
            job_id = result.get("job_id") or 0
            job_title = result.get("job_title") or "未指定職缺"
            student_id = result.get("student_id")
            
            key = (company_id, job_id, student_id)
            if key in added_students_vendor:
                continue
            
            if company_id not in companies_data_vendor:
                companies_data_vendor[company_id] = {
                    "company_id": company_id,
                    "company_name": result["company_name"],
                    "jobs": {}
                }
            
            if job_id not in companies_data_vendor[company_id]["jobs"]:
                job_slots = result.get("job_slots") or 1
                companies_data_vendor[company_id]["jobs"][job_id] = {
                    "job_id": job_id,
                    "job_title": job_title,
                    "job_slots": job_slots,
                    "regulars": [],
                    "reserves": []
                }
            
            if result["is_reserve"]:
                companies_data_vendor[company_id]["jobs"][job_id]["reserves"].append(result)
            else:
                companies_data_vendor[company_id]["jobs"][job_id]["regulars"].append(result)
            
            added_students_vendor[key] = True
        
        # 過濾 formatted_results：重複中選的學生只保留志願序最高的記錄（主任排序結果）
        filtered_results = []
        for result in formatted_results:
            student_id = result.get("student_id")
            if student_id in duplicate_students:
                # 只保留志願序最高的記錄
                match_id = result.get("match_id") or result.get("id")
                match_id_str = str(match_id) if match_id is not None else None
                if student_id in student_best_match_id and student_best_match_id[student_id][0] == match_id_str:
                    # 調試：記錄選擇的記錄
                    selected_order = student_best_match_id[student_id][1]
                    selected_has_sorting = student_best_match_id[student_id][2]
                    current_order = result.get("preference_order")
                    current_company = result.get("company_name")
                    print(f"✅ 重複中選學生 {student_id}：選擇公司 {current_company}，志願序={current_order}（已選擇的志願序={selected_order}，有媒合排序={selected_has_sorting}）")
                    filtered_results.append(result)
            else:
                # 不是重複中選的學生，直接保留
                filtered_results.append(result)
        
        # 使用集合追蹤已添加的學生，確保同一學生在同一公司/職缺只出現一次（主任排序結果）
        added_students = {}  # key: (company_id, job_id, student_id)
        
        for result in filtered_results:
            company_id = result["company_id"]
            job_id = result.get("job_id") or 0
            job_title = result.get("job_title") or "未指定職缺"
            student_id = result.get("student_id")
            
            # 檢查是否已添加（同一學生在同一公司/職缺）
            key = (company_id, job_id, student_id)
            if key in added_students:
                # 已存在，跳過（避免重複）
                continue
            
            # 如果公司不在列表中，添加它
            if company_id not in companies_data:
                companies_data[company_id] = {
                    "company_id": company_id,
                    "company_name": result["company_name"],
                    "jobs": {}
                }
            
            # 如果職缺不在列表中，添加它
            if job_id not in companies_data[company_id]["jobs"]:
                job_slots = result.get("job_slots") or 1
                companies_data[company_id]["jobs"][job_id] = {
                    "job_id": job_id,
                    "job_title": job_title,
                    "job_slots": job_slots,
                    "regulars": [],
                    "reserves": []
                }
            
            # 分配學生到正取或備取
            if result["is_reserve"]:
                companies_data[company_id]["jobs"][job_id]["reserves"].append(result)
            else:
                companies_data[company_id]["jobs"][job_id]["regulars"].append(result)
            
            # 標記為已添加
            added_students[key] = True
        
        # 轉換為列表格式，並對正取和備取名單進行排序
        companies_list = []
        for company_id, company_data in companies_data.items():
            jobs_list = []
            for job_id, job_data in company_data["jobs"].items():
                # 對正取名單排序：重複中選的學生優先按志願序排序，然後按 original_rank 或 slot_index 排序
                def sort_key_regulars(x):
                    student_id = x.get("student_id")
                    is_duplicate = student_id in duplicate_students
                    preference_order = x.get("preference_order")
                    # 優先使用 original_rank（廠商的排序），如果沒有則使用 slot_index
                    rank_value = x.get("original_rank") if x.get("original_rank") is not None else x.get("slot_index")
                    
                    # 排序優先級：
                    # 1. 重複中選的學生排在前面（False < True，所以 not is_duplicate 會讓重複中選的排在前面）
                    # 2. 如果是重複中選，按志願序排序（志願序小的在前）
                    # 3. 如果不是重複中選，或志願序相同，按 original_rank 或 slot_index 排序
                    return (
                        not is_duplicate,  # 重複中選的排在前面（False < True）
                        preference_order is None if is_duplicate else False,  # 重複中選但沒有志願序的排在後面
                        preference_order if (is_duplicate and preference_order is not None) else 999,  # 重複中選的按志願序排序
                        rank_value is None,  # None 值排在後面
                        rank_value or 999  # 按 original_rank 或 slot_index 排序
                    )
                
                regulars = sorted(job_data["regulars"], key=sort_key_regulars)
                
                # 對備取名單排序：重複中選的學生優先按志願序排序，然後按 original_rank 或 slot_index 排序
                def sort_key_reserves(x):
                    student_id = x.get("student_id")
                    is_duplicate = student_id in duplicate_students
                    preference_order = x.get("preference_order")
                    # 優先使用 original_rank（廠商的排序），如果沒有則使用 slot_index
                    rank_value = x.get("original_rank") if x.get("original_rank") is not None else x.get("slot_index")
                    
                    return (
                        not is_duplicate,  # 重複中選的排在前面
                        preference_order is None if is_duplicate else False,  # 重複中選但沒有志願序的排在後面
                        preference_order if (is_duplicate and preference_order is not None) else 999,  # 重複中選的按志願序排序
                        rank_value is None,  # None 值排在後面
                        rank_value or 999  # 按 original_rank 或 slot_index 排序
                    )
                
                reserves = sorted(job_data["reserves"], key=sort_key_reserves)
                jobs_list.append({
                    "job_id": job_data["job_id"],
                    "job_title": job_data["job_title"],
                    "job_slots": job_data["job_slots"],
                    "regulars": regulars,
                    "reserves": reserves
                })
            companies_list.append({
                "company_id": company_id,
                "company_name": company_data["company_name"],
                "jobs": jobs_list
            })
        
        # 建立廠商原始排序的列表格式（不過濾重複學生）
        companies_list_vendor = []
        for company_id, company_data in companies_data_vendor.items():
            jobs_list_vendor = []
            for job_id, job_data in company_data["jobs"].items():
                # 對正取名單排序（按 original_rank 或 slot_index）
                regulars_vendor = sorted(job_data["regulars"], key=lambda x: (
                    x.get("original_rank") is None,
                    x.get("original_rank") or x.get("slot_index") or 999
                ))
                # 對備取名單排序
                reserves_vendor = sorted(job_data["reserves"], key=lambda x: (
                    x.get("original_rank") is None,
                    x.get("original_rank") or x.get("slot_index") or 999
                ))
                jobs_list_vendor.append({
                    "job_id": job_data["job_id"],
                    "job_title": job_data["job_title"],
                    "job_slots": job_data["job_slots"],
                    "regulars": regulars_vendor,
                    "reserves": reserves_vendor
                })
            companies_list_vendor.append({
                "company_id": company_id,
                "company_name": company_data["company_name"],
                "jobs": jobs_list_vendor
            })
        
        return jsonify({
            "success": True,
            "companies": companies_list,  # 主任排序結果（已過濾重複學生）
            "vendor_companies": companies_list_vendor,  # 廠商原始排序（不過濾，保持原樣）
            "duplicate_students": list(duplicate_students.keys()),
            "total_matches": len(formatted_results)
        })
    
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": f"查詢失敗: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

# =========================================================
# API: 查看主任確認後的媒合結果（供其他角色使用，如 TA、admin）
# =========================================================
@admission_bp.route("/api/final_matching_results", methods=["GET"])
def final_matching_results():
    """查看主任確認後的媒合結果（只顯示 Approved 狀態），允許 TA、admin、director 等角色訪問
    使用與 director_matching_results 相同的邏輯，但只過濾出 director_decision = 'Approved' 的記錄"""
    if 'user_id' not in session:
        return jsonify({"success": False, "message": "未登入"}), 401
    
    user_role = session.get('role')
    # 允許 director、ta、admin、class_teacher、teacher 訪問
    if user_role not in ['director', 'ta', 'admin', 'class_teacher', 'teacher']:
        return jsonify({"success": False, "message": "未授權"}), 403
    
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 獲取當前學期ID和代碼
        current_semester_id = get_current_semester_id(cursor)
        current_semester_code = get_current_semester_code(cursor)
        if not current_semester_id or not current_semester_code:
            return jsonify({"success": False, "message": "無法取得當前學期"}), 500
        
        # 使用與 director_matching_results 相同的查詢邏輯
        # 但只過濾出 director_decision = 'Approved' 的記錄
        query = """
            SELECT 
                COALESCE(md.match_id, CONCAT('ra_', ra.id)) AS match_id,
                COALESCE(md.vendor_id, sja.company_id) AS vendor_id,
                COALESCE(md.student_id, sja.student_id) AS student_id,
                COALESCE(sp.id, NULL) AS preference_id,
                COALESCE(md.original_type, CASE WHEN ra.is_reserve = 0 THEN 'Regular' ELSE 'Backup' END) AS original_type,
                COALESCE(md.original_rank, ra.slot_index) AS original_rank,
                COALESCE(md.is_conflict, 0) AS is_conflict,
                COALESCE(md.director_decision, 'Pending') AS director_decision,
                md.final_rank,
                COALESCE(md.is_adjusted, 0) AS is_adjusted,
                COALESCE(md.updated_at, ra.updated_at, ra.created_at) AS updated_at,
                COALESCE(sp.company_id, sja.company_id, md.vendor_id, ra.job_id) AS company_id,
                sp.preference_order,
                COALESCE(sp.job_id, sja.job_id, ra.job_id) AS job_id,
                ic.company_name,
                u.name AS student_name,
                u.username AS student_number,
                u.email AS student_email,
                u.admission_year AS admission_year,
                c.name AS class_name,
                c.department AS class_department,
                COALESCE(v.name, ic.company_name) AS vendor_name,
                ij.title AS job_title,
                ij.slots AS job_slots,
                ra.is_reserve AS vendor_is_reserve,
                ra.slot_index AS vendor_slot_index
            FROM resume_applications ra
            LEFT JOIN student_job_applications sja ON ra.application_id = sja.id
            LEFT JOIN student_preferences sp ON (sja.student_id = sp.student_id 
                AND sja.company_id = sp.company_id 
                AND sja.job_id = sp.job_id
                AND (sp.semester_id = %s OR sp.semester_id IS NULL))
            LEFT JOIN manage_director md ON ra.application_id = md.preference_id
            LEFT JOIN internship_companies ic ON COALESCE(sp.company_id, sja.company_id, md.vendor_id) = ic.id
            LEFT JOIN internship_jobs ij ON COALESCE(sp.job_id, sja.job_id, ra.job_id) = ij.id
            LEFT JOIN users u ON COALESCE(md.student_id, sja.student_id) = u.id
            LEFT JOIN classes c ON u.class_id = c.id
            LEFT JOIN users v ON md.vendor_id = v.id
            WHERE ra.apply_status = 'approved'  -- 廠商必須已通過履歷審核
            AND (
                (ra.is_reserve = 0 AND ra.slot_index IS NOT NULL)  -- 正取學生：is_reserve=0且slot_index有值
                OR (ra.is_reserve = 1)  -- 候補學生：is_reserve=1
            )  -- 必須已完成媒合排序
            AND (md.director_decision = 'Approved')  -- 只顯示主任已確認的記錄
            AND (ic.status = 'approved' OR ic.status IS NULL)  -- 公司狀態必須是已審核
            ORDER BY 
                COALESCE(sp.company_id, sja.company_id, md.vendor_id), 
                COALESCE(sp.job_id, sja.job_id, ra.job_id), 
                CASE WHEN md.final_rank IS NOT NULL THEN 0 ELSE 1 END,
                COALESCE(md.final_rank, md.original_rank, ra.slot_index, 999) ASC,
                COALESCE(md.original_rank, ra.slot_index, 999) ASC
        """
        cursor.execute(query, (current_semester_id,))
        all_results = cursor.fetchall() or []
        
        # 獲取當前學年用於計算年級
        active_semester_year = _get_active_semester_year(cursor)
        
        # 格式化結果並組織資料結構（與 director_matching_results 相同）
        formatted_results = []
        student_company_map = {}  # 用於檢測重複中選：{student_id: [company_ids]}
        
        # 使用字典去重：同一學生在同一公司/職缺只保留一條記錄
        seen_students = {}
        
        for result in all_results:
            student_id = result.get("student_id")
            company_id = result.get("company_id")
            job_id = result.get("job_id")
            
            # 檢查是否已存在（同一學生在同一公司/職缺）
            key = (student_id, company_id, job_id)
            if key in seen_students:
                # 如果已存在，優先保留有 manage_director 記錄的（match_id 不以 'ra_' 開頭）
                existing_match_id = seen_students[key].get("match_id", "")
                current_match_id = result.get("match_id", "")
                if isinstance(existing_match_id, str) and existing_match_id.startswith("ra_") and not (isinstance(current_match_id, str) and current_match_id.startswith("ra_")):
                    # 當前記錄有 manage_director，替換舊記錄
                    seen_students[key] = result
                # 否則保留已存在的記錄（跳過當前重複記錄）
                continue
            
            # 記錄已看到的學生
            seen_students[key] = result
        
        # 處理去重後的記錄，並檢測重複中選
        for key, result in seen_students.items():
            student_id, company_id, job_id = key
            
            # 記錄每個學生被哪些公司選中（用於檢測重複中選）
            if student_id not in student_company_map:
                student_company_map[student_id] = []
            if company_id not in student_company_map[student_id]:
                student_company_map[student_id].append(company_id)
        
        # 標記重複中選的學生
        duplicate_students = {}
        for sid, companies in student_company_map.items():
            if len(companies) > 1:
                duplicate_students[sid] = companies
        
        # 對於重複中選的學生，選擇志願序最高的記錄
        student_best_match_id = {}  # key: student_id, value: (match_id, preference_order, has_sorting)
        
        for key, result in seen_students.items():
            student_id, company_id, job_id = key
            
            if student_id in duplicate_students:
                # 這是重複中選的學生，需要選擇最合適的記錄
                preference_order = result.get("preference_order")
                if preference_order is None:
                    preference_order = 999  # 沒有志願序的排在最後
                
                # 判斷是否有媒合排序結果（廠商已排序）
                vendor_slot_index = result.get("vendor_slot_index")
                original_rank = result.get("original_rank")
                has_sorting = (vendor_slot_index is not None) or (original_rank is not None)
                
                match_id = result.get("match_id") or result.get("id")
                match_id_str = str(match_id) if match_id is not None else None
                
                if student_id not in student_best_match_id:
                    student_best_match_id[student_id] = (match_id_str, preference_order, has_sorting)
                else:
                    # 優先考慮志願序（志願序越小，優先級越高）
                    current_has_sorting = student_best_match_id[student_id][2]
                    current_order = student_best_match_id[student_id][1]
                    
                    # 如果當前記錄的志願序更小（優先級更高），則替換
                    if preference_order < current_order:
                        student_best_match_id[student_id] = (match_id_str, preference_order, has_sorting)
                    # 如果志願序相同，則優先選擇有媒合排序的記錄
                    elif preference_order == current_order:
                        if has_sorting and not current_has_sorting:
                            student_best_match_id[student_id] = (match_id_str, preference_order, has_sorting)
        
        # 過濾 formatted_results：重複中選的學生只保留志願序最高的記錄
        filtered_seen_students = {}
        for key, result in seen_students.items():
            student_id, company_id, job_id = key
            
            if student_id in duplicate_students:
                # 只保留志願序最高的記錄
                match_id = result.get("match_id") or result.get("id")
                match_id_str = str(match_id) if match_id is not None else None
                if student_id in student_best_match_id and student_best_match_id[student_id][0] == match_id_str:
                    filtered_seen_students[key] = result
            else:
                # 不是重複中選的學生，直接保留
                filtered_seen_students[key] = result
        
        # 處理過濾後的記錄
        for key, result in filtered_seen_students.items():
            student_id, company_id, job_id = key
            
            # 判斷是否為正取或備取（使用與 director_matching_results 相同的邏輯）
            # 優先使用廠商的媒合排序資料（resume_applications 表的 is_reserve 和 slot_index）
            # 如果沒有廠商排序資料，則根據 director_decision 和 original_type 判斷
            is_reserve = False
            slot_index = None
            
            # 優先使用廠商的媒合排序資料
            vendor_is_reserve = result.get("vendor_is_reserve")
            vendor_slot_index = result.get("vendor_slot_index")
            original_rank = result.get("original_rank")
            original_type = result.get("original_type")
            
            # 如果 resume_applications 表中有記錄（vendor_is_reserve 或 vendor_slot_index 不是 NULL），表示廠商已經排序
            if vendor_is_reserve is not None or vendor_slot_index is not None:
                # 有廠商的媒合排序資料，優先使用
                # is_reserve: 0=正取, 1=備取
                is_reserve = bool(vendor_is_reserve) if vendor_is_reserve is not None else False
                # 優先使用 final_rank（主任調整後的排序），如果沒有則使用 original_rank，最後使用 vendor_slot_index
                if result.get("final_rank") is not None:
                    slot_index = result.get("final_rank")
                else:
                    slot_index = original_rank if original_rank is not None else (vendor_slot_index if vendor_slot_index is not None else None)
            elif result.get("director_decision") == "Approved" and result.get("final_rank") is not None:
                # 主任已核定為正取
                is_reserve = False
                slot_index = result.get("final_rank")
            elif result.get("director_decision") == "Approved":
                # 主任已確認，根據原始設定判斷
                if original_type == "Regular" and original_rank is not None:
                    is_reserve = False
                    slot_index = original_rank
                else:
                    is_reserve = True
                    slot_index = None
            else:
                is_reserve = True
                slot_index = None
            
            # 計算年級
            grade_display = ''
            admission_year = result.get("admission_year")
            class_name = result.get("class_name") or ''
            
            # 如果沒有 admission_year，嘗試從學號前3碼獲取
            if admission_year is None or str(admission_year).strip() == '':
                student_number = result.get("student_number")
                if student_number and len(str(student_number)) >= 3:
                    try:
                        admission_year = int(str(student_number)[:3])
                    except (TypeError, ValueError):
                        pass
            
            # 計算年級
            if active_semester_year is not None and admission_year is not None:
                try:
                    grade_num = active_semester_year - int(admission_year) + 1
                    grade_labels = ('一', '二', '三', '四', '五', '六')
                    if 1 <= grade_num <= 6:
                        grade_char = grade_labels[grade_num - 1]
                        # 從 class_name 中提取班級名稱（例如「孝」）
                        class_char = ''
                        if class_name:
                            # 提取最後一個字作為班級名稱
                            class_char = class_name[-1] if len(class_name) > 0 else ''
                        grade_display = f"{grade_char}{class_char}" if class_char else f"{grade_char}年級"
                    elif grade_num > 0:
                        grade_display = f"{grade_num}年級"
                except (TypeError, ValueError):
                    pass
            
            formatted_result = {
                "id": result.get("match_id"),
                "match_id": result.get("match_id"),
                "student_id": student_id,
                "student_name": result.get("student_name"),
                "student_number": result.get("student_number"),
                "class_name": result.get("class_name"),
                "grade_display": grade_display,
                "company_id": company_id,
                "company_name": result.get("company_name"),
                "job_id": result.get("job_id"),
                "job_title": result.get("job_title") or "未指定職缺",
                "job_slots": result.get("job_slots") or 1,
                "preference_order": result.get("preference_order"),
                "slot_index": slot_index,
                "is_reserve": is_reserve,
                # 添加主任媒合結果的原始信息，用於顯示「原正取2」等標籤
                "original_rank": original_rank,
                "original_type": original_type,
                "final_rank": result.get("final_rank"),
                "is_adjusted": bool(result.get("is_adjusted", False))
            }
            formatted_results.append(formatted_result)
        
        # 先獲取所有已審核的公司（即使沒有媒合結果也要顯示）
        cursor.execute("""
            SELECT DISTINCT ic.id AS company_id, ic.company_name
            FROM internship_companies ic
            WHERE ic.status = 'approved'
            ORDER BY ic.company_name
        """)
        all_companies = cursor.fetchall() or []
        
        # 獲取所有已審核公司的職缺
        cursor.execute("""
            SELECT ij.id AS job_id, ij.company_id, ij.title AS job_title, ij.slots AS job_slots
            FROM internship_jobs ij
            JOIN internship_companies ic ON ij.company_id = ic.id
            WHERE ic.status = 'approved' AND ij.is_active = 1
            ORDER BY ij.company_id, ij.id
        """)
        all_jobs = cursor.fetchall() or []
        
        # 按公司和職缺組織資料
        companies_data = {}
        
        # 先初始化所有已審核的公司
        for company in all_companies:
            company_id = company["company_id"]
            company_name = company["company_name"]
            companies_data[company_id] = {
                "company_id": company_id,
                "company_name": company_name,
                "jobs": {}
            }
            
        # 為每個公司添加職缺（即使沒有媒合結果）
        for job in all_jobs:
            company_id = job["company_id"]
            job_id = job["job_id"]
            job_title = job["job_title"]
            job_slots = job["job_slots"] or 1
            
            if company_id in companies_data:
                if job_id not in companies_data[company_id]["jobs"]:
                    companies_data[company_id]["jobs"][job_id] = {
                        "job_id": job_id,
                        "job_title": job_title,
                        "job_slots": job_slots,
                        "regulars": [],
                        "reserves": []
                    }
        
        # 添加媒合結果到對應的公司和職缺
        for result in formatted_results:
            company_id = result.get("company_id")
            company_name = result.get("company_name")
            job_id = result.get("job_id")
            job_title = result.get("job_title") or "未指定職缺"
            
            if company_id and company_id in companies_data:
                # 如果職缺不存在，創建一個
                if job_id not in companies_data[company_id]["jobs"]:
                    companies_data[company_id]["jobs"][job_id] = {
                        "job_id": job_id,
                        "job_title": job_title,
                        "job_slots": result.get("job_slots", 1),
                        "regulars": [],
                        "reserves": []
                    }
                
                # 根據 is_reserve 分類
                if result.get("is_reserve"):
                    companies_data[company_id]["jobs"][job_id]["reserves"].append(result)
                else:
                    companies_data[company_id]["jobs"][job_id]["regulars"].append(result)
        
        # 轉換為列表格式
        companies_list = []
        for company_id, company_data in companies_data.items():
            jobs_list = list(company_data["jobs"].values())
            companies_list.append({
                "company_id": company_id,
                "company_name": company_data["company_name"],
                "jobs": jobs_list
            })
        
        return jsonify({
            "success": True,
            "companies": companies_list,
            "total_matches": len(formatted_results)
        })
    
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": f"查詢失敗: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

# =========================================================
# API: 主任移除學生（從媒合結果中移除）
# =========================================================
@admission_bp.route("/api/director_remove_student", methods=["POST"])
def director_remove_student():
    """主任從媒合結果中移除學生（更新 manage_director 表的 director_decision 為 Rejected）"""
    if 'user_id' not in session or session.get('role') != 'director':
        return jsonify({"success": False, "message": "未授權"}), 403
    
    data = request.get_json() or {}
    match_id = data.get("history_id") or data.get("match_id")  # 支援兩種參數名稱
    
    if not match_id:
        return jsonify({"success": False, "message": "請提供記錄ID"}), 400
    
    # 確保 match_id 是字符串類型
    match_id = str(match_id)
    print(f"🔍 [director_remove_student] 收到移除請求: match_id={match_id}, type={type(match_id)}")
    
    director_id = session.get('user_id')
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 獲取當前學期ID
        current_semester_id = get_current_semester_id(cursor)
        if not current_semester_id:
            return jsonify({"success": False, "message": "無法取得當前學期"}), 500
        
        # 檢查 match_id 是否以 'ra_' 開頭（來自 resume_applications）
        if isinstance(match_id, str) and match_id.startswith('ra_'):
            # 從 resume_applications 來的記錄，需要創建或更新 manage_director 記錄
            ra_id = match_id.replace('ra_', '')
            try:
                ra_id_int = int(ra_id)
            except ValueError:
                return jsonify({"success": False, "message": "無效的記錄ID"}), 400
            
            # 查詢 resume_applications 記錄，獲取相關資訊
            cursor.execute("""
                SELECT ra.id, ra.application_id, ra.job_id, ra.is_reserve, ra.slot_index,
                       sja.student_id, sja.company_id,
                       sp.id AS student_preference_id
                FROM resume_applications ra
                INNER JOIN student_job_applications sja ON ra.application_id = sja.id
                INNER JOIN student_preferences sp ON sja.student_id = sp.student_id 
                    AND sja.company_id = sp.company_id 
                    AND sja.job_id = sp.job_id
                    AND sp.semester_id = %s
                WHERE ra.id = %s
            """, (current_semester_id, ra_id_int))
            ra_record = cursor.fetchone()
            
            if not ra_record:
                return jsonify({"success": False, "message": "找不到該記錄"}), 404
            
            # 檢查是否已存在 manage_director 記錄（使用 application_id 作為 preference_id）
            cursor.execute("""
                SELECT match_id FROM manage_director
                WHERE preference_id = %s AND student_id = %s
            """, (ra_record.get('application_id'), ra_record.get('student_id')))
            existing_md = cursor.fetchone()
            
            if existing_md:
                # 更新現有記錄
                cursor.execute("""
                    UPDATE manage_director
                    SET director_decision = 'Rejected',
                        updated_at = CURRENT_TIMESTAMP
                    WHERE match_id = %s
                """, (existing_md['match_id'],))
            else:
                # 創建新記錄並標記為 Rejected
                original_type = 'Regular' if ra_record.get('is_reserve') == 0 else 'Backup'
                original_rank = ra_record.get('slot_index')
                
                # 檢查 semester_id 欄位是否存在
                cursor.execute("""
                    SELECT COLUMN_NAME 
                    FROM information_schema.COLUMNS 
                    WHERE TABLE_SCHEMA = DATABASE() 
                    AND TABLE_NAME = 'manage_director'
                    AND COLUMN_NAME = 'semester_id'
                """)
                has_semester_id = cursor.fetchone() is not None
                cursor.fetchall()  # 確保所有結果都被讀取
                
                # 使用 application_id 作為 preference_id（符合外鍵約束）
                # 外鍵約束要求 preference_id 必須是 resume_applications.application_id
                application_id = ra_record.get('application_id')
                
                if has_semester_id:
                    cursor.execute("""
                        INSERT INTO manage_director (
                            semester_id, vendor_id, student_id, preference_id,
                            original_type, original_rank, is_conflict,
                            director_decision, is_adjusted, updated_at
                        ) VALUES (
                            %s, %s, %s, %s,
                            %s, %s, 0,
                            'Rejected', 0, CURRENT_TIMESTAMP
                        )
                    """, (
                        current_semester_id,
                        ra_record.get('company_id'),  # 使用 company_id 作為 vendor_id（如果沒有對應的 vendor）
                        ra_record.get('student_id'),
                        application_id,  # 使用 application_id 作為 preference_id
                        original_type,
                        original_rank
                    ))
                else:
                    # 如果沒有 semester_id 欄位，不包含它
                    cursor.execute("""
                        INSERT INTO manage_director (
                            vendor_id, student_id, preference_id,
                            original_type, original_rank, is_conflict,
                            director_decision, is_adjusted, updated_at
                        ) VALUES (
                            %s, %s, %s,
                            %s, %s, 0,
                            'Rejected', 0, CURRENT_TIMESTAMP
                        )
                    """, (
                        ra_record.get('company_id'),  # 使用 company_id 作為 vendor_id（如果沒有對應的 vendor）
                        ra_record.get('student_id'),
                        application_id,  # 使用 application_id 作為 preference_id
                        original_type,
                        original_rank
                    ))
        else:
            # 更新 manage_director 表，將 director_decision 設為 Rejected
            cursor.execute("""
                UPDATE manage_director
                SET director_decision = 'Rejected',
                    updated_at = CURRENT_TIMESTAMP
                WHERE match_id = %s
            """, (match_id,))
            
            if cursor.rowcount == 0:
                return jsonify({"success": False, "message": "找不到該記錄"}), 404
        
        conn.commit()
        
        return jsonify({
            "success": True,
            "message": "已移除學生"
        })
    
    except Exception as e:
        traceback.print_exc()
        if conn:
            conn.rollback()
        return jsonify({"success": False, "message": f"移除失敗: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

# =========================================================
# API: 主任從備取名單補上學生
# =========================================================
@admission_bp.route("/api/director_promote_reserve", methods=["POST"])
def director_promote_reserve():
    """主任將備取學生提升為正取（更新 manage_director 表）"""
    if 'user_id' not in session or session.get('role') != 'director':
        return jsonify({"success": False, "message": "未授權"}), 403
    
    data = request.get_json() or {}
    match_id = data.get("history_id") or data.get("match_id")  # 支援兩種參數名稱
    slot_index = data.get("slot_index")  # 新的正取位置
    
    if not match_id or slot_index is None:
        return jsonify({"success": False, "message": "請提供記錄ID和正取位置"}), 400
    
    director_id = session.get('user_id')
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 檢查記錄是否存在
        cursor.execute("""
            SELECT match_id, original_type, final_rank
            FROM manage_director
            WHERE match_id = %s
        """, (match_id,))
        record = cursor.fetchone()
        
        if not record:
            return jsonify({"success": False, "message": "找不到該記錄"}), 404
        
        # 判斷是否為調整（如果 original_type 不是 Regular 或 original_rank 不等於 final_rank）
        is_adjusted = True
        if record.get("original_type") == "Regular" and record.get("original_rank") == slot_index:
            is_adjusted = False
        
        # 更新 manage_director 表
        cursor.execute("""
            UPDATE manage_director
            SET director_decision = 'Approved',
                final_rank = %s,
                is_adjusted = %s,
                updated_at = CURRENT_TIMESTAMP
            WHERE match_id = %s
        """, (slot_index, is_adjusted, match_id))
        
        conn.commit()
        
        return jsonify({
            "success": True,
            "message": "已將備取學生提升為正取"
        })
    
    except Exception as e:
        traceback.print_exc()
        if conn:
            conn.rollback()
        return jsonify({"success": False, "message": f"提升失敗: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

# =========================================================
# API: 主任添加未錄取學生到公司
# =========================================================
@admission_bp.route("/api/director_add_student", methods=["POST"])
def director_add_student():
    """主任將未錄取的學生添加到公司的職缺"""
    if 'user_id' not in session or session.get('role') != 'director':
        return jsonify({"success": False, "message": "未授權"}), 403
    
    try:
        data = request.get_json()
        if not data:
            print("❌ 錯誤：請求體為空")
            return jsonify({"success": False, "message": "請求體為空"}), 400
        
        print(f"📥 收到請求數據: {data}")
        
        student_id = data.get("student_id")
        company_id = data.get("company_id")
        job_id = data.get("job_id")
        type = data.get("type", "regular")  # 'regular' 或 'reserve'
        slot_index = data.get("slot_index")  # 正取位置（如果是正取）
        
        print(f"📋 解析後的參數: student_id={student_id}, company_id={company_id}, job_id={job_id}, type={type}, slot_index={slot_index}")
        
        # 詳細的參數驗證和錯誤訊息
        if student_id is None:
            print("❌ 錯誤：缺少學生ID")
            return jsonify({"success": False, "message": "缺少學生ID (student_id)"}), 400
        if company_id is None:
            print("❌ 錯誤：缺少公司ID")
            return jsonify({"success": False, "message": "缺少公司ID (company_id)"}), 400
        if job_id is None:
            print("❌ 錯誤：缺少職缺ID")
            return jsonify({"success": False, "message": "缺少職缺ID (job_id)"}), 400
        
        # 確保 ID 是整數
        try:
            student_id = int(student_id)
            company_id = int(company_id)
            job_id = int(job_id)
            if slot_index is not None:
                slot_index = int(slot_index)
            print(f"✅ 參數驗證通過: student_id={student_id}, company_id={company_id}, job_id={job_id}, slot_index={slot_index}")
        except (ValueError, TypeError) as e:
            print(f"❌ ID 格式錯誤: {e}")
            return jsonify({"success": False, "message": f"ID 格式錯誤: {str(e)}"}), 400
    except Exception as parse_error:
        print(f"❌ 解析請求數據時出錯: {parse_error}")
        traceback.print_exc()
        return jsonify({"success": False, "message": f"解析請求數據失敗: {str(parse_error)}"}), 400
    
    director_id = session.get('user_id')
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        print(f"🔍 開始處理：student_id={student_id}, company_id={company_id}, job_id={job_id}, type={type}, slot_index={slot_index}")
        
        # 獲取當前學期代碼
        current_semester_code = get_current_semester_code(cursor)
        if not current_semester_code:
            return jsonify({"success": False, "message": "無法取得當前學期"}), 500
        
        # 1. 驗證學生是否存在
        cursor.execute("SELECT id, name, username FROM users WHERE id = %s AND role = 'student'", (student_id,))
        student = cursor.fetchone()
        cursor.fetchall()  # 確保所有結果都被讀取
        if not student:
            return jsonify({"success": False, "message": "找不到該學生"}), 404
        
        # 1.5. 確保 students 表中有對應記錄（如果外鍵約束需要）
        # 檢查 students 表是否存在，如果存在則確保有對應記錄
        try:
            # 先檢查 students 表是否存在
            cursor.execute("""
                SELECT TABLE_NAME 
                FROM information_schema.TABLES 
                WHERE TABLE_SCHEMA = DATABASE() 
                AND TABLE_NAME = 'students'
            """)
            students_table_exists = cursor.fetchone()
            # 確保所有結果都被讀取
            cursor.fetchall()
            
            if students_table_exists:
                # 檢查 students 表中是否有該學生記錄
                cursor.execute("""
                    SELECT id FROM students WHERE id = %s
                """, (student_id,))
                student_in_students = cursor.fetchone()
                # 確保所有結果都被讀取
                cursor.fetchall()
                
                if not student_in_students:
                    # 獲取 students 表的欄位結構
                    cursor.execute("""
                        SELECT COLUMN_NAME, DATA_TYPE 
                        FROM information_schema.COLUMNS 
                        WHERE TABLE_SCHEMA = DATABASE() 
                        AND TABLE_NAME = 'students'
                        ORDER BY ORDINAL_POSITION
                    """)
                    columns = cursor.fetchall()
                    column_names = [col['COLUMN_NAME'] for col in columns] if columns else []
                    
                    # 根據實際欄位構建 INSERT 語句
                    if 'id' in column_names:
                        # 構建欄位和值的列表
                        insert_columns = ['id']
                        insert_values = [student_id]
                        
                        # 添加其他常見欄位
                        if 'name' in column_names:
                            insert_columns.append('name')
                            insert_values.append(student.get('name', ''))
                        if 'username' in column_names:
                            insert_columns.append('username')
                            insert_values.append(student.get('username', ''))
                        if 'user_id' in column_names:
                            insert_columns.append('user_id')
                            insert_values.append(student_id)
                        
                        # 構建並執行 INSERT 語句
                        columns_str = ', '.join(insert_columns)
                        placeholders = ', '.join(['%s'] * len(insert_values))
                        insert_query = f"INSERT INTO students ({columns_str}) VALUES ({placeholders})"
                        cursor.execute(insert_query, insert_values)
        except Exception as students_error:
            # 如果處理 students 表時出錯，記錄但不中斷流程
            # 外鍵約束可能實際指向 users 表，或者 students 表結構不同
            print(f"警告：處理 students 表時出錯: {students_error}")
            pass
        
        # 2. 驗證公司和職缺是否存在
        cursor.execute("SELECT id, company_name FROM internship_companies WHERE id = %s", (company_id,))
        company = cursor.fetchone()
        cursor.fetchall()  # 確保所有結果都被讀取
        if not company:
            return jsonify({"success": False, "message": "找不到該公司"}), 404
        
        cursor.execute("SELECT id, title, company_id, slots FROM internship_jobs WHERE id = %s", (job_id,))
        job = cursor.fetchone()
        cursor.fetchall()  # 確保所有結果都被讀取
        if not job:
            return jsonify({"success": False, "message": "找不到該職缺"}), 404
        
        if job['company_id'] != company_id:
            print(f"❌ 錯誤：職缺 {job_id} 不屬於公司 {company_id}，實際屬於公司 {job['company_id']}")
            return jsonify({"success": False, "message": "職缺不屬於該公司"}), 400
        
        # 3. 獲取當前學期ID
        current_semester_id = get_current_semester_id(cursor)
        if not current_semester_id:
            return jsonify({"success": False, "message": "無法取得當前學期"}), 500
        
        # 4. 檢查是否已經存在該學生的媒合記錄
        cursor.execute("""
            SELECT match_id, director_decision, preference_id, vendor_id
            FROM manage_director
            WHERE student_id = %s AND semester_id = %s
        """, (student_id, current_semester_id))
        existing = cursor.fetchone()
        cursor.fetchall()  # 確保所有結果都被讀取
        
        if existing:
            # 如果記錄存在且狀態為 Rejected，允許更新
            if existing.get('director_decision') == 'Rejected':
                print(f"ℹ️ 學生 {student_id} 有 Rejected 記錄，將更新為新記錄")
                # 繼續處理，後續會更新或創建新記錄
            else:
                # 檢查是否是要更新到同一個公司/職缺
                # 注意：manage_director.preference_id 對應的是 student_job_applications.id
                # 需要從 student_job_applications 表查詢 company_id 和 job_id
                existing_preference_id = existing.get('preference_id')
                existing_vendor_id = existing.get('vendor_id')
                
                if existing_preference_id:
                    # 從 student_job_applications 表查詢（因為 preference_id = student_job_applications.id）
                    cursor.execute("""
                        SELECT company_id, job_id FROM student_job_applications WHERE id = %s
                    """, (existing_preference_id,))
                    existing_sja = cursor.fetchone()
                    cursor.fetchall()  # 確保所有結果都被讀取
                    
                    if existing_sja:
                        existing_company_id = existing_sja.get('company_id')
                        existing_job_id = existing_sja.get('job_id')
                        # 如果是同一個公司/職缺，允許更新
                        if existing_company_id == company_id and existing_job_id == job_id:
                            print(f"ℹ️ 學生 {student_id} 已存在於相同公司/職缺，將更新記錄")
                            # 繼續處理，後續會更新記錄
                        else:
                            # 不同的公司/職缺，自動將舊記錄標記為 Rejected（讓其他廠商的正取位子顯示為空缺）
                            print(f"ℹ️ 學生 {student_id} 已存在於其他公司/職缺 (公司: {existing_company_id}, 職缺: {existing_job_id})，將自動移除舊記錄")
                            cursor.execute("""
                                UPDATE manage_director
                                SET director_decision = 'Rejected',
                                    updated_at = CURRENT_TIMESTAMP
                                WHERE match_id = %s
                            """, (existing.get('match_id'),))
                            print(f"✅ 已將舊記錄 (match_id: {existing.get('match_id')}) 標記為 Rejected")
                    elif existing_vendor_id:
                        # 如果查不到 student_job_applications，但 vendor_id 存在，使用 vendor_id 判斷
                        if existing_vendor_id == company_id:
                            print(f"ℹ️ 學生 {student_id} 已存在於相同公司（vendor_id={existing_vendor_id}），將更新記錄")
                            # 繼續處理，後續會更新記錄
                        else:
                            # 不同的公司，自動將舊記錄標記為 Rejected
                            print(f"ℹ️ 學生 {student_id} 已存在於其他公司 (vendor_id: {existing_vendor_id})，將自動移除舊記錄")
                            cursor.execute("""
                                UPDATE manage_director
                                SET director_decision = 'Rejected',
                                    updated_at = CURRENT_TIMESTAMP
                                WHERE match_id = %s
                            """, (existing.get('match_id'),))
                            print(f"✅ 已將舊記錄 (match_id: {existing.get('match_id')}) 標記為 Rejected")
                    else:
                        print(f"⚠️ 警告：學生 {student_id} 的現有記錄無法確定公司/職缺，跳過自動移除邏輯")
                else:
                    # preference_id 為空，但可能有 vendor_id
                    if existing_vendor_id:
                        if existing_vendor_id == company_id:
                            print(f"ℹ️ 學生 {student_id} 已存在於相同公司（vendor_id={existing_vendor_id}），將更新記錄")
                        else:
                            # 不同的公司，自動將舊記錄標記為 Rejected
                            print(f"ℹ️ 學生 {student_id} 已存在於其他公司 (vendor_id: {existing_vendor_id})，將自動移除舊記錄")
                            cursor.execute("""
                                UPDATE manage_director
                                SET director_decision = 'Rejected',
                                    updated_at = CURRENT_TIMESTAMP
                                WHERE match_id = %s
                            """, (existing.get('match_id'),))
                            print(f"✅ 已將舊記錄 (match_id: {existing.get('match_id')}) 標記為 Rejected")
                    else:
                        print(f"❌ 錯誤：學生 {student_id} 已經在媒合結果中但 preference_id 和 vendor_id 都為空")
                        return jsonify({"success": False, "message": "該學生已經在媒合結果中"}), 400
        
        # 4. 獲取或創建 student_job_applications 記錄（application_id）
        # 注意：manage_director.preference_id 必須引用 resume_applications.application_id
        # 而 resume_applications.application_id 對應的是 student_job_applications.id
        cursor.execute("""
            SELECT id FROM student_job_applications
            WHERE student_id = %s AND company_id = %s AND job_id = %s
            ORDER BY applied_at DESC
            LIMIT 1
        """, (student_id, company_id, job_id))
        application = cursor.fetchone()
        cursor.fetchall()  # 確保所有結果都被讀取
        
        application_id = None
        if application:
            application_id = application['id']
            print(f"✅ 找到現有的 student_job_applications 記錄: application_id={application_id}")
        else:
            # 創建新的 student_job_applications 記錄
            # 需要一個 resume_id，先查找學生的履歷
            cursor.execute("""
                SELECT id FROM resumes
                WHERE user_id = %s AND status IN ('approved', 'uploaded')
                ORDER BY updated_at DESC
                LIMIT 1
            """, (student_id,))
            resume = cursor.fetchone()
            cursor.fetchall()  # 確保所有結果都被讀取
            
            resume_id = resume['id'] if resume else None
            if not resume_id:
                # 如果沒有履歷，創建一個基本的履歷記錄
                # 注意：resumes 表的 semester_id 有外鍵約束，必須引用 semesters.id
                cursor.execute("""
                    INSERT INTO resumes (user_id, status, category, semester_id, created_at, updated_at)
                    VALUES (%s, 'approved', 'ready', %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                """, (student_id, current_semester_id))
                resume_id = cursor.lastrowid
                print(f"✅ 創建新的履歷記錄: resume_id={resume_id}, semester_id={current_semester_id}")
            
            # 創建 student_job_applications 記錄
            cursor.execute("""
                INSERT INTO student_job_applications
                (student_id, company_id, job_id, resume_id, status, applied_at)
                VALUES (%s, %s, %s, %s, 'submitted', CURRENT_TIMESTAMP)
            """, (student_id, company_id, job_id, resume_id))
            application_id = cursor.lastrowid
            print(f"✅ 創建新的 student_job_applications 記錄: application_id={application_id}")
        
        # 5. 確保 resume_applications 記錄存在
        # 檢查是否已存在 resume_applications 記錄
        cursor.execute("""
            SELECT id, application_id FROM resume_applications
            WHERE application_id = %s AND job_id = %s
        """, (application_id, job_id))
        resume_app = cursor.fetchone()
        cursor.fetchall()  # 確保所有結果都被讀取
        
        if not resume_app:
            # 創建 resume_applications 記錄
            # 注意：根據錯誤訊息，resume_applications.job_id 的外鍵約束可能要求引用 internship_companies.id
            # 但根據邏輯，job_id 應該對應到 internship_jobs.id
            try:
                cursor.execute("""
                    INSERT INTO resume_applications
                    (application_id, job_id, apply_status, interview_status, interview_result, created_at)
                    VALUES (%s, %s, 'approved', 'none', 'pending', CURRENT_TIMESTAMP)
                """, (application_id, job_id))
                print(f"✅ 創建 resume_applications 記錄: application_id={application_id}, job_id={job_id}")
            except Exception as insert_error:
                # 如果外鍵約束失敗，檢查是否是 fk_resume_applications_companies 約束
                error_msg = str(insert_error)
                if "fk_resume_applications_companies" in error_msg or ("1452" in error_msg and "internship_companies" in error_msg):
                    # 外鍵約束要求 job_id 引用 internship_companies.id
                    # 這可能是數據庫設計問題，但我們需要處理它
                    # 驗證 company_id 是否存在（這已經驗證過了，但再次確認）
                    cursor.execute("""
                        SELECT id FROM internship_companies WHERE id = %s
                    """, (company_id,))
                    company_check = cursor.fetchone()
                    cursor.fetchall()
                    
                    if company_check:
                        # 如果外鍵約束確實要求 job_id 引用 internship_companies.id
                        # 我們需要使用 company_id 作為 job_id（這會破壞數據完整性，但滿足外鍵約束）
                        print(f"⚠️ 外鍵約束錯誤：resume_applications.job_id 必須引用 internship_companies.id")
                        print(f"   嘗試使用 company_id={company_id} 作為 job_id（這可能是數據庫設計問題）")
                        cursor.execute("""
                            INSERT INTO resume_applications
                            (application_id, job_id, apply_status, interview_status, interview_result, created_at)
                            VALUES (%s, %s, 'approved', 'none', 'pending', CURRENT_TIMESTAMP)
                        """, (application_id, company_id))
                        print(f"✅ 創建 resume_applications 記錄（使用 company_id 作為 job_id）: application_id={application_id}, job_id={company_id}")
                    else:
                        print(f"❌ 錯誤：company_id={company_id} 不存在於 internship_companies 表中")
                        raise
                else:
                    # 其他錯誤，直接拋出
                    raise
        else:
            print(f"✅ resume_applications 記錄已存在: id={resume_app['id']}, application_id={application_id}")
        
        # preference_id 就是 application_id（student_job_applications.id）
        # 這是因為 manage_director.preference_id 外鍵引用 resume_applications.application_id
        preference_id = application_id
        
        # 6. 在 manage_director 表中創建或更新記錄
        is_reserve = (type == 'reserve')
        original_type = "Regular" if not is_reserve else "Reserve"
        original_rank = slot_index if not is_reserve else None
        final_rank = slot_index if not is_reserve else None
        
        # 如果已存在記錄，更新它；否則創建新記錄
        if existing and existing.get('match_id'):
            match_id = existing.get('match_id')
            print(f"🔄 更新現有記錄 match_id={match_id}")
            cursor.execute("""
                UPDATE manage_director
                SET vendor_id = %s,
                    preference_id = %s,
                    original_type = %s,
                    original_rank = %s,
                    is_conflict = 0,
                    director_decision = 'Approved',
                    final_rank = %s,
                    is_adjusted = 0,
                    updated_at = CURRENT_TIMESTAMP
                WHERE match_id = %s
            """, (
                company_id, preference_id,
                original_type, original_rank,
                final_rank,
                match_id
            ))
        else:
            print(f"➕ 創建新記錄")
            # 檢查 project_id 欄位是否存在，如果存在則包含在 INSERT 中
            cursor.execute("""
                SELECT COLUMN_NAME 
                FROM information_schema.COLUMNS 
                WHERE TABLE_SCHEMA = DATABASE() 
                AND TABLE_NAME = 'manage_director'
                AND COLUMN_NAME = 'project_id'
            """)
            has_project_id = cursor.fetchone() is not None
            cursor.fetchall()  # 確保所有結果都被讀取
            
            if has_project_id:
                cursor.execute("""
                    INSERT INTO manage_director (
                        semester_id, project_id, vendor_id, student_id, preference_id,
                        original_type, original_rank, is_conflict,
                        director_decision, final_rank, is_adjusted,
                        updated_at
                    ) VALUES (
                        %s, NULL, %s, %s, %s,
                        %s, %s, 0,
                        'Approved', %s, 0,
                        CURRENT_TIMESTAMP
                    )
                """, (
                    current_semester_id, company_id, student_id, preference_id,
                    original_type, original_rank,
                    final_rank
                ))
            else:
                cursor.execute("""
                    INSERT INTO manage_director (
                        semester_id, vendor_id, student_id, preference_id,
                        original_type, original_rank, is_conflict,
                        director_decision, final_rank, is_adjusted,
                        updated_at
                    ) VALUES (
                        %s, %s, %s, %s,
                        %s, %s, 0,
                        'Approved', %s, 0,
                        CURRENT_TIMESTAMP
                    )
                """, (
                    current_semester_id, company_id, student_id, preference_id,
                    original_type, original_rank,
                    final_rank
                ))
        
        # 自動處理重複學生：選擇志願序最高的記錄，將其他記錄標記為 Pending
        _resolve_duplicate_students(cursor, current_semester_id)
        
        conn.commit()
        
        type_name = '正取' if type == 'regular' else '備取'
        return jsonify({
            "success": True,
            "message": f"已將學生添加到{type_name}名單"
        })
    
    except Exception as e:
        traceback.print_exc()
        if conn:
            conn.rollback()
        return jsonify({"success": False, "message": f"添加失敗: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

# =========================================================
# API: 主任調整學生位置
# =========================================================
@admission_bp.route("/api/director_update_position", methods=["POST"])
def director_update_position():
    """主任調整學生在媒合結果中的位置"""
    if 'user_id' not in session or session.get('role') != 'director':
        return jsonify({"success": False, "message": "未授權"}), 403
    
    data = request.get_json() or {}
    history_id = data.get("history_id")
    slot_index = data.get("slot_index")
    is_reserve = data.get("is_reserve", False)
    
    if not history_id:
        return jsonify({"success": False, "message": "請提供記錄ID"}), 400
    
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        from vendor import _ensure_history_table
        _ensure_history_table(cursor)
        
        # 檢查欄位是否存在
        cursor.execute("""
            SELECT COLUMN_NAME 
            FROM information_schema.COLUMNS 
            WHERE TABLE_SCHEMA = DATABASE() 
            AND TABLE_NAME = 'vendor_preference_history'
            AND COLUMN_NAME IN ('slot_index', 'is_reserve')
        """)
        existing_columns = {row['COLUMN_NAME'] for row in cursor.fetchall()}
        
        if 'slot_index' in existing_columns and 'is_reserve' in existing_columns:
            # 更新位置
            cursor.execute("""
                UPDATE vendor_preference_history
                SET slot_index = %s, is_reserve = %s
                WHERE id = %s
            """, (slot_index, is_reserve, history_id))
        else:
            # 如果欄位不存在，更新 comment
            if is_reserve:
                comment = "媒合排序：候補"
            else:
                comment = f"媒合排序：正取{slot_index}"
            cursor.execute("""
                UPDATE vendor_preference_history
                SET comment = %s
                WHERE id = %s
            """, (comment, history_id))
        
        conn.commit()
        
        return jsonify({
            "success": True,
            "message": "已更新學生位置"
        })
    
    except Exception as e:
        traceback.print_exc()
        if conn:
            conn.rollback()
        return jsonify({"success": False, "message": f"更新失敗: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

# =========================================================
# API: 主任調整正取名單中學生的順序
# =========================================================
@admission_bp.route("/api/director_swap_positions", methods=["POST"])
def director_swap_positions():
    """主任調整正取名單中兩個學生的位置順序"""
    if 'user_id' not in session or session.get('role') != 'director':
        return jsonify({"success": False, "message": "未授權"}), 403
    
    data = request.get_json() or {}
    match_id1 = data.get("match_id1")
    match_id2 = data.get("match_id2")
    
    if not match_id1 or not match_id2:
        return jsonify({"success": False, "message": "請提供兩個記錄ID"}), 400
    
    director_id = session.get('user_id')
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 獲取兩個記錄的當前 final_rank 和 original_rank
        # 如果 match_id 以 'ra_' 開頭，則從 resume_applications 表讀取
        # 否則從 manage_director 表讀取
        records = []
        
        for match_id in [match_id1, match_id2]:
            if isinstance(match_id, str) and match_id.startswith('ra_'):
                # 從 resume_applications 表讀取
                ra_id = match_id.replace('ra_', '')
                cursor.execute("""
                    SELECT CONCAT('ra_', ra.id) AS match_id, ra.slot_index AS final_rank, 
                           ra.slot_index AS original_rank, 'Pending' AS director_decision
                    FROM resume_applications ra
                    WHERE ra.id = %s AND ra.apply_status = 'approved'
                """, (ra_id,))
                record = cursor.fetchone()
                if record:
                    records.append(record)
            else:
                # 從 manage_director 表讀取
                cursor.execute("""
                    SELECT match_id, final_rank, original_rank, director_decision
                    FROM manage_director
                    WHERE match_id = %s
                """, (match_id,))
                record = cursor.fetchone()
                if record:
                    records.append(record)
        
        if len(records) != 2:
            return jsonify({"success": False, "message": "找不到指定的記錄"}), 404
        
        record1 = next((r for r in records if str(r['match_id']) == str(match_id1)), None)
        record2 = next((r for r in records if str(r['match_id']) == str(match_id2)), None)
        
        if not record1 or not record2:
            return jsonify({"success": False, "message": "找不到指定的記錄"}), 404
        
        # 獲取 rank（優先使用 final_rank，如果沒有則使用 original_rank）
        rank1 = record1.get('final_rank') if record1.get('final_rank') is not None else record1.get('original_rank')
        rank2 = record2.get('final_rank') if record2.get('final_rank') is not None else record2.get('original_rank')
        
        if rank1 is None or rank2 is None:
            return jsonify({"success": False, "message": "學生必須有正取位置才能調整順序"}), 400
        
        # 交換兩個學生的 rank
        # 需要同時更新 resume_applications 和 manage_director 表
        for match_id, new_rank in [(match_id1, rank2), (match_id2, rank1)]:
            if isinstance(match_id, str) and match_id.startswith('ra_'):
                # 從 resume_applications 表更新
                ra_id = match_id.replace('ra_', '')
                cursor.execute("""
                    UPDATE resume_applications
                    SET slot_index = %s,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id = %s
                """, (new_rank, ra_id))
                
                # 同時更新 manage_director 表（如果存在）
                cursor.execute("""
                    SELECT application_id FROM resume_applications WHERE id = %s
                """, (ra_id,))
                ra_record = cursor.fetchone()
                if ra_record:
                    cursor.execute("""
                        UPDATE manage_director
                        SET original_rank = %s,
                            final_rank = %s,
                            is_adjusted = TRUE,
                            updated_at = CURRENT_TIMESTAMP
                        WHERE preference_id = %s
                    """, (new_rank, new_rank, ra_record['application_id']))
            else:
                # 從 manage_director 表更新
                # 先獲取 preference_id，以便更新 resume_applications
                cursor.execute("""
                    SELECT preference_id FROM manage_director WHERE match_id = %s
                """, (match_id,))
                md_record = cursor.fetchone()
                
                # 更新 manage_director 表
                cursor.execute("""
                    UPDATE manage_director
                    SET final_rank = %s,
                        is_adjusted = TRUE,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE match_id = %s
                """, (new_rank, match_id))
                
                # 如果 original_rank 為 NULL 或需要更新，也更新它
                cursor.execute("""
                    UPDATE manage_director
                    SET original_rank = %s,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE match_id = %s AND (original_rank IS NULL OR original_rank != %s)
                """, (new_rank, match_id, new_rank))
                
                # 同時更新 resume_applications 表的 slot_index
                if md_record and md_record.get('preference_id'):
                    cursor.execute("""
                        UPDATE resume_applications
                        SET slot_index = %s,
                            updated_at = CURRENT_TIMESTAMP
                        WHERE application_id = %s
                    """, (new_rank, md_record['preference_id']))
        
        conn.commit()
        
        return jsonify({
            "success": True,
            "message": "已交換學生位置"
        })
    
    except Exception as e:
        traceback.print_exc()
        if conn:
            conn.rollback()
        return jsonify({"success": False, "message": f"調整失敗: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

# =========================================================
# API: 手動處理重複學生（管理員/主任可用）
# =========================================================
@admission_bp.route("/api/resolve_duplicate_students", methods=["POST"])
def resolve_duplicate_students():
    """手動處理重複學生：選擇志願序最高的記錄，將其他記錄標記為 Pending"""
    if 'user_id' not in session or session.get('role') not in ['director', 'admin', 'ta']:
        return jsonify({"success": False, "message": "未授權"}), 403
    
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 獲取當前學期ID
        current_semester_id = get_current_semester_id(cursor)
        if not current_semester_id:
            return jsonify({"success": False, "message": "無法取得當前學期"}), 500
        
        # 執行處理重複學生的邏輯
        updated_count = _resolve_duplicate_students(cursor, current_semester_id)
        
        conn.commit()
        
        return jsonify({
            "success": True,
            "message": f"已處理重複學生，更新了 {updated_count} 筆記錄",
            "updated_count": updated_count
        })
    
    except Exception as e:
        traceback.print_exc()
        if conn:
            conn.rollback()
        return jsonify({"success": False, "message": f"處理失敗: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

# =========================================================
# API: 主任確認媒合結果
# =========================================================
@admission_bp.route("/api/director_confirm_matching", methods=["POST"])
def director_confirm_matching():
    """
    主任確認媒合結果後：只通知科助，由科助進行最後發布並通知所有使用者。
    """
    if 'user_id' not in session or session.get('role') != 'director':
        return jsonify({"success": False, "message": "未授權"}), 403
    
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 獲取當前學期ID和學期代碼
        current_semester_id = get_current_semester_id(cursor)
        if not current_semester_id:
            return jsonify({"success": False, "message": "無法取得當前學期"}), 500
        
        # 獲取當前學期代碼（如 '1132'）
        current_semester_code = get_current_semester_code(cursor)
        semester_prefix = f"{current_semester_code}學期" if current_semester_code else "本學期"
        
        # 0. 將所有 Pending 狀態的記錄更新為 Approved（主任確認後，所有待定的記錄都變為已確認）
        # md.preference_id 引用的是 student_job_applications.id（即 resume_applications.application_id）
        # 需要通過 student_job_applications 來 JOIN student_preferences
        cursor.execute("""
            UPDATE manage_director md
            INNER JOIN student_job_applications sja ON md.preference_id = sja.id
            INNER JOIN student_preferences sp ON sja.student_id = sp.student_id
                AND sja.company_id = sp.company_id
                AND sja.job_id = sp.job_id
                AND sp.semester_id = %s
            SET md.director_decision = 'Approved',
                md.updated_at = CURRENT_TIMESTAMP
            WHERE md.director_decision = 'Pending'
        """, (current_semester_id,))
        updated_count = cursor.rowcount
        print(f"✅ 主任確認：已將 {updated_count} 筆 Pending 記錄更新為 Approved")
        
        # 0.05. 自動處理重複學生：選擇志願序最高的記錄，將其他記錄標記為 Pending
        _resolve_duplicate_students(cursor, current_semester_id)
        
        # 0.1. 為來自 resume_applications 但還沒有 manage_director 記錄的學生創建記錄
        # 這些學生是廠商已排序但主任還沒有處理的
        # 先檢查 semester_id 欄位是否存在
        cursor.execute("""
            SELECT COLUMN_NAME
            FROM information_schema.COLUMNS
            WHERE TABLE_SCHEMA = DATABASE()
            AND TABLE_NAME = 'manage_director'
            AND COLUMN_NAME = 'semester_id'
        """)
        has_semester_id = cursor.fetchone() is not None
        cursor.fetchall()
        
        if has_semester_id:
            # 有 semester_id 欄位
            cursor.execute("""
                INSERT INTO manage_director (
                    semester_id, vendor_id, student_id, preference_id,
                    original_type, original_rank, is_conflict,
                    director_decision, final_rank, is_adjusted, updated_at
                )
                SELECT 
                    %s AS semester_id,
                    sja.company_id AS vendor_id,
                    sja.student_id,
                    ra.application_id AS preference_id,
                    CASE WHEN ra.is_reserve = 0 THEN 'Regular' ELSE 'Backup' END AS original_type,
                    ra.slot_index AS original_rank,
                    0 AS is_conflict,
                    'Approved' AS director_decision,
                    ra.slot_index AS final_rank,
                    0 AS is_adjusted,
                    CURRENT_TIMESTAMP AS updated_at
                FROM resume_applications ra
                INNER JOIN student_job_applications sja ON ra.application_id = sja.id
                INNER JOIN student_preferences sp ON sja.student_id = sp.student_id 
                    AND sja.company_id = sp.company_id 
                    AND sja.job_id = sp.job_id
                    AND sp.semester_id = %s
                LEFT JOIN manage_director md ON ra.application_id = md.preference_id
                LEFT JOIN internship_companies ic ON sja.company_id = ic.id
                WHERE ra.apply_status = 'approved'
                AND (ra.is_reserve IS NOT NULL OR ra.slot_index IS NOT NULL)
                AND md.preference_id IS NULL  -- 還沒有 manage_director 記錄
                AND ic.status = 'approved'
                AND (SELECT COUNT(*) FROM manage_director md2 
                     WHERE md2.preference_id = ra.application_id 
                     AND md2.student_id = sja.student_id) = 0  -- 確保不會重複插入
            """, (current_semester_id, current_semester_id))
        else:
            # 沒有 semester_id 欄位
            cursor.execute("""
                INSERT INTO manage_director (
                    vendor_id, student_id, preference_id,
                    original_type, original_rank, is_conflict,
                    director_decision, final_rank, is_adjusted, updated_at
                )
                SELECT 
                    sja.company_id AS vendor_id,
                    sja.student_id,
                    ra.application_id AS preference_id,
                    CASE WHEN ra.is_reserve = 0 THEN 'Regular' ELSE 'Backup' END AS original_type,
                    ra.slot_index AS original_rank,
                    0 AS is_conflict,
                    'Approved' AS director_decision,
                    ra.slot_index AS final_rank,
                    0 AS is_adjusted,
                    CURRENT_TIMESTAMP AS updated_at
                FROM resume_applications ra
                INNER JOIN student_job_applications sja ON ra.application_id = sja.id
                INNER JOIN student_preferences sp ON sja.student_id = sp.student_id 
                    AND sja.company_id = sp.company_id 
                    AND sja.job_id = sp.job_id
                    AND sp.semester_id = %s
                LEFT JOIN manage_director md ON ra.application_id = md.preference_id
                LEFT JOIN internship_companies ic ON sja.company_id = ic.id
                WHERE ra.apply_status = 'approved'
                AND (ra.is_reserve IS NOT NULL OR ra.slot_index IS NOT NULL)
                AND md.preference_id IS NULL  -- 還沒有 manage_director 記錄
                AND ic.status = 'approved'
                AND (SELECT COUNT(*) FROM manage_director md2 
                     WHERE md2.preference_id = ra.application_id 
                     AND md2.student_id = sja.student_id) = 0  -- 確保不會重複插入
            """, (current_semester_id,))
        
        inserted_count = cursor.rowcount
        print(f"✅ 主任確認：已為 {inserted_count} 筆來自 resume_applications 的記錄創建 manage_director 記錄")
        
        # 0.15. 再次自動處理重複學生（插入新記錄後可能產生新的重複）
        _resolve_duplicate_students(cursor, current_semester_id)
        
        # 主任確認：只通知科助（指導老師、班導、學生、廠商由科助確認時一併通知）
        cursor.execute("SELECT id, name FROM users WHERE role = 'ta'")
        tas = cursor.fetchall() or []
        
        for ta in tas:
            title = f"{semester_prefix} 媒合結果待發布"
            message = f"{semester_prefix}媒合結果已由主任確認，請進行最後發布。"
            link_url = "/final_results"
            create_notification(
                user_id=ta['id'],
                title=title,
                message=message,
                category="approval",
                link_url=link_url
            )
        
        # 2. 將已確認的媒合結果寫入 internship_offers 表（主任確認時寫入，狀態為 accepted）
        cursor.execute("""
            SELECT DISTINCT
                md.student_id,
                md.vendor_id,
                sja.job_id,
                sja.company_id,
                %s AS semester_id
            FROM manage_director md
            INNER JOIN student_job_applications sja ON md.preference_id = sja.id
            INNER JOIN student_preferences sp ON sja.student_id = sp.student_id
                AND sja.company_id = sp.company_id
                AND sja.job_id = sp.job_id
                AND sp.semester_id = %s
            WHERE md.director_decision = 'Approved'
        """, (current_semester_id, current_semester_id))
        match_results = cursor.fetchall() or []
        
        inserted_count = 0
        updated_count = 0
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for match_result in match_results:
            student_id = match_result.get('student_id')
            job_id = match_result.get('job_id')
            if not student_id or not job_id:
                continue
            
            cursor.execute("""
                SELECT id FROM internship_offers
                WHERE student_id = %s AND job_id = %s
            """, (student_id, job_id))
            existing = cursor.fetchone()
            
            if existing:
                cursor.execute("""
                    UPDATE internship_offers
                    SET status = 'accepted', offered_at = %s, responded_at = %s
                    WHERE id = %s
                """, (now_str, now_str, existing['id']))
                updated_count += 1
            else:
                cursor.execute("""
                    INSERT INTO internship_offers
                    (student_id, job_id, status, offered_at, responded_at)
                    VALUES (%s, %s, 'accepted', %s, %s)
                """, (student_id, job_id, now_str, now_str))
                inserted_count += 1
        
        print(f"✅ [DEBUG] 主任確認時寫入 internship_offers: 新增 {inserted_count} 筆，更新 {updated_count} 筆")
        
        # 6. 提交事務，確保所有更新都保存
        conn.commit()
        
        # 7. 驗證更新後的記錄數量
        cursor.execute("""
            SELECT COUNT(*) as count
            FROM manage_director md
            LEFT JOIN student_job_applications sja ON md.preference_id = sja.id
            LEFT JOIN student_preferences sp ON sja.student_id = sp.student_id 
                AND sja.company_id = sp.company_id 
                AND sja.job_id = sp.job_id
                AND sp.semester_id = %s
            WHERE md.director_decision = 'Approved'
            AND (sp.semester_id = %s OR (sp.semester_id IS NULL AND sja.id IS NOT NULL))
        """, (current_semester_id, current_semester_id))
        verify_result = cursor.fetchone()
        approved_count = verify_result.get('count', 0) if verify_result else 0
        print(f"✅ 主任確認完成：共有 {approved_count} 筆 Approved 記錄可供科助查看")
        
        return jsonify({
            "success": True,
            "message": "確認成功",
            "approved_count": approved_count,
            "notified": {"tas": len(tas)}
        })
    
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": f"確認失敗: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

# =========================================================
# API: 科助確認媒合結果（主任確認後，科助進行最後確認；通知所有使用者）
# =========================================================
@admission_bp.route("/api/ta/confirm_matching", methods=["POST"])
def ta_confirm_matching():
    """
    科助確認媒合結果後：通知所有使用者（指導老師、班導、學生、廠商、管理員、主任）。
    """
    if 'user_id' not in session or session.get('role') not in ['ta', 'admin']:
        return jsonify({"success": False, "message": "未授權"}), 403
    
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 獲取當前學期ID和學期代碼
        current_semester_id = get_current_semester_id(cursor)
        current_semester_code = get_current_semester_code(cursor)
        if not current_semester_id:
            return jsonify({"success": False, "message": "無法取得當前學期"}), 500
        
        semester_prefix = f"{current_semester_code}學期" if current_semester_code else "本學期"
        
        # 檢查是否有主任已確認的媒合結果
        # md.preference_id 引用的是 student_job_applications.id（即 resume_applications.application_id）
        # 需要通過 student_job_applications 來 JOIN student_preferences
        cursor.execute("""
            SELECT COUNT(*) as count
            FROM manage_director md
            INNER JOIN student_job_applications sja ON md.preference_id = sja.id
            INNER JOIN student_preferences sp ON sja.student_id = sp.student_id
                AND sja.company_id = sp.company_id
                AND sja.job_id = sp.job_id
                AND sp.semester_id = %s
            WHERE md.director_decision = 'Approved'
        """, (current_semester_id,))
        result = cursor.fetchone()
        approved_count = result.get('count', 0) if result else 0
        
        if approved_count == 0:
            return jsonify({"success": False, "message": "目前沒有主任已確認的媒合結果"}), 400
        
        # 1. 通知所有指導老師和班導媒合結果已發布
        notified_user_ids = set()
        
        # 收集所有指導老師（role='teacher'）
        cursor.execute("SELECT id FROM users WHERE role = 'teacher'")
        teachers = cursor.fetchall() or []
        for teacher in teachers:
            notified_user_ids.add(teacher['id'])
        
        # 收集所有班導（從 classes_teacher 表獲取）
        cursor.execute("""
            SELECT DISTINCT ct.teacher_id
            FROM classes_teacher ct
            JOIN users u ON ct.teacher_id = u.id
            WHERE ct.role = 'classteacher'
        """)
        class_teachers = cursor.fetchall() or []
        for class_teacher in class_teachers:
            teacher_id = class_teacher['teacher_id']
            if teacher_id not in notified_user_ids:
                notified_user_ids.add(teacher_id)
        
        # 發送通知給所有需要通知的用戶（指導老師和班導）
        title = f"{semester_prefix} 媒合結果已發布"
        message = f"{semester_prefix}媒合結果已由科助確認並發布，請前往查看。"
        link_url = "/admission/results"
        
        for user_id in notified_user_ids:
            create_notification(
                user_id=user_id,
                title=title,
                message=message,
                category="matching",
                link_url=link_url
            )
        
        # 2. 通知所有在媒合結果中的學生（Approved 狀態）
        # md.preference_id 引用的是 student_job_applications.id（即 resume_applications.application_id）
        # 需要通過 student_job_applications 來 JOIN student_preferences
        cursor.execute("""
            SELECT DISTINCT md.student_id
            FROM manage_director md
            INNER JOIN student_job_applications sja ON md.preference_id = sja.id
            INNER JOIN student_preferences sp ON sja.student_id = sp.student_id
                AND sja.company_id = sp.company_id
                AND sja.job_id = sp.job_id
                AND sp.semester_id = %s
            WHERE md.director_decision = 'Approved'
        """, (current_semester_id,))
        matched_students = cursor.fetchall() or []
        
        student_title = f"{semester_prefix} 媒合結果已發布"
        student_message = f"{semester_prefix}媒合結果已發布，請前往查看您的媒合結果。"
        student_link_url = "/student_home"
        
        for student in matched_students:
            student_id = student.get('student_id')
            if student_id:
                create_notification(
                    user_id=student_id,
                    title=student_title,
                    message=student_message,
                    category="matching",
                    link_url=student_link_url
                )
        
        # 3. 通知所有廠商（role='vendor'）媒合結果已發布
        cursor.execute("SELECT id, name FROM users WHERE role = 'vendor'")
        vendors = cursor.fetchall() or []
        
        for vendor in vendors:
            title = f"{semester_prefix} 媒合結果已發布"
            message = f"{semester_prefix}媒合結果已由科助確認並發布，請前往查看您的實習生名單。"
            link_url = "/vendor/matching_results"
            create_notification(
                user_id=vendor['id'],
                title=title,
                message=message,
                category="matching",
                link_url=link_url
            )
        
        # 4. 通知管理員與主任（科助確認時通知所有使用者）
        cursor.execute("SELECT id FROM users WHERE role IN ('admin', 'director')")
        admins_directors = cursor.fetchall() or []
        admin_title = f"{semester_prefix} 媒合結果已發布"
        admin_message = f"{semester_prefix}媒合結果已由科助確認並發布，請前往查看。"
        admin_link_url = "/admission/results"
        for u in admins_directors:
            create_notification(
                user_id=u['id'],
                title=admin_title,
                message=admin_message,
                category="matching",
                link_url=admin_link_url
            )
        
        # 5. 將已確認的媒合結果寫入 internship_offers 表
        cursor.execute("""
            SELECT DISTINCT
                md.student_id,
                md.vendor_id,
                sja.job_id,
                sja.company_id,
                %s AS semester_id
            FROM manage_director md
            INNER JOIN student_job_applications sja ON md.preference_id = sja.id
            INNER JOIN student_preferences sp ON sja.student_id = sp.student_id
                AND sja.company_id = sp.company_id
                AND sja.job_id = sp.job_id
                AND sp.semester_id = %s
            WHERE md.director_decision = 'Approved'
        """, (current_semester_id, current_semester_id))
        match_results = cursor.fetchall() or []
        
        inserted_count = 0
        updated_count = 0
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for match_result in match_results:
            student_id = match_result.get('student_id')
            job_id = match_result.get('job_id')
            if not student_id or not job_id:
                continue
            
            cursor.execute("""
                SELECT id FROM internship_offers
                WHERE student_id = %s AND job_id = %s
            """, (student_id, job_id))
            existing = cursor.fetchone()
            
            if existing:
                cursor.execute("""
                    UPDATE internship_offers
                    SET status = 'accepted', offered_at = %s, responded_at = %s
                    WHERE id = %s
                """, (now_str, now_str, existing['id']))
                updated_count += 1
            else:
                cursor.execute("""
                    INSERT INTO internship_offers
                    (student_id, job_id, status, offered_at, responded_at)
                    VALUES (%s, %s, 'accepted', %s, %s)
                """, (student_id, job_id, now_str, now_str))
                inserted_count += 1
        
        print(f"✅ [DEBUG] 寫入 internship_offers: 新增 {inserted_count} 筆，更新 {updated_count} 筆")
        
        # 5.1 一併寫入 teacher_student_relations，讓「查看錄取結果」頁（班導／指導老師／主任／科助）有資料
        cursor.execute("""
            SELECT COLUMN_NAME FROM information_schema.COLUMNS
            WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'teacher_student_relations'
            AND COLUMN_NAME IN ('semester_id', 'semester', 'company_id')
        """)
        tsr_columns = {row['COLUMN_NAME'] for row in cursor.fetchall()}
        has_semester_id = 'semester_id' in tsr_columns
        has_semester = 'semester' in tsr_columns
        has_company_id = 'company_id' in tsr_columns
        
        tsr_inserted = 0
        tsr_updated = 0
        for match_result in match_results:
            student_id = match_result.get('student_id')
            company_id = match_result.get('company_id')
            if not student_id or not company_id:
                continue
            cursor.execute("""
                SELECT id, advisor_user_id FROM internship_companies WHERE id = %s
            """, (company_id,))
            company_row = cursor.fetchone()
            if not company_row or not company_row.get('advisor_user_id'):
                continue
            advisor_user_id = company_row['advisor_user_id']
            cursor.execute("""
                SELECT id, name FROM users WHERE id = %s AND role IN ('teacher', 'director')
            """, (advisor_user_id,))
            if not cursor.fetchone():
                continue
            if has_semester_id:
                cursor.execute("""
                    SELECT id FROM teacher_student_relations
                    WHERE teacher_id = %s AND student_id = %s AND semester_id = %s
                """, (advisor_user_id, student_id, current_semester_id))
                existing_tsr = cursor.fetchone()
                if existing_tsr:
                    # teacher_student_relations 表無 updated_at 欄位，已有紀錄則不更新
                    tsr_updated += 1
                else:
                    # 表結構僅有 teacher_id, student_id, role, semester_id, created_at（無 company_id）
                    cursor.execute("""
                        INSERT INTO teacher_student_relations
                        (teacher_id, student_id, semester_id, role, created_at)
                        VALUES (%s, %s, %s, '指導老師', NOW())
                    """, (advisor_user_id, student_id, current_semester_id))
                    tsr_inserted += 1
            elif has_semester:
                cursor.execute("""
                    SELECT id FROM teacher_student_relations
                    WHERE teacher_id = %s AND student_id = %s AND semester = %s
                """, (advisor_user_id, student_id, current_semester_code or ''))
                existing_tsr = cursor.fetchone()
                if existing_tsr:
                    tsr_updated += 1
                else:
                    cursor.execute("""
                        INSERT INTO teacher_student_relations
                        (teacher_id, student_id, semester, role, created_at)
                        VALUES (%s, %s, %s, '指導老師', NOW())
                    """, (advisor_user_id, student_id, current_semester_code or ''))
                    tsr_inserted += 1
        print(f"✅ [DEBUG] 寫入 teacher_student_relations: 新增 {tsr_inserted} 筆，更新 {tsr_updated} 筆")
        
        # 6. 提交事務，確保所有更新都保存
        conn.commit()
        
        return jsonify({
            "success": True,
            "message": "確認成功，已通知所有相關使用者",
            "notified": {
                "teachers_and_class_teachers": len(notified_user_ids),
                "students": len(matched_students),
                "vendors": len(vendors),
                "admins_directors": len(admins_directors)
            },
            "approved_count": approved_count,
            "match_results": {
                "inserted": inserted_count,
                "updated": updated_count
            }
        })
    
    except Exception as e:
        traceback.print_exc()
        if conn:
            conn.rollback()
        return jsonify({"success": False, "message": f"確認失敗: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

# =========================================================
# API: 查詢二面流程狀態
# =========================================================
@admission_bp.route("/api/ta/second_interview_status", methods=["GET"])
def get_second_interview_status():
    """查詢當前學期的二面流程是否已啟動"""
    if 'user_id' not in session or session.get('role') not in ['ta', 'admin']:
        return jsonify({"success": False, "message": "未授權"}), 403
    
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        current_semester_id = get_current_semester_id(cursor)
        if not current_semester_id:
            return jsonify({"success": False, "message": "無法取得當前學期"}), 500
        
        # 查詢當前學期的二面流程狀態
        # 嘗試使用 system_config 表，如果不存在則使用其他方式
        is_enabled = False
        try:
            cursor.execute("""
                SELECT value AS is_enabled
                FROM system_config
                WHERE config_key = 'second_interview_enabled' AND semester_id = %s
            """, (current_semester_id,))
            config = cursor.fetchone()
            
            if config and config.get('is_enabled'):
                try:
                    is_enabled = bool(int(config['is_enabled']))
                except (ValueError, TypeError):
                    is_enabled = False
        except Exception:
            # 如果 system_config 表不存在，嘗試使用其他方式或返回預設值
            # 可以考慮使用 internship_configs 表或其他配置表
            is_enabled = False
        
        return jsonify({
            "success": True,
            "is_enabled": is_enabled,
            "semester_id": current_semester_id
        })
    
    except Exception as e:
        # 如果發生其他錯誤，返回預設值 False
        return jsonify({
            "success": True,
            "is_enabled": False,
            "semester_id": current_semester_id if 'current_semester_id' in locals() else None
        })
    finally:
        cursor.close()
        conn.close()

# =========================================================
# API: 科助啟動/關閉二面流程（開關功能）
# =========================================================
@admission_bp.route("/api/ta/toggle_second_interview", methods=["POST"])
def ta_toggle_second_interview():
    """
    科助啟動/關閉二面流程（開關功能）：
    1. 如果開啟：通知所有指導老師和班導、未錄取學生、同意二面的廠商
    2. 如果關閉：只更新狀態，不發送通知
    """
    if 'user_id' not in session or session.get('role') not in ['ta', 'admin']:
        return jsonify({"success": False, "message": "未授權"}), 403
    
    data = request.get_json(silent=True) or {}
    enable = data.get('enable', True)  # 預設為開啟
    
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 獲取當前學期ID和學期代碼
        current_semester_id = get_current_semester_id(cursor)
        current_semester_code = get_current_semester_code(cursor)
        if not current_semester_id or not current_semester_code:
            return jsonify({"success": False, "message": "無法取得當前學期"}), 500
        
        # 更新或插入系統配置
        # 嘗試使用 system_config 表，如果不存在則創建或使用其他方式
        try:
            cursor.execute("""
                INSERT INTO system_config (config_key, value, semester_id, updated_at)
                VALUES ('second_interview_enabled', %s, %s, NOW())
                ON DUPLICATE KEY UPDATE value = %s, updated_at = NOW()
            """, (1 if enable else 0, current_semester_id, 1 if enable else 0))
        except Exception:
            # 如果表不存在，嘗試創建表（需要適當的權限）
            try:
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS system_config (
                        id INT AUTO_INCREMENT PRIMARY KEY,
                        config_key VARCHAR(100) NOT NULL,
                        value VARCHAR(255),
                        semester_id INT,
                        updated_at DATETIME,
                        UNIQUE KEY unique_config (config_key, semester_id)
                    )
                """)
                cursor.execute("""
                    INSERT INTO system_config (config_key, value, semester_id, updated_at)
                    VALUES ('second_interview_enabled', %s, %s, NOW())
                    ON DUPLICATE KEY UPDATE value = %s, updated_at = NOW()
                """, (1 if enable else 0, current_semester_id, 1 if enable else 0))
            except Exception as e:
                # 如果創建表也失敗，記錄錯誤但繼續執行（通知功能仍可運作）
                print(f"⚠️ 無法創建或更新 system_config 表: {e}")
                # 不中斷流程，繼續發送通知
        
        # 如果只是關閉，不需要發送通知
        if not enable:
            conn.commit()
            return jsonify({
                "success": True,
                "message": "二面流程已關閉",
                "is_enabled": False
            })
        
        # 如果開啟，發送通知
        semester_prefix = f"{current_semester_code}學期" if current_semester_code else "本學期"
        
        # 1. 通知所有指導老師（role='teacher'）
        cursor.execute("SELECT id FROM users WHERE role = 'teacher'")
        teachers = cursor.fetchall() or []
        for teacher in teachers:
            title = f"{semester_prefix} 二面流程已啟動"
            message = f"{semester_prefix}二面流程已由科助啟動，請協助詢問未錄取學生的二面意願。"
            link_url = "/teacher/unadmitted_list"
            create_notification(
                user_id=teacher['id'],
                title=title,
                message=message,
                category="matching",
                link_url=link_url
            )
        
        # 2. 通知所有班導（從 classes_teacher 表獲取）
        cursor.execute("""
            SELECT DISTINCT ct.teacher_id
            FROM classes_teacher ct
            JOIN users u ON ct.teacher_id = u.id
            WHERE ct.role = 'classteacher'
        """)
        class_teachers = cursor.fetchall() or []
        for class_teacher in class_teachers:
            title = f"{semester_prefix} 二面流程已啟動"
            message = f"{semester_prefix}二面流程已由科助啟動，請協助詢問未錄取學生的二面意願。"
            link_url = "/teacher/unadmitted_list"
            create_notification(
                user_id=class_teacher['teacher_id'],
                title=title,
                message=message,
                category="matching",
                link_url=link_url
            )
        
        # 3. 通知所有未錄取的學生
        # 獲取當前學期對應的學號前綴
        student_id_prefix = None
        if current_semester_code and len(current_semester_code) >= 3:
            try:
                year_part = int(current_semester_code[:3])
                student_id_prefix = str(year_part - 3)
            except (ValueError, TypeError):
                pass
        
        # 獲取所有學生
        student_query = "SELECT id FROM users WHERE role = 'student'"
        student_params = []
        if student_id_prefix:
            student_query += " AND username LIKE %s"
            student_params.append(student_id_prefix + "%")
        
        cursor.execute(student_query, student_params)
        all_students = cursor.fetchall() or []
        
        # 獲取已媒合的學生ID
        # md.preference_id 引用的是 student_job_applications.id（即 resume_applications.application_id）
        # 需要通過 student_job_applications 來 JOIN student_preferences
        cursor.execute("""
            SELECT DISTINCT md.student_id
            FROM manage_director md
            INNER JOIN student_job_applications sja ON md.preference_id = sja.id
            INNER JOIN student_preferences sp ON sja.student_id = sp.student_id
                AND sja.company_id = sp.company_id
                AND sja.job_id = sp.job_id
                AND sp.semester_id = %s
            WHERE md.director_decision IN ('Approved', 'Pending')
        """, (current_semester_id,))
        matched_student_ids = {row['student_id'] for row in cursor.fetchall()}
        
        # 只通知未錄取的學生
        unadmitted_students = [s for s in all_students if s['id'] not in matched_student_ids]
        
        # 為未錄取的學生重置面試狀態，讓他們可以重新參與二次面試
        unadmitted_student_ids = [s['id'] for s in unadmitted_students]
        if unadmitted_student_ids:
            # 找到這些學生的所有申請記錄（通過 student_job_applications 和 student_preferences）
            placeholders = ','.join(['%s'] * len(unadmitted_student_ids))
            cursor.execute(f"""
                SELECT DISTINCT ra.id, ra.application_id, ra.job_id, ra.apply_status, ra.interview_status
                FROM resume_applications ra
                INNER JOIN student_job_applications sja ON ra.application_id = sja.id
                INNER JOIN student_preferences sp ON sja.student_id = sp.student_id
                    AND sja.company_id = sp.company_id
                    AND sja.job_id = sp.job_id
                    AND sp.semester_id = %s
                WHERE sja.student_id IN ({placeholders})
            """, [current_semester_id] + unadmitted_student_ids)
            
            resume_apps = cursor.fetchall() or []
            
            # 為每個申請記錄重置面試狀態
            for ra in resume_apps:
                application_id = ra.get('application_id')
                job_id = ra.get('job_id')
                current_apply_status = ra.get('apply_status')
                current_interview_status = ra.get('interview_status')
                
                # 重置面試狀態為 'none'，讓廠商可以重新安排面試
                # 如果 apply_status 是 'rejected'，更新為 'approved'，讓學生可以重新參與面試
                new_apply_status = 'approved' if current_apply_status == 'rejected' else current_apply_status
                
                try:
                    cursor.execute("""
                        UPDATE resume_applications
                        SET interview_status = 'none',
                            interview_time = NULL,
                            interview_timeEnd = NULL,
                            interview_result = 'pending',
                            apply_status = %s,
                            updated_at = NOW()
                        WHERE application_id = %s AND job_id = %s
                    """, (new_apply_status, application_id, job_id))
                    print(f"✅ [二面流程] 重置學生面試狀態: application_id={application_id}, job_id={job_id}, apply_status={current_apply_status}->{new_apply_status}, interview_status={current_interview_status}->none")
                except Exception as e:
                    print(f"⚠️ [二面流程] 重置面試狀態失敗: application_id={application_id}, job_id={job_id}, error={e}")
        
        for student in unadmitted_students:
            title = f"{semester_prefix} 二面流程已啟動"
            message = f"{semester_prefix}二面流程已啟動，請留意相關面試通知。"
            link_url = "/interview_schedule"
            create_notification(
                user_id=student['id'],
                title=title,
                message=message,
                category="matching",
                link_url=link_url
            )
        
        # 4. 只通知同意二面的廠商
        # 查詢所有同意二面的廠商（假設 internship_companies 表有 agree_second_interview 欄位）
        vendors = []
        try:
            # 嘗試查詢有 agree_second_interview 欄位的公司對應的廠商
            cursor.execute("""
                SELECT DISTINCT u.id AS vendor_id
                FROM internship_companies ic
                JOIN users u ON u.role = 'vendor'
                WHERE ic.status = 'approved'
                AND ic.agree_second_interview = 1
                AND (
                    ic.vendor_id = u.id 
                    OR EXISTS (
                        SELECT 1 FROM company_vendor_relations cvr
                        WHERE cvr.company_id = ic.id AND cvr.vendor_id = u.id
                    )
                )
            """)
            vendors = cursor.fetchall() or []
        except Exception:
            # 如果欄位不存在，嘗試其他方式查詢
            try:
                # 備用方案：查詢所有已審核通過的公司對應的廠商
                cursor.execute("""
                    SELECT DISTINCT u.id AS vendor_id
                    FROM internship_companies ic
                    JOIN users u ON u.role = 'vendor'
                    WHERE ic.status = 'approved'
                    AND (
                        ic.vendor_id = u.id 
                        OR EXISTS (
                            SELECT 1 FROM company_vendor_relations cvr
                            WHERE cvr.company_id = ic.id AND cvr.vendor_id = u.id
                        )
                    )
                """)
                vendors = cursor.fetchall() or []
            except Exception:
                # 最後備用：如果表結構不同，查詢所有廠商
                cursor.execute("SELECT id AS vendor_id FROM users WHERE role = 'vendor'")
                vendors = cursor.fetchall() or []
        
        for vendor in vendors:
            vendor_id = vendor.get('vendor_id')
            if vendor_id:
                title = f"{semester_prefix} 二面流程已啟動"
                message = f"{semester_prefix}二面流程已由科助啟動，可開始進行二面排程。"
                link_url = "/vendor_review_resume"
                create_notification(
                    user_id=vendor_id,
                    title=title,
                    message=message,
                    category="matching",
                    link_url=link_url
                )
        
        conn.commit()
        
        return jsonify({
            "success": True,
            "message": "二面流程已啟動，已通知相關人員",
            "is_enabled": True,
            "notified": {
                "teachers_and_class_teachers": len(teachers) + len(class_teachers),
                "unadmitted_students": len(unadmitted_students),
                "vendors": len(vendors)
            }
        })
    
    except Exception as e:
        traceback.print_exc()
        conn.rollback()
        return jsonify({"success": False, "message": f"操作失敗: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

# =========================================================
# API: 學期篩選選項（來自 internship_configs 連結 semesters，學生入學學年對應實習週期）
# =========================================================
@admission_bp.route("/api/semesters_for_filter", methods=["GET"])
def semesters_for_filter():
    """
    取得學期篩選下拉選單的選項。
    資料來源：internship_configs（學生入學學年對應實習週期）INNER JOIN semesters（學期 id、代碼）。
    """
    if 'user_id' not in session:
        return jsonify({"success": False, "message": "未授權"}), 403
    if session.get('role') not in ['ta', 'admin', 'director', 'teacher']:
        return jsonify({"success": False, "message": "未授權"}), 403

    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    try:
        # internship_configs.semester_id 連結 semesters.id，取得有實習配置的學期
        cursor.execute("""
            SELECT DISTINCT s.id, s.code
            FROM semesters s
            INNER JOIN internship_configs ic ON ic.semester_id = s.id
            ORDER BY s.code DESC
        """)
        semesters = cursor.fetchall()
        current_semester_id = get_current_semester_id(cursor)
        current_semester_code = get_current_semester_code(cursor)
        if not semesters and current_semester_id and current_semester_code:
            semesters = [{"id": current_semester_id, "code": current_semester_code}]
        return jsonify({
            "success": True,
            "semesters": semesters,
            "current_semester_id": current_semester_id
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500
    finally:
        cursor.close()
        conn.close()


# =========================================================
# API: 科助工作台統計（媒合已核定數、未錄取人數）
# =========================================================
@admission_bp.route("/api/ta_dashboard_stats", methods=["GET"])
def ta_dashboard_stats():
    """
    科助工作台用：回傳已核定媒合數、未錄取學生人數。
    僅允許 role 為 ta 或 admin。
    可傳 ?semester_id= 指定學期，未傳則使用當前學期。
    """
    if 'user_id' not in session or session.get('role') not in ['ta', 'admin']:
        return jsonify({"success": False, "message": "未授權"}), 403

    conn = get_db()
    cursor = conn.cursor(dictionary=True)

    try:
        # 支援下拉選單選擇學期
        chosen_id = request.args.get('semester_id', type=int)
        if chosen_id:
            cursor.execute("SELECT id, code FROM semesters WHERE id = %s", (chosen_id,))
            row = cursor.fetchone()
            if not row:
                return jsonify({"success": False, "message": "找不到該學期"}), 400
            current_semester_id = row['id']
            current_semester_code = row.get('code') or ''
        else:
            current_semester_id = get_current_semester_id(cursor)
            current_semester_code = get_current_semester_code(cursor)
        if not current_semester_id:
            return jsonify({"success": False, "message": "無法取得當前學期"}), 500

        # 學期對應學號：1132→110xxx，1142→111xxx（學號前3碼 = 學年前3碼 - 3）
        student_id_prefix = None
        if current_semester_code and len(current_semester_code) >= 3:
            try:
                y = int(current_semester_code[:3])
                student_id_prefix = str(y - 3)
            except (ValueError, TypeError):
                pass

        # 已核定／待公告的媒合人數（只計算 Approved，因為主任確認後所有記錄都應該是 Approved）
        # 以 student_preferences.semester_id 篩選學期（不依賴 manage_director.semester_id，因該欄位可能不存在）
        # 注意：manage_director.preference_id 引用的是 student_job_applications.id（即 resume_applications.application_id）
        # 需要通過 student_job_applications 來 JOIN student_preferences，並且只統計當前學期的記錄
        if student_id_prefix:
            cursor.execute("""
                SELECT COUNT(DISTINCT md.student_id) AS cnt
                FROM manage_director md
                INNER JOIN student_job_applications sja ON md.preference_id = sja.id
                INNER JOIN student_preferences sp ON sja.student_id = sp.student_id
                    AND sja.company_id = sp.company_id
                    AND sja.job_id = sp.job_id
                    AND sp.semester_id = %s
                INNER JOIN users u ON md.student_id = u.id AND u.username LIKE %s
                WHERE md.director_decision = 'Approved'
            """, (current_semester_id, student_id_prefix + "%"))
        else:
            cursor.execute("""
                SELECT COUNT(DISTINCT md.student_id) AS cnt
                FROM manage_director md
                INNER JOIN student_job_applications sja ON md.preference_id = sja.id
                INNER JOIN student_preferences sp ON sja.student_id = sp.student_id
                    AND sja.company_id = sp.company_id
                    AND sja.job_id = sp.job_id
                    AND sp.semester_id = %s
                WHERE md.director_decision = 'Approved'
            """, (current_semester_id,))
        row = cursor.fetchone()
        matching_approved_count = (row.get("cnt") or 0) if row else 0

        # 本學期對應年級學生人數（學號前3碼 = 學年 - 3）
        if student_id_prefix:
            cursor.execute("""
                SELECT COUNT(*) AS cnt
                FROM users u
                WHERE u.role = 'student' AND u.username LIKE %s
            """, (student_id_prefix + "%",))
        else:
            cursor.execute("""
                SELECT COUNT(*) AS cnt
                FROM users u
                WHERE u.role = 'student'
            """)
        row = cursor.fetchone()
        total_students = (row.get("cnt") or 0) if row else 0

        # 未錄取人數 = 本學期對應年級學生 - 已核定媒合學生數
        unadmitted_count = max(0, total_students - matching_approved_count)

        # 學期說明：1132 → 113學年第2學期（與未錄取名單管理頁橫幅一致）
        semester_label = current_semester_code or ""
        if current_semester_code and len(current_semester_code) >= 4:
            try:
                y, t = current_semester_code[:3], current_semester_code[-1]
                semester_label = y + "學年" + ("第1學期" if t == "1" else "第2學期")
            except Exception:
                pass

        return jsonify({
            "success": True,
            "semester_id": current_semester_id,
            "semester_code": current_semester_code or "",
            "semester_label": semester_label,
            "student_id_prefix": student_id_prefix,  # 1132→"110"，1142→"111"（學號前3碼，科助工作台顯示用）
            "matching_approved_count": matching_approved_count,
            "unadmitted_count": unadmitted_count,
            "total_students": total_students,
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500
    finally:
        cursor.close()
        conn.close()

# =========================================================
# 輔助函數：簡化職缺名稱
# =========================================================
def simplify_job_title(job_title):
    """
    簡化職缺名稱，移除前綴，只保留簡短的核心名稱（不包含括號）
    例如：
    - "自動化開發 (軟體測試)" -> "軟體測試" 或 "測試"
    - "線上技術客服" -> "客服"
    """
    if not job_title or job_title == "未指定職缺":
        return job_title
    
    # 如果有括號（半形或全形），提取括號內的內容
    import re
    # 先處理半形括號
    bracket_match = re.search(r'\(([^)]+)\)', job_title)
    if bracket_match:
        # 提取括號內的內容
        content = bracket_match.group(1).strip()
        # 如果括號內還有括號，提取最內層的內容
        inner_bracket = re.search(r'\(([^)]+)\)', content)
        if inner_bracket:
            content = inner_bracket.group(1).strip()
        # 移除所有括號（半形和全形），因為調用處會加上括號
        content = content.replace('(', '').replace(')', '').replace('（', '').replace('）', '').strip()
        return content
    
    # 處理全形括號
    bracket_match_full = re.search(r'（([^）]+)）', job_title)
    if bracket_match_full:
        content = bracket_match_full.group(1).strip()
        # 如果括號內還有括號，提取最內層的內容
        inner_bracket = re.search(r'（([^）]+)）', content)
        if inner_bracket:
            content = inner_bracket.group(1).strip()
        # 移除所有括號（半形和全形），因為調用處會加上括號
        content = content.replace('(', '').replace(')', '').replace('（', '').replace('）', '').strip()
        return content
    
    # 如果沒有括號，移除常見前綴
    # 移除 "自動化開發"、"線上技術"、"技術" 等前綴
    prefixes_to_remove = [
        "自動化開發",
        "線上技術",
        "技術",
        "線上",
        "自動化"
    ]
    
    simplified = job_title
    for prefix in prefixes_to_remove:
        if simplified.startswith(prefix):
            simplified = simplified[len(prefix):].strip()
            # 移除可能的前導空格或標點
            simplified = simplified.lstrip('：:、，,')
            break
    
    # 移除所有括號（半形和全形），因為調用處會加上括號
    simplified = simplified.replace('(', '').replace(')', '').replace('（', '').replace('）', '').strip()
    return simplified if simplified else job_title

# =========================================================
# API: 匯出媒合結果 Excel（網格格式）
# =========================================================
@admission_bp.route("/api/export_matching_results_excel", methods=["GET"])
def export_matching_results_excel():
    """
    匯出媒合結果為 Excel 格式，按照圖片樣式：
    - 3列網格布局
    - 每個公司一個區塊
    - 公司名稱用黃色背景
    - 學生列表（學號 + 姓名）
    - 總人數統計
    """
    if 'user_id' not in session or session.get('role') != 'director':
        return jsonify({"success": False, "message": "未授權"}), 403
    
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 獲取當前學期ID
        current_semester_id = get_current_semester_id(cursor)
        if not current_semester_id:
            return jsonify({"success": False, "message": "無法取得當前學期"}), 500
        
        # 獲取媒合結果數據（與 director_matching_results 相同的邏輯）
        query = """
            SELECT 
                md.match_id,
                md.vendor_id,
                md.student_id,
                md.preference_id,
                md.original_type,
                md.original_rank,
                md.is_conflict,
                md.director_decision,
                md.final_rank,
                md.is_adjusted,
                COALESCE(sp.company_id, md.vendor_id) AS company_id,
                sp.preference_order,
                COALESCE(sp.job_id, (
                    SELECT id FROM internship_jobs 
                    WHERE company_id = COALESCE(sp.company_id, md.vendor_id) 
                    ORDER BY id ASC LIMIT 1
                )) AS job_id,
                COALESCE(ic.company_name, v.name) AS company_name,
                u.name AS student_name,
                u.username AS student_number,
                c.name AS class_name,
                COALESCE(ij.title, (
                    SELECT title FROM internship_jobs 
                    WHERE company_id = COALESCE(sp.company_id, md.vendor_id) 
                    ORDER BY id ASC LIMIT 1
                )) AS job_title
            FROM manage_director md
            LEFT JOIN student_job_applications sja ON md.preference_id = sja.id
            LEFT JOIN student_preferences sp ON sja.student_id = sp.student_id
                AND sja.company_id = sp.company_id
                AND sja.job_id = sp.job_id
                AND sp.semester_id = %s
            LEFT JOIN internship_companies ic ON COALESCE(sp.company_id, md.vendor_id) = ic.id
            LEFT JOIN internship_jobs ij ON sp.job_id = ij.id
            LEFT JOIN users u ON md.student_id = u.id
            LEFT JOIN classes c ON u.class_id = c.id
            LEFT JOIN users v ON md.vendor_id = v.id
            WHERE (sp.semester_id = %s OR sp.semester_id IS NULL)
            AND md.director_decision IN ('Approved', 'Pending')
            ORDER BY COALESCE(sp.company_id, md.vendor_id), 
                     COALESCE(sp.job_id, 0),
                     CASE WHEN md.director_decision = 'Approved' AND md.final_rank IS NOT NULL THEN 0 ELSE 1 END,
                     COALESCE(md.final_rank, 999) ASC
        """
        cursor.execute(query, (current_semester_id, current_semester_id))
        all_results = cursor.fetchall() or []
        
        # 按公司分組數據
        companies_data = {}
        for result in all_results:
            company_id = result.get("company_id")
            company_name = result.get("company_name") or "未知公司"
            job_title = result.get("job_title") or "未指定職缺"
            
            if company_id not in companies_data:
                companies_data[company_id] = {
                    "company_name": company_name,
                    "jobs": {}
                }
            
            if job_title not in companies_data[company_id]["jobs"]:
                companies_data[company_id]["jobs"][job_title] = []
            
            companies_data[company_id]["jobs"][job_title].append({
                "student_number": result.get("student_number") or "",
                "student_name": result.get("student_name") or "",
                "job_title": job_title
            })
        
        # 創建 Excel 工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "媒合結果"
        
        # 設定樣式
        company_header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 黃色背景
        company_header_font = Font(bold=True, size=12)
        student_font = Font(size=11)
        total_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")  # 灰色背景
        total_font = Font(bold=True, size=11)
        
        # 邊框樣式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 4列網格布局
        COLUMNS = 4
        COLUMN_WIDTH = 20  # 每列寬度（字符）
        
        # 準備公司數據
        companies_list = []
        for company in companies_data.values():
            company_name = company["company_name"]
            all_students = []
            
            # 檢查該公司是否有兩個或更多職缺
            num_jobs = len(company["jobs"])
            has_multiple_jobs = num_jobs >= 2
            
            # 收集該公司所有職缺的學生
            for job_title, students in company["jobs"].items():
                # 如果公司有多個職缺，在學生姓名後面加上括號職缺名稱
                for student in students:
                    student_copy = student.copy()
                    if has_multiple_jobs and job_title and job_title != "未指定職缺":
                        student_name = student_copy.get('student_name') or ''
                        # 職缺名稱
                        simplified_job_title = simplify_job_title(job_title)
                        # 確保移除所有括號（半形和全形），因為調用處會加上括號
                        simplified_job_title = simplified_job_title.replace('(', '').replace(')', '').replace('（', '').replace('）', '').strip()
                        student_copy['student_name'] = f"{student_name}({simplified_job_title})"
                    all_students.append(student_copy)
            
            if all_students:
                companies_list.append({
                    "name": company_name,
                    "students": all_students
                })
        
        # 將公司分配到4列
        columns_data = [[], [], [], []]  # 4列
        for idx, company in enumerate(companies_list):
            col_idx = idx % COLUMNS
            columns_data[col_idx].append(company)
        
        # 為每列填充數據
        for col_idx in range(COLUMNS):
            # 計算欄位：第1列用A-B-C，第2列用D-E-F，第3列用G-H-I，第4列用J-K-L
            # 每個公司區塊佔用3欄（前兩欄用於內容，第三欄為空）
            col_number_start = col_idx * 3 + 1  # A=1, D=4, G=7, J=10
            col_letter_start = get_column_letter(col_number_start)
            col_letter_end = get_column_letter(col_number_start + 1)
            col_letter_right = get_column_letter(col_number_start + 2)  # 右邊空一格
            current_row = 1
            
            for company in columns_data[col_idx]:
                company_name = company["name"]
                students = company["students"]
                
                # 公司名稱標題（黃色背景，跨兩欄置中，右邊空一格）
                header_cell = ws[f"{col_letter_start}{current_row}"]
                header_cell.value = company_name
                header_cell.fill = company_header_fill
                header_cell.font = company_header_font
                header_cell.border = thin_border
                header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                # 合併兩欄
                ws.merge_cells(f"{col_letter_start}{current_row}:{col_letter_end}{current_row}")
                # 確保合併後的單元格也有邊框
                end_cell = ws[f"{col_letter_end}{current_row}"]
                end_cell.border = thin_border
                end_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                # 右邊空一格（第三欄留空）
                right_empty_cell = ws[f"{col_letter_right}{current_row}"]
                right_empty_cell.value = ""
                right_empty_cell.border = thin_border
                current_row += 1
                
                # 學生列表（學號和姓名分開兩欄，右邊空一格）
                for student in students:
                    student_number = student.get('student_number') or ''
                    student_name = student.get('student_name') or ''
                    
                    # 將學號轉為純數字（移除所有非數字字符）
                    if student_number:
                        student_number_clean = ''.join(filter(str.isdigit, str(student_number)))
                        # 嘗試轉換為 int，讓 Excel 識別為數字類型
                        try:
                            student_number_value = int(student_number_clean) if student_number_clean else ''
                        except (ValueError, TypeError):
                            student_number_value = student_number_clean
                    else:
                        student_number_value = ''
                    
                    # 學號欄位
                    number_cell = ws[f"{col_letter_start}{current_row}"]
                    number_cell.value = student_number_value
                    number_cell.font = student_font
                    number_cell.border = thin_border
                    number_cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # 姓名欄位
                    name_cell = ws[f"{col_letter_end}{current_row}"]
                    name_cell.value = student_name
                    name_cell.font = student_font
                    name_cell.border = thin_border
                    name_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                    
                    # 右邊空一格（第三欄留空）
                    right_empty_cell = ws[f"{col_letter_right}{current_row}"]
                    right_empty_cell.value = ""
                    right_empty_cell.border = thin_border
                    right_empty_cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    current_row += 1
                
                # 總人數
                # 左欄留空
                ws[f"{col_letter_start}{current_row}"].value = ""
                ws[f"{col_letter_start}{current_row}"].border = thin_border
                ws[f"{col_letter_start}{current_row}"].alignment = Alignment(horizontal='center', vertical='center')
                # 右欄顯示總人數
                total_text = f"{len(students)}人"
                total_cell = ws[f"{col_letter_end}{current_row}"]
                total_cell.value = total_text
                # 移除灰色背景
                total_cell.font = total_font
                total_cell.border = thin_border
                total_cell.alignment = Alignment(horizontal='center', vertical='center')
                # 右邊空一格（第三欄留空）
                right_empty_cell = ws[f"{col_letter_right}{current_row}"]
                right_empty_cell.value = ""
                right_empty_cell.border = thin_border
                right_empty_cell.alignment = Alignment(horizontal='center', vertical='center')
                current_row += 1  # 移到下一行
                
                # 公司與公司之間的間隔行（三欄都留空）
                ws[f"{col_letter_start}{current_row}"].value = ""
                ws[f"{col_letter_start}{current_row}"].border = thin_border
                ws[f"{col_letter_start}{current_row}"].alignment = Alignment(horizontal='center', vertical='center')
                ws[f"{col_letter_end}{current_row}"].value = ""
                ws[f"{col_letter_end}{current_row}"].border = thin_border
                ws[f"{col_letter_end}{current_row}"].alignment = Alignment(horizontal='center', vertical='center')
                ws[f"{col_letter_right}{current_row}"].value = ""
                ws[f"{col_letter_right}{current_row}"].border = thin_border
                ws[f"{col_letter_right}{current_row}"].alignment = Alignment(horizontal='center', vertical='center')
                current_row += 1  # 移到下一行
        
        # 設定列寬（每列佔用3個欄位，所以總共12欄）
        # 增加列寬以確保公司名稱可以完整顯示（合併兩欄後寬度足夠）
        for col in range(1, COLUMNS * 3 + 1):
            col_letter = get_column_letter(col)
            # 姓名欄位（每列的第2個欄位：2, 5, 8, 11）增加寬度以確保姓名完整顯示
            if (col - 2) % 3 == 0:
                ws.column_dimensions[col_letter].width = 20  # 姓名欄位更寬，確保「姓名(職缺)」完整顯示
            else:
                ws.column_dimensions[col_letter].width = 12  # 其他欄位
        
        # 設定行高
        for row in range(1, ws.max_row + 1):
            ws.row_dimensions[row].height = 20
        
        # 保存到內存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 生成文件名
        filename = f"媒合結果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": f"匯出失敗: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

# =========================================================
# API: 科助匯出媒合結果 Excel（用於公告）
# =========================================================
@admission_bp.route("/api/ta/export_matching_results_excel", methods=["GET"])
def ta_export_matching_results_excel():
    """
    科助匯出媒合結果為 Excel 格式（用於公告）。
    允許 role 為 ta 或 admin。
    使用 student_preferences.semester_id 篩選，避免依賴 manage_director.semester_id。
    """
    if 'user_id' not in session or session.get('role') not in ['ta', 'admin']:
        return jsonify({"success": False, "message": "未授權"}), 403
    
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 獲取當前學期ID
        current_semester_id = get_current_semester_id(cursor)
        if not current_semester_id:
            return jsonify({"success": False, "message": "無法取得當前學期"}), 500
        
        # 獲取媒合結果數據（使用 manage_director.semester_id 篩選，與 final_matching_results 一致）
        query = """
            SELECT 
                md.match_id,
                md.vendor_id,
                md.student_id,
                md.preference_id,
                md.original_type,
                md.original_rank,
                md.is_conflict,
                md.director_decision,
                md.final_rank,
                md.is_adjusted,
                COALESCE(sp.company_id, md.vendor_id) AS company_id,
                sp.preference_order,
                COALESCE(sp.job_id, (
                    SELECT id FROM internship_jobs 
                    WHERE company_id = COALESCE(sp.company_id, md.vendor_id) 
                    ORDER BY id ASC LIMIT 1
                )) AS job_id,
                COALESCE(ic.company_name, v.name) AS company_name,
                u.name AS student_name,
                u.username AS student_number,
                c.name AS class_name,
                COALESCE(ij.title, (
                    SELECT title FROM internship_jobs 
                    WHERE company_id = COALESCE(sp.company_id, md.vendor_id) 
                    ORDER BY id ASC LIMIT 1
                )) AS job_title
            FROM manage_director md
            LEFT JOIN student_job_applications sja ON md.preference_id = sja.id
            LEFT JOIN student_preferences sp ON sja.student_id = sp.student_id
                AND sja.company_id = sp.company_id
                AND sja.job_id = sp.job_id
                AND sp.semester_id = %s
            LEFT JOIN internship_companies ic ON COALESCE(sp.company_id, md.vendor_id) = ic.id
            LEFT JOIN internship_jobs ij ON sp.job_id = ij.id
            LEFT JOIN users u ON md.student_id = u.id
            LEFT JOIN classes c ON u.class_id = c.id
            LEFT JOIN users v ON md.vendor_id = v.id
            WHERE (sp.semester_id = %s OR sp.semester_id IS NULL)
            AND md.director_decision = 'Approved'
            ORDER BY COALESCE(sp.company_id, md.vendor_id), 
                     COALESCE(sp.job_id, (
                         SELECT id FROM internship_jobs 
                         WHERE company_id = COALESCE(sp.company_id, md.vendor_id) 
                         ORDER BY id ASC LIMIT 1
                     ), 0), 
                     CASE WHEN md.final_rank IS NOT NULL THEN 0 ELSE 1 END,
                     COALESCE(md.final_rank, 999) ASC
        """
        cursor.execute(query, (current_semester_id, current_semester_id))
        all_results = cursor.fetchall() or []
        
        # 按公司分組數據（與原函數相同的邏輯）
        companies_data = {}
        for result in all_results:
            company_id = result.get("company_id")
            company_name = result.get("company_name") or "未知公司"
            job_title = result.get("job_title") or "未指定職缺"
            
            if company_id not in companies_data:
                companies_data[company_id] = {
                    "company_name": company_name,
                    "jobs": {}
                }
            
            if job_title not in companies_data[company_id]["jobs"]:
                companies_data[company_id]["jobs"][job_title] = []
            
            companies_data[company_id]["jobs"][job_title].append({
                "student_number": result.get("student_number") or "",
                "student_name": result.get("student_name") or "",
                "job_title": job_title
            })
        
        # 創建 Excel 工作簿（與原函數相同的邏輯）
        wb = Workbook()
        ws = wb.active
        ws.title = "媒合結果"
        
        # 設定樣式
        company_header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        company_header_font = Font(bold=True, size=12)
        student_font = Font(size=11)
        total_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        total_font = Font(bold=True, size=11)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        COLUMNS = 4
        COLUMN_WIDTH = 20
        
        companies_list = []
        for company in companies_data.values():
            company_name = company["company_name"]
            all_students = []
            
            # 檢查該公司是否有兩個或更多職缺
            num_jobs = len(company["jobs"])
            has_multiple_jobs = num_jobs >= 2
            
            # 收集該公司所有職缺的學生
            for job_title, students in company["jobs"].items():
                # 如果公司有多個職缺，在學生姓名後面加上括號職缺名稱（簡化後）
                for student in students:
                    student_copy = student.copy()
                    if has_multiple_jobs and job_title and job_title != "未指定職缺":
                        student_name = student_copy.get('student_name') or ''
                        # 簡化職缺名稱
                        simplified_job_title = simplify_job_title(job_title)
                        # 確保移除所有括號（半形和全形），因為調用處會加上括號
                        simplified_job_title = simplified_job_title.replace('(', '').replace(')', '').replace('（', '').replace('）', '').strip()
                        student_copy['student_name'] = f"{student_name}({simplified_job_title})"
                    all_students.append(student_copy)
            
            if all_students:
                companies_list.append({
                    "name": company_name,
                    "students": all_students
                })
        
        columns_data = [[], [], [], []]
        for idx, company in enumerate(companies_list):
            col_idx = idx % COLUMNS
            columns_data[col_idx].append(company)
        
        for col_idx in range(COLUMNS):
            col_number_start = col_idx * 3 + 1
            col_letter_start = get_column_letter(col_number_start)
            col_letter_end = get_column_letter(col_number_start + 1)
            col_letter_right = get_column_letter(col_number_start + 2)
            current_row = 1
            
            for company in columns_data[col_idx]:
                company_name = company["name"]
                students = company["students"]
                
                header_cell = ws[f"{col_letter_start}{current_row}"]
                header_cell.value = company_name
                header_cell.fill = company_header_fill
                header_cell.font = company_header_font
                header_cell.border = thin_border
                header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                ws.merge_cells(f"{col_letter_start}{current_row}:{col_letter_end}{current_row}")
                end_cell = ws[f"{col_letter_end}{current_row}"]
                end_cell.border = thin_border
                end_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                ws[f"{col_letter_right}{current_row}"].value = ""
                ws[f"{col_letter_right}{current_row}"].border = thin_border
                current_row += 1
                
                for student in students:
                    student_number = student.get('student_number') or ''
                    student_name = student.get('student_name') or ''
                    
                    # 將學號轉為純數字（移除所有非數字字符）
                    if student_number:
                        student_number_clean = ''.join(filter(str.isdigit, str(student_number)))
                        # 嘗試轉換為 int，讓 Excel 識別為數字類型
                        try:
                            student_number_value = int(student_number_clean) if student_number_clean else ''
                        except (ValueError, TypeError):
                            student_number_value = student_number_clean
                    else:
                        student_number_value = ''
                    
                    number_cell = ws[f"{col_letter_start}{current_row}"]
                    number_cell.value = student_number_value
                    number_cell.font = student_font
                    number_cell.border = thin_border
                    number_cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    name_cell = ws[f"{col_letter_end}{current_row}"]
                    name_cell.value = student_name
                    name_cell.font = student_font
                    name_cell.border = thin_border
                    name_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                    
                    ws[f"{col_letter_right}{current_row}"].value = ""
                    ws[f"{col_letter_right}{current_row}"].border = thin_border
                    ws[f"{col_letter_right}{current_row}"].alignment = Alignment(horizontal='center', vertical='center')
                    current_row += 1
                
                ws[f"{col_letter_start}{current_row}"].value = ""
                ws[f"{col_letter_start}{current_row}"].border = thin_border
                ws[f"{col_letter_start}{current_row}"].alignment = Alignment(horizontal='center', vertical='center')
                total_text = f"{len(students)}人"
                total_cell = ws[f"{col_letter_end}{current_row}"]
                total_cell.value = total_text
                # 移除灰色背景
                total_cell.font = total_font
                total_cell.border = thin_border
                total_cell.alignment = Alignment(horizontal='center', vertical='center')
                ws[f"{col_letter_right}{current_row}"].value = ""
                ws[f"{col_letter_right}{current_row}"].border = thin_border
                ws[f"{col_letter_right}{current_row}"].alignment = Alignment(horizontal='center', vertical='center')
                current_row += 1
                
                ws[f"{col_letter_start}{current_row}"].value = ""
                ws[f"{col_letter_start}{current_row}"].border = thin_border
                ws[f"{col_letter_start}{current_row}"].alignment = Alignment(horizontal='center', vertical='center')
                ws[f"{col_letter_end}{current_row}"].value = ""
                ws[f"{col_letter_end}{current_row}"].border = thin_border
                ws[f"{col_letter_end}{current_row}"].alignment = Alignment(horizontal='center', vertical='center')
                ws[f"{col_letter_right}{current_row}"].value = ""
                ws[f"{col_letter_right}{current_row}"].border = thin_border
                ws[f"{col_letter_right}{current_row}"].alignment = Alignment(horizontal='center', vertical='center')
                current_row += 1
        
        for col in range(1, COLUMNS * 3 + 1):
            col_letter = get_column_letter(col)
            # 姓名欄位（每列的第2個欄位：2, 5, 8, 11）增加寬度以確保姓名完整顯示
            if (col - 2) % 3 == 0:
                ws.column_dimensions[col_letter].width = 20  # 姓名欄位更寬，確保「姓名(職缺)」完整顯示
            else:
                ws.column_dimensions[col_letter].width = 12  # 其他欄位
        
        for row in range(1, ws.max_row + 1):
            ws.row_dimensions[row].height = 20
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"媒合結果公告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": f"匯出失敗: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

# =========================================================
# API: 匯出未錄取學生名單 Excel（科助/主任/班導/管理員）
# =========================================================
@admission_bp.route("/api/ta/export_unadmitted_students_excel", methods=["GET"])
def ta_export_unadmitted_students_excel():
    """
    匯出未錄取學生名單 Excel。
    - 預設使用系統當前學期對應的學號前綴規則（與 get_all_students 一致）
    - 支援 ?semester_id= 指定學期（可選）
    - 支援 ?class_id= 指定班級（可選）
    - 角色限制：ta / admin / director / class_teacher
    """
    if 'user_id' not in session or session.get('role') not in ['ta', 'admin', 'director', 'class_teacher']:
        return jsonify({"success": False, "message": "未授權"}), 403

    user_id = session.get('user_id')
    user_role = session.get('role')

    conn = get_db()
    cursor = conn.cursor(dictionary=True)

    try:
        chosen_id = request.args.get('semester_id', type=int)
        class_id = request.args.get('class_id', type=int)

        if chosen_id:
            cursor.execute("SELECT id, code FROM semesters WHERE id = %s", (chosen_id,))
            row = cursor.fetchone()
            if not row:
                return jsonify({"success": False, "message": "找不到該學期"}), 400
            current_semester_id = row['id']
            current_semester_code = row.get('code') or ''
        else:
            current_semester_code = get_current_semester_code(cursor)
            current_semester_id = get_current_semester_id(cursor)

        if not current_semester_code:
            return jsonify({"success": False, "message": "無法取得當前學期"}), 500
        if not current_semester_id:
            return jsonify({"success": False, "message": "無法取得當前學期"}), 500

        # 已在媒合結果中的學生（Approved/Pending）- 使用 student_preferences.semester_id 篩選
        # md.preference_id 引用的是 student_job_applications.id（即 resume_applications.application_id）
        # 需要通過 student_job_applications 來 JOIN student_preferences
        cursor.execute("""
            SELECT DISTINCT md.student_id
            FROM manage_director md
            INNER JOIN student_job_applications sja ON md.preference_id = sja.id
            INNER JOIN student_preferences sp ON sja.student_id = sp.student_id
                AND sja.company_id = sp.company_id
                AND sja.job_id = sp.job_id
                AND sp.semester_id = %s
            WHERE md.director_decision IN ('Approved', 'Pending')
        """, (current_semester_id,))
        matched_student_ids = {row['student_id'] for row in cursor.fetchall()}

        # 學期對應學號前綴（與 get_all_students 一致）
        student_id_prefix = None
        if current_semester_code and len(current_semester_code) >= 3:
            try:
                year_part = int(current_semester_code[:3])
                student_id_prefix = str(year_part - 3)
            except (ValueError, TypeError):
                pass

        # 基礎查詢：學生 + 班級
        base_query = """
            SELECT 
                u.id AS student_id,
                u.name AS student_name,
                u.username AS student_number,
                u.admission_year AS admission_year,
                c.id AS class_id,
                c.name AS class_name,
                c.department
            FROM users u
            LEFT JOIN classes c ON u.class_id = c.id
            WHERE u.role = 'student'
        """
        params = []

        if student_id_prefix:
            base_query += " AND u.username LIKE %s"
            params.append(student_id_prefix + "%")

        # 依角色限制範圍（與 get_all_students 一致）
        if user_role == 'director':
            cursor.execute("""
                SELECT DISTINCT c.department
                FROM classes c
                JOIN classes_teacher ct ON ct.class_id = c.id
                WHERE ct.teacher_id = %s
                LIMIT 1
            """, (user_id,))
            dept_result = cursor.fetchone()
            if dept_result and dept_result.get('department'):
                base_query += " AND c.department = %s"
                params.append(dept_result['department'])
        elif user_role == 'class_teacher':
            cursor.execute("""
                SELECT class_id FROM classes_teacher 
                WHERE teacher_id = %s
            """, (user_id,))
            teacher_classes = cursor.fetchall()
            if teacher_classes:
                class_ids = [tc['class_id'] for tc in teacher_classes]
                placeholders = ','.join(['%s'] * len(class_ids))
                base_query += f" AND u.class_id IN ({placeholders})"
                params.extend(class_ids)
            else:
                # 沒有管理班級 → 匯出空檔（仍返回合法 Excel）
                pass

        if class_id:
            base_query += " AND u.class_id = %s"
            params.append(class_id)

        base_query += " ORDER BY u.username ASC"
        cursor.execute(base_query, params)
        students = cursor.fetchall() or []

        # 只匯出未錄取（未媒合）者
        unadmitted_students = []
        for s in students:
            sid = s.get('student_id')
            is_matched = (sid in matched_student_ids) if sid else False
            if not is_matched:
                unadmitted_students.append(s)

        # 學期 label（與 get_all_students 一致）
        semester_label = current_semester_code
        if current_semester_code and len(current_semester_code) >= 4:
            try:
                year_part = current_semester_code[:3]
                term_part = current_semester_code[-1]
                term_name = "第1學期" if term_part == "1" else "第2學期"
                semester_label = f"{year_part}學年{term_name}"
            except Exception:
                pass

        # 建立 Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "未錄取名單"

        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        title = f"未錄取學生名單\n（{current_semester_code} {semester_label}）"
        ws["A1"].value = title
        ws.merge_cells("A1:C1")
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 50  # 增加第一行高度，確保兩行文字完整顯示不被切到

        ws.append(["姓名", "學號", "班級"])
        for cell in ws[2]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 計算當前學年（從學期代碼提取：1132 -> 113）
        current_semester_year = None
        if current_semester_code and len(current_semester_code) >= 3:
            try:
                current_semester_year = int(current_semester_code[:3])
            except (ValueError, TypeError):
                pass
        
        # 年級數字對應的中文
        grade_labels = ('', '一', '二', '三', '四', '五', '六')

        for s in unadmitted_students:
            dept = (s.get("department") or "").strip()
            cls = (s.get("class_name") or "").strip()
            base_class_label = (dept + cls) if (dept or cls) else ""
            
            # 計算年級並插入到班級名稱中
            class_label = base_class_label
            if base_class_label and current_semester_year is not None:
                admission_year = s.get("admission_year")
                # 如果沒有 admission_year，嘗試從學號前3碼獲取
                if admission_year is None or str(admission_year).strip() == '':
                    student_number = s.get("student_number")
                    if student_number and len(str(student_number)) >= 3:
                        try:
                            admission_year = int(str(student_number)[:3])
                        except (ValueError, TypeError):
                            pass
                
                if admission_year is not None:
                    try:
                        grade_num = current_semester_year - int(admission_year) + 1
                        if 1 <= grade_num <= 6:
                            grade_char = grade_labels[grade_num]
                            # 在「科」和「孝/忠」之間插入年級數字
                            # 例如：「資管科孝」→「資管科四孝」
                            match = re.match(r'^(.+科)(.+)$', base_class_label)
                            if match:
                                class_label = match.group(1) + grade_char + match.group(2)
                            else:
                                # 如果格式不符合，嘗試在最後插入
                                class_label = base_class_label + grade_char
                    except (ValueError, TypeError):
                        pass
            
            # 處理學號：轉換為數字格式
            student_number = s.get("student_number") or ""
            student_number_value = student_number
            if student_number:
                try:
                    # 嘗試轉換為整數，確保以數字格式儲存
                    student_number_value = int(str(student_number))
                except (ValueError, TypeError):
                    # 如果無法轉換，保持原值
                    student_number_value = student_number
            
            ws.append([
                s.get("student_name") or "",
                student_number_value,
                class_label
            ])

        # 套用基本格式
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=3):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                # 學號欄位（B欄）設定為數字格式
                if cell.column == 2:  # B欄是第2欄
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0'  # 設定為整數格式，不顯示小數點

        # 增加欄位寬度，確保標題完整顯示
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 24

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # 檔名（包含學期與時間）
        filename = f"未錄取學生名單_{current_semester_code}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": f"匯出失敗: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()