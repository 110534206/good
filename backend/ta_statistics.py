from flask import Blueprint, request, jsonify, session,render_template,redirect, send_file
from config import get_db
from datetime import datetime
from semester import get_current_semester_code
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook
import traceback
import io 
import os

ta_statistics_bp = Blueprint("ta_statistics_bp", __name__, )

# =========================================================
# API: 取得全系統統計總覽
# =========================================================
@ta_statistics_bp.route("/api/overview", methods=["GET"])
def get_overview():
    """科助端取得全系統統計總覽"""
    if 'user_id' not in session or session.get('role') not in ['ta', 'admin']:
        return jsonify({"success": False, "message": "未授權"}), 403
    
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 獲取當前學期
        current_semester_code = get_current_semester_code(cursor)
        
        # 1. 學生總數統計
        cursor.execute("SELECT COUNT(*) AS total FROM users WHERE role = 'student'")
        total_students = cursor.fetchone()['total']
        
        # 2. 履歷統計
        cursor.execute("""
            SELECT 
                COUNT(DISTINCT r.user_id) AS students_with_resume,
                COUNT(DISTINCT CASE WHEN r.status = 'approved' THEN r.user_id END) AS students_approved,
                COUNT(DISTINCT CASE WHEN r.status = 'rejected' THEN r.user_id END) AS students_rejected,
                COUNT(DISTINCT CASE WHEN r.status = 'uploaded' THEN r.user_id END) AS students_pending,
                COUNT(*) AS total_resumes
            FROM resumes r
            WHERE r.semester_id = (SELECT id FROM semesters WHERE is_active = 1 LIMIT 1)
        """)
        resume_stats = cursor.fetchone()
        
        # 3. 志願序統計
        cursor.execute("""
            SELECT 
                COUNT(DISTINCT sp.student_id) AS students_with_preferences,
                COUNT(DISTINCT CASE WHEN sp.status = 'approved' THEN sp.student_id END) AS students_approved,
                COUNT(DISTINCT CASE WHEN sp.status = 'rejected' THEN sp.student_id END) AS students_rejected,
                COUNT(DISTINCT CASE WHEN sp.status = 'pending' THEN sp.student_id END) AS students_pending,
                COUNT(*) AS total_preferences
            FROM student_preferences sp
            WHERE sp.semester_id = (SELECT id FROM semesters WHERE is_active = 1 LIMIT 1)
        """)
        preference_stats = cursor.fetchone()
        
        # 4. 公司統計
        cursor.execute("""
            SELECT 
                COUNT(*) AS total_companies,
                COUNT(CASE WHEN status = 'approved' THEN 1 END) AS approved_companies,
                COUNT(CASE WHEN status = 'pending' THEN 1 END) AS pending_companies,
                COUNT(CASE WHEN status = 'rejected' THEN 1 END) AS rejected_companies
            FROM internship_companies
        """)
        company_stats = cursor.fetchone()
        
        # 5. 各公司被選擇次數（前10名）
        cursor.execute("""
            SELECT 
                ic.company_name,
                COUNT(sp.id) AS preference_count
            FROM internship_companies ic
            LEFT JOIN student_preferences sp ON ic.id = sp.company_id
            GROUP BY ic.id, ic.company_name
            ORDER BY preference_count DESC
            LIMIT 10
        """)
        top_companies = cursor.fetchall()
        
        # 6. 計算完成率
        resume_completion_rate = round(
            (resume_stats['students_with_resume'] or 0) * 100.0 / total_students 
            if total_students > 0 else 0, 2
        )
        preference_completion_rate = round(
            (preference_stats['students_with_preferences'] or 0) * 100.0 / total_students 
            if total_students > 0 else 0, 2
        )
        resume_approval_rate = round(
            (resume_stats['students_approved'] or 0) * 100.0 / (resume_stats['students_with_resume'] or 1)
            if resume_stats['students_with_resume'] else 0, 2
        )
        preference_approval_rate = round(
            (preference_stats['students_approved'] or 0) * 100.0 / (preference_stats['students_with_preferences'] or 1)
            if preference_stats['students_with_preferences'] else 0, 2
        )
        
        return jsonify({
            "success": True,
            "current_semester": current_semester_code,
            "total_students": total_students,
            "resume_stats": {
                **resume_stats,
                "completion_rate": resume_completion_rate,
                "approval_rate": resume_approval_rate
            },
            "preference_stats": {
                **preference_stats,
                "completion_rate": preference_completion_rate,
                "approval_rate": preference_approval_rate
            },
            "company_stats": company_stats,
            "top_companies": top_companies
        })
    
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": f"查詢失敗: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

# =========================================================
# API: 取得各班級統計列表
# =========================================================
@ta_statistics_bp.route("/api/classes", methods=["GET"])
def get_classes_statistics():
    """科助端取得各班級統計列表"""
    if 'user_id' not in session or session.get('role') not in ['ta', 'admin']:
        return jsonify({"success": False, "message": "未授權"}), 403

    conn = get_db()
    cursor = conn.cursor(dictionary=True)

    try:
        # 取得目前學期
        cursor.execute("SELECT id FROM semesters WHERE is_active = 1 LIMIT 1")
        current_semester = cursor.fetchone()
        semester_id = current_semester['id'] if current_semester else None

        # 查詢所有班級統計 (SQL中已移除 ORDER BY)
        cursor.execute("""
            SELECT 
                c.id AS class_id,
                c.name AS class_name,
                c.department,
                COUNT(DISTINCT u.id) AS total_students,
                -- 履歷統計
                COUNT(DISTINCT r.user_id) AS students_with_resume,
                COUNT(DISTINCT CASE WHEN r.status = 'approved' THEN r.user_id END) AS students_resume_approved,
                COUNT(DISTINCT CASE WHEN r.status = 'rejected' THEN r.user_id END) AS students_resume_rejected,
                COUNT(DISTINCT CASE WHEN r.status = 'uploaded' THEN r.user_id END) AS students_resume_pending,
                -- 志願序統計
                COUNT(DISTINCT sp.student_id) AS students_with_preferences,
                COUNT(DISTINCT CASE WHEN sp.status = 'approved' THEN sp.student_id END) AS students_preferences_approved,
                COUNT(DISTINCT CASE WHEN sp.status = 'rejected' THEN sp.student_id END) AS students_preferences_rejected,
                COUNT(DISTINCT CASE WHEN sp.status = 'pending' THEN sp.student_id END) AS students_preferences_pending
            FROM classes c
            LEFT JOIN users u ON u.class_id = c.id AND u.role = 'student'
            LEFT JOIN resumes r ON r.user_id = u.id AND r.semester_id = %s
            LEFT JOIN student_preferences sp ON sp.student_id = u.id AND sp.semester_id = %s
            GROUP BY c.id, c.name, c.department
        """, (semester_id, semester_id))

        classes_stats = cursor.fetchall()

        # ✅ 改成依班級名稱中數字遞增排序 (Python 邏輯)
        def extract_grade_num(name):
            """從班級名稱取出年級數字 (ex: 資一孝 → 1)"""
            # 增加 '六' 以應對五專，可根據貴校情況調整
            mapping = {'一': 1, '二': 2, '三': 3, '四': 4, '五': 5, '六': 6} 
            
            # 遍歷班級名稱，尋找第一個匹配的年級數字
            for ch, num in mapping.items():
                if ch in name:
                    return num
            return 99  # 沒找到則放列表的最後面

        # 核心排序：主要依據年級數字 (數字遞增)，次要依據完整班級名稱 (字串遞增，確保甲班排在乙班前)
        classes_stats.sort(
            key=lambda x: (extract_grade_num(x['class_name']), x['class_name'])
        )

        # 計算完成率
        for stat in classes_stats:
            total = stat['total_students'] or 0
            # 履歷完成率
            stat['resume_completion_rate'] = round(
                (stat['students_with_resume'] or 0) * 100.0 / total if total > 0 else 0, 2
            )
            # 履歷通過率
            stat['resume_approval_rate'] = round(
                (stat['students_resume_approved'] or 0) * 100.0 / (stat['students_with_resume'] or 1)
                if stat['students_with_resume'] else 0, 2
            )
            # 志願序完成率
            stat['preference_completion_rate'] = round(
                (stat['students_with_preferences'] or 0) * 100.0 / total if total > 0 else 0, 2
            )
            # 志願序通過率
            stat['preference_approval_rate'] = round(
                (stat['students_preferences_approved'] or 0) * 100.0 / (stat['students_with_preferences'] or 1)
                if stat['students_with_preferences'] else 0, 2
            )

        return jsonify({
            "success": True,
            "classes": classes_stats
        })

    except Exception as e:
        # 請確保檔案頂部有 import traceback
        import traceback 
        traceback.print_exc()
        return jsonify({"success": False, "message": f"查詢失敗: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

# --------------------------------
# 班級列表
# --------------------------------
@ta_statistics_bp.route('/api/get_classes')
def get_classes():
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute(
            "SELECT id, name, department, admission_year FROM classes "
            "ORDER BY department ASC, admission_year DESC, name ASC"
        )
        classes = cursor.fetchall()
        
        # 為每個班級計算學期代碼（假設是第2學期，所以 113 -> 1132, 114 -> 1142）
        for c in classes:
            if c.get('admission_year'):
                admission_year = c['admission_year']
                # 如果 admission_year 已經是 4 位數（如 1132、1142），直接使用
                if isinstance(admission_year, int):
                    if admission_year >= 1000:  # 已經是完整學期代碼
                        c['semester_code'] = str(admission_year)
                    else:  # 只有年度（如 113、114），轉換為學期代碼（假設是第2學期）
                        c['semester_code'] = f"{admission_year}2"
                else:
                    admission_str = str(admission_year)
                    if len(admission_str) >= 4:  # 已經是完整學期代碼
                        c['semester_code'] = admission_str
                    elif admission_str.isdigit():  # 只有年度，轉換為學期代碼
                        c['semester_code'] = admission_str + "2"
                    else:
                        c['semester_code'] = None
            else:
                c['semester_code'] = None
        
        return jsonify({"success": True, "classes": classes})
    except Exception as e:
        print(f"取得班級列表錯誤: {e}")
        return jsonify({"success": False, "message": "取得班級列表失敗"}), 500
    finally:
        cursor.close()
        conn.close()

# =========================================================
# API: 取得學生資料（依班級篩選）
# =========================================================
@ta_statistics_bp.route('/api/get_students_by_class', methods=['GET'])
def get_students_by_class():
    # 權限檢查：統計功能通常允許 ta, admin 訪問
    if 'user_id' not in session or session.get('role') not in ['ta', 'admin']:
        return jsonify({"success": False, "message": "未授權"}), 403

    class_id = request.args.get('class_id')
    semester_code = request.args.get('semester_code')
    
    print(f"[DEBUG] 收到請求: class_id={class_id} (type: {type(class_id)}), semester_code={semester_code}")
    
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    try:
        semester_id = None
        if semester_code:
            cursor.execute("SELECT id FROM semesters WHERE code = %s LIMIT 1", (semester_code,))
            semester_row = cursor.fetchone()
            semester_id = semester_row['id'] if semester_row else None
            print(f"[DEBUG] semester_code={semester_code} -> semester_id={semester_id}")

        # 確保 class_id 是整數類型
        if class_id and class_id != "all":
            try:
                class_id = int(class_id)
                print(f"[DEBUG] class_id 轉換為整數: {class_id}")
            except (ValueError, TypeError):
                print(f"[ERROR] class_id 無法轉換為整數: {class_id}")
                return jsonify({"success": False, "message": f"無效的班級ID: {class_id}"}), 400

        # 直接使用管理員 API 的方式查詢所有學生（不限制班級）
        # 這樣可以確保能抓到所有學生資料
        query = """
            SELECT 
                u.id, u.username, u.name, u.email, u.class_id, u.role,
                c.name AS class_name,
                c.department
            FROM users u
            LEFT JOIN classes c ON u.class_id = c.id
            WHERE u.role = 'student'
            ORDER BY u.username
        """
        cursor.execute(query)
        all_students = cursor.fetchall()
        print(f"[DEBUG] 查詢到所有學生: {len(all_students)} 位")
        
        # 如果指定了班級，根據班級類型（忠/孝）過濾
        if class_id and class_id != "all":
            try:
                class_id_int = int(class_id)
                # 先查詢該班級的名稱，判斷是「忠」還是「孝」
                cursor.execute("SELECT name, department FROM classes WHERE id = %s", (class_id_int,))
                class_info = cursor.fetchone()
                
                if class_info:
                    class_name = class_info.get('name', '')
                    department = class_info.get('department', '').replace('管科', '')
                    full_class_name = f"{department}{class_name}"
                    
                    # 判斷是「忠」還是「孝」
                    if '忠' in full_class_name or '忠' in class_name:
                        class_type = '忠'
                    elif '孝' in full_class_name or '孝' in class_name:
                        class_type = '孝'
                    else:
                        class_type = None
                    
                    print(f"[DEBUG] 班級 {class_id_int} ({full_class_name}) 類型: {class_type}")
                    
                    if class_type:
                        # 查詢所有「忠」或「孝」班的學生
                        students = []
                        for s in all_students:
                            s_class_name = s.get('class_name', '')
                            s_department = (s.get('department', '') or '').replace('管科', '')
                            s_full_name = f"{s_department}{s_class_name}"
                            if class_type in s_full_name or class_type in s_class_name:
                                students.append(s)
                        print(f"[DEBUG] 查詢到所有「{class_type}」班的學生: {len(students)} 位")
                    else:
                        # 如果無法判斷類型，只查詢該班級的學生
                        students = [s for s in all_students if s.get('class_id') == class_id_int]
                        print(f"[DEBUG] 查詢到 class_id={class_id_int} 的學生: {len(students)} 位")
                else:
                    students = []
                    print(f"[DEBUG] 找不到 class_id={class_id_int} 的班級")
            except (ValueError, TypeError) as e:
                students = []
                print(f"[DEBUG] class_id 轉換失敗: {e}")
        else:
            students = all_students
        
        print(f"[DEBUG] 最終查詢結果: {len(students)} 位學生")
        
        # 嚴格按照選擇的班級顯示，如果沒有學生就返回空列表（前端會顯示「無資料」）
        
        # 顯示所有學生的 class_id 供調試
        if len(all_students) > 0:
            print(f"[DEBUG] 所有學生的 class_id 分佈: {[{'id': s['id'], 'username': s['username'], 'class_id': s.get('class_id'), 'class_name': s.get('class_name')} for s in all_students[:10]]}")
        
        print(f"[DEBUG] 查詢到 {len(students)} 位學生，class_id={class_id}, semester_code={semester_code}")
        
        # 如果沒有找到學生，進行詳細診斷
        if len(students) == 0:
            # 1. 檢查系統中是否有任何學生
            cursor.execute("SELECT COUNT(*) as total FROM users WHERE role = 'student'", ())
            total_students = cursor.fetchone()
            print(f"[DEBUG] 系統總學生數: {total_students.get('total', 0)}")
            
            # 2. 如果指定了班級，檢查該班級
            if class_id and class_id != "all":
                cursor.execute("SELECT id, name, department FROM classes WHERE id = %s", (class_id,))
                class_info = cursor.fetchone()
                if class_info:
                    print(f"[DEBUG] 班級存在: {class_info}")
                    # 檢查該班級是否有任何使用者（不限制 role）
                    cursor.execute("SELECT COUNT(*) as total FROM users WHERE class_id = %s", (class_id,))
                    total_users = cursor.fetchone()
                    print(f"[DEBUG] 該班級總使用者數（不限角色）: {total_users.get('total', 0)}")
                    
                    # 檢查該班級是否有學生（使用字串比較，以防類型問題）
                    cursor.execute("SELECT COUNT(*) as total FROM users WHERE class_id = %s AND role = 'student'", (class_id,))
                    class_students = cursor.fetchone()
                    print(f"[DEBUG] 該班級學生數: {class_students.get('total', 0)}")
                    
                    # 列出該班級的所有使用者（用於調試）
                    cursor.execute("SELECT id, username, name, role, class_id FROM users WHERE class_id = %s LIMIT 5", (class_id,))
                    sample_users = cursor.fetchall()
                    print(f"[DEBUG] 該班級前5位使用者範例: {sample_users}")
                else:
                    print(f"[DEBUG] 警告：找不到 class_id={class_id} 的班級")
            
            # 3. 列出所有班級及其學生數（用於調試）
            cursor.execute("""
                SELECT c.id, c.name, c.department, COUNT(u.id) as student_count
                FROM classes c
                LEFT JOIN users u ON c.id = u.class_id AND u.role = 'student'
                GROUP BY c.id, c.name, c.department
                ORDER BY c.id
                LIMIT 10
            """)
            class_stats = cursor.fetchall()
            print(f"[DEBUG] 前10個班級的學生分佈: {class_stats}")

        # 為每個學生查詢詳細的履歷和志願序資訊
        for student in students:
            student_id = student['id']
            
            # 查詢履歷資訊 - 不限制學期，抓取所有履歷
            try:
                resume_query = """
                    SELECT r.id, r.filepath, r.status, r.created_at, r.updated_at,
                           r.reviewed_by, r.comment, r.semester_id, r.original_filename
                    FROM resumes r
                    WHERE r.user_id = %s
                    ORDER BY r.created_at DESC
                """
                resume_params = [student_id]
                
                cursor.execute(resume_query, resume_params)
                resumes = cursor.fetchall()
                
                print(f"[DEBUG] 學生 {student['username']} (ID: {student_id}) 有 {len(resumes)} 筆履歷")
                if len(resumes) > 0:
                    print(f"[DEBUG] 第一筆履歷: ID={resumes[0].get('id')}, status={resumes[0].get('status')}")
            except Exception as e:
                print(f"[ERROR] 查詢履歷失敗 (學生 {student['username']}): {str(e)}")
                resumes = []
            
            # 格式化履歷日期
            for resume in resumes:
                if isinstance(resume.get('created_at'), datetime):
                    resume['created_at'] = resume['created_at'].strftime("%Y-%m-%d %H:%M:%S")
                if isinstance(resume.get('updated_at'), datetime):
                    resume['reviewed_at'] = resume['updated_at'].strftime("%Y-%m-%d %H:%M:%S")
                else:
                    resume['reviewed_at'] = None
                # 將 comment 映射為 reject_reason（用於前端顯示）
                resume['reject_reason'] = resume.get('comment', '')
            
            student['resumes'] = resumes
            student['resume_count'] = len(resumes)
            
            # 查詢志願序資訊 - 不限制學期，抓取所有志願序
            try:
                preference_query = """
                    SELECT sp.id, sp.preference_order, sp.status, sp.submitted_at,
                           sp.company_id, sp.job_id, sp.job_title, sp.semester_id,
                           ic.company_name, ic.location AS company_address,
                           ic.contact_person, ic.contact_email, ic.contact_phone,
                           ij.title AS job_title_full, ij.description AS job_description
                    FROM student_preferences sp
                    LEFT JOIN internship_companies ic ON sp.company_id = ic.id
                    LEFT JOIN internship_jobs ij ON sp.job_id = ij.id
                    WHERE sp.student_id = %s
                    ORDER BY sp.preference_order ASC
                """
                preference_params = [student_id]
                
                cursor.execute(preference_query, preference_params)
                preferences = cursor.fetchall()
                
                print(f"[DEBUG] 學生 {student['username']} (ID: {student_id}) 有 {len(preferences)} 筆志願序")
                if len(preferences) > 0:
                    print(f"[DEBUG] 第一筆志願序: order={preferences[0].get('preference_order')}, company={preferences[0].get('company_name')}")
            except Exception as e:
                print(f"[ERROR] 查詢志願序失敗 (學生 {student['username']}): {str(e)}")
                preferences = []
            
            # 格式化志願序日期
            for pref in preferences:
                if isinstance(pref.get('submitted_at'), datetime):
                    pref['submitted_at'] = pref['submitted_at'].strftime("%Y-%m-%d %H:%M:%S")
            
            student['preferences'] = preferences
            student['preference_count'] = len(preferences)

        return jsonify({"success": True, "students": students})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": f"取得學生資料失敗: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

# --------------------------------
# 單一班級統計 (新增部分)
# --------------------------------
@ta_statistics_bp.route('/api/get_class_stats/<int:class_id>', methods=['GET'])
def get_class_stats(class_id):
    """取得單一班級的實習進度統計資料"""
    # 這裡假設只有科助或管理員可以查看
    if 'user_id' not in session or session.get('role') not in ['admin', 'ta']:
        return jsonify({"success": False, "message": "未授權"}), 403

    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    try:
        semester_code = request.args.get('semester_code')
        semester_id = None
        if semester_code:
            cursor.execute("SELECT id FROM semesters WHERE code=%s LIMIT 1", (semester_code,))
            semester = cursor.fetchone()
            semester_id = semester['id'] if semester else None

        # 1. 查詢班級名稱和部門
        cursor.execute("SELECT name, department FROM classes WHERE id = %s", (class_id,))
        class_info = cursor.fetchone()
        if not class_info:
            return jsonify({"success": False, "message": "找不到該班級資料"}), 404

        # 根據班級類型（忠/孝）計算統計
        class_name = class_info.get('name', '')
        department = class_info.get('department', '').replace('管科', '')
        full_class_name = f"{department}{class_name}"
        
        # 判斷是「忠」還是「孝」
        if '忠' in full_class_name or '忠' in class_name:
            class_type = '忠'
            class_pattern = '%忠%'
        elif '孝' in full_class_name or '孝' in class_name:
            class_type = '孝'
            class_pattern = '%孝%'
        else:
            class_type = None
            class_pattern = None
        
        if class_type:
            # 查詢所有「忠」或「孝」班的統計
            cursor.execute("""
                SELECT
                    COUNT(u.id) AS total_students,
                    SUM(
                        CASE WHEN EXISTS (
                            SELECT 1 FROM resumes r WHERE r.user_id = u.id
                        ) THEN 1 ELSE 0 END
                    ) AS students_with_resume,
                    SUM(
                        CASE WHEN EXISTS (
                            SELECT 1 FROM student_preferences sp WHERE sp.student_id = u.id
                        ) THEN 1 ELSE 0 END
                    ) AS students_with_preference
                FROM users u
                LEFT JOIN classes c ON u.class_id = c.id
                WHERE u.role = 'student' 
                  AND (c.name LIKE %s OR CONCAT(REPLACE(c.department, '管科', ''), c.name) LIKE %s)
            """, (class_pattern, class_pattern))
            stats = cursor.fetchone()
            class_name = f"所有{class_type}班"
        else:
            # 如果無法判斷類型，只查詢該班級的統計
            cursor.execute("""
                SELECT
                    COUNT(u.id) AS total_students,
                    SUM(
                        CASE WHEN EXISTS (
                            SELECT 1 FROM resumes r WHERE r.user_id = u.id
                        ) THEN 1 ELSE 0 END
                    ) AS students_with_resume,
                    SUM(
                        CASE WHEN EXISTS (
                            SELECT 1 FROM student_preferences sp WHERE sp.student_id = u.id
                        ) THEN 1 ELSE 0 END
                    ) AS students_with_preference
                FROM users u
                WHERE u.class_id = %s AND u.role = 'student'
            """, (class_id,))
            stats = cursor.fetchone()

        # 組合結果
        result = {
            "class_name": class_name,
            "total_students": stats['total_students'] if stats and stats.get('total_students') is not None else 0,
            "students_with_resume": stats['students_with_resume'] if stats and stats.get('students_with_resume') is not None else 0,
            "students_with_preference": stats['students_with_preference'] if stats and stats.get('students_with_preference') is not None else 0
        }
        
        print(f"[DEBUG] 班級統計: {result}")
        
        return jsonify({"success": True, "stats": result})
            
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500
    finally:
        cursor.close()
        conn.close()

# --------------------------------
# 公司統計
# --------------------------------
@ta_statistics_bp.route('/api/manage_companies_stats')
def manage_companies_stats():
    class_id = request.args.get("class_id")
    semester_code = request.args.get("semester_code")
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    try:
        semester_id = None
        if semester_code:
            cursor.execute("SELECT id FROM semesters WHERE code=%s LIMIT 1", (semester_code,))
            semester = cursor.fetchone()
            semester_id = semester['id'] if semester else None

        # 各公司被選志願次數 - 根據班級類型（忠/孝）過濾
        company_params = []
        sp_semester_clause = ""
        u_class_clause = ""
        
        if semester_id:
            sp_semester_clause = " AND sp.semester_id = %s"
            company_params.append(semester_id)
        
        if class_id and class_id != "all":
            try:
                class_id_int = int(class_id)
                # 先查詢該班級的名稱，判斷是「忠」還是「孝」
                cursor.execute("SELECT name, department FROM classes WHERE id = %s", (class_id_int,))
                class_info = cursor.fetchone()
                
                if class_info:
                    class_name = class_info.get('name', '')
                    department = class_info.get('department', '').replace('管科', '')
                    full_class_name = f"{department}{class_name}"
                    
                    # 判斷是「忠」還是「孝」
                    if '忠' in full_class_name or '忠' in class_name:
                        class_pattern = '%忠%'
                    elif '孝' in full_class_name or '孝' in class_name:
                        class_pattern = '%孝%'
                    else:
                        class_pattern = None
                    
                    if class_pattern:
                        # 使用班級類型過濾
                        u_class_clause = " AND (cls.name LIKE %s OR CONCAT(REPLACE(cls.department, '管科', ''), cls.name) LIKE %s)"
                        company_params.append(class_pattern)
                        company_params.append(class_pattern)
                    else:
                        # 使用 class_id 過濾
                        u_class_clause = " AND u.class_id = %s"
                        company_params.append(class_id_int)
            except (ValueError, TypeError):
                pass

        cursor.execute(f"""
            SELECT c.company_name, COUNT(sp.id) AS preference_count
            FROM internship_companies c
            LEFT JOIN student_preferences sp ON c.id = sp.company_id {sp_semester_clause}
            LEFT JOIN users u ON sp.student_id = u.id AND u.role='student'
            LEFT JOIN classes cls ON u.class_id = cls.id {u_class_clause}
            GROUP BY c.id, c.company_name
            HAVING COUNT(sp.id) > 0
            ORDER BY preference_count DESC
            LIMIT 5
        """, company_params)
        top_companies = cursor.fetchall()

        # 履歷繳交率 - 根據班級類型（忠/孝）計算
        if class_id and class_id != "all":
            try:
                class_id_int = int(class_id)
                # 先查詢該班級的名稱，判斷是「忠」還是「孝」
                cursor.execute("SELECT name, department FROM classes WHERE id = %s", (class_id_int,))
                class_info = cursor.fetchone()
                
                if class_info:
                    class_name = class_info.get('name', '')
                    department = class_info.get('department', '').replace('管科', '')
                    full_class_name = f"{department}{class_name}"
                    
                    # 判斷是「忠」還是「孝」
                    if '忠' in full_class_name or '忠' in class_name:
                        class_type = '忠'
                        class_pattern = '%忠%'
                    elif '孝' in full_class_name or '孝' in class_name:
                        class_type = '孝'
                        class_pattern = '%孝%'
                    else:
                        class_type = None
                        class_pattern = None
                    
                    print(f"[DEBUG] 統計圖表 - 班級 {class_id_int} ({full_class_name}) 類型: {class_type}")
                    
                    if class_type:
                        # 查詢所有「忠」或「孝」班的學生
                        cursor.execute("""
                            SELECT COUNT(*) AS total
                            FROM users u
                            LEFT JOIN classes c ON u.class_id = c.id
                            WHERE u.role = 'student' 
                              AND (c.name LIKE %s OR CONCAT(REPLACE(c.department, '管科', ''), c.name) LIKE %s)
                        """, (class_pattern, class_pattern))
                        total_students = cursor.fetchone()["total"] or 0
                        class_condition = " AND (c.name LIKE %s OR CONCAT(REPLACE(c.department, '管科', ''), c.name) LIKE %s)"
                        class_params = [class_pattern, class_pattern]
                    else:
                        # 如果無法判斷類型，只查詢該班級的學生
                        cursor.execute("""
                            SELECT COUNT(*) AS total
                            FROM users u
                            WHERE u.role='student' AND u.class_id = %s
                        """, (class_id_int,))
                        total_students = cursor.fetchone()["total"] or 0
                        class_condition = " AND u.class_id=%s"
                        class_params = [class_id_int]
                else:
                    total_students = 0
                    class_condition = ""
                    class_params = []
            except (ValueError, TypeError):
                total_students = 0
                class_condition = ""
                class_params = []
        else:
            # 查詢所有學生
            cursor.execute("SELECT COUNT(*) AS total FROM users u WHERE u.role='student'", ())
            total_students = cursor.fetchone()["total"] or 0
            class_condition = ""
            class_params = []

        # 構建履歷查詢 - 根據班級類型過濾
        if class_condition and class_params and len(class_params) == 2 and '%' in str(class_params[0]):
            # 使用班級類型過濾（忠/孝）
            resume_query = """
                SELECT COUNT(DISTINCT r.user_id) AS uploaded
                FROM resumes r
                JOIN users u ON r.user_id = u.id
                LEFT JOIN classes c ON u.class_id = c.id
                WHERE u.role='student' AND (c.name LIKE %s OR CONCAT(REPLACE(c.department, '管科', ''), c.name) LIKE %s)
            """
            cursor.execute(resume_query, class_params)
        else:
            # 使用 class_id 過濾或所有學生
            resume_params = list(class_params)
            resume_query = f"""
                SELECT COUNT(DISTINCT r.user_id) AS uploaded
                FROM resumes r
                JOIN users u ON r.user_id = u.id
                WHERE u.role='student'{class_condition}
            """
            cursor.execute(resume_query, resume_params)
        
        result = cursor.fetchone()
        uploaded = result["uploaded"] if result and result.get("uploaded") is not None else 0
        resume_stats = {"uploaded": uploaded, "not_uploaded": max(total_students - uploaded, 0)}
        
        print(f"[DEBUG] 履歷統計: 總學生={total_students}, 已上傳={uploaded}, 未上傳={resume_stats['not_uploaded']}")

        # 志願序填寫率 - 根據班級類型過濾
        if class_condition and class_params and len(class_params) == 2 and '%' in str(class_params[0]):
            # 使用班級類型過濾（忠/孝）
            pref_query = """
                SELECT COUNT(DISTINCT sp.student_id) AS filled
                FROM student_preferences sp
                JOIN users u ON sp.student_id = u.id
                LEFT JOIN classes c ON u.class_id = c.id
                WHERE u.role='student' AND (c.name LIKE %s OR CONCAT(REPLACE(c.department, '管科', ''), c.name) LIKE %s)
            """
            cursor.execute(pref_query, class_params)
        else:
            # 使用 class_id 過濾或所有學生
            pref_params = list(class_params)
            pref_query = f"""
                SELECT COUNT(DISTINCT sp.student_id) AS filled
                FROM student_preferences sp
                JOIN users u ON sp.student_id = u.id
                WHERE u.role='student'{class_condition}
            """
            cursor.execute(pref_query, pref_params)
        
        result = cursor.fetchone()
        filled = result["filled"] if result and result.get("filled") is not None else 0
        preference_stats = {"filled": filled, "not_filled": max(total_students - filled, 0)}
        
        print(f"[DEBUG] 志願序統計: 總學生={total_students}, 已填寫={filled}, 未填寫={preference_stats['not_filled']}")
        
        resume_stats = {"uploaded": uploaded, "not_uploaded": max(total_students - uploaded, 0)}
        preference_stats = {"filled": filled, "not_filled": max(total_students - filled, 0)}
        
        print(f"[DEBUG] 履歷統計: 總學生={total_students}, 已上傳={uploaded}, 未上傳={resume_stats['not_uploaded']}")
        print(f"[DEBUG] 志願序統計: 總學生={total_students}, 已填寫={filled}, 未填寫={preference_stats['not_filled']}")

        return jsonify({
            "success": True,
            "top_companies": top_companies,
            "resume_stats": resume_stats,
            "preference_stats": preference_stats
        })
    except Exception as e:
        print("❌ manage_companies_stats error:", e)
        return jsonify({"success": False, "message": str(e)}), 500
    finally:
        cursor.close()
        conn.close()

# =========================================================
# API: 匯出 Excel (公司志願統計)
# =========================================================
@ta_statistics_bp.route("/api/export_companies_stats", methods=["GET"])
def export_companies_stats():
    # 權限檢查：統計功能通常允許 ta, admin 訪問
    if 'user_id' not in session or session.get('role') not in ['ta', 'admin']:
        return jsonify({"success": False, "message": "未授權"}), 403
        
    class_id = request.args.get("class_id")
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    try:
        from openpyxl import Workbook

        params = []
        query = """
            SELECT c.company_name, COUNT(sp.id) AS preference_count 
            FROM internship_companies c 
            LEFT JOIN student_preferences sp ON c.id=sp.company_id 
        """
        
        # 處理班級篩選邏輯
        if class_id and class_id != "all":
            # 如果有指定班級，需要 JOIN users 表進行過濾
            query += "JOIN users u ON sp.student_id = u.id WHERE u.role='student' AND u.class_id=%s "
            params.append(class_id)
        else:
            # 針對所有學生進行統計
            query += "LEFT JOIN users u ON sp.student_id = u.id WHERE u.role='student' "

        query += "GROUP BY c.id, c.company_name ORDER BY preference_count DESC"
            
        cursor.execute(query, params)
        company_data = cursor.fetchall() or []

        # 取得履歷與志願統計（與前端圖表一致）
        student_filter = "WHERE u.role='student'"
        student_params = []
        if class_id and class_id != "all":
            student_filter += " AND u.class_id=%s"
            student_params.append(class_id)

        cursor.execute(f"SELECT COUNT(*) AS total FROM users u {student_filter}", student_params)
        total_students = cursor.fetchone()["total"] or 0

        cursor.execute(f"""
            SELECT COUNT(DISTINCT r.user_id) AS uploaded
            FROM resumes r
            JOIN users u ON r.user_id = u.id
            {student_filter}
        """, student_params)
        uploaded = cursor.fetchone()["uploaded"] or 0
        resume_stats = {
            "已上傳": uploaded,
            "未上傳": max(total_students - uploaded, 0)
        }

        cursor.execute(f"""
            SELECT COUNT(DISTINCT sp.student_id) AS filled
            FROM student_preferences sp
            JOIN users u ON sp.student_id = u.id
            {student_filter}
        """, student_params)
        filled = cursor.fetchone()["filled"] or 0
        preference_stats = {
            "已填寫": filled,
            "未填寫": max(total_students - filled, 0)
        }

        # 建立 Excel 檔，分成三個工作表
        wb = Workbook()

        # 工作表 1：公司志願統計
        ws1 = wb.active
        ws1.title = "公司志願統計"
        ws1.append(["排名", "公司名稱", "志願次數"])
        if company_data:
            for idx, row in enumerate(company_data, start=1):
                ws1.append([idx, row.get("company_name", "-"), row.get("preference_count", 0)])
        else:
            ws1.append(["-", "目前沒有志願資料", 0])

        # 工作表 2：履歷繳交率
        ws2 = wb.create_sheet("履歷繳交率")
        ws2.append(["項目", "人數", "比例"])
        resume_total = sum(resume_stats.values()) or 0
        for label, value in resume_stats.items():
            percent = f"{(value / resume_total * 100):.0f}%" if resume_total else "0%"
            ws2.append([label, value, percent])

        # 工作表 3：志願填寫率
        ws3 = wb.create_sheet("志願填寫率")
        ws3.append(["項目", "人數", "比例"])
        pref_total = sum(preference_stats.values()) or 0
        for label, value in preference_stats.items():
            percent = f"{(value / pref_total * 100):.0f}%" if pref_total else "0%"
            ws3.append([label, value, percent])

        # 工作表 4：班級實習進度統計
        ws4 = wb.create_sheet("班級實習進度統計")
        ws4.append(["項目", "人數"])
        ws4.append(["總學生數", total_students])
        ws4.append(["已上傳履歷", uploaded])
        ws4.append(["已填寫志願", filled])

        # 調整欄寬以便閱讀
        for ws in [ws1, ws2, ws3, ws4]:
            for column_cells in ws.columns:
                max_length = 0
                for cell in column_cells:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = max_length + 4
                ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # 生成檔案名稱
        filename = f"公司志願統計_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output, 
            download_name=filename, 
            as_attachment=True, 
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": f"匯出 Excel 失敗: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

# =========================================================
# API: 取得錄取統計
# =========================================================
@ta_statistics_bp.route("/api/admissions", methods=["GET"])
def get_admission_statistics():
    """科助端取得錄取統計"""
    if 'user_id' not in session or session.get('role') not in ['ta', 'admin']:
        return jsonify({"success": False, "message": "未授權"}), 403
    
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 獲取當前學期代碼
        current_semester_code = get_current_semester_code(cursor)
        
        # 錄取統計
        cursor.execute("""
            SELECT 
                COUNT(DISTINCT tsr.student_id) AS total_admitted_students,
                COUNT(DISTINCT tsr.company_id) AS total_companies_with_admissions,
                COUNT(DISTINCT tsr.teacher_id) AS total_teachers_involved
            FROM teacher_student_relations tsr
            WHERE tsr.semester = %s
        """, (current_semester_code,))
        admission_stats = cursor.fetchone()
        
        # 各公司錄取人數
        cursor.execute("""
            SELECT 
                ic.company_name,
                COUNT(DISTINCT tsr.student_id) AS admitted_count
            FROM internship_companies ic
            LEFT JOIN teacher_student_relations tsr ON ic.id = tsr.company_id AND tsr.semester = %s
            GROUP BY ic.id, ic.company_name
            HAVING admitted_count > 0
            ORDER BY admitted_count DESC
            LIMIT 10
        """, (current_semester_code,))
        company_admissions = cursor.fetchall()
        
        # 各指導老師負責學生數
        cursor.execute("""
            SELECT 
                u.name AS teacher_name,
                COUNT(DISTINCT tsr.student_id) AS student_count
            FROM teacher_student_relations tsr
            JOIN users u ON tsr.teacher_id = u.id
            WHERE tsr.semester = %s
            GROUP BY tsr.teacher_id, u.name
            ORDER BY student_count DESC
            LIMIT 10
        """, (current_semester_code,))
        teacher_stats = cursor.fetchall()
        
        return jsonify({
            "success": True,
            "current_semester": current_semester_code,
            "admission_stats": admission_stats,
            "company_admissions": company_admissions,
            "teacher_stats": teacher_stats
        })
    
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": f"查詢失敗: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

# =========================================================
# API: 取得時間序列統計（履歷/志願序提交趨勢）
# =========================================================
@ta_statistics_bp.route("/api/trends", methods=["GET"])
def get_trends():
    """科助端取得時間序列統計（履歷/志願序提交趨勢）"""
    if 'user_id' not in session or session.get('role') not in ['ta', 'admin']:
        return jsonify({"success": False, "message": "未授權"}), 403
    
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 獲取當前學期ID
        cursor.execute("SELECT id FROM semesters WHERE is_active = 1 LIMIT 1")
        current_semester = cursor.fetchone()
        semester_id = current_semester['id'] if current_semester else None
        
        # 履歷提交趨勢（按日期）
        cursor.execute("""
            SELECT 
                DATE(r.created_at) AS date,
                COUNT(*) AS count
            FROM resumes r
            WHERE r.semester_id = %s
            GROUP BY DATE(r.created_at)
            ORDER BY date ASC
        """, (semester_id,))
        resume_trends = cursor.fetchall()
        
        # 志願序提交趨勢（按日期）
        cursor.execute("""
            SELECT 
                DATE(sp.submitted_at) AS date,
                COUNT(*) AS count
            FROM student_preferences sp
            WHERE sp.semester_id = %s
            GROUP BY DATE(sp.submitted_at)
            ORDER BY date ASC
        """, (semester_id,))
        preference_trends = cursor.fetchall()
        
        # 格式化日期
        for trend in resume_trends + preference_trends:
            if isinstance(trend.get('date'), datetime):
                trend['date'] = trend['date'].strftime("%Y-%m-%d")
        
        return jsonify({
            "success": True,
            "resume_trends": resume_trends,
            "preference_trends": preference_trends
        })
    
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": f"查詢失敗: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

# =========================================================
# API: 匯出統計報表（Excel）
# =========================================================
@ta_statistics_bp.route("/api/export", methods=["GET"])
def export_statistics():
    """科助端匯出統計報表（Excel）"""
    if 'user_id' not in session or session.get('role') not in ['ta', 'admin']:
        return jsonify({"success": False, "message": "未授權"}), 403
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        from flask import send_file
        import io
        
        conn = get_db()
        cursor = conn.cursor(dictionary=True)
        
        # 獲取當前學期
        current_semester_code = get_current_semester_code(cursor)
        
        # 創建 Excel 工作簿
        wb = Workbook()
        
        # Sheet 1: 總覽統計
        ws1 = wb.active
        ws1.title = "總覽統計"
        
        # 標題
        ws1['A1'] = f"智慧實習系統統計報表 - {current_semester_code or '當前學期'}"
        ws1['A1'].font = Font(bold=True, size=14)
        ws1.merge_cells('A1:D1')
        
        # 獲取總覽數據（使用現有的 API 邏輯）
        cursor.execute("SELECT COUNT(*) AS total FROM users WHERE role = 'student'")
        total_students = cursor.fetchone()['total']
        
        cursor.execute("""
            SELECT 
                COUNT(DISTINCT r.user_id) AS students_with_resume,
                COUNT(DISTINCT CASE WHEN r.status = 'approved' THEN r.user_id END) AS students_approved
            FROM resumes r
            WHERE r.semester_id = (SELECT id FROM semesters WHERE is_active = 1 LIMIT 1)
        """)
        resume_stats = cursor.fetchone()
        
        cursor.execute("""
            SELECT 
                COUNT(DISTINCT sp.student_id) AS students_with_preferences,
                COUNT(DISTINCT CASE WHEN sp.status = 'approved' THEN sp.student_id END) AS students_approved
            FROM student_preferences sp
            WHERE sp.semester_id = (SELECT id FROM semesters WHERE is_active = 1 LIMIT 1)
        """)
        preference_stats = cursor.fetchone()
        
        # 寫入總覽數據
        row = 3
        ws1[f'A{row}'] = "項目"
        ws1[f'B{row}'] = "數量"
        ws1[f'C{row}'] = "完成率"
        row += 1
        
        ws1[f'A{row}'] = "總學生數"
        ws1[f'B{row}'] = total_students
        row += 1
        
        ws1[f'A{row}'] = "已上傳履歷人數"
        ws1[f'B{row}'] = resume_stats['students_with_resume'] or 0
        ws1[f'C{row}'] = f"{round((resume_stats['students_with_resume'] or 0) * 100.0 / total_students if total_students > 0 else 0, 2)}%"
        row += 1
        
        ws1[f'A{row}'] = "已填寫志願序人數"
        ws1[f'B{row}'] = preference_stats['students_with_preferences'] or 0
        ws1[f'C{row}'] = f"{round((preference_stats['students_with_preferences'] or 0) * 100.0 / total_students if total_students > 0 else 0, 2)}%"
        
        # Sheet 2: 各班級統計
        ws2 = wb.create_sheet("各班級統計")
        
        # 獲取各班級統計
        cursor.execute("""
            SELECT 
                c.name AS class_name,
                c.department,
                COUNT(DISTINCT u.id) AS total_students,
                COUNT(DISTINCT r.user_id) AS students_with_resume,
                COUNT(DISTINCT sp.student_id) AS students_with_preferences
            FROM classes c
            LEFT JOIN users u ON u.class_id = c.id AND u.role = 'student'
            LEFT JOIN resumes r ON r.user_id = u.id
            LEFT JOIN student_preferences sp ON sp.student_id = u.id
            GROUP BY c.id, c.name, c.department
            ORDER BY c.name ASC, c.department ASC
        """)
        classes_stats = cursor.fetchall()
        
        # 寫入表頭
        headers = ['班級名稱', '系所', '總學生數', '已上傳履歷', '已填寫志願序', '履歷完成率', '志願序完成率']
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
        
        # 寫入數據
        for idx, stat in enumerate(classes_stats, 2):
            ws2.cell(row=idx, column=1, value=stat['class_name'])
            ws2.cell(row=idx, column=2, value=stat['department'])
            ws2.cell(row=idx, column=3, value=stat['total_students'] or 0)
            ws2.cell(row=idx, column=4, value=stat['students_with_resume'] or 0)
            ws2.cell(row=idx, column=5, value=stat['students_with_preferences'] or 0)
            
            total = stat['total_students'] or 0
            resume_rate = round((stat['students_with_resume'] or 0) * 100.0 / total if total > 0 else 0, 2)
            pref_rate = round((stat['students_with_preferences'] or 0) * 100.0 / total if total > 0 else 0, 2)
            ws2.cell(row=idx, column=6, value=f"{resume_rate}%")
            ws2.cell(row=idx, column=7, value=f"{pref_rate}%")
        
        # 調整欄寬
        for col in range(1, len(headers) + 1):
            ws2.column_dimensions[get_column_letter(col)].width = 15
        
        # 保存到內存
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # 生成檔案名稱
        filename = f"實習系統統計報表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": f"匯出失敗: {str(e)}"}), 500
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'conn' in locals():
            conn.close()

# -------------------------
# 匯入核心科目 (Excel)
# -------------------------
@ta_statistics_bp.route('/api/import_standard_courses', methods=['POST'])
def import_standard_courses():
    if 'user_id' not in session or session.get('role') != 'ta':
        return jsonify({"success": False, "message": "未授權"}), 403

    if 'file' not in request.files:
        return jsonify({"success": False, "message": "缺少文件"}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({"success": False, "message": "未選擇文件"}), 400

    allowed_extensions = {'xlsx', 'xls'}
    if '.' not in file.filename or file.filename.rsplit('.', 1)[1].lower() not in allowed_extensions:
        return jsonify({"success": False, "message": "不支援的文件類型"}), 400
    
    file_stream = io.BytesIO(file.read())
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        workbook = load_workbook(file_stream)
        sheet = workbook.active
        
        headers = [cell.value for cell in sheet[1]]
        
        course_name_col = None
        credits_col = None
        
        for i, header in enumerate(headers):
            if header and ('課程名稱' in str(header) or '科目名稱' in str(header)):
                course_name_col = i + 1
            elif header and '學分' in str(header):
                credits_col = i + 1

        if not course_name_col or not credits_col:
            return jsonify({"success": False, "message": "Excel 檔案缺少必要的欄位（課程名稱/科目名稱、學分）"}), 400

        # 清空現有核心科目（避免重複或過時資料）
        cursor.execute("UPDATE standard_courses SET is_active = 0")

        imported_count = 0
        for row_index in range(2, sheet.max_row + 1):
            try:
                course_name = str(sheet.cell(row=row_index, column=course_name_col).value or '').strip()
                credits_value = str(sheet.cell(row=row_index, column=credits_col).value or '').strip()

                if not course_name or not credits_value:
                    continue

                # 嘗試將學分轉換為數字
                try:
                    credits = float(credits_value)
                except ValueError:
                    credits = 0.0 # 無效學分設為 0

                # 檢查是否已存在，如果存在則更新 is_active 和 credits
                cursor.execute("""
                    SELECT id FROM standard_courses WHERE course_name = %s LIMIT 1
                """, (course_name,))
                existing_course = cursor.fetchone()
                
                if existing_course:
                    cursor.execute("""
                        UPDATE standard_courses 
                        SET credits = %s, is_active = 1, updated_at = NOW() 
                        WHERE id = %s
                    """, (credits, existing_course['id']))
                else:
                    cursor.execute("""
                        INSERT INTO standard_courses 
                            (course_name, credits, is_active, uploaded_by, uploaded_at)
                        VALUES (%s, %s, 1, %s, NOW())
                    """, (course_name, credits, session['username']))
                
                imported_count += 1
                
            except Exception as row_e:
                print(f"⚠️ 處理 Excel 第 {row_index} 行錯誤: {row_e}")
                continue

        conn.commit()
        return jsonify({"success": True, "message": f"成功匯入 {imported_count} 筆核心科目資料"})
        
    except Exception as e:
        conn.rollback()
        print("❌ 匯入核心科目 Excel 錯誤:", e)
        traceback.print_exc()
        return jsonify({"success": False, "message": f"伺服器錯誤: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

# -------------------------
# 科助上傳標準課程Excel（預覽）
# -------------------------
@ta_statistics_bp.route('/api/ta/preview_standard_courses', methods=['POST'])
def preview_standard_courses():
    """科助預覽標準課程Excel文件"""
    if 'user_id' not in session or session.get('role') != 'ta':
        return jsonify({"success": False, "message": "未授權"}), 403
    
    if 'file' not in request.files:
        return jsonify({"success": False, "message": "未找到上傳文件"}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({"success": False, "message": "文件名稱不能為空"}), 400
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({"success": False, "message": "只支援Excel文件(.xlsx, .xls)"}), 400
    
    try:
        file_content = file.read()
        wb = load_workbook(io.BytesIO(file_content), data_only=False)
        ws = wb.active
        
        def get_cell_value(cell):
            """獲取單元格值，處理日期格式問題"""
            if cell is None or cell.value is None:
                return None
            value = cell.value
            if isinstance(value, datetime):
                month = value.month
                day = value.day
                return f"{month}/{day}"
            return value
        
        courses = []
        for row_idx in range(2, ws.max_row + 1):
            cell_name = ws.cell(row=row_idx, column=1)
            cell_credits = ws.cell(row=row_idx, column=2)
            
            course_name = get_cell_value(cell_name)
            credits_raw = cell_credits.value
            
            if not course_name or str(course_name).strip() == '':
                continue
            
            course_name = str(course_name).strip()
            
            # 處理學分數
            credits_str = ''
            if credits_raw is not None:
                if isinstance(credits_raw, datetime):
                    month = credits_raw.month
                    day = credits_raw.day
                    credits_str = f"{month}/{day}"
                elif isinstance(credits_raw, str):
                    credits_str = credits_raw.strip()
                    if ('2025-' in credits_str or '2024-' in credits_str or '2026-' in credits_str) and ('-' in credits_str):
                        try:
                            date_part = credits_str.split()[0] if ' ' in credits_str else credits_str
                            date_obj = datetime.strptime(date_part, '%Y-%m-%d')
                            month = date_obj.month
                            day = date_obj.day
                            credits_str = f"{month}/{day}"
                        except:
                            # 解析失敗，使用format_credits格式化
                            credits_str = format_credits(credits_str)
                    else:
                        # 不是日期格式，使用format_credits格式化
                        credits_str = format_credits(credits_str)
                else:
                    credits_str = format_credits(credits_raw)
            
            courses.append({
                'name': course_name,
                'credits': credits_str
            })
        
        return jsonify({
            "success": True,
            "courses": courses,
            "message": f"成功解析 {len(courses)} 門課程"
        })
    except Exception as e:
        print("❌ 預覽Excel錯誤:", e)
        traceback.print_exc()
        return jsonify({"success": False, "message": f"解析Excel失敗: {str(e)}"}), 500

# -------------------------
# 輔助函數：格式化學分數（整數顯示為整數，如2而不是2.0）
# -------------------------
def format_credits(credits_value):
    """格式化學分數，整數顯示為整數格式"""
    if credits_value is None:
        return ''
    
    # 如果是字符串，嘗試解析
    if isinstance(credits_value, str):
        credits_value = credits_value.strip()
        # 如果包含分數符號（如"2/2"），直接返回
        if '/' in credits_value:
            return credits_value
        # 嘗試轉換為數字
        try:
            num_value = float(credits_value)
            # 如果是整數，返回整數格式
            if num_value.is_integer():
                return str(int(num_value))
            return str(num_value)
        except (ValueError, TypeError):
            # 無法轉換為數字，返回原字符串
            return credits_value
    
    # 如果是數字類型
    if isinstance(credits_value, (int, float)):
        # 如果是整數，返回整數格式
        if isinstance(credits_value, float) and credits_value.is_integer():
            return str(int(credits_value))
        elif isinstance(credits_value, int):
            return str(credits_value)
        else:
            return str(credits_value)
    
    # 其他類型，轉換為字符串
    return str(credits_value)

# -------------------------
# 科助上傳標準課程Excel（寫入資料庫）
# -------------------------
@ta_statistics_bp.route('/api/ta/upload_standard_courses', methods=['POST'])
def upload_standard_courses():
    """科助上傳標準課程Excel並寫入standard_courses表"""
    if 'user_id' not in session or session.get('role') != 'ta':
        return jsonify({"success": False, "message": "未授權"}), 403
    
    if 'file' not in request.files:
        return jsonify({"success": False, "message": "未找到上傳文件"}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({"success": False, "message": "文件名稱不能為空"}), 400
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({"success": False, "message": "只支援Excel文件(.xlsx, .xls)"}), 400
    
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        file_content = file.read()
        wb = load_workbook(io.BytesIO(file_content), data_only=False)
        ws = wb.active
        
        def get_cell_value(cell):
            if cell is None or cell.value is None:
                return None
            value = cell.value
            if isinstance(value, datetime):
                month = value.month
                day = value.day
                return f"{month}/{day}"
            return value
        
        courses = []
        for row_idx in range(2, ws.max_row + 1):
            cell_name = ws.cell(row=row_idx, column=1)
            cell_credits = ws.cell(row=row_idx, column=2)
            
            course_name = get_cell_value(cell_name)
            credits_raw = cell_credits.value
            
            if not course_name or str(course_name).strip() == '':
                continue
            
            course_name = str(course_name).strip()
            
            # 處理學分數
            credits_str = ''
            if credits_raw is not None:
                if isinstance(credits_raw, datetime):
                    month = credits_raw.month
                    day = credits_raw.day
                    credits_str = f"{month}/{day}"
                elif isinstance(credits_raw, str):
                    credits_str = credits_raw.strip()
                    if ('2025-' in credits_str or '2024-' in credits_str or '2026-' in credits_str) and ('-' in credits_str):
                        try:
                            date_part = credits_str.split()[0] if ' ' in credits_str else credits_str
                            date_obj = datetime.strptime(date_part, '%Y-%m-%d')
                            month = date_obj.month
                            day = date_obj.day
                            credits_str = f"{month}/{day}"
                        except:
                            # 解析失敗，使用format_credits格式化
                            credits_str = format_credits(credits_str)
                    else:
                        # 不是日期格式，使用format_credits格式化
                        credits_str = format_credits(credits_str)
                else:
                    credits_str = format_credits(credits_raw)
            
            courses.append({
                'name': course_name,
                'credits': credits_str
            })
        
        if len(courses) == 0:
            return jsonify({"success": False, "message": "Excel文件中沒有找到課程資料"}), 400
        
        # 保存上傳的Excel文件
        # 獲取項目根目錄（backend的父目錄）
        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        upload_base_dir = os.path.join(project_root, 'uploads', 'standard_courses')
        os.makedirs(upload_base_dir, exist_ok=True)
        
        print(f"📁 項目根目錄: {project_root}")
        print(f"📁 上傳目錄: {upload_base_dir}")
        
        # 生成安全的文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # 先從原始文件名提取擴展名
        original_filename = file.filename if file.filename else 'upload.xlsx'
        original_ext = os.path.splitext(original_filename)[1].lower()
        if not original_ext or original_ext not in ['.xlsx', '.xls']:
            original_ext = '.xlsx'  # 默認使用 .xlsx
        
        # 處理文件名：移除擴展名，使用secure_filename處理，然後重新添加擴展名
        filename_without_ext = os.path.splitext(original_filename)[0]
        if not filename_without_ext or filename_without_ext.strip() == '':
            filename_without_ext = 'upload'
        
        safe_basename = secure_filename(filename_without_ext)
        if not safe_basename or safe_basename.strip() == '':
            safe_basename = 'upload'
        
        # 確保最終文件名包含擴展名
        safe_filename = safe_basename + original_ext
        filename = f"{timestamp}_{safe_filename}"
        
        # 完整的絕對路徑（用於保存文件）
        abs_file_path = os.path.join(upload_base_dir, filename)
        
        # 相對路徑（用於存儲到數據庫）
        db_file_path = os.path.join('uploads', 'standard_courses', filename).replace('\\', '/')
        
        print(f"📝 文件上傳信息:")
        print(f"  - 原始文件名: {original_filename}")
        print(f"  - 提取的擴展名: {original_ext}")
        print(f"  - 安全的文件名: {safe_filename}")
        print(f"  - 最終文件名: {filename}")
        print(f"  - 絕對保存路徑: {abs_file_path}")
        print(f"  - 數據庫路徑: {db_file_path}")
        
        # 保存文件
        file.seek(0)  # 重置文件指針
        os.makedirs(os.path.dirname(abs_file_path), exist_ok=True)
        with open(abs_file_path, 'wb') as f:
            f.write(file_content)
        
        print(f"✅ 文件已保存到: {abs_file_path}")
        # 驗證文件是否真的保存成功
        if os.path.exists(abs_file_path):
            file_size = os.path.getsize(abs_file_path)
            print(f"✅ 文件保存成功，大小: {file_size} bytes")
        else:
            print(f"❌ 警告：文件保存後無法找到！")
        
        # 檢查並創建 uploaded_course_templates 表（如果不存在）
        cursor.execute("SHOW TABLES LIKE 'uploaded_course_templates'")
        has_template_table = cursor.fetchone() is not None
        
        if not has_template_table:
            # 創建 uploaded_course_templates 表
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS uploaded_course_templates (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    file_path VARCHAR(500) NOT NULL,
                    uploaded_by INT NULL,
                    uploaded_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    INDEX idx_uploaded_at (uploaded_at),
                    INDEX idx_file_path (file_path)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
            """)
            print("✅ 已創建 uploaded_course_templates 表")
        
        # 先將舊資料標記為非活躍（不直接刪除，保留歷史）
        cursor.execute("UPDATE standard_courses SET is_active = 0")
        
        # 重新插入Excel中的課程（不包含文件路徑）
        insert_count = 0
        for idx, course in enumerate(courses, 1):
            try:
                cursor.execute("""
                    INSERT INTO standard_courses (course_name, credits, order_index, is_active, created_at)
                    VALUES (%s, %s, %s, 1, NOW())
                """, (course['name'], course['credits'], idx))
                insert_count += 1
            except Exception as e:
                print(f"⚠️ 插入課程失敗: {course['name']}, 錯誤: {e}")
                # 繼續插入其他課程，不中斷
                continue
        
        # 將文件路徑保存到 uploaded_course_templates 表
        template_id = None
        try:
            cursor.execute("""
                INSERT INTO uploaded_course_templates (file_path, uploaded_by, uploaded_at)
                VALUES (%s, %s, NOW())
            """, (db_file_path, session['user_id']))
            cursor.execute("SELECT LAST_INSERT_ID() as id")
            result = cursor.fetchone()
            if result:
                template_id = result['id']
            print(f"✅ 已保存文件路徑到 uploaded_course_templates 表，ID: {template_id}, 文件路徑: {db_file_path}, 課程數: {insert_count}")
        except Exception as e:
            print(f"⚠️ 保存文件路徑失敗: {e}")
            traceback.print_exc()
        
        print(f"✅ 已插入 {insert_count} 門課程到 standard_courses 表")
        
        # 確保事務提交
        try:
            conn.commit()
            print(f"✅ 成功更新 standard_courses 表，插入 {insert_count} 門課程")
            print(f"✅ 文件已保存到: {abs_file_path}")
            
            # 驗證更新是否成功
            cursor.execute("SELECT COUNT(*) as count FROM standard_courses WHERE is_active = 1")
            verify_result = cursor.fetchone()
            active_count = verify_result['count'] if verify_result else 0
            print(f"✅ 驗證：standard_courses 表中 is_active=1 的記錄數: {active_count}")
            
            # 驗證文件路徑是否正確保存到 uploaded_course_templates 表
            if template_id:
                cursor.execute("SELECT * FROM uploaded_course_templates WHERE id = %s", (template_id,))
                verify_template = cursor.fetchone()
                if verify_template:
                    print(f"✅ 驗證：文件路徑已保存到 uploaded_course_templates 表，ID: {template_id}, 文件路徑: {verify_template.get('file_path', 'N/A')}")
                else:
                    print(f"⚠️ 警告：uploaded_course_templates 表記錄ID {template_id} 未找到")
            
            return jsonify({
                "success": True,
                "count": insert_count,
                "message": f"成功上傳 {insert_count} 門課程",
                "file_path": db_file_path
            })
        except Exception as commit_error:
            conn.rollback()
            print(f"❌ 提交事務失敗: {commit_error}")
            traceback.print_exc()
            raise commit_error
    except Exception as e:
        conn.rollback()
        print("❌ 上傳標準課程錯誤:", e)
        traceback.print_exc()
        return jsonify({"success": False, "message": f"上傳失敗: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

# -------------------------
# 科助取得標準課程上傳歷史
# -------------------------
@ta_statistics_bp.route('/api/ta/get_standard_courses_history', methods=['GET'])
def get_standard_courses_history():
    """取得標準課程上傳歷史記錄"""
    if 'user_id' not in session or session.get('role') != 'ta':
        return jsonify({"success": False, "message": "未授權"}), 403
    
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 檢查 uploaded_course_templates 表是否存在
        cursor.execute("SHOW TABLES LIKE 'uploaded_course_templates'")
        has_template_table = cursor.fetchone() is not None
        
        if has_template_table:
            # 從 uploaded_course_templates 表獲取歷史記錄
            # 並從 standard_courses 表計算每次上傳的課程數量（根據上傳日期匹配）
            cursor.execute("""
                SELECT 
                    t.id,
                    t.file_path,
                    t.uploaded_by,
                    t.uploaded_at,
                    COALESCE(COUNT(DISTINCT s.id), 0) as course_count
                FROM uploaded_course_templates t
                LEFT JOIN standard_courses s ON DATE(s.created_at) = DATE(t.uploaded_at)
                    AND s.is_active = 1
                GROUP BY t.id, t.file_path, t.uploaded_by, t.uploaded_at
                ORDER BY t.uploaded_at DESC
                LIMIT 20
            """)
            history = cursor.fetchall()
            # 調試：打印查詢結果
            print(f"🔍 從 uploaded_course_templates 表查詢到 {len(history)} 筆歷史記錄")
            for record in history:
                print(f"  - ID: {record.get('id')}, 文件路徑: {record.get('file_path', 'NULL')}, 課程數: {record.get('course_count', 0)}")
        else:
            # 如果表不存在，返回空列表
            print("⚠️ uploaded_course_templates 表不存在")
            history = []
        
        return jsonify({
            "success": True,
            "history": history
        })
    except Exception as e:
        print("❌ 取得上傳歷史錯誤:", e)
        traceback.print_exc()
        return jsonify({"success": False, "message": f"取得歷史記錄失敗: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

# -------------------------
# 科助下載標準課程Excel文件
# -------------------------
@ta_statistics_bp.route('/api/ta/download_standard_course_file/<int:history_id>', methods=['GET'])
def download_standard_course_file(history_id):
    """下載上傳的Excel文件（從uploaded_course_templates表）"""
    if 'user_id' not in session or session.get('role') != 'ta':
        return jsonify({"success": False, "message": "未授權"}), 403
    
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 從 uploaded_course_templates 表獲取文件路徑
        cursor.execute("""
            SELECT file_path 
            FROM uploaded_course_templates 
            WHERE id = %s
        """, (history_id,))
        record = cursor.fetchone()
        
        if not record or not record.get('file_path'):
            return jsonify({"success": False, "message": "找不到文件"}), 404
        
        file_path = record.get('file_path')
        
        # 處理相對路徑 - 從項目根目錄開始
        if not os.path.isabs(file_path):
            # 獲取項目根目錄（backend的父目錄）
            project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            abs_file_path = os.path.join(project_root, file_path)
        else:
            abs_file_path = file_path
        
        # 標準化路徑分隔符
        # abs_file_path = os.path.normpath(abs_file_path)
        
        print(f"🔍 嘗試下載文件: {abs_file_path}")
        
        # 檢查文件是否存在，如果不存在，嘗試多種方式查找
        if not os.path.exists(abs_file_path):
            print(f"⚠️ 文件不存在，嘗試查找相似文件...")
            
            # 方法1：嘗試添加 .xlsx 擴展名
            abs_file_path_xlsx = abs_file_path + '.xlsx'
            abs_file_path_xls = abs_file_path + '.xls'
            
            if os.path.exists(abs_file_path_xlsx):
                print(f"✅ 找到文件（添加.xlsx後）: {abs_file_path_xlsx}")
                abs_file_path = abs_file_path_xlsx
            elif os.path.exists(abs_file_path_xls):
                print(f"✅ 找到文件（添加.xls後）: {abs_file_path_xls}")
                abs_file_path = abs_file_path_xls
            else:
                # 方法2：在目錄中查找以該文件名開頭的文件
                file_dir = os.path.dirname(abs_file_path)
                file_basename = os.path.basename(abs_file_path)
                
                if os.path.isdir(file_dir):
                    print(f"🔍 在目錄中搜索: {file_dir}, 文件名前綴: {file_basename}")
                    try:
                        files_in_dir = os.listdir(file_dir)
                        print(f"📁 目錄中的文件: {files_in_dir}")
                        
                        # 查找以該文件名開頭的Excel文件
                        matching_files = [f for f in files_in_dir 
                                        if f.startswith(file_basename) 
                                        and (f.lower().endswith('.xlsx') or f.lower().endswith('.xls'))]
                        
                        if matching_files:
                            # 找到匹配的文件，使用第一個
                            found_file = matching_files[0]
                            abs_file_path = os.path.join(file_dir, found_file)
                            print(f"✅ 找到匹配文件: {abs_file_path}")
                        else:
                            # 方法3：查找所有Excel文件，看是否有相似的時間戳
                            excel_files = [f for f in files_in_dir 
                                         if f.lower().endswith('.xlsx') or f.lower().endswith('.xls')]
                            print(f"📊 目錄中的Excel文件: {excel_files}")
                            
                            # 嘗試提取時間戳部分進行匹配
                            if file_basename and '_' in file_basename:
                                timestamp_part = file_basename.split('_')[0] + '_' + file_basename.split('_')[1] if len(file_basename.split('_')) >= 2 else file_basename
                                matching_by_timestamp = [f for f in excel_files if timestamp_part in f]
                                
                                if matching_by_timestamp:
                                    abs_file_path = os.path.join(file_dir, matching_by_timestamp[0])
                                    print(f"✅ 根據時間戳找到文件: {abs_file_path}")
                                else:
                                    print(f"❌ 無法找到匹配的文件")
                                    print(f"❌ 嘗試過: {abs_file_path}")
                                    print(f"❌ 嘗試過: {abs_file_path_xlsx}")
                                    print(f"❌ 嘗試過: {abs_file_path_xls}")
                                    return jsonify({"success": False, "message": f"文件不存在: {os.path.basename(file_path)}"}), 404
                            else:
                                print(f"❌ 無法找到匹配的文件")
                                print(f"❌ 嘗試過: {abs_file_path}")
                                print(f"❌ 嘗試過: {abs_file_path_xlsx}")
                                print(f"❌ 嘗試過: {abs_file_path_xls}")
                                return jsonify({"success": False, "message": f"文件不存在: {os.path.basename(file_path)}"}), 404
                    except Exception as e:
                        print(f"❌ 搜索文件時發生錯誤: {e}")
                        return jsonify({"success": False, "message": f"搜索文件失敗: {str(e)}"}), 500
                else:
                    print(f"❌ 目錄不存在: {file_dir}")
                    return jsonify({"success": False, "message": f"目錄不存在: {file_dir}"}), 404
        
        # 獲取原始文件名（從路徑中提取）
        original_filename = os.path.basename(file_path)
        # 如果文件名包含時間戳，嘗試提取原始文件名
        if '_' in original_filename and original_filename[0].isdigit():
            # 檢查是否是時間戳格式 (YYYYMMDD_HHMMSS_)
            parts = original_filename.split('_', 2)
            if len(parts) >= 3 and len(parts[0]) == 8 and len(parts[1]) == 6:
                original_filename = '_'.join(parts[2:])  # 保留後面的部分
        
        # 確保文件名有正確的擴展名（從實際文件路徑獲取）
        actual_filename = os.path.basename(abs_file_path)
        if actual_filename.lower().endswith('.xlsx'):
            ext = '.xlsx'
        elif actual_filename.lower().endswith('.xls'):
            ext = '.xls'
        else:
            ext = '.xlsx'  # 默認使用 .xlsx
        
        # 如果原始文件名沒有擴展名，添加擴展名
        if not original_filename.lower().endswith(('.xlsx', '.xls')):
            original_filename = original_filename + ext
        elif not original_filename.lower().endswith(ext):
            # 如果擴展名不匹配，使用實際文件的擴展名
            original_filename = os.path.splitext(original_filename)[0] + ext
        
        # 設置正確的MIME類型
        if original_filename.lower().endswith('.xlsx'):
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        elif original_filename.lower().endswith('.xls'):
            mimetype = 'application/vnd.ms-excel'
        else:
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        print(f"✅ 下載文件: {abs_file_path}, 文件名: {original_filename}, MIME: {mimetype}")
        return send_file(abs_file_path, as_attachment=True, download_name=original_filename, mimetype=mimetype)
    except Exception as e:
        print(f"❌ 下載文件錯誤: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "message": f"下載失敗: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()

# -------------------------
# 科助刪除標準課程上傳記錄
# -------------------------
@ta_statistics_bp.route('/api/ta/delete_standard_course_history/<int:history_id>', methods=['DELETE'])
def delete_standard_course_history(history_id):
    """刪除上傳歷史記錄及對應的文件（從uploaded_course_templates表）"""
    if 'user_id' not in session or session.get('role') != 'ta':
        return jsonify({"success": False, "message": "未授權"}), 403
    
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 從 uploaded_course_templates 表獲取文件路徑
        cursor.execute("""
            SELECT file_path
            FROM uploaded_course_templates 
            WHERE id = %s
        """, (history_id,))
        record = cursor.fetchone()
        
        if not record:
            return jsonify({"success": False, "message": "找不到記錄"}), 404
        
        file_path = record.get('file_path')
        
        # 刪除文件（如果存在）
        if file_path:
            abs_file_path = os.path.abspath(file_path)
            if os.path.exists(abs_file_path):
                try:
                    os.remove(abs_file_path)
                    print(f"✅ 已刪除文件: {abs_file_path}")
                except Exception as e:
                    print(f"⚠️ 刪除文件失敗: {e}")
        
        # 刪除 uploaded_course_templates 表中的記錄
        cursor.execute("DELETE FROM uploaded_course_templates WHERE id = %s", (history_id,))
        conn.commit()
        
        print(f"✅ 已刪除 uploaded_course_templates 表記錄，ID: {history_id}")
        
        return jsonify({
            "success": True,
            "message": "已成功刪除記錄"
        })
    except Exception as e:
        conn.rollback()
        print(f"❌ 刪除記錄錯誤: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "message": f"刪除失敗: {str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()    

# --------------------------------
# 實習廠商管理頁面
# --------------------------------
@ta_statistics_bp.route('/manage_companies')
def manage_companies():
    # 權限檢查：允許 ta, admin 訪問
    if 'user_id' not in session or session.get('role') not in ['ta', 'admin']:
        from flask import redirect, url_for
        return redirect(url_for('auth_bp.login_page'))
    return render_template('user_shared/manage_companies.html')

# --------------------------------
# 學生管理頁面
# --------------------------------
@ta_statistics_bp.route('/manage_students')
def manage_students():
    # 權限檢查：允許 ta, admin 訪問
    if 'user_id' not in session or session.get('role') not in ['ta', 'admin']:
        from flask import redirect, url_for
        return redirect(url_for('auth_bp.login_page'))
    return render_template('user_shared/manage_students.html')               

# --------------------------------
# API: 獲取所有面試排程（從 vendor_preference_history）
# --------------------------------
@ta_statistics_bp.route('/api/interview_schedules', methods=['GET'])
def get_interview_schedules():
    """獲取所有廠商的面試排程（用於 TA 查看所有面試排程）"""
    if 'user_id' not in session or session.get('role') not in ['ta', 'admin']:
        return jsonify({"success": False, "message": "未授權"}), 403
    
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 檢查 vendor_preference_history 表是否存在
        cursor.execute("""
            SELECT COUNT(*) as count
            FROM information_schema.tables
            WHERE table_schema = DATABASE()
            AND table_name = 'vendor_preference_history'
        """)
        table_exists = cursor.fetchone().get('count', 0) > 0
        
        if not table_exists:
            return jsonify({
                "success": True,
                "schedules": []
            })
        
        # 查詢所有廠商的面試排程（從 vendor_preference_history 表中）
        # 只查詢 interview_status = 'in interview' 的記錄（廠商新增面試排程時會記錄到此狀態）
        cursor.execute("""
            SELECT DISTINCT
                vph.id,
                vph.reviewer_id,
                vph.student_id,
                vph.preference_id,
                vph.comment,
                vph.created_at,
                u.name AS vendor_name,
                ic.company_name,
                ic.id AS company_id,
                sp.student_id AS pref_student_id,
                su.name AS student_name
            FROM vendor_preference_history vph
            JOIN users u ON vph.reviewer_id = u.id
            LEFT JOIN student_preferences sp ON vph.preference_id = sp.id
            LEFT JOIN internship_companies ic ON sp.company_id = ic.id
            LEFT JOIN users su ON COALESCE(vph.student_id, sp.student_id) = su.id
            WHERE vph.interview_status = 'in interview'
            AND vph.comment LIKE '%面試日期：%'
            ORDER BY vph.created_at DESC
        """)
        
        all_schedules = cursor.fetchall()
        
        # 解析面試資訊
        import re
        parsed_schedules = []
        
        for schedule in all_schedules:
            comment = schedule.get('comment', '')
            reviewer_id = schedule.get('reviewer_id')
            vendor_name = schedule.get('vendor_name', '未知廠商')
            company_id = schedule.get('company_id')
            student_id = schedule.get('student_id') or schedule.get('pref_student_id')
            student_name = schedule.get('student_name', '未知學生')
            
            # 從 vendor 的 name 欄位提取公司名稱
            # vendor name 格式通常是 "公司名稱 聯絡人姓名"，例如 "人人人 周建羽"
            # 取第一個詞作為公司名稱
            company_name = '未知公司'
            if vendor_name:
                # 如果包含空格，取第一個詞（公司名稱）
                if ' ' in vendor_name:
                    company_name = vendor_name.split()[0]
                else:
                    # 如果沒有空格，整個作為公司名稱
                    company_name = vendor_name
            
            # 提取日期
            date_match = re.search(r'面試日期：(\d{4}-\d{2}-\d{2})', comment)
            if not date_match:
                continue
            
            interview_date = date_match.group(1)
            
            # 提取時間
            time_match = re.search(r'時間：(\d{2}:\d{2})\s*-\s*(\d{2}:\d{2})', comment)
            time_start = None
            time_end = None
            if time_match:
                time_start = time_match.group(1)
                time_end = time_match.group(2)
            else:
                # 兼容只有開始時間的情況
                time_match = re.search(r'時間：(\d{2}:\d{2})', comment)
                if time_match:
                    time_start = time_match.group(1)
            
            # 提取地點
            location_match = re.search(r'地點：([^，\n]+)', comment)
            location = location_match.group(1).strip() if location_match else ''
            
            parsed_schedules.append({
                'id': schedule.get('id'),
                'date': interview_date,
                'time_start': time_start,
                'time_end': time_end,
                'location': location,
                'vendor_id': reviewer_id,
                'vendor_name': vendor_name,
                'company_id': company_id,
                'company_name': company_name,
                'student_id': student_id,
                'student_name': student_name,
                'created_at': schedule.get('created_at').strftime('%Y-%m-%d %H:%M:%S') if schedule.get('created_at') else None
            })
        
        return jsonify({
            "success": True,
            "schedules": parsed_schedules
        })
        
    except Exception as exc:
        traceback.print_exc()
        return jsonify({"success": False, "message": f"獲取面試排程失敗：{exc}"}), 500
    finally:
        cursor.close()
        conn.close()

# --------------------------------
# API: 獲取志願序和履歷截止時間
# --------------------------------
@ta_statistics_bp.route('/api/deadlines', methods=['GET'])
def get_deadlines():
    """獲取志願序和履歷的截止時間（用於面試排程頁面顯示）"""
    if 'user_id' not in session or session.get('role') not in ['ta', 'admin']:
        return jsonify({"success": False, "message": "未授權"}), 403
    
    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # 查詢志願序截止時間（最新的公告）
        cursor.execute("""
            SELECT end_time 
            FROM announcement 
            WHERE title LIKE '[作業]%填寫志願序截止時間' AND is_published = 1
            ORDER BY created_at DESC 
            LIMIT 1
        """)
        preference_result = cursor.fetchone()
        
        # 查詢履歷截止時間（最新的公告）
        cursor.execute("""
            SELECT end_time 
            FROM announcement 
            WHERE title LIKE '[作業]%上傳履歷截止時間' AND is_published = 1
            ORDER BY created_at DESC 
            LIMIT 1
        """)
        resume_result = cursor.fetchone()
        
        # 格式化日期和時間
        preference_deadline_date = None
        preference_deadline_time = None
        resume_deadline_date = None
        resume_deadline_time = None
        
        if preference_result and preference_result.get('end_time'):
            deadline = preference_result['end_time']
            if isinstance(deadline, datetime):
                deadline_dt = deadline
            else:
                # 如果是字符串，解析
                try:
                    deadline_dt = datetime.strptime(str(deadline), '%Y-%m-%d %H:%M:%S')
                except:
                    deadline_dt = datetime.strptime(str(deadline), '%Y-%m-%d %H:%M')
            preference_deadline_date = deadline_dt.strftime('%Y-%m-%d')
            preference_deadline_time = deadline_dt.strftime('%H:%M')
        
        if resume_result and resume_result.get('end_time'):
            deadline = resume_result['end_time']
            if isinstance(deadline, datetime):
                deadline_dt = deadline
            else:
                # 如果是字符串，解析
                try:
                    deadline_dt = datetime.strptime(str(deadline), '%Y-%m-%d %H:%M:%S')
                except:
                    deadline_dt = datetime.strptime(str(deadline), '%Y-%m-%d %H:%M')
            resume_deadline_date = deadline_dt.strftime('%Y-%m-%d')
            resume_deadline_time = deadline_dt.strftime('%H:%M')
        
        return jsonify({
            "success": True,
            "preference_deadline_date": preference_deadline_date,
            "preference_deadline_time": preference_deadline_time,
            "resume_deadline_date": resume_deadline_date,
            "resume_deadline_time": resume_deadline_time
        })
    
    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": f"獲取截止時間失敗：{str(e)}"}), 500
    finally:
        cursor.close()
        conn.close()


